import os
import logging
from models import LancamentoOrganiza
from datetime import datetime, date, time, timedelta, timezone
from typing import Optional
from pathlib import Path
from io import BytesIO, StringIO
import shutil
import csv
import uuid
import hashlib
import re
import math
import json
import zipfile
import unicodedata
from xml.sax.saxutils import escape as xml_escape
from difflib import SequenceMatcher
from urllib.parse import quote, urlparse, parse_qsl, urlencode, urlunparse
from urllib.request import Request as UrlRequest, urlopen
from urllib.error import URLError, HTTPError
import time as time_module
import threading

logger = logging.getLogger("conect")
geo_logger = logging.getLogger("conect.geocodificacao")

from fastapi import FastAPI, Depends, Form, Request, HTTPException, File, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, Response, JSONResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy.orm import Session, joinedload, selectinload, make_transient_to_detached
from sqlalchemy import func, text, inspect, or_, case

from config import APP_NOME, APP_VERSION, SECRET_KEY, ADMIN_NOME, ADMIN_SENHA
from database import Base, engine, get_db, SessionLocal
from performance_monitor import PerformanceMiddleware, install_sql_monitor, perf_stage, recent_records, monitor_status, clear_records, performance_summary
from models import Agenda, CampoEmpresa, CampoGlobal, Cliente, EnderecoCliente, Contrato, Empresa, EquipamentoCliente, Pagamento, Equipe, UsuarioEquipe, \
    ProdutoServico, ReservaItem, Solicitacao, UsuarioEmpresa, ContaFinanceira, LancamentoBanco, \
    LancamentoManualFinanceiro, VinculoRepasseBanco, HumiatMovimento, VeiculoLogistico, ConfiguracaoRotaInteligente, RotaInteligente, RotaInteligenteParada, VeiculoPerfilCarga
from seed import inicializar_dados
from utils import limpar_identificador, somar_horas, somar_minutos, hora_meia_em_meia_valida, texto_para_float, \
    cpf_valido, cnpj_valido, aplicar_variaveis_mensagem

from fastapi.templating import Jinja2Templates

class ControleAcessoMiddleware:
    """Bloqueia a entrada nos módulos sem esconder cards, alertas ou pendências."""
    def __init__(self, app):
        self.app = app

    async def __call__(self, scope, receive, send):
        if scope.get("type") == "http":
            path = scope.get("path", "")
            session = scope.get("session") or {}
            if session.get("empresa_id") and not session.get("acesso_total"):
                area = self.area_da_rota(path)
                acessos = session.get("acessos") or {}
                if area and not acessos.get(area, False):
                    response = RedirectResponse(f"/painel/acesso-negado?area={area}", status_code=303)
                    await response(scope, receive, send)
                    return
        await self.app(scope, receive, send)

    @staticmethod
    def area_da_rota(path: str):
        # A permissão protege o módulo de destino. Pendências e dados exibidos no painel continuam visíveis.
        if path == "/painel/agenda" or path.startswith("/painel/agenda/"):
            return "agenda"
        if path == "/painel/reservas" or path.startswith("/painel/reservas/"):
            return "operacao"
        if path == "/painel/inteligencia-logistica" or path.startswith("/painel/inteligencia-logistica/"):
            return "operacao"
        if path == "/painel/clientes" or path.startswith("/painel/cliente/"):
            return "buscar_cliente"
        if path == "/painel/financeiro" or path.startswith("/painel/financeiro/"):
            return "financeiro"
        if path == "/painel/relatorios" or path.startswith("/painel/relatorios/"):
            return "relatorios"
        prefixos_cadastro = (
            "/painel/configuracoes", "/painel/produtos", "/painel/produto/",
            "/painel/contratos", "/painel/contrato/", "/painel/disponibilidade"
        )
        if any(path == p or path.startswith(p) for p in prefixos_cadastro):
            return "cadastros"
        return None


app = FastAPI(title=APP_NOME, version=APP_VERSION)
install_sql_monitor(engine)
app.add_middleware(PerformanceMiddleware)
app.add_middleware(ControleAcessoMiddleware)
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY)
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")
templates.env.globals["APP_VERSION"] = APP_VERSION
Path("static/uploads/logos").mkdir(parents=True, exist_ok=True)

FUSO_EMPRESA = timezone(timedelta(hours=-3))


def agora_utc() -> datetime:
    """Salva horários em UTC para não depender do fuso do servidor."""
    return datetime.now(timezone.utc).replace(tzinfo=None)


def redirect_preservando_filtros(request: Request, fallback: str = "/painel/financeiro",
                                 extras: dict | None = None) -> RedirectResponse:
    url = request.headers.get("referer") or fallback
    if extras:
        partes = urlparse(url)
        qs = dict(parse_qsl(partes.query, keep_blank_values=True))
        qs.update({k: str(v) for k, v in extras.items()})
        url = urlunparse((partes.scheme, partes.netloc, partes.path, partes.params, urlencode(qs), partes.fragment))
    return RedirectResponse(url, status_code=303)


def datahora_local(valor):
    """Mostra horários no fuso do Brasil/RJ."""
    if not valor:
        return "-"
    try:
        return valor.replace(tzinfo=timezone.utc).astimezone(FUSO_EMPRESA).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return "-"


templates.env.filters["datahora_local"] = datahora_local


def valor_falta(item) -> float:
    return max(float(getattr(item, "valor", 0) or 0) - float(getattr(item, "valor_pago", 0) or 0), 0)


def resumo_financeiro(itens):
    total = sum(float(getattr(i, "valor", 0) or 0) for i in itens)
    recebido = sum(float(getattr(i, "valor_pago", 0) or 0) for i in itens)
    falta = sum(valor_falta(i) for i in itens)
    return {"qtd": len(itens), "total": total, "recebido": recebido, "falta": falta}


def pagamento_sem_conciliar(item) -> bool:
    return any(not getattr(p, "conciliado_em", None) for p in getattr(item, "pagamentos", []) or [])


def somente_lancamentos_financeiros(itens):
    # Financeiro recebe os contratos que já possuem pelo menos uma linha de pagamento.
    # Cada pagamento fica listado dentro do card para conciliação com banco/cartão/dinheiro.
    return [i for i in itens if getattr(i, "pagamentos", None)]


def pagamentos_pendentes_conciliacao(itens):
    return [p for i in itens for p in (getattr(i, "pagamentos", []) or []) if not getattr(p, "conciliado_em", None)]


def recalcular_pagamento_solicitacao(db: Session, item: Solicitacao):
    # Fonte da verdade do financeiro: tabela de pagamentos.
    # O campo Solicitacao.valor_pago é apenas um resumo/cache usado nos cards.
    # Antes havia casos em que o card mostrava falta receber mesmo com todos
    # os pagamentos lançados/conciliados, porque esse resumo ficou desatualizado.
    db.flush()
    total_pago = sum(
        (p.valor or 0) for p in db.query(Pagamento).filter_by(empresa_id=item.empresa_id, solicitacao_id=item.id).all())
    item.valor_pago = total_pago
    item.sinal_recebido = total_pago > 0
    if total_pago <= 0:
        item.pagamento_confirmado_em = None
    elif not item.pagamento_confirmado_em:
        item.pagamento_confirmado_em = agora_utc()
    return total_pago


def sincronizar_pagamentos_solicitacoes(db: Session, solicitacoes):
    """Recalcula o resumo financeiro exibido nas telas operacionais/detalhe."""
    alterou = False
    vistos = set()
    for item in solicitacoes or []:
        if not item or item.id in vistos:
            continue
        vistos.add(item.id)
        total_pago = sum((p.valor or 0) for p in
                         db.query(Pagamento).filter_by(empresa_id=item.empresa_id, solicitacao_id=item.id).all())
        if round(float(item.valor_pago or 0), 2) != round(float(total_pago or 0), 2):
            item.valor_pago = total_pago
            item.sinal_recebido = total_pago > 0
            if total_pago <= 0:
                item.pagamento_confirmado_em = None
            elif not item.pagamento_confirmado_em:
                item.pagamento_confirmado_em = agora_utc()
            alterou = True
    if alterou:
        db.commit()
    return alterou


def existe_pagamento_conciliado(item: Solicitacao) -> bool:
    return any(getattr(p, "conciliado_em", None) for p in (getattr(item, "pagamentos", None) or []))


def classe_alerta_contrato(status: str) -> str:
    if status in {"pre_reserva"}:
        return "card-rascunho"
    if status in {"aguardando_aceite", "contrato_enviado"}:
        return "card-nao-aceito"
    return ""


templates.env.globals["classe_alerta_contrato"] = classe_alerta_contrato


def validar_total_pagamentos(item: Solicitacao, total_pago: float):
    if item.valor and total_pago > float(item.valor or 0) + 0.009:
        raise HTTPException(400, "A soma dos pagamentos não pode ser maior que o total do contrato.")


STATUS_CONTRATO_APROVADO = {"aceito", "aguardando_pagamento", "reserva_confirmada"}


def status_reserva_confirmada(status: str) -> bool:
    return (status or "") in STATUS_CONTRATO_APROVADO


def contrato_aprovado_para_operacao(item: Solicitacao | None) -> bool:
    """Somente contratos efetivamente aprovados podem entrar na Operação/Inteligência."""
    return bool(item and status_reserva_confirmada(item.status) and reserva_tem_itens(item))


def status_em_contrato(status: str) -> bool:
    return status in {"pre_reserva", "aguardando_aceite", "contrato_enviado"}


def reserva_tem_itens(item) -> bool:
    return bool(getattr(item, "itens", None))


def reserva_pode_aprovar(item) -> bool:
    """Contrato só pode ser aprovado quando já existe pelo menos um item."""
    return reserva_tem_itens(item)


def corrigir_reservas_aprovadas_sem_itens(db: Session):
    """
    Corrige reservas que ficaram em status aprovado/confirmado sem itens.
    Esse estado não é permitido: a próxima ação correta é adicionar itens.
    """
    alterou = False
    reservas = db.query(Solicitacao).filter(
        Solicitacao.status.in_(["reserva_confirmada", "aguardando_pagamento"])).all()
    for item in reservas:
        qtd_itens = db.query(ReservaItem).filter_by(empresa_id=item.empresa_id, solicitacao_id=item.id).count()
        if qtd_itens == 0:
            item.status = "pre_reserva"
            item.aprovado_em = None
            item.sinal_recebido = False
            item.valor_pago = 0
            item.pagamento_confirmado_em = None
            db.query(Pagamento).filter_by(empresa_id=item.empresa_id, solicitacao_id=item.id).delete()
            alterou = True
    if alterou:
        db.commit()


def corrigir_valores_teste(db: Session):
    """
    Corrige valores inflados em bases de teste geradas por máscara monetária antiga.
    Ex.: 310.000,00 salvo como 310000.00 volta para 310.00.
    Regra conservadora para este projeto: valores operacionais acima de 50 mil,
    quando múltiplos de 1000, são reduzidos em 1000.
    """

    def ajustar(valor):
        try:
            numero = float(valor or 0)
        except Exception:
            return valor
        if numero >= 50000 and numero % 1000 == 0:
            return numero / 1000
        return numero

    alterou = False
    for item in db.query(Solicitacao).all():
        novo_valor = ajustar(item.valor)
        novo_sinal = ajustar(item.sinal)
        novo_pago = ajustar(item.valor_pago)
        if (novo_valor, novo_sinal, novo_pago) != (item.valor, item.sinal, item.valor_pago):
            item.valor, item.sinal, item.valor_pago = novo_valor, novo_sinal, novo_pago
            alterou = True

    for linha in db.query(ReservaItem).all():
        novo_unitario = ajustar(linha.valor_unitario)
        novo_total = ajustar(linha.valor_total)
        if (novo_unitario, novo_total) != (linha.valor_unitario, linha.valor_total):
            linha.valor_unitario, linha.valor_total = novo_unitario, novo_total
            alterou = True

    for produto in db.query(ProdutoServico).all():
        novo_base = ajustar(produto.valor_base)
        if novo_base != produto.valor_base:
            produto.valor_base = novo_base
            alterou = True

    for pagamento in db.query(Pagamento).all():
        novo_pagamento = ajustar(pagamento.valor)
        if novo_pagamento != pagamento.valor:
            pagamento.valor = novo_pagamento
            alterou = True

    if alterou:
        db.commit()


def recalcular_valores_reservas(db: Session):
    """Mantém o valor da reserva igual à soma dos itens e corrige bases antigas."""
    alterou = False
    for item in db.query(Solicitacao).all():
        total_itens = sum((linha.valor_total or 0) for linha in item.itens)
        if total_itens > 0 and round(float(item.valor or 0), 2) != round(float(total_itens), 2):
            item.valor = total_itens
            # Valor pago não pode ficar maior que o total da reserva.
            if item.valor_pago and item.valor_pago > item.valor:
                item.valor_pago = item.valor
            alterou = True
    if alterou:
        db.commit()


def limpar_agenda_operacional(db: Session):
    """
    Remove duplicidade operacional.
    Regra atual: a reserva nasce com ENTREGA.
    A RETIRADA nasce ao concluir a entrega, exceto quando o cliente exigiu retirada obrigatória.
    """
    alterou = False
    reservas = db.query(Solicitacao).all()
    for reserva in reservas:
        eventos = (
            db.query(Agenda)
            .filter_by(empresa_id=reserva.empresa_id, solicitacao_id=reserva.id)
            .order_by(Agenda.id)
            .all()
        )
        entregas = [e for e in eventos if (e.tipo_evento or "entrega") == "entrega"]
        retiradas = [e for e in eventos if e.tipo_evento == "retirada"]

        if not entregas and eventos:
            eventos[0].tipo_evento = "entrega"
            entregas = [eventos[0]]
            alterou = True

        if entregas:
            principal = entregas[0]
            principal.tipo_evento = "entrega"
            # Depois de roteirizado, data e hora pertencem exclusivamente à rota.
            # Alterações no contrato ou rotinas de sincronização não podem mudar
            # a posição operacional; somente o botão Salvar da roteirização pode.
            if not principal.roteirizado:
                principal.data = reserva.data_evento
                principal.hora_inicio = reserva.hora_inicio
                principal.hora_fim = reserva.hora_fim
            principal.titulo = f"{nome_item_reserva(reserva)} - {reserva.cliente.nome if reserva.cliente else 'Cliente'}"
            principal.bairro = reserva.bairro
            for duplicado in entregas[1:]:
                db.delete(duplicado)
                alterou = True

        if retirada_obrigatoria_ativa(reserva):
            criar_ou_atualizar_retirada_obrigatoria(db, reserva)
            alterou = True
            if len(retiradas) > 1:
                for duplicada in retiradas[1:]:
                    db.delete(duplicada)
                    alterou = True
            continue

        # Retiradas comuns só devem existir depois que a entrega foi concluída.
        entrega_concluida = bool(entregas and entregas[0].status_operacional == "concluido")
        if not entrega_concluida:
            for retirada in retiradas:
                db.delete(retirada)
                alterou = True
        elif len(retiradas) > 1:
            for duplicada in retiradas[1:]:
                db.delete(duplicada)
                alterou = True

    if alterou:
        db.commit()

def janela_uma_hora(hora) -> str:
    if not hora:
        return "-"
    fim = somar_horas(hora, 1)
    if not fim:
        return hora.strftime("%H:%M")
    return f"{hora.strftime('%H:%M')} às {fim.strftime('%H:%M')}"


def ajustar_hora_texto(hora_texto, horas: int) -> str:
    """Recebe HH:MM e devolve HH:MM somando/subtraindo horas."""
    try:
        if not hora_texto or hora_texto == "--":
            return "-"
        base = datetime.strptime(str(hora_texto), "%H:%M")
        return (base + timedelta(hours=int(horas))).strftime("%H:%M")
    except Exception:
        return "-"


def periodo_semana_atual():
    hoje = date.today()
    inicio = hoje - timedelta(days=hoje.weekday())
    fim = inicio + timedelta(days=6)
    return inicio, fim


def moeda_br(valor) -> str:
    try:
        numero = float(valor or 0)
    except Exception:
        numero = 0.0
    texto = f"{numero:,.2f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


templates.env.filters["moeda_br"] = moeda_br
templates.env.globals["status_reserva_confirmada"] = status_reserva_confirmada
templates.env.globals["status_em_contrato"] = status_em_contrato
templates.env.globals["janela_uma_hora"] = janela_uma_hora
templates.env.globals["ajustar_hora_texto"] = ajustar_hora_texto


def _limpar_tel_whatsapp(valor: str) -> str:
    tel = "".join(ch for ch in str(valor or "") if ch.isdigit())
    if not tel:
        return ""
    if tel.startswith("55"):
        return tel
    return "55" + tel


def _link_absoluto(request: Request, nome_rota: str, **params) -> str:
    return str(request.url_for(nome_rota, **params))


def _solicitacao_tem_endereco_congelado(item: Solicitacao) -> bool:
    """Diferencia contratos novos/migrados dos registros legados.

    Campos vazios ("") contam como snapshot: isso é intencional para impedir que
    um contrato antigo volte a herdar um número posteriormente alterado no cliente.
    """
    return any(
        getattr(item, campo, None) is not None
        for campo in ("local_numero", "local_complemento", "local_cidade", "local_estado", "local_cep")
    )


def dados_endereco_solicitacao(item: Solicitacao) -> dict:
    """Retorna o endereço do EVENTO. O cadastro do cliente é apenas fallback legado.

    Em contratos novos/migrados, nunca consulta endereço/número atual do cliente.
    Isso impede que uma nova locação do mesmo cliente altere a rota de outra reserva.
    """
    cliente = item.cliente
    if _solicitacao_tem_endereco_congelado(item):
        return {
            "endereco": (item.local or "").strip(),
            "numero": (item.local_numero or "").strip(),
            "complemento": (item.local_complemento or "").strip(),
            "bairro": (item.bairro or "").strip(),
            "cidade": (item.local_cidade or "").strip(),
            "estado": (item.local_estado or "").strip(),
            "cep": (item.local_cep or "").strip(),
        }

    # Fallback de emergência para registro legado ainda não migrado. Se a reserva
    # já possui um ``local``, nunca combina esse logradouro com o número atual do
    # cliente: é preferível navegar sem número a enviar um número de outro evento.
    local_legado = (item.local or "").strip()
    if local_legado:
        mesma_rua_atual = bool(
            cliente and cliente.endereco
            and _normalizar_chave_endereco(cliente.endereco) == _normalizar_chave_endereco(local_legado)
        )
        return {
            "endereco": local_legado,
            "numero": "",
            "complemento": "",
            "bairro": (item.bairro or (cliente.bairro if cliente and mesma_rua_atual else "") or "").strip(),
            "cidade": ((cliente.cidade if cliente and mesma_rua_atual else "") or "").strip(),
            "estado": ((cliente.estado if cliente and mesma_rua_atual else "") or "").strip(),
            "cep": "",
        }

    # Registros muito antigos sem endereço na própria solicitação só têm o cadastro
    # atual como última referência disponível. A migração de startup congela isso.
    return {
        "endereco": ((cliente.endereco if cliente else "") or "").strip(),
        "numero": ((cliente.numero if cliente else "") or "").strip(),
        "complemento": ((cliente.complemento if cliente else "") or "").strip(),
        "bairro": ((cliente.bairro if cliente else "") or "").strip(),
        "cidade": ((cliente.cidade if cliente else "") or "").strip(),
        "estado": ((cliente.estado if cliente else "") or "").strip(),
        "cep": ((cliente.cep if cliente else "") or "").strip(),
    }


def endereco_rota_solicitacao(item: Solicitacao) -> str:
    """Monta exatamente o destino textual exibido na Operação e enviado à navegação.

    Para permitir comparação direta em produção, Waze e Google Maps recebem a mesma
    string mostrada em ``Destino``: logradouro + número + bairro + CEP. Complemento
    (apto, bloco, sala etc.) fica sempre fora da navegação.
    """
    dados = dados_endereco_solicitacao(item)
    partes = [
        dados["endereco"],
        dados["numero"],
        dados["bairro"],
        f"CEP {dados['cep']}" if dados["cep"] else "",
    ]
    return ", ".join(str(valor).strip() for valor in partes if valor and str(valor).strip())


def coordenadas_rota_solicitacao(item: Solicitacao) -> str:
    """Retorna coordenadas verificadas da própria solicitação para navegação.

    O Waze com ``q=`` navega para o primeiro resultado da pesquisa e pode escolher
    um estabelecimento/endereço vizinho. Quando a geocodificação do contrato está
    marcada como localizada, usamos ``ll=latitude,longitude`` para eliminar essa
    segunda pesquisa dentro do Waze.
    """
    if (item.status_geocodificacao or "").strip().lower() != "localizado":
        return ""
    if not _coordenadas_validas_brasil(item.latitude, item.longitude):
        return ""
    return f"{float(item.latitude):.7f},{float(item.longitude):.7f}"


def endereco_referencia_solicitacao(item: Solicitacao) -> str:
    """Retorna nome/ponto de referência sem misturar referência com o endereço da rota."""
    if (item.local_nome or "").strip():
        return item.local_nome.strip()
    # Compatibilidade com registros muito antigos em que ``local`` podia ser referência.
    if not _solicitacao_tem_endereco_congelado(item) and item.cliente:
        local = (item.local or "").strip()
        endereco_cliente = (item.cliente.endereco or "").strip()
        if local and endereco_cliente and local.casefold() != endereco_cliente.casefold():
            return local
    return ""


templates.env.globals["dados_endereco_solicitacao"] = dados_endereco_solicitacao
templates.env.globals["endereco_rota_solicitacao"] = endereco_rota_solicitacao
templates.env.globals["coordenadas_rota_solicitacao"] = coordenadas_rota_solicitacao
templates.env.globals["endereco_referencia_solicitacao"] = endereco_referencia_solicitacao


def linhas_endereco_reserva(item: Solicitacao) -> list[str]:
    """Monta o endereço completo a partir do snapshot do contrato/reserva."""
    dados = dados_endereco_solicitacao(item)
    local_nome = (item.local_nome or "").strip()
    endereco = dados["endereco"]
    numero = dados["numero"]
    bairro = dados["bairro"]
    cidade = dados["cidade"]
    estado = dados["estado"]
    cep = dados["cep"]
    referencia = (item.observacoes or (item.cliente.observacoes if item.cliente else "") or "").strip()

    linhas = []
    if local_nome:
        linhas.append(f"*Local:* {local_nome}")

    endereco_partes = []
    if endereco:
        endereco_partes.append(endereco)
    if numero:
        endereco_partes.append(f"nº {numero}")
    if complemento:
        endereco_partes.append(complemento)
    if endereco_partes:
        linhas.append(f"*Endereço:* {', '.join(endereco_partes)}")

    if bairro:
        linhas.append(f"*Bairro:* {bairro}")

    cidade_uf = " / ".join([p for p in [cidade, estado] if p])
    if cidade_uf:
        linhas.append(f"*Cidade:* {cidade_uf}")

    if cep:
        linhas.append(f"*CEP:* {cep}")

    if referencia:
        linhas.append(f"*Observação:* {referencia}")

    if not linhas:
        linhas.append("*Endereço:* -")

    return linhas


def linhas_informacoes_preenchidas_contrato(item: Solicitacao, formato: str = "texto") -> list[str]:
    """Lista todas as informações preenchidas do contrato/reserva para PDF e WhatsApp.
    formato='whatsapp' usa negrito com *campo*.
    """
    cliente = item.cliente

    def fmt_data(v):
        return v.strftime("%d/%m/%Y") if v else ""

    def fmt_hora(v):
        return v.strftime("%H:%M") if v else ""

    def add(linhas, rotulo, valor):
        if valor is None:
            return
        valor_txt = str(valor).strip()
        if not valor_txt:
            return
        if formato == "whatsapp":
            linhas.append(f"*{rotulo}:* {valor_txt}")
        else:
            linhas.append(f"{rotulo}: {valor_txt}")

    linhas = []

    add(linhas, "Cliente", getattr(cliente, "nome", ""))
    add(linhas, "Telefone", (getattr(cliente, "telefone", "") or getattr(cliente, "identificador", "")))
    add(linhas, "CPF", getattr(cliente, "cpf", ""))
    add(linhas, "CNPJ", getattr(cliente, "cnpj", ""))
    add(linhas, "E-mail", getattr(cliente, "email", ""))
    add(linhas, "Nascimento", fmt_data(getattr(cliente, "data_nascimento", None)))

    add(linhas, "Data do evento", fmt_data(item.data_evento))
    add(linhas, "Hora de início", fmt_hora(item.hora_inicio))
    add(linhas, "Hora de fim", fmt_hora(item.hora_fim))

    add(linhas, "Nome do local", item.local_nome)
    add(linhas, "Endereço do evento", item.local)
    add(linhas, "Bairro do evento", item.bairro)
    add(linhas, "Acesso ao local", item.acesso_local)
    add(linhas, "Responsável no local", item.local_responsavel_nome)
    add(linhas, "Telefone do responsável", item.local_responsavel_telefone)

    add(linhas, "Endereço do evento", item.local or getattr(cliente, "endereco", ""))
    add(linhas, "Número", getattr(cliente, "numero", ""))
    add(linhas, "Complemento", getattr(cliente, "complemento", ""))
    add(linhas, "Bairro do cliente", getattr(cliente, "bairro", ""))
    cidade_uf = " - ".join([p for p in [getattr(cliente, "cidade", ""), getattr(cliente, "estado", "")] if p])
    add(linhas, "Cidade/UF", cidade_uf)
    add(linhas, "CEP", getattr(cliente, "cep", ""))
    add(linhas, "Observações do cliente", getattr(cliente, "observacoes", ""))
    add(linhas, "Observações da reserva", item.observacoes)

    add(linhas, "Valor total", f"R$ {moeda_br(item.valor or 0)}")
    add(linhas, "Valor recebido", f"R$ {moeda_br(item.valor_pago or 0)}")
    add(linhas, "Sinal previsto", f"R$ {moeda_br(item.sinal or 0)}")
    add(linhas, "Falta", f"R$ {moeda_br(max((item.valor or 0) - (item.valor_pago or 0), 0))}")

    return linhas


def _resumo_reserva_whatsapp(empresa: Empresa, item: Solicitacao, itens_reserva) -> list[str]:
    """Monta o resumo principal da reserva para mensagens de WhatsApp."""
    total = float(item.valor or 0)
    pago = float(item.valor_pago or 0)
    falta = max(total - pago, 0)
    data_txt = item.data_evento.strftime("%d/%m/%Y") if item.data_evento else "-"
    hora_txt = item.hora_inicio.strftime("%H:%M") if item.hora_inicio else "-"

    equipamentos = []
    if itens_reserva:
        for ri in itens_reserva:
            prefixo = f"{ri.quantidade or 1}x " if (ri.quantidade or 1) > 1 else ""
            equipamentos.append(f"• {prefixo}{ri.nome}")
    elif item.produto:
        equipamentos.append(f"• {item.produto.nome}")
    else:
        equipamentos.append("• Itens da reserva")

    endereco_linhas = linhas_endereco_reserva(item)
    endereco_texto = "\n".join(
        l.replace("*Endereço:* ", "").replace("*Local:* ", "").replace("*Bairro:* ", "Bairro: ")
        for l in endereco_linhas
    )

    return [
        f"*{empresa.nome or 'Karaokê RJ'}*",
        "",
        f"*Cliente:* {item.cliente.nome if item.cliente else '-'}",
        "",
        "*📅 Entrega*",
        f"{data_txt} às {hora_txt}",
        "",
        "*📍 Local*",
        endereco_texto or "-",
        "",
        "*🎤 Equipamentos*",
        *equipamentos,
        "",
        "*💰 Financeiro*",
        f"*Total:* R$ {moeda_br(total)}",
        f"*Pago:* R$ {moeda_br(pago)}",
        f"*Saldo:* R$ {moeda_br(falta)}",
    ]


def montar_mensagem_whatsapp_aceite(request: Request, empresa: Empresa, item: Solicitacao, db: Session) -> str:
    """Mensagem para o cliente aceitar a reserva. Usa o texto do cadastro da empresa e complementos do sistema."""
    link_aceite = _link_absoluto(request, "contrato_cliente", slug=empresa.slug, solicitacao_id=item.id)
    cliente_nome = item.cliente.nome if item.cliente else "cliente"

    texto_base = aplicar_variaveis_mensagem(
        mensagens_empresa(empresa).get("aceite", ""),
        link=link_aceite,
        empresa=empresa.nome,
        cliente=cliente_nome,
        valor_sinal=moeda_br(item.sinal or 0),
        pix=empresa.pix_copia_cola or "",
    ).strip()

    linhas = [texto_base] if texto_base else []

    if getattr(empresa, "exige_sinal", False):
        linhas.extend([
            "",
            "Para concluir a confirmação, realize o PIX do sinal para a chave abaixo e envie o comprovante.",
            "",
            f"*PIX:* {empresa.pix_copia_cola or '-'}",
            "",
            "Assim que o aceite do pré-contrato e a confirmação do pagamento do sinal forem concluídos, sua reserva será efetivada.",
        ])
    else:
        linhas.extend([
            "",
            "Assim que o aceite do pré-contrato for concluído, sua reserva será efetivada.",
        ])

    linhas.extend([
        "",
        "Em seguida, você receberá:",
        "• O resumo da reserva;",
        "• O contrato em PDF;",
        "• As cláusulas do contrato para sua consulta.",
    ])

    return "\n".join(linhas).strip()


def montar_mensagem_whatsapp_contrato(request: Request, empresa: Empresa, item: Solicitacao, db: Session) -> str:
    """Mensagem enviada somente depois do aceite, com o link do contrato final."""
    itens_reserva = db.query(ReservaItem).filter_by(empresa_id=empresa.id, solicitacao_id=item.id).all()
    link_contrato = _link_absoluto(request, "contrato_cliente_pdf", slug=empresa.slug, solicitacao_id=item.id)
    link_clausulas = _link_absoluto(request, "contrato_cliente_clausulas", slug=empresa.slug, solicitacao_id=item.id)

    linhas = _resumo_reserva_whatsapp(empresa, item, itens_reserva)
    linhas.extend([
        "",
        "*📄 Contrato final:*",
        link_contrato,
        "",
        "*📄 Cláusulas do contrato:*",
        link_clausulas,
        "",
    ])

    mensagem_final = mensagens_empresa(empresa).get("confirmacao", "").strip()
    if mensagem_final:
        linhas.append(mensagem_final)

    return "\n".join(linhas).strip()



MENSAGEM_OPERACAO_PREPARACAO_APROVADA = (
    "Olá, {{cliente}}.\n\n"
    "Estamos nos preparando para sair e, em breve, iniciaremos o deslocamento até você.\n\n"
    "Nossa previsão de chegada é entre {{hora_previsao_inicio}} e {{hora_previsao_fim}}.\n\n"
    "Caso esse horário não seja adequado ou aconteça algum imprevisto, por favor nos avise.\n\n"
    "Se houver qualquer alteração em nossa programação, entraremos em contato imediatamente.\n\n"
    "Equipe {{empresa}}"
)

MENSAGEM_OPERACAO_A_CAMINHO_APROVADA = (
    "Olá, {{cliente}}.\n\n"
    "Nossa equipe já está a caminho.\n\n"
    "Em breve estaremos no local informado.\n\n"
    "Caso precise falar conosco, basta responder esta mensagem.\n\n"
    "Equipe {{empresa}}"
)

def garantir_colunas_novas():
    """Migração simples para bases locais/teste já existentes."""
    insp = inspect(engine)
    try:
        tabelas = insp.get_table_names()
    except Exception:
        return
    if "empresas" not in tabelas:
        return

    def colunas(tabela):
        return {c["name"] for c in insp.get_columns(tabela)}

    comandos = []

    if "usuarios_empresa" not in tabelas:
        comandos.append("""
        CREATE TABLE usuarios_empresa (
            id INTEGER PRIMARY KEY,
            empresa_id INTEGER NOT NULL,
            nome VARCHAR(120) NOT NULL,
            usuario VARCHAR(80) NOT NULL,
            senha VARCHAR(120) NOT NULL,
            ativo BOOLEAN DEFAULT true,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(empresa_id) REFERENCES empresas (id)
        )
        """)
        comandos.append("CREATE INDEX IF NOT EXISTS ix_usuarios_empresa_usuario ON usuarios_empresa (usuario)")
    cols_emp = colunas("empresas")
    if "pix_copia_cola" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN pix_copia_cola TEXT")
    if "whatsapp_retorno" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN whatsapp_retorno VARCHAR(30)")
    if "exige_sinal" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN exige_sinal BOOLEAN DEFAULT false")
    if "suporte_inicio" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN suporte_inicio VARCHAR(5)")
    if "suporte_fim" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN suporte_fim VARCHAR(5)")
    if "mostrar_suporte_contrato" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN mostrar_suporte_contrato BOOLEAN DEFAULT false")
    if "logo_url" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN logo_url VARCHAR(300)")
    if "tema" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN tema VARCHAR(30) DEFAULT 'azul'")
    if "mensagem_reserva" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN mensagem_reserva TEXT")
    if "mensagem_preparacao" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN mensagem_preparacao TEXT")
    if "mensagem_a_caminho" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN mensagem_a_caminho TEXT")
    if "mensagem_localizacao" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN mensagem_localizacao TEXT")
    if "logo_idb_url" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN logo_idb_url VARCHAR(300)")
    if "mensagem_hora_fim" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN mensagem_hora_fim TEXT")
    if "mostrar_mensagem_hora_fim" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN mostrar_mensagem_hora_fim BOOLEAN DEFAULT true")
    if "mensagem_aceite" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN mensagem_aceite TEXT")
    if "mensagem_pagamento" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN mensagem_pagamento TEXT")
    if "mensagem_confirmacao" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN mensagem_confirmacao TEXT")
    if "humiat_saldo" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN humiat_saldo INTEGER DEFAULT 0 NOT NULL")
    if "humiat_gratis_mes" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN humiat_gratis_mes INTEGER DEFAULT 4 NOT NULL")
    if "humiat_custo_contrato" not in cols_emp:
        comandos.append("ALTER TABLE empresas ADD COLUMN humiat_custo_contrato INTEGER DEFAULT 1 NOT NULL")

    if "clientes" in tabelas:
        cols_cli = colunas("clientes")
        if "data_nascimento" not in cols_cli:
            comandos.append("ALTER TABLE clientes ADD COLUMN data_nascimento DATE")

    if "produtos_servicos" in tabelas:
        cols_prod = colunas("produtos_servicos")
        if "duracao_minutos" not in cols_prod:
            comandos.append("ALTER TABLE produtos_servicos ADD COLUMN duracao_minutos INTEGER DEFAULT 240")
        if "prazo_retirada_dias" not in cols_prod:
            comandos.append("ALTER TABLE produtos_servicos ADD COLUMN prazo_retirada_dias INTEGER DEFAULT 1")
        if "carga_pontos" not in cols_prod:
            comandos.append("ALTER TABLE produtos_servicos ADD COLUMN carga_pontos INTEGER DEFAULT 1 NOT NULL")
        if "volume_logistico" not in cols_prod:
            comandos.append("ALTER TABLE produtos_servicos ADD COLUMN volume_logistico INTEGER DEFAULT 1 NOT NULL")
        if "permite_interno" not in cols_prod:
            comandos.append("ALTER TABLE produtos_servicos ADD COLUMN permite_interno BOOLEAN DEFAULT true NOT NULL")
        if "permite_mala" not in cols_prod:
            comandos.append("ALTER TABLE produtos_servicos ADD COLUMN permite_mala BOOLEAN DEFAULT true NOT NULL")
        if "permite_teto" not in cols_prod:
            comandos.append("ALTER TABLE produtos_servicos ADD COLUMN permite_teto BOOLEAN DEFAULT false NOT NULL")

    if "veiculos_logisticos" in tabelas:
        cols_vei = colunas("veiculos_logisticos")
        if "capacidade_interno" not in cols_vei:
            comandos.append("ALTER TABLE veiculos_logisticos ADD COLUMN capacidade_interno INTEGER DEFAULT 4 NOT NULL")
        if "capacidade_mala" not in cols_vei:
            comandos.append("ALTER TABLE veiculos_logisticos ADD COLUMN capacidade_mala INTEGER DEFAULT 1 NOT NULL")
        if "capacidade_teto" not in cols_vei:
            comandos.append("ALTER TABLE veiculos_logisticos ADD COLUMN capacidade_teto INTEGER DEFAULT 3 NOT NULL")

    if "veiculos_perfis_carga" not in tabelas:
        comandos.append("""CREATE TABLE veiculos_perfis_carga (
            id INTEGER PRIMARY KEY,
            veiculo_id INTEGER NOT NULL REFERENCES veiculos_logisticos(id) ON DELETE CASCADE,
            produto_id INTEGER NOT NULL REFERENCES produtos_servicos(id) ON DELETE CASCADE,
            volumes INTEGER DEFAULT 1 NOT NULL,
            permite_interno BOOLEAN DEFAULT true NOT NULL,
            permite_mala BOOLEAN DEFAULT false NOT NULL,
            permite_teto BOOLEAN DEFAULT false NOT NULL,
            ativo BOOLEAN DEFAULT true NOT NULL,
            CONSTRAINT uq_veiculo_produto_carga UNIQUE (veiculo_id, produto_id)
        )""")

    if "solicitacoes" in tabelas:
        cols_sol = colunas("solicitacoes")
        if "valor_pago" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN valor_pago FLOAT DEFAULT 0")
        if "sinal_recebido" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN sinal_recebido BOOLEAN DEFAULT false")
        if "pagamento_confirmado_em" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN pagamento_confirmado_em TIMESTAMP")
        if "aprovado_em" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN aprovado_em TIMESTAMP")
        if "contrato_enviado_em" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN contrato_enviado_em TIMESTAMP")
        if "responsavel_contrato" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN responsavel_contrato VARCHAR(120)")
        if "responsavel_operacao" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN responsavel_operacao VARCHAR(120)")
        if "cancelado_em" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN cancelado_em TIMESTAMP")
        if "retirada_obrigatoria" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN retirada_obrigatoria BOOLEAN DEFAULT false")
        if "retirada_data" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN retirada_data DATE")
        if "retirada_hora" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN retirada_hora TIME")
        if "local_nome" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN local_nome VARCHAR(160)")
        if "local_numero" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN local_numero VARCHAR(30)")
        if "local_complemento" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN local_complemento VARCHAR(120)")
        if "local_cidade" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN local_cidade VARCHAR(120)")
        if "local_estado" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN local_estado VARCHAR(40)")
        if "local_cep" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN local_cep VARCHAR(20)")
        if "local_responsavel_nome" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN local_responsavel_nome VARCHAR(160)")
        if "local_responsavel_telefone" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN local_responsavel_telefone VARCHAR(40)")
        if "retirada_responsavel_nome" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN retirada_responsavel_nome VARCHAR(160)")
        if "retirada_responsavel_telefone" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN retirada_responsavel_telefone VARCHAR(40)")

        if "acesso_local" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN acesso_local VARCHAR(40)")

        if "empresa_transferida_id" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN empresa_transferida_id INTEGER")
        if "valor_repasse" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN valor_repasse FLOAT DEFAULT 0")
        if "transferencia_origem_id" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN transferencia_origem_id INTEGER")
        if "transferencia_copia_id" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN transferencia_copia_id INTEGER")
        if "transferida_em" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN transferida_em TIMESTAMP")
        if "repasse_pago_em" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN repasse_pago_em TIMESTAMP")
        if "repasse_pago_por" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN repasse_pago_por VARCHAR(120)")
        if "humiat_processado" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN humiat_processado BOOLEAN DEFAULT false NOT NULL")
        if "humiat_competencia" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN humiat_competencia VARCHAR(7)")
        if "humiat_custo" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN humiat_custo INTEGER DEFAULT 0 NOT NULL")
        if "humiat_status" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN humiat_status VARCHAR(30)")
        if "latitude" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN latitude FLOAT")
        if "longitude" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN longitude FLOAT")
        if "status_geocodificacao" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN status_geocodificacao VARCHAR(20) DEFAULT 'pendente'")
        if "data_geocodificacao" not in cols_sol:
            comandos.append("ALTER TABLE solicitacoes ADD COLUMN data_geocodificacao TIMESTAMP")

    if "humiat_movimentos" not in tabelas:
        comandos.append("""
        CREATE TABLE humiat_movimentos (
            id INTEGER PRIMARY KEY,
            empresa_id INTEGER NOT NULL,
            solicitacao_id INTEGER,
            tipo VARCHAR(30) NOT NULL,
            quantidade INTEGER NOT NULL,
            saldo_anterior INTEGER DEFAULT 0 NOT NULL,
            saldo_posterior INTEGER DEFAULT 0 NOT NULL,
            motivo VARCHAR(200),
            observacao TEXT,
            usuario VARCHAR(120),
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP NOT NULL,
            FOREIGN KEY(empresa_id) REFERENCES empresas (id),
            FOREIGN KEY(solicitacao_id) REFERENCES solicitacoes (id)
        )
        """)
        comandos.append("CREATE INDEX IF NOT EXISTS ix_humiat_movimentos_empresa_id ON humiat_movimentos (empresa_id)")
        comandos.append("CREATE INDEX IF NOT EXISTS ix_humiat_movimentos_solicitacao_id ON humiat_movimentos (solicitacao_id)")

    if "pagamentos" in tabelas:
        cols_pag = colunas("pagamentos")
        if "usuario_registro" not in cols_pag:
            comandos.append("ALTER TABLE pagamentos ADD COLUMN usuario_registro VARCHAR(120)")
        if "conciliado_por" not in cols_pag:
            comandos.append("ALTER TABLE pagamentos ADD COLUMN conciliado_por VARCHAR(120)")
        if "conciliado_em" not in cols_pag:
            comandos.append("ALTER TABLE pagamentos ADD COLUMN conciliado_em TIMESTAMP")

    if "equipes" not in tabelas:
        comandos.append("""
        CREATE TABLE equipes (
            id INTEGER PRIMARY KEY,
            empresa_id INTEGER NOT NULL,
            nome VARCHAR(80) NOT NULL,
            ativa BOOLEAN DEFAULT true,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(empresa_id) REFERENCES empresas (id),
            UNIQUE(empresa_id, nome)
        )
        """)
    if "usuarios_equipes" not in tabelas:
        comandos.append("""
        CREATE TABLE usuarios_equipes (
            id INTEGER PRIMARY KEY,
            usuario_id INTEGER NOT NULL,
            equipe_id INTEGER NOT NULL,
            FOREIGN KEY(usuario_id) REFERENCES usuarios_empresa (id),
            FOREIGN KEY(equipe_id) REFERENCES equipes (id),
            UNIQUE(usuario_id, equipe_id)
        )
        """)

    if "usuarios_empresa" in tabelas:
        cols_usu = colunas("usuarios_empresa")
        novas_permissoes = {
            "acesso_agenda": "BOOLEAN DEFAULT false",
            "acesso_operacao": "BOOLEAN DEFAULT false",
            "acesso_buscar_cliente": "BOOLEAN DEFAULT false",
            "acesso_financeiro": "BOOLEAN DEFAULT false",
            "acesso_cadastros": "BOOLEAN DEFAULT false",
            "acesso_relatorios": "BOOLEAN DEFAULT false",
            "acesso_nao_roteirizados": "BOOLEAN DEFAULT false",
        }
        for coluna, tipo in novas_permissoes.items():
            if coluna not in cols_usu:
                comandos.append(f"ALTER TABLE usuarios_empresa ADD COLUMN {coluna} {tipo}")

    if "contas_financeiras" not in tabelas:
        comandos.append("""
        CREATE TABLE contas_financeiras (
            id INTEGER PRIMARY KEY,
            empresa_id INTEGER NOT NULL,
            nome VARCHAR(80) NOT NULL,
            tipo VARCHAR(20) DEFAULT 'banco',
            saldo_inicial FLOAT DEFAULT 0,
            ativa BOOLEAN DEFAULT true,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(empresa_id) REFERENCES empresas (id)
        )
        """)

    if "lancamentos_banco" not in tabelas:
        comandos.append("""
        CREATE TABLE lancamentos_banco (
            id INTEGER PRIMARY KEY,
            empresa_id INTEGER NOT NULL,
            conta_id INTEGER NOT NULL,
            data DATE NOT NULL,
            historico TEXT NOT NULL,
            documento VARCHAR(80),
            valor FLOAT DEFAULT 0,
            saldo FLOAT DEFAULT 0,
            categoria VARCHAR(20) DEFAULT 'aluguel',
            categoria_confirmada BOOLEAN DEFAULT false,
            pagamento_id INTEGER,
            hash_importacao VARCHAR(64),
            origem_importacao VARCHAR(120),
            ordem INTEGER DEFAULT 0,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(empresa_id) REFERENCES empresas (id),
            FOREIGN KEY(conta_id) REFERENCES contas_financeiras (id),
            FOREIGN KEY(pagamento_id) REFERENCES pagamentos (id),
            FOREIGN KEY(repasse_solicitacao_id) REFERENCES solicitacoes (id)
        )
        """)

    if "lancamentos_banco" in tabelas:
        cols_lb = colunas("lancamentos_banco")
        if "hash_importacao" not in cols_lb:
            comandos.append("ALTER TABLE lancamentos_banco ADD COLUMN hash_importacao VARCHAR(64)")
        if "categoria_confirmada" not in cols_lb:
            comandos.append("ALTER TABLE lancamentos_banco ADD COLUMN categoria_confirmada BOOLEAN DEFAULT false")
        if "ordem" not in cols_lb:
            comandos.append("ALTER TABLE lancamentos_banco ADD COLUMN ordem INTEGER DEFAULT 0")
        if "repasse_solicitacao_id" not in cols_lb:
            comandos.append("ALTER TABLE lancamentos_banco ADD COLUMN repasse_solicitacao_id INTEGER")
        if "organiza_lancamento_id" not in cols_lb:
            comandos.append("ALTER TABLE lancamentos_banco ADD COLUMN organiza_lancamento_id INTEGER")

    if "lancamentos_manuais_financeiros" not in tabelas:
        comandos.append("""
        CREATE TABLE lancamentos_manuais_financeiros (
            id INTEGER PRIMARY KEY,
            empresa_id INTEGER NOT NULL,
            conta_id INTEGER NOT NULL,
            data DATE NOT NULL,
            descricao TEXT NOT NULL,
            valor FLOAT DEFAULT 0,
            categoria VARCHAR(20) DEFAULT 'empresa',
            tipo VARCHAR(20) DEFAULT 'real',
            recebido BOOLEAN DEFAULT false,
            pagamento_id INTEGER,
            repasse_solicitacao_id INTEGER,
            ordem INTEGER DEFAULT 0,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(empresa_id) REFERENCES empresas (id),
            FOREIGN KEY(conta_id) REFERENCES contas_financeiras (id),
            FOREIGN KEY(pagamento_id) REFERENCES pagamentos (id)
        )
        """)

    if "lancamentos_manuais_financeiros" in tabelas:
        cols_lmf = colunas("lancamentos_manuais_financeiros")
        if "pagamento_id" not in cols_lmf:
            comandos.append("ALTER TABLE lancamentos_manuais_financeiros ADD COLUMN pagamento_id INTEGER")
        if "ordem" not in cols_lmf:
            comandos.append("ALTER TABLE lancamentos_manuais_financeiros ADD COLUMN ordem INTEGER DEFAULT 0")
        if "repasse_solicitacao_id" not in cols_lmf:
            comandos.append("ALTER TABLE lancamentos_manuais_financeiros ADD COLUMN repasse_solicitacao_id INTEGER")

    if "app_migrations" not in tabelas:
        comandos.append("""
        CREATE TABLE app_migrations (
            chave VARCHAR(120) PRIMARY KEY,
            executado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

    if "agenda" in tabelas:
        cols_ag = colunas("agenda")
        if "previsao_entrega" not in cols_ag:
            comandos.append("ALTER TABLE agenda ADD COLUMN previsao_entrega VARCHAR(5)")
        if "link_localizacao" not in cols_ag:
            comandos.append("ALTER TABLE agenda ADD COLUMN link_localizacao TEXT")
        if "tipo_evento" not in cols_ag:
            comandos.append("ALTER TABLE agenda ADD COLUMN tipo_evento VARCHAR(20) DEFAULT 'entrega'")
        if "status_operacional" not in cols_ag:
            comandos.append("ALTER TABLE agenda ADD COLUMN status_operacional VARCHAR(20) DEFAULT 'pendente'")
        if "observacoes_operacionais" not in cols_ag:
            comandos.append("ALTER TABLE agenda ADD COLUMN observacoes_operacionais TEXT")
        if "equipe_id" not in cols_ag:
            comandos.append("ALTER TABLE agenda ADD COLUMN equipe_id INTEGER")
        if "roteirizado" not in cols_ag:
            comandos.append("ALTER TABLE agenda ADD COLUMN roteirizado BOOLEAN DEFAULT false")

    if "vinculos_repasse_banco" not in tabelas:
        comandos.append("""
        CREATE TABLE vinculos_repasse_banco (
            id INTEGER PRIMARY KEY,
            empresa_id INTEGER NOT NULL,
            lancamento_banco_id INTEGER NOT NULL,
            solicitacao_id INTEGER NOT NULL,
            valor FLOAT DEFAULT 0,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            criado_por VARCHAR(120),
            FOREIGN KEY(empresa_id) REFERENCES empresas (id),
            FOREIGN KEY(lancamento_banco_id) REFERENCES lancamentos_banco (id),
            FOREIGN KEY(solicitacao_id) REFERENCES solicitacoes (id),
            CONSTRAINT uq_vinculo_repasse_banco UNIQUE (lancamento_banco_id, solicitacao_id)
        )
        """)

    if "lancamentos_manuais_financeiros" in tabelas:
        cols_manual_fin = colunas("lancamentos_manuais_financeiros")
        if "organiza_lancamento_id" not in cols_manual_fin:
            comandos.append(
                "ALTER TABLE lancamentos_manuais_financeiros "
                "ADD COLUMN organiza_lancamento_id INTEGER"
            )

    if "lancamentos_organiza" in tabelas:
        cols_org = colunas("lancamentos_organiza")
        if "falta_receber" not in cols_org:
            comandos.append(
                "ALTER TABLE lancamentos_organiza "
                "ADD COLUMN falta_receber NUMERIC(12, 2) DEFAULT 0 NOT NULL"
            )
        if "empresa_id" not in cols_org:
            comandos.append(
                "ALTER TABLE lancamentos_organiza "
                "ADD COLUMN empresa_id INTEGER"
            )
            comandos.append(
                "CREATE INDEX IF NOT EXISTS ix_lancamentos_organiza_empresa_id "
                "ON lancamentos_organiza (empresa_id)"
            )

    if "configuracoes_rota_inteligente" in tabelas:
        cols_cfg_ri = colunas("configuracoes_rota_inteligente")
        if "horario_minimo_cliente" not in cols_cfg_ri:
            comandos.append("ALTER TABLE configuracoes_rota_inteligente ADD COLUMN horario_minimo_cliente TIME DEFAULT '08:00:00' NOT NULL")
        if "raio_retirada_estrategica_km" not in cols_cfg_ri:
            comandos.append("ALTER TABLE configuracoes_rota_inteligente ADD COLUMN raio_retirada_estrategica_km FLOAT DEFAULT 10 NOT NULL")
        if "desvio_max_retirada_estrategica_min" not in cols_cfg_ri:
            comandos.append("ALTER TABLE configuracoes_rota_inteligente ADD COLUMN desvio_max_retirada_estrategica_min INTEGER DEFAULT 60 NOT NULL")
        if "custo_km" not in cols_cfg_ri:
            comandos.append("ALTER TABLE configuracoes_rota_inteligente ADD COLUMN custo_km FLOAT DEFAULT 0 NOT NULL")
        if "custo_hora_equipe" not in cols_cfg_ri:
            comandos.append("ALTER TABLE configuracoes_rota_inteligente ADD COLUMN custo_hora_equipe FLOAT DEFAULT 0 NOT NULL")
    if "rotas_inteligentes" in tabelas:
        cols_ri = colunas("rotas_inteligentes")
        if "carga_maxima_pontos" not in cols_ri:
            comandos.append("ALTER TABLE rotas_inteligentes ADD COLUMN carga_maxima_pontos INTEGER DEFAULT 0 NOT NULL")
        if "custo_estimado" not in cols_ri:
            comandos.append("ALTER TABLE rotas_inteligentes ADD COLUMN custo_estimado FLOAT DEFAULT 0 NOT NULL")
    if "rotas_inteligentes_paradas" in tabelas:
        cols_rip = colunas("rotas_inteligentes_paradas")
        if "carga_movimento" not in cols_rip:
            comandos.append("ALTER TABLE rotas_inteligentes_paradas ADD COLUMN carga_movimento INTEGER DEFAULT 0 NOT NULL")
        if "carga_apos_parada" not in cols_rip:
            comandos.append("ALTER TABLE rotas_inteligentes_paradas ADD COLUMN carga_apos_parada INTEGER DEFAULT 0 NOT NULL")

    if comandos:
        with engine.begin() as conn:
            for comando in comandos:
                conn.execute(text(comando))

    # As paradas inteligentes preservam o histórico da rota mesmo quando um
    # registro operacional da Agenda é limpo ou recriado. No PostgreSQL, a FK
    # antiga usava RESTRICT/NO ACTION e impedia o startup ao excluir duplicatas.
    # ON DELETE SET NULL mantém a parada salva e apenas remove o vínculo obsoleto.
    if engine.dialect.name == "postgresql" and "rotas_inteligentes_paradas" in tabelas:
        with engine.begin() as conn:
            conn.execute(text("""
                DO $$
                BEGIN
                    IF EXISTS (
                        SELECT 1
                          FROM pg_constraint
                         WHERE conname = 'rotas_inteligentes_paradas_agenda_id_fkey'
                           AND confdeltype <> 'n'
                    ) THEN
                        ALTER TABLE rotas_inteligentes_paradas
                            DROP CONSTRAINT rotas_inteligentes_paradas_agenda_id_fkey;
                        ALTER TABLE rotas_inteligentes_paradas
                            ADD CONSTRAINT rotas_inteligentes_paradas_agenda_id_fkey
                            FOREIGN KEY (agenda_id) REFERENCES agenda(id)
                            ON DELETE SET NULL;
                    END IF;
                END $$;
            """))

    # Lançamentos antigos do Organiza pertencem à empresa Karaoke RJ.
    # O preenchimento é idempotente e evita que apareçam em outras empresas.
    if "lancamentos_organiza" in tabelas:
        with engine.begin() as conn:
            conn.execute(text("""
                UPDATE lancamentos_organiza
                   SET empresa_id = (
                       SELECT id FROM empresas
                        WHERE lower(slug) IN ('karaokerj', 'karaoke-rj')
                           OR lower(nome) = 'karaoke rj'
                        ORDER BY CASE WHEN lower(slug) = 'karaokerj' THEN 0 ELSE 1 END
                        LIMIT 1
                   )
                 WHERE empresa_id IS NULL
            """))

    # Regra Humiat 1:1: 4 contratos grátis por mês e 1 Humiat por contrato excedente.
    # Corrige instalações que receberam temporariamente a configuração de 40 gratuitos.
    with engine.begin() as conn:
        conn.execute(text(
            "UPDATE empresas SET humiat_gratis_mes = 4, humiat_custo_contrato = 1 "
            "WHERE humiat_gratis_mes = 40"
        ))
        conn.execute(text(
            "UPDATE empresas SET humiat_custo_contrato = 1 "
            "WHERE humiat_gratis_mes = 4 AND humiat_custo_contrato = 10"
        ))




def _normalizar_chave_endereco(valor: str) -> str:
    texto_norm = unicodedata.normalize("NFKD", str(valor or ""))
    texto_norm = "".join(ch for ch in texto_norm if not unicodedata.combining(ch)).casefold()
    return " ".join(re.sub(r"[^a-z0-9]+", " ", texto_norm).split())


def migrar_enderecos_contratos_legados():
    """Congela o melhor endereço conhecido nos contratos antigos, uma única vez.

    Regra de segurança: quando não existe histórico suficiente para provar o número
    de um contrato antigo, o número fica vazio em vez de herdar silenciosamente o
    número atual do cliente. Para o contrato mais recente do cliente, o endereço
    atual pode ser usado quando o logradouro coincide.
    """
    chave_migracao = "20260830_endereco_evento_por_contrato_v1"
    try:
        with engine.begin() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS app_migrations (
                    chave VARCHAR(120) PRIMARY KEY,
                    executado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))
            if conn.execute(
                text("SELECT chave FROM app_migrations WHERE chave = :chave"),
                {"chave": chave_migracao},
            ).first():
                return
    except Exception:
        logger.exception("Falha ao preparar migração de endereço por contrato")
        return

    db = SessionLocal()
    try:
        pendentes = (
            db.query(Solicitacao)
            .filter(
                Solicitacao.local_numero.is_(None),
                Solicitacao.local_complemento.is_(None),
                Solicitacao.local_cidade.is_(None),
                Solicitacao.local_estado.is_(None),
                Solicitacao.local_cep.is_(None),
            )
            .order_by(Solicitacao.cliente_id, Solicitacao.criado_em, Solicitacao.id)
            .all()
        )
        if not pendentes:
            with engine.begin() as conn:
                conn.execute(text("INSERT INTO app_migrations (chave) VALUES (:chave)"), {"chave": chave_migracao})
            return

        cliente_ids = {s.cliente_id for s in pendentes if s.cliente_id}
        historicos = (
            db.query(EnderecoCliente)
            .filter(EnderecoCliente.cliente_id.in_(cliente_ids or [-1]))
            .order_by(EnderecoCliente.cliente_id, EnderecoCliente.criado_em, EnderecoCliente.id)
            .all()
        )
        por_cliente = {}
        for e in historicos:
            por_cliente.setdefault(e.cliente_id, []).append(e)

        # O cadastro atual é fallback somente para a reserva mais recente do cliente.
        ids_mais_recentes = dict(
            db.query(Solicitacao.cliente_id, func.max(Solicitacao.id))
            .filter(Solicitacao.cliente_id.in_(cliente_ids or [-1]))
            .group_by(Solicitacao.cliente_id)
            .all()
        )

        migrados_historico = 0
        migrados_atual = 0
        sem_numero_confiavel = 0

        for sol in pendentes:
            cliente = sol.cliente
            chave_rua = _normalizar_chave_endereco(sol.local)
            chave_bairro = _normalizar_chave_endereco(sol.bairro)
            candidatos = [
                e for e in por_cliente.get(sol.cliente_id, [])
                if chave_rua and _normalizar_chave_endereco(e.endereco) == chave_rua
            ]
            if chave_bairro:
                candidatos_bairro = [e for e in candidatos if _normalizar_chave_endereco(e.bairro) == chave_bairro]
                if candidatos_bairro:
                    candidatos = candidatos_bairro

            # Se o contrato já tinha nome do local, ele é um ótimo desempate para
            # endereços repetidos do mesmo cliente (condomínio, salão, empresa etc.).
            chave_apelido = _normalizar_chave_endereco(sol.local_nome)
            if chave_apelido:
                candidatos_apelido = [e for e in candidatos if _normalizar_chave_endereco(e.apelido) == chave_apelido]
                if candidatos_apelido:
                    candidatos = candidatos_apelido

            escolhido = None
            # Nunca "adivinha" entre dois números diferentes na mesma rua. Esse é
            # justamente o tipo de ambiguidade que causava o Waze ir ao local errado.
            numeros_distintos = {_normalizar_chave_endereco(e.numero) for e in candidatos if (e.numero or "").strip()}
            historico_ambiguo = len(numeros_distintos) > 1
            if candidatos and not historico_ambiguo:
                if sol.criado_em:
                    def distancia_tempo(e):
                        momento = e.criado_em or e.atualizado_em
                        if not momento:
                            return float("inf")
                        try:
                            return abs((momento - sol.criado_em).total_seconds())
                        except Exception:
                            return float("inf")
                    escolhido = min(candidatos, key=lambda e: (distancia_tempo(e), -(e.id or 0)))
                else:
                    escolhido = candidatos[-1]

            if escolhido:
                sol.local = (sol.local or escolhido.endereco or "").strip()
                sol.local_numero = (escolhido.numero or "").strip()
                sol.local_complemento = (escolhido.complemento or "").strip()
                sol.bairro = (sol.bairro or escolhido.bairro or "").strip()
                sol.local_cidade = (escolhido.cidade or "").strip()
                sol.local_estado = (escolhido.estado or "").strip()
                sol.local_cep = (escolhido.cep or "").strip()
                migrados_historico += 1
                continue

            endereco_atual_igual = bool(
                cliente and cliente.endereco and chave_rua
                and _normalizar_chave_endereco(cliente.endereco) == chave_rua
            )
            eh_mais_recente = ids_mais_recentes.get(sol.cliente_id) == sol.id

            if cliente and (not sol.local or (endereco_atual_igual and eh_mais_recente and not historico_ambiguo)):
                sol.local = (sol.local or cliente.endereco or "").strip()
                sol.local_numero = (cliente.numero or "").strip()
                sol.local_complemento = (cliente.complemento or "").strip()
                sol.bairro = (sol.bairro or cliente.bairro or "").strip()
                sol.local_cidade = (cliente.cidade or "").strip()
                sol.local_estado = (cliente.estado or "").strip()
                sol.local_cep = (cliente.cep or "").strip()
                migrados_atual += 1
            else:
                # Congela explicitamente sem número para nunca herdar outro número depois.
                sol.local_numero = ""
                sol.local_complemento = ""
                sol.local_cidade = (cliente.cidade or "").strip() if cliente and endereco_atual_igual else ""
                sol.local_estado = (cliente.estado or "").strip() if cliente and endereco_atual_igual else ""
                sol.local_cep = ""
                sem_numero_confiavel += 1

        db.commit()
        with engine.begin() as conn:
            conn.execute(text("INSERT INTO app_migrations (chave) VALUES (:chave)"), {"chave": chave_migracao})
        logger.info(
            "Migração endereço por contrato concluída: historico=%s atual=%s sem_numero=%s",
            migrados_historico, migrados_atual, sem_numero_confiavel,
        )
    except Exception:
        db.rollback()
        logger.exception("Falha na migração de endereço por contrato")
    finally:
        db.close()

def migrar_vinculos_repasse_legados():
    """Converte vínculos antigos 1:1 para o novo rateio N:N sem duplicar dados."""
    db = SessionLocal()
    try:
        antigos = db.query(LancamentoBanco).filter(LancamentoBanco.repasse_solicitacao_id != None).all()
        alterou = False
        for lanc in antigos:
            repasse = db.get(Solicitacao, lanc.repasse_solicitacao_id)
            if not repasse:
                lanc.repasse_solicitacao_id = None
                alterou = True
                continue
            existente = db.query(VinculoRepasseBanco).filter(
                VinculoRepasseBanco.lancamento_banco_id == lanc.id,
                VinculoRepasseBanco.solicitacao_id == repasse.id
            ).first()
            if not existente:
                valor = min(abs(float(lanc.valor or 0)), float(repasse.valor_repasse or 0))
                if valor > 0.01:
                    db.add(VinculoRepasseBanco(
                        empresa_id=lanc.empresa_id,
                        lancamento_banco_id=lanc.id,
                        solicitacao_id=repasse.id,
                        valor=valor,
                        criado_por="Migração automática"
                    ))
            lanc.categoria = "repasse"
            lanc.repasse_solicitacao_id = None
            alterou = True
        if alterou:
            db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()

def atualizar_mensagem_previsao_padrao():
    """Copia uma única vez para o cadastro as mensagens aprovadas que estavam fixas nos botões da operação.

    Depois dessa migração, os botões Previsão e A caminho passam a usar o texto cadastrado
    na empresa. O controle por chave evita sobrescrever edições futuras feitas em Configurações.
    """
    chave_migracao = "20260706_mensagens_operacao_aprovadas"
    try:
        with engine.begin() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS app_migrations (
                    chave VARCHAR(120) PRIMARY KEY,
                    executado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))
            ja_executou = conn.execute(
                text("SELECT chave FROM app_migrations WHERE chave = :chave"),
                {"chave": chave_migracao},
            ).first()
            if ja_executou:
                return

            conn.execute(
                text("""
                    UPDATE empresas
                       SET mensagem_preparacao = :preparacao,
                           mensagem_a_caminho = :a_caminho
                """),
                {
                    "preparacao": MENSAGEM_OPERACAO_PREPARACAO_APROVADA,
                    "a_caminho": MENSAGEM_OPERACAO_A_CAMINHO_APROVADA,
                },
            )
            conn.execute(
                text("INSERT INTO app_migrations (chave) VALUES (:chave)"),
                {"chave": chave_migracao},
            )
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Integração Organiza -> Financeiro
# Um registro enviado pelo Organiza corresponde a um lançamento do Sistema.
# id_externo é idempotente: reenvio atualiza o registro, sem duplicar.
# ---------------------------------------------------------------------------
@app.post("/api/integracoes/organiza/lancamentos")
async def receber_lancamento_organiza(request: Request, db: Session = Depends(get_db)):
    from datetime import date
    from decimal import Decimal, InvalidOperation

    # Chave opcional: se ORGANIZA_API_KEY estiver configurada, passa a ser obrigatória.
    chave_esperada = os.getenv("ORGANIZA_API_KEY", "").strip()
    if chave_esperada:
        chave_recebida = (request.headers.get("X-API-Key") or "").strip()
        if chave_recebida != chave_esperada:
            raise HTTPException(status_code=401, detail="Chave de integração inválida.")

    try:
        dados = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="JSON inválido.")

    obrigatorios = ("id_externo", "tipo", "valor", "data_pagamento", "banco")
    faltando = [campo for campo in obrigatorios if dados.get(campo) in (None, "")]
    if faltando:
        raise HTTPException(status_code=422, detail=f"Campos obrigatórios: {', '.join(faltando)}")

    tipo = str(dados["tipo"]).strip().lower()
    # Aceita também grafias comuns, mas grava de forma padronizada.
    aliases = {
        "venda": "venda",
        "manutencao": "manutencao",
        "manutenção": "manutencao",
    }
    if tipo not in aliases:
        raise HTTPException(status_code=422, detail="tipo deve ser 'venda' ou 'manutencao'.")
    tipo = aliases[tipo]

    try:
        valor = Decimal(str(dados["valor"]).replace("R$", "").replace(" ", "").replace(",", "."))
        falta_receber = Decimal(
            str(dados.get("falta_receber", 0))
            .replace("R$", "")
            .replace(" ", "")
            .replace(",", ".")
        )
    except (InvalidOperation, ValueError):
        raise HTTPException(status_code=422, detail="valor ou falta_receber inválido.")

    if valor <= 0:
        raise HTTPException(status_code=422, detail="valor deve ser maior que zero.")
    if falta_receber < 0:
        falta_receber = Decimal("0")

    try:
        data_pagamento = date.fromisoformat(str(dados["data_pagamento"])[:10])
    except ValueError:
        raise HTTPException(status_code=422, detail="data_pagamento deve usar AAAA-MM-DD.")

    id_externo = str(dados["id_externo"]).strip()

    # A integração do Organiza é vinculada à empresa de destino.
    # Por compatibilidade, quando o Organiza não envia a empresa, usa Karaoke RJ.
    empresa_slug = str(dados.get("empresa_slug") or "karaokerj").strip().lower()
    empresa_destino = None
    if dados.get("empresa_id") not in (None, ""):
        try:
            empresa_destino = db.get(Empresa, int(dados["empresa_id"]))
        except (TypeError, ValueError):
            raise HTTPException(status_code=422, detail="empresa_id inválido.")
    if not empresa_destino:
        empresa_destino = db.query(Empresa).filter(func.lower(Empresa.slug) == empresa_slug).first()
    if not empresa_destino and empresa_slug in ("karaokerj", "karaoke-rj"):
        empresa_destino = db.query(Empresa).filter(func.lower(Empresa.nome) == "karaoke rj").first()
    if not empresa_destino:
        raise HTTPException(status_code=422, detail="Empresa de destino da integração não encontrada.")

    registro = db.query(LancamentoOrganiza).filter(
        LancamentoOrganiza.id_externo == id_externo
    ).first()

    criado = registro is None
    if criado:
        registro = LancamentoOrganiza(id_externo=id_externo)
        db.add(registro)

    registro.empresa_id = empresa_destino.id
    registro.tipo = tipo
    registro.cliente = (str(dados.get("cliente") or "").strip() or None)
    registro.descricao = (str(dados.get("descricao") or "").strip() or None)
    registro.valor = valor
    registro.falta_receber = falta_receber
    registro.data_pagamento = data_pagamento
    registro.banco = str(dados["banco"]).strip()

    db.commit()
    db.refresh(registro)

    return {
        "ok": True,
        "acao": "criado" if criado else "atualizado",
        "id": registro.id,
        "id_externo": registro.id_externo,
        "tipo": registro.tipo,
    }


@app.get("/api/integracoes/organiza/lancamentos")
def listar_lancamentos_organiza(request: Request, db: Session = Depends(get_db)):
    """Consulta simples para conferência da integração."""
    chave_esperada = os.getenv("ORGANIZA_API_KEY", "").strip()
    if chave_esperada:
        chave_recebida = (request.headers.get("X-API-Key") or "").strip()
        if chave_recebida != chave_esperada:
            raise HTTPException(status_code=401, detail="Chave de integração inválida.")

    empresa_slug = (request.query_params.get("empresa_slug") or "karaokerj").strip().lower()
    empresa_destino = db.query(Empresa).filter(func.lower(Empresa.slug) == empresa_slug).first()
    consulta = db.query(LancamentoOrganiza)
    if empresa_destino:
        consulta = consulta.filter(LancamentoOrganiza.empresa_id == empresa_destino.id)
    registros = (
        consulta
        .order_by(LancamentoOrganiza.data_pagamento.desc(), LancamentoOrganiza.id.desc())
        .limit(500)
        .all()
    )
    return [{
        "empresa_id": r.empresa_id,
        "id_externo": r.id_externo,
        "tipo": r.tipo,
        "cliente": r.cliente,
        "descricao": r.descricao,
        "valor": float(r.valor or 0),
        "falta_receber": float(r.falta_receber or 0),
        "data_pagamento": r.data_pagamento.isoformat(),
        "banco": r.banco,
    } for r in registros]


@app.on_event("startup")
def startup():
    Base.metadata.create_all(bind=engine)
    garantir_colunas_novas()
    migrar_enderecos_contratos_legados()
    migrar_vinculos_repasse_legados()
    atualizar_mensagem_previsao_padrao()
    db = SessionLocal()
    try:
        inicializar_dados(db)
        for emp in db.query(Empresa).all():
            configurar_campos_empresa(db, emp.id)
            criar_modelos_iniciais_empresa(db, emp)
            if db.query(Equipe).filter_by(empresa_id=emp.id).count() == 0:
                db.add_all([Equipe(empresa_id=emp.id, nome="Equipe 1", ativa=True), Equipe(empresa_id=emp.id, nome="Equipe 2", ativa=True)])
                db.commit()
            corrigir_valores_teste(db)
            recalcular_valores_reservas(db)
            corrigir_reservas_aprovadas_sem_itens(db)
            limpar_agenda_operacional(db)
            garantir_agenda_reservas(db, emp.id)
    finally:
        db.close()


def nome_item_reserva(item: Solicitacao) -> str:
    """Nome operacional da reserva, sempre refletindo todos os itens do contrato.

    Solicitacao.produto representa apenas o produto principal/legado. Quando a
    reserva possui ReservaItem, a Operacao precisa mostrar o conjunto completo
    para evitar que um equipamento deixe de ser separado ou entregue.
    """
    itens_reserva = list(getattr(item, "itens", None) or [])
    if itens_reserva:
        nomes = []
        for ri in itens_reserva:
            nome = (getattr(ri, "nome", None) or "Item").strip()
            quantidade = int(getattr(ri, "quantidade", None) or 1)
            nomes.append(f"{quantidade}x {nome}" if quantidade > 1 else nome)
        return " + ".join(nomes)
    if item.produto:
        return item.produto.nome
    return "Reserva"


def retirada_obrigatoria_ativa(item: Solicitacao) -> bool:
    return bool(getattr(item, "retirada_obrigatoria", False))


def criar_ou_atualizar_retirada_obrigatoria(db: Session, item: Solicitacao):
    """
    Cria o card de BUSCA antes da entrega ser concluída quando o cliente exigiu retirada.
    Esse card fica com data/hora do contrato, não deve ser duplicado depois da entrega.
    """
    if not item or not item.id:
        return

    retirada = (
        db.query(Agenda)
        .filter_by(empresa_id=item.empresa_id, solicitacao_id=item.id, tipo_evento="retirada")
        .first()
    )

    if not retirada_obrigatoria_ativa(item):
        # Se a busca obrigatória foi removida e a busca ainda não foi executada, remove o card especial.
        if retirada and retirada.status_operacional != "concluido":
            db.delete(retirada)
        return

    data_retirada = item.retirada_data or item.data_evento
    hora_retirada = item.retirada_hora or item.hora_fim or item.hora_inicio
    titulo_base = f"{nome_item_reserva(item)} - {item.cliente.nome if item.cliente else 'Cliente'}"

    if not retirada:
        retirada = Agenda(
            empresa_id=item.empresa_id,
            solicitacao_id=item.id,
            tipo_evento="retirada",
            status_operacional="pendente",
        )
        db.add(retirada)

    # Retirada obrigatória é uma exigência contratual. Ao alterar a data/hora
    # no contrato, a rota deve acompanhar imediatamente, mesmo que já estivesse
    # roteirizada. Assim as duas horas exibidas permanecem iguais.
    retirada.data = data_retirada
    retirada.hora_inicio = hora_retirada
    retirada.hora_fim = None
    retirada.previsao_entrega = hora_retirada.strftime("%H:%M") if hora_retirada else ""
    retirada.titulo = titulo_base
    retirada.bairro = item.bairro


def criar_eventos_operacionais(db: Session, item: Solicitacao):
    """Cria operação somente para contrato aprovado com itens."""
    if not item or not item.id or not contrato_aprovado_para_operacao(item):
        return

    titulo_base = f"{nome_item_reserva(item)} - {item.cliente.nome if item.cliente else 'Cliente'}"
    entrega = (
        db.query(Agenda)
        .filter_by(empresa_id=item.empresa_id, solicitacao_id=item.id, tipo_evento="entrega")
        .first()
    )
    if not entrega:
        entrega = Agenda(
            empresa_id=item.empresa_id,
            solicitacao_id=item.id,
            tipo_evento="entrega",
            status_operacional="pendente",
            data=item.data_evento,
            hora_inicio=item.hora_inicio,
            hora_fim=item.hora_fim,
            titulo=titulo_base,
            bairro=item.bairro,
        )
        db.add(entrega)
    else:
        # Não sobrescreve data/hora operacional já roteirizada.
        # A data/hora do contrato continua em Solicitacao; a operação usa Agenda.
        ja_roteirizado = bool(
            (entrega.previsao_entrega or "").strip()
            or (entrega.observacoes_operacionais and "Roteirização salva" in entrega.observacoes_operacionais)
            or (entrega.data and item.data_evento and entrega.data != item.data_evento)
            or (entrega.hora_inicio and item.hora_inicio and entrega.hora_inicio != item.hora_inicio)
        )
        if not ja_roteirizado:
            entrega.data = item.data_evento
            entrega.hora_inicio = item.hora_inicio
            entrega.hora_fim = item.hora_fim
        entrega.titulo = titulo_base
        entrega.bairro = item.bairro

    criar_ou_atualizar_retirada_obrigatoria(db, item)

def retirar_solicitacao_da_operacao(db: Session, item: Solicitacao):
    """Retira entrega/busca da operação sem apagar contrato, itens ou pagamentos.

    O contrato continua disponível como crédito. Os pagamentos já registrados
    permanecem vinculados e serão reaproveitados quando uma nova data for definida.
    """
    if not item or not item.id:
        return

    agendas = (
        db.query(Agenda)
        .filter_by(empresa_id=item.empresa_id, solicitacao_id=item.id)
        .all()
    )
    agenda_ids = [agenda.id for agenda in agendas if agenda.id]

    # Preserva o histórico das rotas inteligentes, removendo apenas o vínculo
    # com os cards operacionais que deixarão de existir.
    if agenda_ids:
        db.query(RotaInteligenteParada).filter(
            RotaInteligenteParada.agenda_id.in_(agenda_ids)
        ).update({RotaInteligenteParada.agenda_id: None}, synchronize_session=False)

    db.query(Agenda).filter_by(
        empresa_id=item.empresa_id,
        solicitacao_id=item.id,
    ).delete(synchronize_session=False)


def garantir_agenda_reservas(db: Session, empresa_id: int | None = None):
    """
    Garante que toda reserva operacionalmente ativa apareça na Agenda.
    Contratos em crédito ou cancelados não podem recriar Entregar/Buscar.
    """
    q = db.query(Solicitacao)
    if empresa_id:
        q = q.filter(Solicitacao.empresa_id == empresa_id)

    alterou = False
    for reserva in q.all():
        if not contrato_aprovado_para_operacao(reserva):
            # Limpa cards antigos que possam ter sido criados quando o contrato ainda era rascunho.
            removidos = db.query(Agenda).filter_by(empresa_id=reserva.empresa_id, solicitacao_id=reserva.id).delete(synchronize_session=False)
            alterou = alterou or bool(removidos)
            continue
        existe = (
            db.query(Agenda)
            .filter_by(empresa_id=reserva.empresa_id, solicitacao_id=reserva.id, tipo_evento="entrega")
            .first()
        )
        if not existe or retirada_obrigatoria_ativa(reserva):
            criar_eventos_operacionais(db, reserva)
            alterou = True

    if alterou:
        db.commit()


def _previsao_retirada_operacional(reserva: Solicitacao, entrega: Agenda | None = None) -> tuple[date, time]:
    """Calcula a busca prevista mesmo antes de a entrega ser concluída.

    A Operação continua sendo a fonte da verdade: se a entrega foi movida manualmente,
    a previsão parte da data/hora operacional. Retirada obrigatória mantém exatamente
    a data/hora contratada.
    """
    if retirada_obrigatoria_ativa(reserva):
        data_prevista = reserva.retirada_data or reserva.data_evento
        hora_prevista = reserva.retirada_hora or reserva.hora_fim or reserva.hora_inicio
        return data_prevista, hora_prevista

    prazo_dias = 1
    if reserva.produto and reserva.produto.prazo_retirada_dias is not None:
        prazo_dias = max(0, int(reserva.produto.prazo_retirada_dias or 0))

    data_entrega = (entrega.data if entrega and entrega.data else reserva.data_evento)
    hora_entrega = (
        (entrega.hora_fim or entrega.hora_inicio) if entrega
        else (reserva.hora_fim or reserva.hora_inicio)
    )
    return data_entrega + timedelta(days=prazo_dias), hora_entrega


def _criar_ou_obter_retirada_prevista(db: Session, reserva: Solicitacao, entrega: Agenda | None = None) -> Agenda:
    """Materializa na Operação a busca que a Inteligência já conseguia prever."""
    retirada = (
        db.query(Agenda)
        .filter_by(empresa_id=reserva.empresa_id, solicitacao_id=reserva.id, tipo_evento="retirada")
        .first()
    )
    if retirada:
        return retirada

    data_prevista, hora_prevista = _previsao_retirada_operacional(reserva, entrega)
    titulo = entrega.titulo if entrega else f"{nome_item_reserva(reserva)} - {reserva.cliente.nome if reserva.cliente else 'Cliente'}"
    retirada = Agenda(
        empresa_id=reserva.empresa_id,
        solicitacao_id=reserva.id,
        tipo_evento="retirada",
        status_operacional="pendente",
        data=data_prevista,
        hora_inicio=hora_prevista,
        hora_fim=None,
        titulo=titulo,
        bairro=(entrega.bairro if entrega else reserva.bairro),
    )
    db.add(retirada)
    db.flush()
    return retirada


def criar_retirada_apos_entrega(db: Session, entrega: Agenda):
    """Ao concluir uma entrega, transforma a previsão da Inteligência em Operação real."""
    reserva = entrega.solicitacao
    if not reserva:
        return

    if retirada_obrigatoria_ativa(reserva):
        criar_ou_atualizar_retirada_obrigatoria(db, reserva)
        return

    _criar_ou_obter_retirada_prevista(db, reserva, entrega)


def mensagens_empresa(empresa: Empresa) -> dict:
    """Mensagens prontas. A empresa pode editar sem precisar entender o sistema."""
    return {
        "reserva": empresa.mensagem_reserva or (
            "Olá!\n\n"
            "Para agilizar sua reserva, preencha este formulário:\n"
            "{{link}}\n\n"
            "Importante: favor preencher utilizando o seu próprio WhatsApp, pois ele será usado para identificar e confirmar sua solicitação.\n\n"
            "Após o envio, nossa equipe irá preparar os equipamentos, valores e o pré-contrato.\n\n"
            "Assim que estiver tudo pronto, você receberá o contrato para análise e aceite."
        ),
        "aceite": empresa.mensagem_aceite or (
            "Olá, {{cliente}}!\n\n"
            "Seu pré-contrato está pronto.\n\n"
            "Confira atentamente as informações e, se estiver tudo correto, efetue o aceite pelo link abaixo:\n"
            "{{link}}"
        ),
        # Mantido apenas por compatibilidade com bancos antigos. Não é mais exibido nem utilizado no fluxo.
        "pagamento": "",
        "confirmacao": empresa.mensagem_confirmacao or (
            "Sua reserva foi efetivada com sucesso.\n\n"
            "Obrigado pela confiança!"
        ),
        "hora_fim": empresa.mensagem_hora_fim or (
            "A locação padrão tem duração de 4 horas. Após esse período, o equipamento poderá permanecer no local, porém sem acesso ao suporte técnico."
        ),
        "preparacao": empresa.mensagem_preparacao or MENSAGEM_OPERACAO_PREPARACAO_APROVADA,
        "a_caminho": empresa.mensagem_a_caminho or MENSAGEM_OPERACAO_A_CAMINHO_APROVADA,
    }


def url_publica(request: Request, caminho: str) -> str:
    base = str(request.base_url).rstrip("/")
    return f"{base}{caminho}"


# Cache leve de empresa para evitar uma consulta ao Neon em toda página de leitura.
# O cache guarda apenas colunas escalares e nunca substitui o banco em gravações.
_EMPRESA_CACHE_TTL_SECONDS = max(60, int(os.getenv("EMPRESA_CACHE_TTL_SECONDS", "3600")))
_empresa_cache: dict[int, tuple[float, dict]] = {}
_empresa_cache_lock = threading.Lock()


def _empresa_cache_snapshot(empresa: Empresa) -> dict:
    return {col.name: getattr(empresa, col.name) for col in Empresa.__table__.columns}


def empresa_cache_salvar(empresa: Empresa) -> None:
    if not empresa or not getattr(empresa, "id", None):
        return
    expira_em = time_module.monotonic() + _EMPRESA_CACHE_TTL_SECONDS
    with _empresa_cache_lock:
        _empresa_cache[int(empresa.id)] = (expira_em, _empresa_cache_snapshot(empresa))


def empresa_cache_invalidar(empresa_id: int | None = None) -> None:
    with _empresa_cache_lock:
        if empresa_id is None:
            _empresa_cache.clear()
        else:
            _empresa_cache.pop(int(empresa_id), None)


def _empresa_cache_obter(db: Session, empresa_id: int) -> Empresa | None:
    agora = time_module.monotonic()
    with _empresa_cache_lock:
        entrada = _empresa_cache.get(int(empresa_id))
        if not entrada:
            return None
        expira_em, snapshot = entrada
        if expira_em <= agora:
            _empresa_cache.pop(int(empresa_id), None)
            return None
        dados = dict(snapshot)

    # Recria uma instância destacada e a anexa à sessão sem SELECT.
    # Assim o restante do sistema continua recebendo um objeto Empresa normal.
    empresa = Empresa(**dados)
    make_transient_to_detached(empresa)
    return db.merge(empresa, load=False)


def empresa_logada(request: Request, db: Session = Depends(get_db)) -> Empresa:
    empresa_id = request.session.get("empresa_id")
    if not empresa_id:
        raise HTTPException(status_code=303, headers={"Location": "/empresa/login"})

    # GET/HEAD são consultas: reutilizam a empresa já conhecida da sessão.
    # Qualquer gravação invalida o cache e lê o registro atual antes de salvar.
    empresa = None
    if request.method in {"GET", "HEAD"}:
        empresa = _empresa_cache_obter(db, int(empresa_id))
    else:
        empresa_cache_invalidar(int(empresa_id))

    if empresa is None:
        empresa = db.get(Empresa, empresa_id)
        if empresa:
            empresa_cache_salvar(empresa)

    if not empresa:
        empresa_cache_invalidar(int(empresa_id))
        request.session.clear()
        raise HTTPException(status_code=303, headers={"Location": "/empresa/login"})
    return empresa


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    if request.session.get("empresa_id"):
        return RedirectResponse("/painel", status_code=303)
    if request.session.get("admin_geral"):
        return RedirectResponse("/admin", status_code=303)
    return templates.TemplateResponse("home.html", {"request": request})


def admin_geral_logado(request: Request):
    if not request.session.get("admin_geral"):
        raise HTTPException(status_code=303, headers={"Location": "/admin/login"})
    return True


@app.get("/admin/performance", response_class=HTMLResponse)
def admin_performance(
        request: Request,
        limit: int = 200,
        ok: bool = Depends(admin_geral_logado),
):
    """Painel temporário de diagnóstico, sem dados pessoais ou parâmetros SQL."""
    resumo = performance_summary(limit)
    return templates.TemplateResponse("admin/performance.html", {
        "request": request,
        "monitor": monitor_status(),
        "ranking": resumo["ranking"],
        "registros": resumo["records"],
        "limite": limit,
        "limpos": request.query_params.get("limpos"),
    })


@app.get("/admin/performance/dados", response_class=JSONResponse)
def admin_performance_dados(
        limit: int = 200,
        ok: bool = Depends(admin_geral_logado),
):
    resumo = performance_summary(limit)
    return {"monitor": monitor_status(), **resumo}


@app.post("/admin/performance/limpar")
def admin_performance_limpar(ok: bool = Depends(admin_geral_logado)):
    total = clear_records()
    return RedirectResponse(f"/admin/performance?limpos={total}", status_code=303)


@app.get("/admin/login", response_class=HTMLResponse)
def admin_login_form(request: Request):
    return templates.TemplateResponse("admin/login.html", {
        "request": request,
        "erro": request.query_params.get("erro"),
        "titulo_login": "Administrador Geral",
        "action_login": "/admin/login"
    })


@app.post("/admin/login")
def admin_login(request: Request, usuario: str = Form(...), senha: str = Form(...)):
    if usuario.strip() == ADMIN_NOME and senha.strip() == ADMIN_SENHA:
        request.session.clear()
        request.session["admin_geral"] = True
        return RedirectResponse("/admin", status_code=303)
    return RedirectResponse("/admin/login?erro=Usuário ou senha inválidos", status_code=303)


@app.get("/admin/sair")
def admin_sair(request: Request):
    request.session.clear()
    return RedirectResponse("/", status_code=303)


def _registrar_movimento_humiat(db: Session, empresa: Empresa, quantidade: int, tipo: str, motivo: str = "", observacao: str = "", usuario: str = "", solicitacao_id: int | None = None):
    anterior = int(empresa.humiat_saldo or 0)
    posterior = anterior + int(quantidade)
    empresa.humiat_saldo = posterior
    mov = HumiatMovimento(
        empresa_id=empresa.id, solicitacao_id=solicitacao_id, tipo=tipo, quantidade=int(quantidade),
        saldo_anterior=anterior, saldo_posterior=posterior, motivo=(motivo or "")[:200],
        observacao=observacao or None, usuario=(usuario or "")[:120]
    )
    db.add(mov)
    return mov


def _processar_humiat_aceite(db: Session, empresa: Empresa, item: Solicitacao):
    """Consome 1 Humiat por contrato aceito, usando primeiro a franquia mensal gratuita."""
    if item.humiat_processado:
        return
    aceite = item.aceite_em or agora_utc()
    competencia = aceite.strftime("%Y-%m")
    # Regra 1:1: cada contrato aceito representa exatamente 1 Humiat.
    gratis_mes = max(0, int(empresa.humiat_gratis_mes or 4))
    custo = 1
    usados_gratis = db.query(Solicitacao).filter(
        Solicitacao.empresa_id == empresa.id,
        Solicitacao.humiat_processado == True,
        Solicitacao.humiat_competencia == competencia,
        Solicitacao.humiat_status == "gratuito",
    ).count()
    item.humiat_processado = True
    item.humiat_competencia = competencia
    item.humiat_custo = custo
    if usados_gratis < gratis_mes:
        item.humiat_status = "gratuito"
        return
    if int(empresa.humiat_saldo or 0) >= custo:
        _registrar_movimento_humiat(
            db, empresa, -custo, "consumo_contrato",
            f"Contrato #{item.id} aceito", solicitacao_id=item.id
        )
        item.humiat_status = "debitado"
    else:
        item.humiat_status = "pendente_saldo"

def _quitar_humiats_pendentes(db: Session, empresa: Empresa):
    pendentes = db.query(Solicitacao).filter(
        Solicitacao.empresa_id == empresa.id,
        Solicitacao.humiat_processado == True,
        Solicitacao.humiat_status == "pendente_saldo",
        Solicitacao.humiat_custo > 0,
    ).order_by(Solicitacao.aceite_em.asc(), Solicitacao.id.asc()).all()
    quitados = 0
    for item in pendentes:
        custo = int(item.humiat_custo or 0)
        if int(empresa.humiat_saldo or 0) < custo:
            break
        _registrar_movimento_humiat(db, empresa, -custo, "consumo_contrato", f"Contrato #{item.id} aceito - quitação automática", solicitacao_id=item.id)
        item.humiat_status = "debitado"
        quitados += 1
    return quitados


@app.get("/admin", response_class=HTMLResponse)
def admin_geral(request: Request, db: Session = Depends(get_db), ok: bool = Depends(admin_geral_logado)):
    empresas = db.query(Empresa).order_by(Empresa.nome).all()
    return templates.TemplateResponse("admin/empresas.html",
                                      {"request": request, "empresas": empresas, "empresa": None})


@app.post("/admin/empresas")
def admin_criar_empresa(
        nome: str = Form(...),
        slug: str = Form(...),
        usuario_admin: str = Form(...),
        senha_admin: str = Form(...),
        identificador_principal: str = Form("telefone"),
        pix_copia_cola: str = Form(""),
        whatsapp_retorno: str = Form(""),
        exige_sinal: Optional[str] = Form(None),
        suporte_inicio: str = Form(""),
        suporte_fim: str = Form(""),
        mostrar_suporte_contrato: Optional[str] = Form(None),
        logo_url: str = Form(""),
        logo_idb_url: str = Form(""),
        logo_arquivo: UploadFile | None = File(None),
        tema: str = Form("azul"),
        mensagem_reserva: str = Form(""),
        mensagem_aceite: str = Form(""),
        mensagem_pagamento: str = Form(""),
        mensagem_confirmacao: str = Form(""),
        mensagem_hora_fim: str = Form(""),
        mostrar_mensagem_hora_fim: Optional[str] = Form(None),
        mensagem_preparacao: str = Form(""),
        mensagem_a_caminho: str = Form(""),
        mensagem_localizacao: str = Form(""),
        db: Session = Depends(get_db),
        ok: bool = Depends(admin_geral_logado)
):
    empresa = Empresa(
        nome=nome.strip(),
        slug=slug.strip().lower().replace(" ", "-"),
        identificador_principal=identificador_principal,
        usuario_admin=usuario_admin.strip(),
        senha_admin=senha_admin.strip(),
        pix_copia_cola=pix_copia_cola.strip(),
        whatsapp_retorno=_limpar_tel_whatsapp(whatsapp_retorno),
        exige_sinal=bool(exige_sinal),
        suporte_inicio=suporte_inicio.strip(),
        suporte_fim=suporte_fim.strip(),
        mostrar_suporte_contrato=bool(mostrar_suporte_contrato),
        logo_url="",
        logo_idb_url="",
        tema=tema,
        mensagem_reserva=mensagem_reserva.strip(),
        mensagem_aceite=mensagem_aceite.strip(),
        mensagem_pagamento=mensagem_pagamento.strip(),
        mensagem_confirmacao=mensagem_confirmacao.strip(),
        mensagem_preparacao=mensagem_preparacao.strip(),
        mensagem_a_caminho=mensagem_a_caminho.strip(),
        mensagem_localizacao=mensagem_localizacao.strip(),
        ativa=True
    )
    db.add(empresa)
    db.commit()
    db.refresh(empresa)

    # Logo no cadastro inicial da empresa.
    if logo_arquivo and logo_arquivo.filename:
        extensao = Path(logo_arquivo.filename).suffix.lower()
        if extensao not in [".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg"]:
            raise HTTPException(400, "Formato de logo inválido. Use PNG, JPG, WEBP, GIF ou SVG.")
        nome_arquivo = f"empresa_{empresa.id}_{uuid.uuid4().hex}{extensao}"
        destino = Path("static/uploads/logos") / nome_arquivo
        with destino.open("wb") as buffer:
            shutil.copyfileobj(logo_arquivo.file, buffer)
        empresa.logo_url = f"/static/uploads/logos/{nome_arquivo}"
        empresa.logo_idb_url = ""
    elif logo_url.strip():
        empresa.logo_url = logo_url.strip()
        empresa.logo_idb_url = ""
    elif logo_idb_url.strip():
        empresa.logo_idb_url = logo_idb_url.strip()
        empresa.logo_url = ""
    db.commit()
    db.refresh(empresa)
    configurar_campos_empresa(db, empresa.id)
    criar_modelos_iniciais_empresa(db, empresa)
    return RedirectResponse("/admin", status_code=303)


@app.get("/admin/empresa/{empresa_id}", response_class=HTMLResponse)
def admin_editar_empresa(empresa_id: int, request: Request, db: Session = Depends(get_db),
                         ok: bool = Depends(admin_geral_logado)):
    empresa = db.get(Empresa, empresa_id)
    if not empresa:
        raise HTTPException(404)
    empresas = db.query(Empresa).order_by(Empresa.nome).all()
    usuarios_empresa = db.query(UsuarioEmpresa).filter_by(empresa_id=empresa.id).order_by(UsuarioEmpresa.nome).all()
    equipes = db.query(Equipe).filter_by(empresa_id=empresa.id).order_by(Equipe.nome).all()
    equipes_usuario = {u.id: [e.id for e in u.equipes] for u in usuarios_empresa}
    competencia_atual = agora_utc().strftime("%Y-%m")
    aceitos_mes = db.query(Solicitacao).filter(
        Solicitacao.empresa_id == empresa.id, Solicitacao.humiat_processado == True,
        Solicitacao.humiat_competencia == competencia_atual
    ).count()
    pendentes_humiat = db.query(Solicitacao).filter_by(empresa_id=empresa.id, humiat_status="pendente_saldo").count()
    movimentos_humiat = db.query(HumiatMovimento).filter_by(empresa_id=empresa.id).order_by(HumiatMovimento.id.desc()).limit(50).all()
    return templates.TemplateResponse("admin/empresas.html",
                                      {"request": request, "empresas": empresas, "empresa": empresa,
                                       "usuarios_empresa": usuarios_empresa, "equipes": equipes,
                                       "equipes_usuario": equipes_usuario, "aceitos_mes": aceitos_mes,
                                       "pendentes_humiat": pendentes_humiat, "movimentos_humiat": movimentos_humiat})


@app.post("/admin/empresa/{empresa_id}")
def admin_salvar_empresa(
        empresa_id: int,
        nome: str = Form(...),
        slug: str = Form(...),
        usuario_admin: str = Form(...),
        senha_admin: str = Form(...),
        identificador_principal: str = Form("telefone"),
        pix_copia_cola: str = Form(""),
        exige_sinal: Optional[str] = Form(None),
        suporte_inicio: str = Form(""),
        suporte_fim: str = Form(""),
        mostrar_suporte_contrato: Optional[str] = Form(None),
        logo_url: str = Form(""),
        logo_idb_url: str = Form(""),
        logo_arquivo: UploadFile | None = File(None),
        tema: str = Form("azul"),
        ativa: Optional[str] = Form(None),
        humiat_gratis_mes: int = Form(4),
        humiat_custo_contrato: int = Form(1),
        db: Session = Depends(get_db),
        ok: bool = Depends(admin_geral_logado)
):
    empresa = db.get(Empresa, empresa_id)
    if not empresa:
        raise HTTPException(404)
    empresa.nome = nome.strip()
    empresa.slug = slug.strip().lower().replace(" ", "-")
    empresa.identificador_principal = identificador_principal
    empresa.usuario_admin = usuario_admin.strip()
    empresa.senha_admin = senha_admin.strip()
    empresa.pix_copia_cola = pix_copia_cola.strip()
    empresa.exige_sinal = bool(exige_sinal)
    empresa.suporte_inicio = suporte_inicio.strip()
    empresa.suporte_fim = suporte_fim.strip()
    empresa.mostrar_suporte_contrato = bool(mostrar_suporte_contrato)
    # Logo: o caminho mais simples para o locador é enviar do próprio PC/celular.
    # Mantemos URL apenas como alternativa técnica.
    if logo_arquivo and logo_arquivo.filename:
        extensao = Path(logo_arquivo.filename).suffix.lower()
        if extensao not in [".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg"]:
            raise HTTPException(400, "Formato de logo inválido. Use PNG, JPG, WEBP, GIF ou SVG.")
        nome_arquivo = f"empresa_{empresa.id}_{uuid.uuid4().hex}{extensao}"
        destino = Path("static/uploads/logos") / nome_arquivo
        with destino.open("wb") as buffer:
            shutil.copyfileobj(logo_arquivo.file, buffer)
        empresa.logo_url = f"/static/uploads/logos/{nome_arquivo}"
        empresa.logo_idb_url = ""
    elif logo_url.strip():
        empresa.logo_url = logo_url.strip()
        empresa.logo_idb_url = ""
    elif logo_idb_url.strip():
        empresa.logo_idb_url = logo_idb_url.strip()
        empresa.logo_url = ""
    empresa.tema = tema
    empresa.ativa = bool(ativa)
    empresa.humiat_gratis_mes = max(0, int(humiat_gratis_mes or 0))
    empresa.humiat_custo_contrato = max(0, int(humiat_custo_contrato or 0))
    db.commit()
    empresa_cache_invalidar(empresa.id)
    return RedirectResponse("/admin", status_code=303)


@app.post("/admin/empresa/{empresa_id}/humiats")
def admin_movimentar_humiats(
        empresa_id: int, request: Request, quantidade: int = Form(...), tipo: str = Form("credito"),
        motivo: str = Form(...), observacao: str = Form(""),
        db: Session = Depends(get_db), ok: bool = Depends(admin_geral_logado)):
    empresa = db.get(Empresa, empresa_id)
    if not empresa:
        raise HTTPException(404)
    qtd = abs(int(quantidade or 0))
    if qtd <= 0:
        return RedirectResponse(f"/admin/empresa/{empresa_id}?erro=Informe uma quantidade maior que zero", status_code=303)
    if tipo == "remover":
        qtd = -qtd
        if int(empresa.humiat_saldo or 0) + qtd < 0:
            return RedirectResponse(f"/admin/empresa/{empresa_id}?erro=O ajuste não pode deixar o saldo negativo", status_code=303)
        mov_tipo = "ajuste"
    else:
        mov_tipo = "credito"
    usuario = (request.session.get("usuario_nome") if request else None) or ADMIN_NOME or "Administrador"
    _registrar_movimento_humiat(db, empresa, qtd, mov_tipo, motivo.strip(), observacao.strip(), usuario)
    quitados = _quitar_humiats_pendentes(db, empresa) if qtd > 0 else 0
    db.commit()
    sufixo = f"&ok={quitados} contrato(s) pendente(s) quitado(s) automaticamente" if quitados else ""
    return RedirectResponse(f"/admin/empresa/{empresa_id}?humiat=ok{sufixo}", status_code=303)


@app.post("/admin/empresa/{empresa_id}/usuarios")
def admin_criar_usuario_empresa(
        empresa_id: int,
        nome: str = Form(...),
        usuario: str = Form(...),
        senha: Optional[str] = Form(None),
        usuario_id: Optional[str] = Form(None),
        ativo: Optional[str] = Form("1"),
        acesso_agenda: Optional[str] = Form(None),
        acesso_operacao: Optional[str] = Form(None),
        acesso_buscar_cliente: Optional[str] = Form(None),
        acesso_financeiro: Optional[str] = Form(None),
        acesso_cadastros: Optional[str] = Form(None),
        acesso_relatorios: Optional[str] = Form(None),
        acesso_nao_roteirizados: Optional[str] = Form(None),
        equipes_permitidas: list[int] = Form([]),
        db: Session = Depends(get_db),
        ok: bool = Depends(admin_geral_logado)
):
    empresa = db.get(Empresa, empresa_id)
    if not empresa:
        raise HTTPException(404)

    usuario_limpo = usuario.strip()
    if not usuario_limpo:
        raise HTTPException(400, "Informe o usuário.")

    usuario_id_int = int(usuario_id) if usuario_id and str(usuario_id).strip() else None
    existente = None
    if usuario_id_int:
        existente = db.get(UsuarioEmpresa, usuario_id_int)
        if not existente or existente.empresa_id != empresa.id:
            raise HTTPException(404, "Usuário não encontrado.")

    conflito = (
        db.query(UsuarioEmpresa)
        .filter(
            UsuarioEmpresa.empresa_id == empresa.id,
            UsuarioEmpresa.usuario == usuario_limpo
        )
        .first()
    )
    if conflito and (not existente or conflito.id != existente.id):
        raise HTTPException(400, "Já existe um usuário com este login nesta empresa.")

    dados = {
        "nome": nome.strip(),
        "usuario": usuario_limpo,
        "ativo": bool(ativo),
        "acesso_agenda": bool(acesso_agenda),
        "acesso_operacao": bool(acesso_operacao),
        "acesso_buscar_cliente": bool(acesso_buscar_cliente),
        "acesso_financeiro": bool(acesso_financeiro),
        "acesso_cadastros": bool(acesso_cadastros),
        "acesso_relatorios": bool(acesso_relatorios),
        "acesso_nao_roteirizados": bool(acesso_nao_roteirizados),
    }

    if existente:
        for campo, valor in dados.items():
            setattr(existente, campo, valor)
        if senha and senha.strip():
            existente.senha = senha.strip()
    else:
        if not senha or not senha.strip():
            raise HTTPException(400, "Informe a senha para criar o usuário.")
        db.add(UsuarioEmpresa(
            empresa_id=empresa.id,
            senha=senha.strip(),
            **dados
        ))

    db.commit()
    alvo = existente or db.query(UsuarioEmpresa).filter_by(empresa_id=empresa.id, usuario=usuario_limpo).first()
    permitidas = db.query(Equipe).filter(Equipe.empresa_id == empresa.id, Equipe.id.in_(equipes_permitidas or [-1])).all()
    alvo.equipes = permitidas
    db.commit()
    return RedirectResponse(f"/admin/empresa/{empresa_id}", status_code=303)


@app.post("/admin/empresa/{empresa_id}/equipes")
def admin_salvar_equipe(empresa_id: int, nome: str = Form(...), equipe_id: Optional[int] = Form(None), ativo: Optional[str] = Form("1"), db: Session = Depends(get_db), ok: bool = Depends(admin_geral_logado)):
    empresa = db.get(Empresa, empresa_id)
    if not empresa: raise HTTPException(404)
    equipe = db.get(Equipe, equipe_id) if equipe_id else None
    if equipe and equipe.empresa_id != empresa.id: raise HTTPException(404)
    if not equipe:
        equipe = Equipe(empresa_id=empresa.id)
        db.add(equipe)
    equipe.nome = nome.strip()
    equipe.ativa = bool(ativo)
    db.commit()
    return RedirectResponse(f"/admin/empresa/{empresa_id}", status_code=303)


@app.get("/admin/empresa/{empresa_id}/equipe/{equipe_id}/excluir")
def admin_excluir_equipe(empresa_id: int, equipe_id: int, db: Session = Depends(get_db), ok: bool = Depends(admin_geral_logado)):
    equipe = db.get(Equipe, equipe_id)
    if equipe and equipe.empresa_id == empresa_id:
        em_uso = db.query(Agenda).filter(Agenda.equipe_id == equipe.id).first()
        if em_uso: equipe.ativa = False
        else: db.delete(equipe)
        db.commit()
    return RedirectResponse(f"/admin/empresa/{empresa_id}", status_code=303)


@app.get("/admin/empresa/{empresa_id}/usuario/{usuario_id}/excluir")
def admin_excluir_usuario_empresa(
        empresa_id: int,
        usuario_id: int,
        db: Session = Depends(get_db),
        ok: bool = Depends(admin_geral_logado)
):
    usuario = db.get(UsuarioEmpresa, usuario_id)
    if usuario and usuario.empresa_id == empresa_id:
        db.delete(usuario)
        db.commit()
    return RedirectResponse(f"/admin/empresa/{empresa_id}", status_code=303)


@app.get("/empresa/login", response_class=HTMLResponse)
def empresa_login_form(request: Request, db: Session = Depends(get_db)):
    if request.session.get("empresa_id"):
        return RedirectResponse("/painel", status_code=303)
    return templates.TemplateResponse("admin/login.html", {
        "request": request,
        "erro": request.query_params.get("erro"),
        "titulo_login": "Acesso da Empresa",
        "action_login": "/empresa/login"
    })


@app.post("/empresa/login")
def empresa_login(request: Request, usuario: str = Form(...), senha: str = Form(...), db: Session = Depends(get_db)):
    if request.session.get("empresa_id"):
        return RedirectResponse("/painel", status_code=303)
    usuario_limpo = usuario.strip()
    usuario_busca = usuario_limpo.lower()
    senha_limpa = senha.strip()

    empresa = db.query(Empresa).filter(
        func.lower(Empresa.usuario_admin) == usuario_busca,
        Empresa.senha_admin == senha_limpa,
        Empresa.ativa == True
    ).first()
    if empresa:
        request.session.clear()
        request.session["empresa_id"] = empresa.id
        request.session["usuario_sistema"] = usuario_limpo
        request.session["usuario_nome"] = empresa.usuario_admin or usuario_limpo
        request.session["acesso_total"] = True
        request.session["acessos"] = {}
        empresa_cache_salvar(empresa)
        return RedirectResponse("/painel", status_code=303)

    usuario_empresa = (
        db.query(UsuarioEmpresa)
        .join(Empresa, Empresa.id == UsuarioEmpresa.empresa_id)
        .filter(func.lower(UsuarioEmpresa.usuario) == usuario_busca, UsuarioEmpresa.senha == senha_limpa,
                UsuarioEmpresa.ativo == True, Empresa.ativa == True)
        .first()
    )
    if usuario_empresa:
        request.session.clear()
        request.session["empresa_id"] = usuario_empresa.empresa_id
        request.session["usuario_sistema"] = usuario_empresa.usuario
        request.session["usuario_nome"] = usuario_empresa.nome
        request.session["usuario_empresa_id"] = usuario_empresa.id
        request.session["acesso_total"] = False
        request.session["acessos"] = {
            "agenda": bool(usuario_empresa.acesso_agenda),
            "operacao": bool(usuario_empresa.acesso_operacao),
            "buscar_cliente": bool(usuario_empresa.acesso_buscar_cliente),
            "financeiro": bool(usuario_empresa.acesso_financeiro),
            "cadastros": bool(usuario_empresa.acesso_cadastros),
            "relatorios": bool(usuario_empresa.acesso_relatorios),
        }
        empresa_usuario = db.get(Empresa, usuario_empresa.empresa_id)
        if empresa_usuario:
            empresa_cache_salvar(empresa_usuario)
        return RedirectResponse("/painel", status_code=303)

    return RedirectResponse(
        "/empresa/login?erro=Usuário ou senha não encontrado. Confira o usuário, a senha e se o celular está acessando o endereço correto da rede local.",
        status_code=303)


@app.get("/empresa/sair")
def empresa_sair(request: Request):
    request.session.clear()
    return RedirectResponse("/", status_code=303)


@app.get("/admin/setup")
def setup_antigo():
    return RedirectResponse("/admin/login", status_code=303)


def configurar_campos_empresa(db: Session, empresa_id: int):
    campos = db.query(CampoGlobal).all()
    obrigatorios = {"telefone", "nome", "bairro", "endereco", "numero", "data_evento", "hora_inicio"}
    for ordem, campo in enumerate(campos, start=1):
        existe = db.query(CampoEmpresa).filter_by(empresa_id=empresa_id, campo_id=campo.id).first()
        if not existe:
            visivel = campo.chave not in ["hora_fim"]
            db.add(CampoEmpresa(empresa_id=empresa_id, campo_id=campo.id, ordem=ordem, visivel=visivel,
                                obrigatorio=campo.chave in obrigatorios))
    db.commit()


def criar_modelos_iniciais_empresa(db: Session, empresa: Empresa):
    """Cria produto, contrato e mensagens padrão para a empresa não começar vazia."""
    contrato = db.query(Contrato).filter_by(empresa_id=empresa.id).first()
    if not contrato:
        contrato = Contrato(
            empresa_id=empresa.id,
            nome="Contrato padrão de locação",
            descricao="Modelo inicial pronto para editar.",
            clausulas="""CONTRATO DE LOCAÇÃO DE EQUIPAMENTOS

A LOCADORA disponibilizará ao CLIENTE os equipamentos e serviços combinados para a data do evento.

O CLIENTE declara que recebeu a proposta com descrição dos itens, endereço, horário, valor total e condições de pagamento antes do aceite.

A reserva somente será considerada confirmada após o aceite digital e, quando exigido, após a confirmação do pagamento do sinal.

O CLIENTE se compromete a informar corretamente endereço, acesso ao local, responsável pelo recebimento e qualquer restrição de entrega, como escadas, elevador, horário de carga e descarga ou necessidade de autorização.

A LOCADORA poderá cancelar ou reagendar a reserva caso as informações do local impeçam a entrega segura dos equipamentos.

Este é um contrato fictício inicial. Edite este texto conforme a política da empresa."""
        )
        db.add(contrato)
        db.commit()
        db.refresh(contrato)

    produto = db.query(ProdutoServico).filter_by(empresa_id=empresa.id).first()
    if not produto:
        db.add(ProdutoServico(
            empresa_id=empresa.id,
            contrato_id=None,
            nome="Jukebox Básico - exemplo",
            descricao="1 jukebox, 2 caixas, 2 microfones e cabos. Edite ou exclua este exemplo.",
            quantidade_disponivel=1,
            valor_base=0,
            duracao_minutos=240,
            ativo=True
        ))
        db.commit()

    mensagens = mensagens_empresa(empresa)
    empresa.mensagem_reserva = empresa.mensagem_reserva or mensagens["reserva"]
    empresa.mensagem_aceite = empresa.mensagem_aceite or mensagens["aceite"]
    empresa.mensagem_confirmacao = empresa.mensagem_confirmacao or mensagens["confirmacao"]
    empresa.mensagem_hora_fim = empresa.mensagem_hora_fim or mensagens["hora_fim"]
    if empresa.mostrar_mensagem_hora_fim is None:
        empresa.mostrar_mensagem_hora_fim = True
    db.commit()


@app.get("/painel/acesso-negado", response_class=HTMLResponse)
def acesso_negado(request: Request, area: str = "", empresa: Empresa = Depends(empresa_logada)):
    nomes = {
        "agenda": "Agenda", "operacao": "Operação", "buscar_cliente": "Buscar cliente",
        "financeiro": "Financeiro", "cadastros": "Cadastros", "relatorios": "Relatórios"
    }
    return templates.TemplateResponse("admin/acesso_negado.html", {
        "request": request, "empresa": empresa, "area": nomes.get(area, area)
    }, status_code=403)


@app.get("/painel/relatorios", response_class=HTMLResponse)
def relatorios(request: Request, empresa: Empresa = Depends(empresa_logada)):
    return templates.TemplateResponse("admin/relatorios.html", {"request": request, "empresa": empresa})


@app.get("/painel", response_class=HTMLResponse)
def painel(request: Request, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    """Home leve: somente leitura, agregações consolidadas e relacionamentos pré-carregados."""
    hoje = date.today()
    inicio_semana, fim_semana = periodo_semana_atual()
    status_pendentes = ["reserva", "pre_reserva", "contrato_enviado", "aguardando_aceite"]
    status_agenda_inativos = {"aguardando_nova_data", "cancelada", "cancelado_cliente", "rejeitada"}
    competencia_humiat = agora_utc().strftime("%Y-%m")

    with perf_stage("home.resumo_solicitacoes"):
        resumo_solicitacoes = db.query(
            func.sum(case((Solicitacao.status.in_(status_pendentes), 1), else_=0)).label("pendentes"),
            func.sum(case((
                (Solicitacao.data_evento >= inicio_semana)
                & (Solicitacao.data_evento <= fim_semana)
                & (~Solicitacao.status.in_(status_agenda_inativos)), 1
            ), else_=0)).label("agenda_periodo"),
            func.sum(case((
                (Solicitacao.humiat_processado == True)
                & (Solicitacao.humiat_competencia == competencia_humiat), 1
            ), else_=0)).label("humiat_aceitos"),
            func.sum(case((
                (Solicitacao.humiat_processado == True)
                & (Solicitacao.humiat_competencia == competencia_humiat)
                & (Solicitacao.humiat_status == "gratuito"), 1
            ), else_=0)).label("humiat_gratis"),
            func.sum(case((
                (Solicitacao.humiat_processado == True)
                & (Solicitacao.humiat_competencia == competencia_humiat)
                & (Solicitacao.humiat_status.in_(["debitado", "pendente_saldo"])), 1
            ), else_=0)).label("humiat_cobrados"),
        ).filter(Solicitacao.empresa_id == empresa.id).one()

    with perf_stage("home.resumo_operacao"):
        resumo_operacao = db.query(
            func.sum(case((Agenda.tipo_evento == "entrega", 1), else_=0)).label("entregas"),
            func.sum(case((Agenda.tipo_evento == "retirada", 1), else_=0)).label("retiradas"),
        ).filter(
            Agenda.empresa_id == empresa.id,
            Agenda.data >= inicio_semana,
            Agenda.data <= fim_semana,
            Agenda.status_operacional != "concluido",
        ).one()

    with perf_stage("home.totais_cadastros"):
        total_clientes = db.query(func.count(Cliente.id)).filter(Cliente.empresa_id == empresa.id).scalar() or 0
        total_produtos = db.query(func.count(ProdutoServico.id)).filter(ProdutoServico.empresa_id == empresa.id).scalar() or 0

    with perf_stage("home.solicitacoes_pendentes"):
        solicitacoes = (
            db.query(Solicitacao)
            .options(joinedload(Solicitacao.cliente), selectinload(Solicitacao.pagamentos))
            .filter(Solicitacao.empresa_id == empresa.id, Solicitacao.status.in_(status_pendentes))
            .order_by(Solicitacao.data_evento.asc(), Solicitacao.hora_inicio.asc())
            .limit(8)
            .all()
        )

    pendencias_agenda = solicitacoes

    with perf_stage("home.pendencias_contrato"):
        pendencias_sinal = []
        if empresa.exige_sinal:
            pendencias_sinal = (
                db.query(Solicitacao)
                .options(joinedload(Solicitacao.cliente))
                .filter(
                    Solicitacao.empresa_id == empresa.id,
                    Solicitacao.status.in_(["aceito", "aguardando_pagamento", "reserva_confirmada"]),
                    Solicitacao.sinal > 0,
                    Solicitacao.valor_pago <= 0,
                )
                .order_by(Solicitacao.data_evento.asc(), Solicitacao.hora_inicio.asc())
                .limit(12)
                .all()
            )

        pendencias_envio_contrato = (
            db.query(Solicitacao)
            .options(joinedload(Solicitacao.cliente))
            .filter(
                Solicitacao.empresa_id == empresa.id,
                Solicitacao.status.in_(["aceito", "aguardando_pagamento", "reserva_confirmada"]),
                Solicitacao.contrato_id.isnot(None),
                Solicitacao.contrato_enviado_em.is_(None),
                Solicitacao.cancelado_em.is_(None),
            )
            .order_by(Solicitacao.data_evento.asc(), Solicitacao.id.asc())
            .limit(12)
            .all()
        )

    with perf_stage("home.pendencias_financeiro_operacao"):
        pendencias_a_receber = (
            db.query(Solicitacao)
            .options(joinedload(Solicitacao.cliente))
            .filter(
                Solicitacao.empresa_id == empresa.id,
                Solicitacao.cancelado_em.is_(None),
                Solicitacao.status.in_(STATUS_CONTRATO_APROVADO),
                Solicitacao.data_evento < hoje,
                (func.coalesce(Solicitacao.valor, 0) - func.coalesce(Solicitacao.valor_pago, 0)) > 0.009,
            )
            .order_by(Solicitacao.data_evento.asc(), Solicitacao.id.asc())
            .limit(12)
            .all()
        )

        pendencias_operacao = (
            db.query(Agenda)
            .join(Solicitacao)
            .options(joinedload(Agenda.solicitacao))
            .filter(
                Agenda.empresa_id == empresa.id,
                Agenda.data < hoje,
                Agenda.status_operacional != "concluido",
                ~Solicitacao.status.in_(status_agenda_inativos),
            )
            .order_by(Agenda.data.asc(), Agenda.hora_inicio.asc())
            .limit(12)
            .all()
        )

        pendencias_financeiras = (
            db.query(Pagamento)
            .options(joinedload(Pagamento.solicitacao).joinedload(Solicitacao.cliente))
            .filter(Pagamento.empresa_id == empresa.id, Pagamento.conciliado_em.is_(None))
            .order_by(Pagamento.data_pagamento.asc(), Pagamento.id.asc())
            .limit(12)
            .all()
        )

    _anexar_responsaveis_exibicao(solicitacoes)

    link_pre_contrato = f"{str(request.base_url).rstrip('/')}/e/{empresa.slug}/pre-contrato"
    mensagem_pre_contrato = aplicar_variaveis_mensagem(
        mensagens_empresa(empresa).get("reserva", ""),
        link=link_pre_contrato,
        empresa=empresa.nome,
        cliente="",
        valor_sinal="",
        pix=empresa.pix_copia_cola or "",
    )

    pendentes = int(resumo_solicitacoes.pendentes or 0)
    agenda_periodo_qtd = int(resumo_solicitacoes.agenda_periodo or 0)
    operacao_entregar_qtd = int(resumo_operacao.entregas or 0)
    operacao_buscar_qtd = int(resumo_operacao.retiradas or 0)
    operacao_periodo_qtd = operacao_entregar_qtd + operacao_buscar_qtd

    aceitos_humiat_mes = int(resumo_solicitacoes.humiat_aceitos or 0)
    gratis_limite = max(0, int(empresa.humiat_gratis_mes or 4))
    gratis_usados = int(resumo_solicitacoes.humiat_gratis or 0)
    gratis_restantes = max(0, gratis_limite - gratis_usados)
    contratos_cobrados_mes = int(resumo_solicitacoes.humiat_cobrados or 0)

    return templates.TemplateResponse("admin/painel.html", {
        "request": request,
        "empresa": empresa,
        "mensagem_pre_contrato": mensagem_pre_contrato,
        "solicitacoes": solicitacoes,
        "pendencias_agenda": pendencias_agenda,
        "pendencias_sinal": pendencias_sinal,
        "pendencias_envio_contrato": pendencias_envio_contrato,
        "pendencias_a_receber": pendencias_a_receber,
        "pendencias_operacao": pendencias_operacao,
        "pendencias_financeiras": pendencias_financeiras,
        "total_clientes": int(total_clientes),
        "total_produtos": int(total_produtos),
        "pendentes": pendentes,
        "agenda_periodo_qtd": agenda_periodo_qtd,
        "operacao_periodo_qtd": operacao_periodo_qtd,
        "operacao_entregar_qtd": operacao_entregar_qtd,
        "operacao_buscar_qtd": operacao_buscar_qtd,
        "inicio_semana": inicio_semana,
        "fim_semana": fim_semana,
        "humiat_aceitos_mes": aceitos_humiat_mes,
        "humiat_gratis_limite": gratis_limite,
        "humiat_gratis_usados": gratis_usados,
        "humiat_gratis_restantes": gratis_restantes,
        "humiats_gratis_restantes": gratis_restantes,
        "humiat_contratos_cobrados_mes": contratos_cobrados_mes,
        "humiats_consumidos_mes": aceitos_humiat_mes,
        "usuario_online": request.session.get("usuario_nome") or request.session.get("usuario") or "Usuário",
    })


@app.get("/painel/humiats/extrato", response_class=HTMLResponse)
def painel_humiats_extrato(request: Request, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    movimentos = db.query(HumiatMovimento).filter_by(empresa_id=empresa.id).order_by(
        HumiatMovimento.criado_em.desc(), HumiatMovimento.id.desc()
    ).limit(300).all()
    contratos_humiat = db.query(Solicitacao).filter(
        Solicitacao.empresa_id == empresa.id,
        Solicitacao.humiat_processado == True,
    ).order_by(Solicitacao.aceite_em.desc(), Solicitacao.id.desc()).limit(500).all()
    total_humiats_usados = len(contratos_humiat)
    total_faturado = sum(float(c.valor or 0) for c in contratos_humiat)
    comprados_consumidos = sum(1 for c in contratos_humiat if c.humiat_status in ("debitado", "pendente_saldo"))
    gratis_consumidos = sum(1 for c in contratos_humiat if c.humiat_status == "gratuito")
    return templates.TemplateResponse("admin/humiat_extrato.html", {
        "request": request, "empresa": empresa, "movimentos": movimentos,
        "contratos_humiat": contratos_humiat,
        "total_humiats_usados": total_humiats_usados,
        "total_faturado": total_faturado,
        "comprados_consumidos": comprados_consumidos,
        "gratis_consumidos": gratis_consumidos,
        "usuario_online": request.session.get("usuario_nome") or request.session.get("usuario") or "Usuário",
    })


@app.get("/painel/humiats/comprar", response_class=HTMLResponse)
def painel_humiats_comprar(request: Request, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    competencia = agora_utc().strftime("%Y-%m")
    aceitos_mes = db.query(Solicitacao).filter(
        Solicitacao.empresa_id == empresa.id,
        Solicitacao.humiat_processado == True,
        Solicitacao.humiat_competencia == competencia,
    ).count()
    gratis_limite = max(0, int(empresa.humiat_gratis_mes or 4))
    custo = 1
    gratis_restantes = max(0, gratis_limite - min(aceitos_mes, gratis_limite))
    return templates.TemplateResponse("admin/humiat_comprar.html", {
        "request": request, "empresa": empresa,
        "gratis_restantes": gratis_restantes, "custo": custo,
        "usuario_online": request.session.get("usuario_nome") or request.session.get("usuario") or "Usuário",
    })


def usuario_empresa_atual(db: Session, empresa: Empresa, request: Request):
    usuario_sessao = (request.session.get("usuario_sistema") or request.session.get("usuario") or "").strip()
    usuario_busca = usuario_sessao.lower()
    usuario = None
    if usuario_busca:
        usuario = (
            db.query(UsuarioEmpresa)
            .filter(
                UsuarioEmpresa.empresa_id == empresa.id,
                func.lower(UsuarioEmpresa.usuario) == usuario_busca,
                UsuarioEmpresa.ativo == True,
            )
            .first()
        )
    if usuario:
        return "usuario", usuario
    return "admin", empresa


@app.get("/painel/perfil", response_class=HTMLResponse)
def perfil_usuario(request: Request, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    tipo, usuario = usuario_empresa_atual(db, empresa, request)
    perfil_nome = usuario.nome if tipo == "usuario" else (request.session.get("usuario_nome") or empresa.usuario_admin or "Administrador")
    perfil_usuario_valor = usuario.usuario if tipo == "usuario" else (empresa.usuario_admin or request.session.get("usuario_sistema") or "")
    return templates.TemplateResponse("admin/perfil.html", {
        "request": request,
        "empresa": empresa,
        "perfil_nome": perfil_nome,
        "perfil_usuario": perfil_usuario_valor,
        "erro": request.query_params.get("erro"),
        "sucesso": request.query_params.get("sucesso"),
    })


@app.post("/painel/perfil")
def salvar_perfil_usuario(
        request: Request,
        nome: str = Form(...),
        usuario: str = Form(...),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)):
    nome_limpo = nome.strip()
    usuario_limpo = usuario.strip()
    if not nome_limpo or not usuario_limpo:
        return RedirectResponse("/painel/perfil?erro=Informe nome e usuário.", status_code=303)

    tipo, registro = usuario_empresa_atual(db, empresa, request)
    usuario_busca = usuario_limpo.lower()

    empresa_com_usuario = (
        db.query(Empresa)
        .filter(func.lower(Empresa.usuario_admin) == usuario_busca, Empresa.id != empresa.id)
        .first()
    )
    usuario_com_usuario = (
        db.query(UsuarioEmpresa)
        .filter(func.lower(UsuarioEmpresa.usuario) == usuario_busca)
        .first()
    )
    if empresa_com_usuario or (usuario_com_usuario and (tipo != "usuario" or usuario_com_usuario.id != registro.id)):
        return RedirectResponse("/painel/perfil?erro=Este usuário já está em uso.", status_code=303)

    if tipo == "usuario":
        registro.nome = nome_limpo
        registro.usuario = usuario_limpo
        request.session["usuario_nome"] = nome_limpo
        request.session["usuario_sistema"] = usuario_limpo
    else:
        empresa.usuario_admin = usuario_limpo
        request.session["usuario_nome"] = nome_limpo
        request.session["usuario_sistema"] = usuario_limpo

    db.commit()
    return RedirectResponse("/painel/perfil?sucesso=Perfil atualizado com sucesso.", status_code=303)


@app.get("/painel/alterar-senha", response_class=HTMLResponse)
def alterar_senha_form(request: Request, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    return templates.TemplateResponse("admin/alterar_senha.html", {
        "request": request,
        "empresa": empresa,
        "erro": request.query_params.get("erro"),
        "sucesso": request.query_params.get("sucesso"),
    })


@app.post("/painel/alterar-senha")
def alterar_senha_salvar(
        request: Request,
        senha_atual: str = Form(...),
        nova_senha: str = Form(...),
        confirmar_senha: str = Form(...),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)):
    senha_atual = senha_atual.strip()
    nova_senha = nova_senha.strip()
    confirmar_senha = confirmar_senha.strip()
    if len(nova_senha) < 6:
        return RedirectResponse("/painel/alterar-senha?erro=A nova senha precisa ter pelo menos 6 caracteres.", status_code=303)
    if nova_senha != confirmar_senha:
        return RedirectResponse("/painel/alterar-senha?erro=A confirmação da senha não confere.", status_code=303)

    tipo, registro = usuario_empresa_atual(db, empresa, request)
    senha_cadastrada = registro.senha if tipo == "usuario" else empresa.senha_admin
    if senha_atual != (senha_cadastrada or ""):
        return RedirectResponse("/painel/alterar-senha?erro=Senha atual incorreta.", status_code=303)

    if tipo == "usuario":
        registro.senha = nova_senha
    else:
        empresa.senha_admin = nova_senha
    db.commit()
    return RedirectResponse("/painel/alterar-senha?sucesso=Senha alterada com sucesso.", status_code=303)


@app.get("/painel/configuracoes", response_class=HTMLResponse)
def configuracoes_empresa(request: Request, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    mensagens_padrao = mensagens_empresa(empresa)
    campos = db.query(CampoEmpresa).join(CampoGlobal).filter(CampoEmpresa.empresa_id == empresa.id).order_by(
        CampoEmpresa.ordem).all()
    return templates.TemplateResponse("admin/configuracoes.html",
                                      {"request": request, "empresa": empresa, "mensagens_padrao": mensagens_padrao,
                                       "campos": campos})


@app.post("/painel/configuracoes")
async def salvar_configuracoes_empresa(
        request: Request,
        pix_copia_cola: str = Form(""),
        whatsapp_retorno: str = Form(""),
        exige_sinal: Optional[str] = Form(None),
        suporte_inicio: str = Form(""),
        suporte_fim: str = Form(""),
        mostrar_suporte_contrato: Optional[str] = Form(None),
        logo_url: str = Form(""),
        logo_idb_url: str = Form(""),
        logo_arquivo: UploadFile | None = File(None),
        tema: str = Form("azul"),
        mensagem_reserva: str = Form(""),
        mensagem_aceite: str = Form(""),
        mensagem_pagamento: str = Form(""),
        mensagem_confirmacao: str = Form(""),
        mensagem_hora_fim: str = Form(""),
        mostrar_mensagem_hora_fim: Optional[str] = Form(None),
        mensagem_preparacao: str = Form(""),
        mensagem_a_caminho: str = Form(""),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    empresa.pix_copia_cola = pix_copia_cola.strip()
    empresa.whatsapp_retorno = _limpar_tel_whatsapp(whatsapp_retorno)
    empresa.exige_sinal = bool(exige_sinal)
    empresa.suporte_inicio = suporte_inicio.strip()
    empresa.suporte_fim = suporte_fim.strip()
    empresa.mostrar_suporte_contrato = bool(mostrar_suporte_contrato)
    # Logo: o caminho mais simples para o locador é enviar do próprio PC/celular.
    # Mantemos URL apenas como alternativa técnica.
    if logo_arquivo and logo_arquivo.filename:
        extensao = Path(logo_arquivo.filename).suffix.lower()
        if extensao not in [".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg"]:
            raise HTTPException(400, "Formato de logo inválido. Use PNG, JPG, WEBP, GIF ou SVG.")
        nome_arquivo = f"empresa_{empresa.id}_{uuid.uuid4().hex}{extensao}"
        destino = Path("static/uploads/logos") / nome_arquivo
        with destino.open("wb") as buffer:
            shutil.copyfileobj(logo_arquivo.file, buffer)
        empresa.logo_url = f"/static/uploads/logos/{nome_arquivo}"
        empresa.logo_idb_url = ""
    elif logo_url.strip():
        empresa.logo_url = logo_url.strip()
        empresa.logo_idb_url = ""
    elif logo_idb_url.strip():
        empresa.logo_idb_url = logo_idb_url.strip()
        empresa.logo_url = ""
    empresa.tema = tema
    empresa.mensagem_reserva = mensagem_reserva.strip()
    empresa.mensagem_aceite = mensagem_aceite.strip()
    empresa.mensagem_confirmacao = mensagem_confirmacao.strip()
    empresa.mensagem_hora_fim = mensagem_hora_fim.strip()
    empresa.mostrar_mensagem_hora_fim = bool(mostrar_mensagem_hora_fim)
    empresa.mensagem_preparacao = mensagem_preparacao.strip()
    empresa.mensagem_a_caminho = mensagem_a_caminho.strip()
    form = await request.form()
    campos = db.query(CampoEmpresa).filter_by(empresa_id=empresa.id).all()
    for ce in campos:
        ce.visivel = f"campo_visivel_{ce.id}" in form
        ce.obrigatorio = f"campo_obrigatorio_{ce.id}" in form
    db.commit()
    return RedirectResponse("/painel", status_code=303)


@app.get("/painel/produtos", response_class=HTMLResponse)
def produtos(request: Request, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    produtos = db.query(ProdutoServico).filter_by(empresa_id=empresa.id).order_by(ProdutoServico.nome).all()
    contratos = db.query(Contrato).filter_by(empresa_id=empresa.id, ativo=True).order_by(Contrato.nome).all()
    return templates.TemplateResponse("admin/produtos.html",
                                      {"request": request, "empresa": empresa, "produtos": produtos, "produto": None,
                                       "contratos": contratos})


@app.get("/painel/produto/{produto_id}", response_class=HTMLResponse)
def produto_editar(produto_id: int, request: Request, db: Session = Depends(get_db),
                   empresa: Empresa = Depends(empresa_logada)):
    produto = db.get(ProdutoServico, produto_id)
    if not produto or produto.empresa_id != empresa.id:
        raise HTTPException(404)
    produtos = db.query(ProdutoServico).filter_by(empresa_id=empresa.id).order_by(ProdutoServico.nome).all()
    contratos = db.query(Contrato).filter_by(empresa_id=empresa.id, ativo=True).order_by(Contrato.nome).all()
    return templates.TemplateResponse("admin/produtos.html",
                                      {"request": request, "empresa": empresa, "produtos": produtos,
                                       "produto": produto, "contratos": contratos})


@app.post("/painel/produto/{produto_id_url}")
def salvar_produto_url(produto_id_url: int, nome: str = Form(...), descricao: str = Form(""),
                       quantidade_disponivel: int = Form(1), valor_base: str = Form("0"),
                       duracao_minutos: int = Form(240), prazo_retirada_dias: int = Form(1),
                       carga_pontos: int = Form(1), volume_logistico: int = Form(1),
                       permite_interno: bool = Form(False), permite_mala: bool = Form(False), permite_teto: bool = Form(False),
                       contrato_id: str = Form(""), db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    return salvar_produto(str(produto_id_url), nome, descricao, quantidade_disponivel, valor_base, duracao_minutos,
                          prazo_retirada_dias, carga_pontos, volume_logistico, permite_interno, permite_mala, permite_teto, contrato_id, db, empresa)


@app.post("/painel/produtos")
def salvar_produto(
        produto_id: str = Form(""),
        nome: str = Form(...), descricao: str = Form(""),
        quantidade_disponivel: int = Form(1), valor_base: str = Form("0"), duracao_minutos: int = Form(240),
        prazo_retirada_dias: int = Form(1), carga_pontos: int = Form(1), volume_logistico: int = Form(1),
        permite_interno: bool = Form(False), permite_mala: bool = Form(False), permite_teto: bool = Form(False),
        contrato_id: str = Form(""), db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)
):
    produto_id_int = int(produto_id) if produto_id else None
    produto = db.get(ProdutoServico, produto_id_int) if produto_id_int else None
    if not produto:
        produto = ProdutoServico(empresa_id=empresa.id)
        db.add(produto)
    produto.nome = nome.strip()
    produto.descricao = descricao
    contrato_id_int = int(contrato_id) if contrato_id and str(contrato_id).isdigit() else None
    contrato = db.get(Contrato, contrato_id_int) if contrato_id_int else None
    produto.contrato_id = contrato.id if contrato and contrato.empresa_id == empresa.id else None
    produto.quantidade_disponivel = quantidade_disponivel
    produto.valor_base = texto_para_float(valor_base)
    produto.duracao_minutos = duracao_minutos
    produto.prazo_retirada_dias = prazo_retirada_dias
    produto.carga_pontos = max(1, int(carga_pontos or 1))
    produto.volume_logistico = max(1, int(volume_logistico or 1))
    produto.permite_interno = bool(permite_interno)
    produto.permite_mala = bool(permite_mala)
    produto.permite_teto = bool(permite_teto)
    if not (produto.permite_interno or produto.permite_mala or produto.permite_teto):
        produto.permite_interno = True
    produto.tipo_locacao = "horas_fixas"
    db.commit()
    return RedirectResponse("/painel/produtos", status_code=303)


@app.get("/painel/produto/{produto_id}/copiar")
def copiar_produto(produto_id: int, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    origem = db.get(ProdutoServico, produto_id)
    if not origem or origem.empresa_id != empresa.id:
        raise HTTPException(404)
    novo = ProdutoServico(empresa_id=empresa.id, contrato_id=origem.contrato_id, nome=f"{origem.nome} - cópia",
                          descricao=origem.descricao, quantidade_disponivel=origem.quantidade_disponivel,
                          valor_base=origem.valor_base, duracao_minutos=origem.duracao_minutos,
                          prazo_retirada_dias=origem.prazo_retirada_dias, carga_pontos=origem.carga_pontos or 1,
                          volume_logistico=origem.volume_logistico or 1, permite_interno=origem.permite_interno,
                          permite_mala=origem.permite_mala, permite_teto=origem.permite_teto, ativo=True)
    db.add(novo)
    db.commit()
    db.refresh(novo)
    return RedirectResponse(f"/painel/produto/{novo.id}", status_code=303)


@app.get("/painel/contratos", response_class=HTMLResponse)
def contratos(request: Request, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    contratos = db.query(Contrato).filter_by(empresa_id=empresa.id).order_by(Contrato.nome).all()
    return templates.TemplateResponse("admin/contratos.html",
                                      {"request": request, "empresa": empresa, "contratos": contratos,
                                       "contrato": None})


@app.get("/painel/contrato/{contrato_id}", response_class=HTMLResponse)
def contrato_editar(contrato_id: int, request: Request, db: Session = Depends(get_db),
                    empresa: Empresa = Depends(empresa_logada)):
    contrato = db.get(Contrato, contrato_id)
    if not contrato or contrato.empresa_id != empresa.id:
        raise HTTPException(404)
    contratos = db.query(Contrato).filter_by(empresa_id=empresa.id).order_by(Contrato.nome).all()
    return templates.TemplateResponse("admin/contratos.html",
                                      {"request": request, "empresa": empresa, "contratos": contratos,
                                       "contrato": contrato})


@app.post("/painel/contrato/{contrato_id}")
@app.post("/painel/contratos")
def salvar_contrato(
        contrato_id: int | None = None,
        contrato_id_form: str = Form("", alias="contrato_id"),
        nome: str = Form(...), descricao: str = Form(""), clausulas: str = Form(...),
        db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)
):
    contrato_id_final = contrato_id or (int(contrato_id_form) if contrato_id_form else None)
    contrato = db.get(Contrato, contrato_id_final) if contrato_id_final else None
    if contrato and contrato.empresa_id != empresa.id:
        raise HTTPException(404)
    if not contrato:
        contrato = Contrato(empresa_id=empresa.id)
        db.add(contrato)
    contrato.nome = nome.strip()
    contrato.descricao = descricao
    contrato.clausulas = clausulas
    db.commit()
    return RedirectResponse("/painel/contratos", status_code=303)


@app.get("/painel/contrato/{contrato_id}/copiar")
def copiar_contrato(contrato_id: int, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    origem = db.get(Contrato, contrato_id)
    if not origem or origem.empresa_id != empresa.id:
        raise HTTPException(404)
    novo = Contrato(
        empresa_id=empresa.id,
        nome=f"{origem.nome} - cópia",
        descricao=origem.descricao,
        clausulas=origem.clausulas,
        ativo=True
    )
    db.add(novo)
    db.commit()
    db.refresh(novo)
    return RedirectResponse(f"/painel/contrato/{novo.id}", status_code=303)


def usuario_pode_ver_nao_roteirizados(request: Request, db: Session) -> bool:
    """Administrador vê tudo; usuário comum depende da permissão cadastrada."""
    if request.session.get("acesso_total"):
        return True
    usuario_id = request.session.get("usuario_empresa_id")
    if not usuario_id:
        return False
    usuario = db.get(UsuarioEmpresa, usuario_id)
    return bool(usuario and usuario.ativo and usuario.acesso_nao_roteirizados)


def equipes_visiveis_usuario(request: Request, db: Session, empresa_id: int):
    q = db.query(Equipe).filter(Equipe.empresa_id == empresa_id, Equipe.ativa == True)
    if request.session.get("acesso_total"):
        return q.order_by(Equipe.nome).all()
    usuario_id = request.session.get("usuario_empresa_id")
    if not usuario_id:
        return []
    return q.join(UsuarioEquipe, UsuarioEquipe.equipe_id == Equipe.id).filter(UsuarioEquipe.usuario_id == usuario_id).order_by(Equipe.nome).all()


@app.get("/painel/reservas", response_class=HTMLResponse)
def preparar_reservas(
        request: Request,
        data_inicial: str = "",
        data_final: str = "",
        data_inicio: str = "",
        data_fim: str = "",
        mostrar_entregas: str = "",
        mostrar_retiradas: str = "",
        mostrar_concluidas: str = "",
        equipe_id: int = 0,
        situacao_rota: str = "todos",
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    # Consulta operacional deve ser somente leitura. Correções históricas ficam fora do GET.
    inicio, fim = periodo_semana_atual()

    # Compatibilidade com links antigos que usam data_inicio/data_fim.
    # A tela oficial continua trabalhando com data_inicial/data_final.
    data_inicial = data_inicial or data_inicio or inicio.isoformat()
    data_final = data_final or data_fim or fim.isoformat()

    # Nunca executar uma consulta com o período invertido. Ao trocar apenas
    # uma das datas no filtro, ajusta a outra para o mesmo dia.
    try:
        inicio_filtro = datetime.strptime(data_inicial, "%Y-%m-%d").date()
        fim_filtro = datetime.strptime(data_final, "%Y-%m-%d").date()
    except ValueError:
        inicio_filtro, fim_filtro = inicio, fim
        data_inicial, data_final = inicio.isoformat(), fim.isoformat()
    if fim_filtro < inicio_filtro:
        fim_filtro = inicio_filtro
        data_final = data_inicial

    equipes = equipes_visiveis_usuario(request, db, empresa.id)
    ids_equipes = {e.id for e in equipes}
    filtro_salvo = request.session.get("operacao_equipe_id", 0)
    if equipe_id not in ids_equipes:
        equipe_id = filtro_salvo if filtro_salvo in ids_equipes else (equipes[0].id if equipes else 0)
    if equipe_id:
        request.session["operacao_equipe_id"] = equipe_id

    # Checkbox desmarcado não vem no GET. Se for o primeiro acesso da tela,
    # começa com Entregar e Retirar ligados. Depois disso, respeita exatamente
    # o que o usuário marcou/desmarcou.
    query = request.query_params
    if not query:
        mostrar_entregas = "1"
        mostrar_retiradas = "1"
        mostrar_concluidas = ""
    else:
        mostrar_entregas = "1" if "mostrar_entregas" in query else ""
        mostrar_retiradas = "1" if "mostrar_retiradas" in query else ""
        mostrar_concluidas = "1" if "mostrar_concluidas" in query else ""

    q = db.query(Agenda).filter_by(empresa_id=empresa.id)
    pode_ver_nao_roteirizados = usuario_pode_ver_nao_roteirizados(request, db)
    if equipe_id:
        # Usuário restrito só passa a enxergar o card quando ele for roteirizado
        # para uma de suas equipes. Cards sem rota ficam apenas para quem possui
        # a permissão "Roteirizados e não roteirizados".
        if pode_ver_nao_roteirizados:
            q = q.filter(
                (Agenda.equipe_id == equipe_id) |
                (Agenda.roteirizado == False) |
                (Agenda.roteirizado == None)
            )
        else:
            q = q.filter(
                Agenda.equipe_id == equipe_id,
                Agenda.roteirizado == True
            )
    if situacao_rota == "roteirizado":
        q = q.filter(Agenda.roteirizado == True)
    elif situacao_rota == "nao_roteirizado":
        q = q.filter((Agenda.roteirizado == False) | (Agenda.roteirizado == None))
    if data_inicial:
        q = q.filter(Agenda.data >= inicio_filtro)
    if data_final:
        q = q.filter(Agenda.data <= fim_filtro)

    tipos = []
    if mostrar_entregas:
        tipos.append("entrega")
    if mostrar_retiradas:
        tipos.append("retirada")
    if tipos:
        q = q.filter(Agenda.tipo_evento.in_(tipos))
    else:
        q = q.filter(text("1=0"))

    if not mostrar_concluidas:
        q = q.filter(Agenda.status_operacional != "concluido")

    # Regra central: a Operação trabalha exclusivamente com contratos aprovados/aceitos.
    # Rascunhos, aguardando aceite, crédito, cancelados e qualquer outro status ficam fora no backend.
    q = q.join(Solicitacao, Agenda.solicitacao_id == Solicitacao.id).filter(
        Solicitacao.status.in_(STATUS_CONTRATO_APROVADO),
    )
    itens = (
        q.options(
            joinedload(Agenda.equipe),
            joinedload(Agenda.solicitacao).joinedload(Solicitacao.cliente),
            joinedload(Agenda.solicitacao).joinedload(Solicitacao.produto),
            joinedload(Agenda.solicitacao).selectinload(Solicitacao.itens),
            joinedload(Agenda.solicitacao).selectinload(Solicitacao.pagamentos),
        )
        .join(Cliente, Solicitacao.cliente_id == Cliente.id)
        .all()
    )
    # valor_pago já é mantido quando o pagamento é gravado. Não recalcular contrato a contrato aqui.

    def hora_roteirizada(a: Agenda):
        """
        Retorna a hora que aparece no card da operação.

        Registros antigos podem ter previsao_entrega diferente de hora_inicio.
        A ordenação precisa usar exatamente a hora roteirizada exibida na tela,
        nunca posição salva, ordem de criação ou nome do cliente.
        """
        if a.roteirizado:
            previsao = (a.previsao_entrega or "").strip()
            if previsao:
                try:
                    return datetime.strptime(previsao, "%H:%M").time()
                except ValueError:
                    pass
            if a.hora_inicio:
                return a.hora_inicio

        sol = a.solicitacao
        return sol.hora_inicio if sol and sol.hora_inicio else (a.hora_inicio or time.max)

    def chave_operacao(a: Agenda):
        sol = a.solicitacao
        data_base = a.data if a.roteirizado and a.data else (sol.data_evento if sol else a.data)
        return (data_base or date.max, hora_roteirizada(a), a.id)

    itens = sorted(itens, key=chave_operacao)

    # Mapa de vínculo entre ENTREGA e RETIRADA da mesma solicitação.
    # A consulta ignora os filtros da tela para permitir localizar também
    # operações concluídas ou fora do período atualmente selecionado.
    ids_solicitacoes = {a.solicitacao_id for a in itens if a.solicitacao_id}
    operacoes_vinculadas = {}
    if ids_solicitacoes:
        eventos_relacionados = (
            db.query(Agenda)
            .filter(
                Agenda.empresa_id == empresa.id,
                Agenda.solicitacao_id.in_(ids_solicitacoes),
            )
            .order_by(Agenda.id.asc())
            .all()
        )
        por_solicitacao = {}
        for evento in eventos_relacionados:
            por_solicitacao.setdefault(evento.solicitacao_id, []).append(evento)

        for agenda in itens:
            tipo_atual = agenda.tipo_evento or "entrega"
            tipo_procurado = "retirada" if tipo_atual == "entrega" else "entrega"
            candidatos = [
                evento for evento in por_solicitacao.get(agenda.solicitacao_id, [])
                if (evento.tipo_evento or "entrega") == tipo_procurado
            ]
            # Em bases antigas pode haver duplicidade. Exibimos o registro mais
            # recente, que é o melhor candidato para auditoria operacional.
            operacoes_vinculadas[agenda.id] = candidatos[-1] if candidatos else None

    return templates.TemplateResponse("admin/preparar.html", {
        "request": request,
        "empresa": empresa,
        "itens": itens,
        "total_itens": len(itens),
        "data_inicial": data_inicial,
        "data_final": data_final,
        "mostrar_entregas": mostrar_entregas,
        "mostrar_retiradas": mostrar_retiradas,
        "mostrar_concluidas": mostrar_concluidas,
        "equipes": equipes, "equipe_id": equipe_id, "situacao_rota": situacao_rota,
        "mensagens": mensagens_empresa(empresa),
        "operacoes_vinculadas": operacoes_vinculadas,
    })


@app.get("/painel/agenda/{agenda_id}/localizar-vinculada")
def localizar_operacao_vinculada(
        agenda_id: int,
        request: Request,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada),
):
    """Abre, na tela Operação, a entrega ou retirada vinculada ao mesmo contrato."""
    agenda = db.query(Agenda).filter_by(id=agenda_id, empresa_id=empresa.id).first()
    if not agenda:
        raise HTTPException(404, "Operação não encontrada.")

    tipo_atual = agenda.tipo_evento or "entrega"
    tipo_procurado = "retirada" if tipo_atual == "entrega" else "entrega"
    vinculada = (
        db.query(Agenda)
        .filter(
            Agenda.empresa_id == empresa.id,
            Agenda.solicitacao_id == agenda.solicitacao_id,
            Agenda.tipo_evento == tipo_procurado,
        )
        .order_by(Agenda.id.desc())
        .first()
    )
    if not vinculada:
        raise HTTPException(
            404,
            "A operação vinculada ainda não existe no banco de dados."
        )

    data_vinculada = vinculada.data or (
        vinculada.solicitacao.data_evento if vinculada.solicitacao else None
    )
    data_texto = data_vinculada.isoformat() if data_vinculada else date.today().isoformat()
    parametros = {
        "data_inicial": data_texto,
        "data_final": data_texto,
        "mostrar_entregas": "1",
        "mostrar_retiradas": "1",
        "mostrar_concluidas": "1",
        "situacao_rota": "todos",
    }
    if vinculada.equipe_id:
        parametros["equipe_id"] = str(vinculada.equipe_id)

    return RedirectResponse(
        url=f"/painel/reservas?{urlencode(parametros)}#agenda-{vinculada.id}",
        status_code=303,
    )


@app.get("/painel/solicitacoes", response_class=HTMLResponse)
def solicitacoes(request: Request, busca: str = "", db: Session = Depends(get_db),
                 empresa: Empresa = Depends(empresa_logada)):
    q = db.query(Solicitacao).filter_by(empresa_id=empresa.id)
    termo = limpar_identificador(busca)
    if termo:
        q = q.join(Cliente).filter((Cliente.cpf.contains(termo)) | (Cliente.telefone.contains(termo)) | (
            Cliente.identificador.contains(termo)))
    itens = q.join(Cliente, Solicitacao.cliente_id == Cliente.id).order_by(Solicitacao.data_evento, Cliente.nome,
                                                                           Solicitacao.hora_inicio,
                                                                           Solicitacao.id).all()
    return templates.TemplateResponse("admin/solicitacoes.html",
                                      {"request": request, "empresa": empresa, "itens": itens, "busca": busca})


@app.get("/painel/solicitacao/{solicitacao_id}", response_class=HTMLResponse)
def detalhe_solicitacao(solicitacao_id: int, request: Request, db: Session = Depends(get_db),
                        empresa: Empresa = Depends(empresa_logada)):
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)
    sincronizar_pagamentos_solicitacoes(db, [item])
    produtos = db.query(ProdutoServico).filter_by(empresa_id=empresa.id, ativo=True).order_by(ProdutoServico.nome).all()
    contratos = db.query(Contrato).filter_by(empresa_id=empresa.id, ativo=True).order_by(Contrato.nome).all()
    empresas_transferencia = (
        db.query(Empresa)
        .filter(Empresa.ativa == True, Empresa.id != empresa.id)
        .order_by(Empresa.nome.asc())
        .all()
    )
    mensagens = mensagens_empresa(empresa)
    return templates.TemplateResponse("admin/solicitacao_detalhe.html",
                                      {"request": request, "item": item, "empresa": empresa, "produtos": produtos,
                                       "contratos": contratos, "empresas_transferencia": empresas_transferencia,
                                       "mensagens": mensagens})



def _copiar_cliente_para_empresa(db: Session, cliente_origem: Cliente, empresa_destino_id: int) -> Cliente:
    """Garante o mesmo cliente na empresa de destino sem misturar os cadastros das empresas."""
    cliente = db.query(Cliente).filter(
        Cliente.empresa_id == empresa_destino_id,
        Cliente.identificador == cliente_origem.identificador,
    ).first()
    if cliente:
        return cliente

    campos = (
        "identificador", "telefone", "cpf", "cnpj", "nome", "data_nascimento", "email",
        "endereco", "numero", "complemento", "bairro", "cidade", "estado", "cep", "observacoes",
    )
    cliente = Cliente(empresa_id=empresa_destino_id, **{
        campo: getattr(cliente_origem, campo) for campo in campos
    })
    db.add(cliente)
    db.flush()

    for endereco in cliente_origem.enderecos:
        db.add(EnderecoCliente(
            empresa_id=empresa_destino_id,
            cliente_id=cliente.id,
            apelido=endereco.apelido,
            endereco=endereco.endereco,
            numero=endereco.numero,
            complemento=endereco.complemento,
            bairro=endereco.bairro,
            cidade=endereco.cidade,
            estado=endereco.estado,
            cep=endereco.cep,
            ativo=endereco.ativo,
        ))
    return cliente


def _copiar_modelo_contrato_para_empresa(db: Session, contrato_origem: Contrato | None, empresa_destino_id: int) -> Contrato | None:
    """Copia o texto contratual para que a empresa de destino tenha um contrato próprio."""
    if not contrato_origem:
        return None
    contrato = db.query(Contrato).filter(
        Contrato.empresa_id == empresa_destino_id,
        Contrato.nome == contrato_origem.nome,
        Contrato.clausulas == contrato_origem.clausulas,
    ).first()
    if contrato:
        return contrato
    contrato = Contrato(
        empresa_id=empresa_destino_id,
        nome=contrato_origem.nome,
        descricao=contrato_origem.descricao,
        clausulas=contrato_origem.clausulas,
        ativo=contrato_origem.ativo,
    )
    db.add(contrato)
    db.flush()
    return contrato


def _sincronizar_copia_transferencia(db: Session, origem: Solicitacao, destino: Empresa) -> Solicitacao:
    """Cria/atualiza a cópia operacional da transferência sem duplicar recebimentos do cliente."""
    copia = db.get(Solicitacao, origem.transferencia_copia_id) if origem.transferencia_copia_id else None
    if copia and copia.empresa_id != destino.id:
        # O contrato mudou de empresa de destino. Preservamos o histórico da cópia anterior,
        # mas ela deixa de participar da operação.
        copia.cancelado_em = copia.cancelado_em or agora_utc()
        copia = None

    cliente_destino = _copiar_cliente_para_empresa(db, origem.cliente, destino.id)
    contrato_destino = _copiar_modelo_contrato_para_empresa(db, origem.contrato, destino.id)

    campos = (
        "data_evento", "hora_inicio", "hora_fim", "retirada_obrigatoria", "retirada_data",
        "retirada_hora", "bairro", "local", "local_numero", "local_complemento",
        "local_cidade", "local_estado", "local_cep", "local_nome", "local_responsavel_nome",
        "local_responsavel_telefone", "retirada_responsavel_nome", "retirada_responsavel_telefone",
        "acesso_local", "valor", "sinal", "valor_pago", "sinal_recebido", "pagamento_confirmado_em",
        "observacoes", "status", "aceite_em", "aprovado_em", "contrato_enviado_em",
        "responsavel_contrato", "responsavel_operacao",
    )
    dados = {campo: getattr(origem, campo) for campo in campos}

    if not copia:
        copia = Solicitacao(
            empresa_id=destino.id,
            cliente_id=cliente_destino.id,
            produto_id=None,
            contrato_id=contrato_destino.id if contrato_destino else None,
            transferencia_origem_id=origem.id,
            empresa_transferida_id=None,
            valor_repasse=0,
            transferida_em=origem.transferida_em,
            # Transferência interna não pode gerar uma segunda cobrança HUMIAT.
            humiat_processado=True,
            humiat_status="transferencia_interna",
            **dados,
        )
        db.add(copia)
        db.flush()
        origem.transferencia_copia_id = copia.id
    else:
        copia.cliente_id = cliente_destino.id
        copia.contrato_id = contrato_destino.id if contrato_destino else None
        copia.transferencia_origem_id = origem.id
        copia.cancelado_em = None
        for campo, valor in dados.items():
            setattr(copia, campo, valor)

    # Itens são cópias descritivas. Não vinculamos produto da empresa de origem
    # ao estoque da empresa de destino.
    db.query(ReservaItem).filter(ReservaItem.solicitacao_id == copia.id).delete(synchronize_session=False)
    for item_origem in origem.itens:
        db.add(ReservaItem(
            empresa_id=destino.id,
            solicitacao_id=copia.id,
            produto_id=None,
            nome=item_origem.nome,
            descricao=item_origem.descricao,
            quantidade=item_origem.quantidade,
            valor_unitario=item_origem.valor_unitario,
            valor_total=item_origem.valor_total,
        ))

    # Agenda/operação acompanha o contrato transferido.
    agenda_origem = origem.agenda
    agenda_destino = db.query(Agenda).filter(Agenda.solicitacao_id == copia.id).first()
    if agenda_origem:
        if not agenda_destino:
            agenda_destino = Agenda(
                empresa_id=destino.id,
                solicitacao_id=copia.id,
                data=agenda_origem.data,
                hora_inicio=agenda_origem.hora_inicio,
                hora_fim=agenda_origem.hora_fim,
                titulo=agenda_origem.titulo,
                bairro=agenda_origem.bairro,
                equipe_id=None,
                roteirizado=False,
                previsao_entrega=agenda_origem.previsao_entrega,
                link_localizacao=agenda_origem.link_localizacao,
                tipo_evento=agenda_origem.tipo_evento,
                status_operacional="pendente",
                observacoes_operacionais=agenda_origem.observacoes_operacionais,
            )
            db.add(agenda_destino)
        else:
            agenda_destino.data = agenda_origem.data
            agenda_destino.hora_inicio = agenda_origem.hora_inicio
            agenda_destino.hora_fim = agenda_origem.hora_fim
            agenda_destino.titulo = agenda_origem.titulo
            agenda_destino.bairro = agenda_origem.bairro
            agenda_destino.previsao_entrega = agenda_origem.previsao_entrega
            agenda_destino.link_localizacao = agenda_origem.link_localizacao
            agenda_destino.tipo_evento = agenda_origem.tipo_evento
    return copia


@app.post("/painel/solicitacao/{solicitacao_id}/transferir")
def transferir_solicitacao_empresa(
    solicitacao_id: int,
    empresa_destino_id: int = Form(0),
    valor_repasse: str = Form(""),
    db: Session = Depends(get_db),
    empresa: Empresa = Depends(empresa_logada),
):
    """Transfere o contrato preservando a origem e, quando possível, cria a cópia na empresa de destino."""
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)

    if not empresa_destino_id:
        # Remoção da transferência: não apagamos a cópia para preservar histórico; apenas a cancelamos.
        if item.transferencia_copia_id:
            copia = db.get(Solicitacao, item.transferencia_copia_id)
            if copia:
                copia.cancelado_em = copia.cancelado_em or agora_utc()
        item.transferencia_copia_id = None
        item.empresa_transferida_id = None
        item.valor_repasse = 0
        item.transferida_em = None
        item.repasse_pago_em = None
        item.repasse_pago_por = None
        db.query(LancamentoBanco).filter(LancamentoBanco.repasse_solicitacao_id == item.id).update({"repasse_solicitacao_id": None})
        db.query(LancamentoManualFinanceiro).filter(LancamentoManualFinanceiro.repasse_solicitacao_id == item.id).update({"repasse_solicitacao_id": None})
        db.commit()
        return RedirectResponse(f"/painel/solicitacao/{solicitacao_id}", status_code=303)

    destino = db.get(Empresa, empresa_destino_id)
    if not destino or not destino.ativa or destino.id == empresa.id:
        raise HTTPException(400, "Empresa de destino inválida.")

    destino_e_interno = (destino.nome or "").strip().lower() not in {"outros", "outro", "externo", "empresa externa"}

    if destino_e_interno:
        # Em transferência interna, o que a Empresa A deve à B é o valor que ela já recebeu
        # do cliente. Esse valor deixa de ser conta a receber do cliente na Empresa B.
        repasse = max(float(item.valor_pago or 0), 0)
        if repasse <= 0 and item.sinal_recebido:
            repasse = max(float(item.sinal or 0), 0)
    else:
        # Para "Outros"/externos, preserva exatamente a regra antiga.
        repasse = max(texto_para_float(valor_repasse or "0"), 0)
        if repasse <= 0:
            repasse = max(float(item.valor or 0), 0)

    mudou_destino_ou_valor = (
        item.empresa_transferida_id != destino.id
        or abs(float(item.valor_repasse or 0) - repasse) > 0.009
    )
    item.empresa_transferida_id = destino.id
    item.valor_repasse = repasse
    item.transferida_em = agora_utc()

    if destino_e_interno:
        _sincronizar_copia_transferencia(db, item, destino)
    elif item.transferencia_copia_id:
        # Se mudou de uma empresa interna para "Outros", desativa a antiga cópia.
        copia = db.get(Solicitacao, item.transferencia_copia_id)
        if copia:
            copia.cancelado_em = copia.cancelado_em or agora_utc()
        item.transferencia_copia_id = None

    if mudou_destino_ou_valor:
        item.repasse_pago_em = None
        item.repasse_pago_por = None
        db.query(LancamentoBanco).filter(LancamentoBanco.repasse_solicitacao_id == item.id).update({"repasse_solicitacao_id": None})
        db.query(LancamentoManualFinanceiro).filter(LancamentoManualFinanceiro.repasse_solicitacao_id == item.id).update({"repasse_solicitacao_id": None})

    db.commit()
    return RedirectResponse(f"/painel/solicitacao/{solicitacao_id}", status_code=303)


@app.get("/painel/solicitacao/{solicitacao_id}/whatsapp")
def compartilhar_aceite_whatsapp(
    solicitacao_id: int,
    request: Request,
    db: Session = Depends(get_db),
    empresa: Empresa = Depends(empresa_logada),
):
    """Envia o link de aceite ao cliente. Não envia o contrato final."""
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)

    if item.status == "aguardando_nova_data":
        return RedirectResponse(
            f"/painel/solicitacao/{solicitacao_id}?erro=Reative o contrato e revise o rascunho antes de enviar para aceite.",
            status_code=303,
        )

    telefone = _limpar_tel_whatsapp(item.cliente.telefone or item.cliente.identificador)
    if not telefone:
        raise HTTPException(400, "Cliente sem telefone para WhatsApp")

    alterou = False
    if item.status == "pre_reserva" and item.contrato_id and len(item.itens) > 0:
        item.status = "contrato_enviado"
        alterou = True

    # Para os novos contratos, quem fizer o primeiro envio para aceite passa a
    # ser o responsável pela comunicação comercial daquele contrato. Edições
    # posteriores não trocam o responsável.
    if not item.responsavel_contrato:
        item.responsavel_contrato = (
            request.session.get("usuario_nome")
            or request.session.get("usuario_sistema")
            or request.session.get("usuario")
            or "Usuário"
        )
        alterou = True
    if alterou:
        db.commit()

    texto = montar_mensagem_whatsapp_aceite(request, empresa, item, db)

    return RedirectResponse(
        f"https://wa.me/{telefone}?text={quote(texto)}",
        status_code=303,
    )


@app.get("/painel/solicitacao/{solicitacao_id}/whatsapp-contrato")
def compartilhar_contrato_whatsapp(
    solicitacao_id: int,
    request: Request,
    db: Session = Depends(get_db),
    empresa: Empresa = Depends(empresa_logada),
):
    """Envia o contrato final somente após aceite do cliente ou aceite manual."""
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)

    if not status_reserva_confirmada(item.status):
        return RedirectResponse(
            f"/painel/solicitacao/{solicitacao_id}?erro=O contrato final só pode ser enviado depois do aceite do cliente ou aceite manual.",
            status_code=303,
        )

    telefone = _limpar_tel_whatsapp(item.cliente.telefone or item.cliente.identificador)
    if not telefone:
        raise HTTPException(400, "Cliente sem telefone para WhatsApp")

    texto = montar_mensagem_whatsapp_contrato(request, empresa, item, db)

    # O clique no envio pelo WhatsApp conclui a pendência do painel.
    # Mantemos o status da reserva intacto para não interferir na agenda/financeiro.
    if not item.contrato_enviado_em:
        item.contrato_enviado_em = agora_utc()
        db.commit()

    return RedirectResponse(
        f"https://wa.me/{telefone}?text={quote(texto)}",
        status_code=303,
    )

@app.get("/painel/solicitacao/{solicitacao_id}/cliente", response_class=HTMLResponse)
def editar_cliente_da_solicitacao(solicitacao_id: int, request: Request, db: Session = Depends(get_db),
                                  empresa: Empresa = Depends(empresa_logada)):
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id or not item.cliente:
        raise HTTPException(404)
    return templates.TemplateResponse(
        "admin/solicitacao_cliente_editar.html",
        {"request": request, "item": item, "cliente": item.cliente, "empresa": empresa},
    )


@app.post("/painel/solicitacao/{solicitacao_id}/cliente")
def salvar_cliente_da_solicitacao(
        solicitacao_id: int,
        nome: str = Form(""),
        telefone: str = Form(""),
        cpf: str = Form(""),
        cnpj: str = Form(""),
        data_nascimento: str = Form(""),
        email: str = Form(""),
        endereco: str = Form(""),
        numero: str = Form(""),
        complemento: str = Form(""),
        bairro: str = Form(""),
        cidade: str = Form(""),
        estado: str = Form(""),
        cep: str = Form(""),
        observacoes: str = Form(""),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id or not item.cliente:
        raise HTTPException(404)

    cliente = item.cliente
    cliente.nome = nome.strip() or cliente.nome
    cliente.telefone = limpar_identificador(telefone)
    cliente.cpf = limpar_identificador(cpf)
    cliente.cnpj = limpar_identificador(cnpj)
    cliente.email = email.strip()
    cliente.endereco = endereco.strip()
    cliente.numero = numero.strip()
    cliente.complemento = complemento.strip()
    cliente.bairro = bairro.strip()
    cliente.cidade = cidade.strip()
    cliente.estado = estado.strip().upper()
    cliente.cep = limpar_identificador(cep)
    cliente.observacoes = observacoes.strip()

    if data_nascimento:
        try:
            cliente.data_nascimento = datetime.strptime(data_nascimento, "%Y-%m-%d").date()
        except ValueError:
            pass

    if empresa.identificador_principal == "cpf" and cliente.cpf:
        cliente.identificador = cliente.cpf
    elif empresa.identificador_principal == "cnpj" and cliente.cnpj:
        cliente.identificador = cliente.cnpj
    elif cliente.telefone:
        cliente.identificador = cliente.telefone

    db.commit()
    return RedirectResponse(f"/painel/solicitacao/{solicitacao_id}", status_code=303)


@app.get("/painel/solicitacao/{solicitacao_id}/editar", response_class=HTMLResponse)
def editar_solicitacao(solicitacao_id: int, request: Request, db: Session = Depends(get_db),
                       empresa: Empresa = Depends(empresa_logada)):
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)
    produtos = db.query(ProdutoServico).filter_by(empresa_id=empresa.id, ativo=True).order_by(ProdutoServico.nome).all()
    contratos = db.query(Contrato).filter_by(empresa_id=empresa.id, ativo=True).order_by(Contrato.nome).all()
    return templates.TemplateResponse("admin/solicitacao_editar.html",
                                      {"request": request, "item": item, "empresa": empresa, "produtos": produtos,
                                       "contratos": contratos})


@app.post("/painel/solicitacao/{solicitacao_id}/editar")
def salvar_edicao_solicitacao(
        solicitacao_id: int,
        data_evento: str = Form(""),
        hora_inicio: str = Form(""),
        hora_fim: str = Form(""),
        bairro: str = Form(""),
        local: str = Form(""),
        acesso_local: str = Form(""),
        valor: str = Form("0"),
        sinal: str = Form("0"),
        status: str = Form(""),
        observacoes: str = Form(""),
        aprovacao_manual: str = Form(""),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)

    if status in ["aguardando_nova_data", "cancelada", "cancelado_cliente", "rejeitada"]:
        item.status = status
        # Crédito ou cancelamento elimina Entregar/Buscar da operação, mas
        # preserva contrato, equipamentos e pagamentos já registrados.
        retirar_solicitacao_da_operacao(db, item)
    else:
        if data_evento:
            item.data_evento = datetime.strptime(data_evento, "%Y-%m-%d").date()
        if hora_inicio:
            item.hora_inicio = datetime.strptime(hora_inicio, "%H:%M").time()
        if hora_fim:
            item.hora_fim = datetime.strptime(hora_fim, "%H:%M").time()
        item.status = status or item.status
        criar_eventos_operacionais(db, item)

    item.bairro = bairro
    item.local = local
    item.acesso_local = acesso_local
    item.valor = texto_para_float(valor)
    item.sinal = texto_para_float(sinal)
    item.observacoes = observacoes

    tem_itens = db.query(ReservaItem).filter_by(empresa_id=empresa.id, solicitacao_id=item.id).count() > 0

    if item.status in ["reserva_confirmada", "aguardando_pagamento"] and not tem_itens:
        # Não deixa salvar uma reserva como aprovada/confirmada sem itens.
        item.status = "pre_reserva"
        item.aprovado_em = None
        item.sinal_recebido = False
        item.valor_pago = 0
        item.pagamento_confirmado_em = None

    if aprovacao_manual and tem_itens:
        item.status = "reserva_confirmada"
        item.aprovado_em = agora_utc()

    _invalidar_geocodificacao(item)
    db.commit()
    _tentar_geocodificar_solicitacao(db, item)
    return RedirectResponse(f"/painel/solicitacao/{solicitacao_id}", status_code=303)


@app.post("/painel/solicitacao/{solicitacao_id}/reativar")
def reativar_solicitacao_em_credito(
        solicitacao_id: int,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    """Reabre um contrato em crédito como rascunho para revisão antes de novo aceite.

    Mantém itens e pagamentos já registrados, mas inicia um novo ciclo de aceite.
    """
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)

    if item.status != "aguardando_nova_data":
        return RedirectResponse(f"/painel/solicitacao/{solicitacao_id}", status_code=303)

    item.status = "pre_reserva"
    item.aprovado_em = None
    item.aceite_em = None
    item.contrato_enviado_em = None
    item.cancelado_em = None
    # Não recria operação neste momento. Ela só volta após o novo aceite.
    retirar_solicitacao_da_operacao(db, item)
    db.commit()

    return RedirectResponse(
        f"/painel/solicitacao/{solicitacao_id}/editar-completo?reativado=1",
        status_code=303,
    )


@app.post("/painel/solicitacao/{solicitacao_id}/credito")
def colocar_solicitacao_em_credito(
        solicitacao_id: int,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    """Coloca o contrato em crédito e remove definitivamente os cards operacionais.

    Endpoint dedicado para não depender do valor enviado por um botão de status.
    Preserva contrato, itens e pagamentos para reutilização em uma nova data.
    """
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)

    item.status = "aguardando_nova_data"
    retirar_solicitacao_da_operacao(db, item)
    db.flush()

    # Garantia adicional: nenhum registro de Entregar/Retirar pode permanecer.
    db.query(Agenda).filter_by(
        empresa_id=item.empresa_id,
        solicitacao_id=item.id,
    ).delete(synchronize_session=False)

    db.commit()
    return RedirectResponse(
        f"/painel/solicitacao/{solicitacao_id}?credito=ok",
        status_code=303,
    )


@app.post("/painel/solicitacao/{solicitacao_id}/status")
def atualizar_status_solicitacao(
        solicitacao_id: int,
        status: str = Form(""),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)

    status_permitidos = ["reserva_confirmada", "aguardando_pagamento", "aguardando_nova_data", "cancelada"]
    if status in status_permitidos:
        tem_itens = db.query(ReservaItem).filter_by(empresa_id=empresa.id, solicitacao_id=item.id).count() > 0
        if status in ["reserva_confirmada", "aguardando_pagamento"] and not tem_itens:
            item.status = "pre_reserva"
            item.aprovado_em = None
            item.sinal_recebido = False
            item.valor_pago = 0
            item.pagamento_confirmado_em = None
        else:
            item.status = status
            if status == "reserva_confirmada" and not item.aprovado_em:
                item.aprovado_em = agora_utc()
            if status in ["aguardando_nova_data", "cancelada", "cancelado_cliente", "rejeitada"]:
                retirar_solicitacao_da_operacao(db, item)
            else:
                criar_eventos_operacionais(db, item)

    db.commit()
    return RedirectResponse(f"/painel/solicitacao/{solicitacao_id}", status_code=303)


@app.post("/painel/solicitacao/{solicitacao_id}/cliente-local")
def salvar_cliente_local_solicitacao(
        solicitacao_id: int,
        nome: str = Form(""),
        telefone: str = Form(""),
        cpf: str = Form(""),
        email: str = Form(""),
        data_evento: str = Form(""),
        hora_inicio: str = Form(""),
        hora_fim: str = Form(""),
        local: str = Form(""),
        numero: str = Form(""),
        bairro: str = Form(""),
        acesso_local: str = Form(""),
        local_nome: str = Form(""),
        local_responsavel_nome: str = Form(""),
        local_responsavel_telefone: str = Form(""),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)

    cliente = item.cliente
    cliente.nome = nome.strip() or cliente.nome
    cliente.telefone = limpar_identificador(telefone) or telefone.strip()
    cliente.cpf = limpar_identificador(cpf)
    cliente.email = email.strip()
    cliente.endereco = local.strip()
    cliente.numero = numero.strip()
    cliente.bairro = bairro.strip()

    if data_evento:
        item.data_evento = datetime.strptime(data_evento, "%Y-%m-%d").date()
    if hora_inicio:
        item.hora_inicio = datetime.strptime(hora_inicio, "%H:%M").time()
    item.hora_fim = datetime.strptime(hora_fim, "%H:%M").time() if hora_fim else None
    endereco_anterior = (item.local or "").strip()
    item.local = local.strip()
    item.local_numero = numero.strip()
    item.bairro = bairro.strip()
    if _normalizar_chave_endereco(endereco_anterior) != _normalizar_chave_endereco(item.local):
        item.local_complemento = ""
        item.local_cidade = ""
        item.local_estado = ""
        item.local_cep = ""
    item.acesso_local = acesso_local.strip()
    item.local_nome = local_nome.strip()
    item.local_responsavel_nome = local_responsavel_nome.strip()
    item.local_responsavel_telefone = limpar_identificador(
        local_responsavel_telefone) or local_responsavel_telefone.strip()
    salvar_endereco_cliente(
        db, empresa.id, cliente.id, item.local, item.local_numero, item.local_complemento or "",
        item.bairro, item.local_cidade or "", item.local_estado or "", item.local_cep or "",
        apelido=item.local_nome,
    )

    if contrato_aprovado_para_operacao(item):
        criar_eventos_operacionais(db, item)

    _invalidar_geocodificacao(item)
    db.commit()
    _tentar_geocodificar_solicitacao(db, item)
    return RedirectResponse(f"/painel/solicitacao/{solicitacao_id}", status_code=303)


@app.post("/painel/solicitacao/{solicitacao_id}/preparar")
async def preparar_contrato(
        solicitacao_id: int,
        request: Request,
        contrato_id: str = Form(""),
        data_evento: str = Form(""),
        hora_inicio: str = Form(""),
        hora_fim: str = Form(""),
        bairro: str = Form(""),
        local: str = Form(""),
        acesso_local: str = Form(""),
        valor: str = Form("0"),
        sinal: str = Form("0"),
        observacoes: str = Form(""),
        acao: str = Form("salvar"),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)
    if item.status in ["cancelada", "cancelado_cliente"]:
        return RedirectResponse(f"/painel/solicitacao/{solicitacao_id}", status_code=303)

    if data_evento:
        item.data_evento = datetime.strptime(data_evento, "%Y-%m-%d").date()
    if hora_inicio:
        item.hora_inicio = datetime.strptime(hora_inicio, "%H:%M").time()
    item.hora_fim = datetime.strptime(hora_fim, "%H:%M").time() if hora_fim else None
    item.bairro = bairro
    item.local = local
    item.acesso_local = acesso_local

    form = await request.form()
    produto_ids = form.getlist("produto_id")
    quantidades = form.getlist("quantidade")
    valores_unitarios = form.getlist("valor_unitario")

    # Regrava os itens da reserva para permitir vários produtos/serviços.
    db.query(ReservaItem).filter_by(empresa_id=empresa.id, solicitacao_id=item.id).delete()
    primeiro_produto = None
    for idx, produto_id in enumerate(produto_ids):
        if not produto_id:
            continue
        produto = db.get(ProdutoServico, int(produto_id))
        if not produto or produto.empresa_id != empresa.id:
            continue
        quantidade = int(quantidades[idx]) if idx < len(quantidades) and str(quantidades[idx]).isdigit() else 1
        valor_unitario = texto_para_float(valores_unitarios[idx]) if idx < len(valores_unitarios) else (
                produto.valor_base or 0)
        total_item = quantidade * valor_unitario
        db.add(ReservaItem(
            empresa_id=empresa.id,
            solicitacao_id=item.id,
            produto_id=produto.id,
            nome=produto.nome,
            descricao=produto.descricao,
            quantidade=quantidade,
            valor_unitario=valor_unitario,
            valor_total=total_item
        ))
        if primeiro_produto is None:
            primeiro_produto = produto

    item.produto_id = primeiro_produto.id if primeiro_produto else None
    contrato_padrao_id = primeiro_produto.contrato_id if primeiro_produto and primeiro_produto.contrato_id else None
    item.contrato_id = int(contrato_id) if contrato_id else contrato_padrao_id
    db.flush()
    total_itens = sum((linha.valor_total or 0) for linha in item.itens)
    valor_manual = texto_para_float(valor)
    item.valor = total_itens if total_itens > 0 else valor_manual
    item.sinal = texto_para_float(sinal)
    item.observacoes = observacoes
    if primeiro_produto and item.hora_inicio:
        item.hora_fim = somar_minutos(item.hora_inicio, primeiro_produto.duracao_minutos or 240)

    # Salvar não significa aceitar nem enviar.
    # Antes do aceite, o contrato continua como rascunho até o usuário liberar o envio.
    # Depois de um contrato aceito, qualquer edição volta para pendente de novo aceite.
    if status_reserva_confirmada(item.status):
        item.status = "aguardando_aceite"
        db.query(Agenda).filter_by(empresa_id=empresa.id, solicitacao_id=item.id).delete()
    elif acao == "enviar" and primeiro_produto and item.contrato_id:
        item.status = "contrato_enviado"
    elif item.status not in ["contrato_enviado", "aguardando_aceite"]:
        item.status = "pre_reserva"

    db.commit()
    return RedirectResponse(f"/painel/solicitacao/{solicitacao_id}", status_code=303)


@app.post("/painel/solicitacao/{solicitacao_id}/excluir")
def excluir_solicitacao_completa(
        solicitacao_id: int,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)
    if existe_pagamento_conciliado(item):
        msg = quote("Pagamento conciliado. Chame o financeiro antes de excluir este contrato.")
        return RedirectResponse(f"/painel/solicitacao/{solicitacao_id}?erro={msg}", status_code=303)

    cliente = item.cliente
    pagamento_ids = [p.id for p in (item.pagamentos or [])]
    if pagamento_ids:
        db.query(LancamentoBanco).filter(
            LancamentoBanco.empresa_id == empresa.id,
            LancamentoBanco.pagamento_id.in_(pagamento_ids)
        ).update({LancamentoBanco.pagamento_id: None}, synchronize_session=False)
        db.query(LancamentoManualFinanceiro).filter(
            LancamentoManualFinanceiro.empresa_id == empresa.id,
            LancamentoManualFinanceiro.pagamento_id.in_(pagamento_ids)
        ).update({LancamentoManualFinanceiro.pagamento_id: None}, synchronize_session=False)

    db.query(Agenda).filter_by(empresa_id=empresa.id, solicitacao_id=item.id).delete()
    db.delete(item)
    db.flush()

    if cliente and db.query(Solicitacao).filter_by(empresa_id=empresa.id, cliente_id=cliente.id).count() == 0:
        db.delete(cliente)

    db.commit()
    return RedirectResponse("/painel", status_code=303)


@app.post("/painel/solicitacao/{solicitacao_id}/aceite-manual")
def aceite_manual_solicitacao(
        request: Request,
        solicitacao_id: int,
        observacao_aceite: str = Form(...),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)

    tem_itens = db.query(ReservaItem).filter_by(empresa_id=empresa.id, solicitacao_id=item.id).count() > 0
    if not item.contrato_id or not tem_itens:
        raise HTTPException(400, "Para aceitar manualmente, o contrato precisa ter modelo e pelo menos um item.")

    motivo = observacao_aceite.strip()
    if not motivo:
        raise HTTPException(400, "Informe o motivo do aceite manual.")

    usuario = request.session.get("usuario_sistema", "Usuário")
    item.status = "reserva_confirmada"
    item.aceite_em = agora_utc()
    item.aprovado_em = item.aceite_em
    registro = f"Aceite manual por {usuario}: {motivo}"
    item.observacoes = (item.observacoes + "\n\n" if item.observacoes else "") + registro

    if item.hora_inicio and not item.hora_fim and item.produto and item.produto.duracao_minutos:
        item.hora_fim = somar_minutos(item.hora_inicio, item.produto.duracao_minutos)
    criar_eventos_operacionais(db, item)
    _processar_humiat_aceite(db, empresa, item)
    db.commit()
    return RedirectResponse(f"/painel/solicitacao/{solicitacao_id}", status_code=303)


@app.get("/painel/contrato-novo", response_class=HTMLResponse)
def contrato_novo_form(request: Request, busca: str = "", db: Session = Depends(get_db),
                       empresa: Empresa = Depends(empresa_logada)):
    produtos = db.query(ProdutoServico).filter_by(empresa_id=empresa.id, ativo=True).order_by(ProdutoServico.nome).all()
    contratos = db.query(Contrato).filter_by(empresa_id=empresa.id, ativo=True).order_by(Contrato.nome).all()
    busca_limpa = limpar_identificador(busca)
    form = {}
    if busca_limpa:
        # A busca da barra inferior normalmente é telefone ou CPF.
        # Deixamos o dado já preenchido para o contrato nascer sem retrabalho.
        if len(busca_limpa) == 11 and not busca_limpa.startswith(("2", "3", "4", "5", "6", "7", "8", "9")):
            form["cpf"] = busca_limpa
        else:
            form["telefone"] = busca_limpa
    return templates.TemplateResponse("admin/contrato_novo.html", {
        "request": request,
        "empresa": empresa,
        "produtos": produtos,
        "contratos": contratos,
        "erro": "",
        "form": form
    })


def celular_brasileiro_valido(valor: str) -> bool:
    numero = limpar_identificador(valor)
    if numero.startswith("55") and len(numero) == 13:
        numero = numero[2:]
    return len(numero) == 11 and numero[2] == "9" and numero[:2] != "00"


def endereco_cliente_payload(item: EnderecoCliente) -> dict:
    return {
        "id": item.id, "apelido": item.apelido or "", "endereco": item.endereco or "",
        "numero": item.numero or "", "complemento": item.complemento or "",
        "bairro": item.bairro or "", "cidade": item.cidade or "",
        "estado": item.estado or "", "cep": item.cep or ""
    }


def salvar_endereco_cliente(db: Session, empresa_id: int, cliente_id: int, endereco: str, numero: str = "",
                            complemento: str = "", bairro: str = "", cidade: str = "",
                            estado: str = "", cep: str = "", apelido: str = ""):
    """Salva um endereço já utilizado pelo cliente como atalho para futuros contratos."""
    dados = {
        "endereco": (endereco or "").strip(), "numero": (numero or "").strip(),
        "complemento": (complemento or "").strip(), "bairro": (bairro or "").strip(),
        "cidade": (cidade or "").strip(), "estado": (estado or "").strip(), "cep": (cep or "").strip(),
    }
    if not dados["endereco"]:
        return None
    existente = db.query(EnderecoCliente).filter_by(empresa_id=empresa_id, cliente_id=cliente_id, **dados).first()
    apelido_limpo = (apelido or "").strip()
    if existente:
        existente.ativo = True
        if apelido_limpo:
            existente.apelido = apelido_limpo
        return existente
    item = EnderecoCliente(empresa_id=empresa_id, cliente_id=cliente_id, apelido=apelido_limpo or None, **dados)
    db.add(item)
    return item


@app.get("/e/{slug}/api/clientes/por-telefone")
def api_publico_cliente_por_telefone(slug: str, telefone: str, db: Session = Depends(get_db)):
    empresa = db.query(Empresa).filter_by(slug=slug).first()
    if not empresa:
        raise HTTPException(404)
    tel = limpar_identificador(telefone)
    if tel.startswith("55") and len(tel) == 13:
        tel = tel[2:]
    if len(tel) != 11 or tel[2] != "9":
        return JSONResponse({"encontrado": False, "enderecos": []})
    clientes = db.query(Cliente).filter(Cliente.empresa_id == empresa.id, or_(Cliente.telefone == tel, Cliente.identificador == tel)).all()
    if not clientes:
        return JSONResponse({"encontrado": False, "enderecos": []})
    cliente = clientes[0]
    ids = [c.id for c in clientes]
    enderecos = db.query(EnderecoCliente).filter(EnderecoCliente.empresa_id == empresa.id, EnderecoCliente.cliente_id.in_(ids), EnderecoCliente.ativo == True).order_by(EnderecoCliente.atualizado_em.desc()).all()
    if not enderecos:
        for c in clientes:
            if c.endereco:
                salvar_endereco_cliente(db, empresa.id, c.id, c.endereco, c.numero, c.complemento, c.bairro, c.cidade, c.estado, c.cep)
        db.commit()
        enderecos = db.query(EnderecoCliente).filter(EnderecoCliente.empresa_id == empresa.id, EnderecoCliente.cliente_id.in_(ids), EnderecoCliente.ativo == True).order_by(EnderecoCliente.atualizado_em.desc()).all()
    return JSONResponse({"encontrado": True, "quantidade": len(clientes), "cliente": {"id": cliente.id, "nome": cliente.nome or '', "cpf": cliente.cpf or '', "cnpj": cliente.cnpj or '', "email": cliente.email or '', "telefone": cliente.telefone or tel}, "enderecos": [endereco_cliente_payload(e) for e in enderecos[:10]]})


@app.get("/api/clientes/por-telefone")
def api_cliente_por_telefone(request: Request, telefone: str, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    tel = limpar_identificador(telefone)
    if tel.startswith("55") and len(tel) == 13:
        tel = tel[2:]
    if len(tel) < 10:
        return JSONResponse({"encontrado": False, "enderecos": []})
    clientes = db.query(Cliente).filter(Cliente.empresa_id == empresa.id, or_(Cliente.telefone == tel, Cliente.identificador == tel)).all()
    if not clientes:
        return JSONResponse({"encontrado": False, "enderecos": []})
    cliente = clientes[0]
    ids = [c.id for c in clientes]
    enderecos = db.query(EnderecoCliente).filter(EnderecoCliente.empresa_id == empresa.id, EnderecoCliente.cliente_id.in_(ids), EnderecoCliente.ativo == True).order_by(EnderecoCliente.atualizado_em.desc()).all()
    # Compatibilidade: transforma o endereço antigo do cliente em endereço oficial na primeira consulta.
    if not enderecos:
        for c in clientes:
            if c.endereco:
                salvar_endereco_cliente(db, empresa.id, c.id, c.endereco, c.numero, c.complemento, c.bairro, c.cidade, c.estado, c.cep)
        db.commit()
        enderecos = db.query(EnderecoCliente).filter(EnderecoCliente.empresa_id == empresa.id, EnderecoCliente.cliente_id.in_(ids), EnderecoCliente.ativo == True).order_by(EnderecoCliente.atualizado_em.desc()).all()
    return JSONResponse({"encontrado": True, "quantidade": len(clientes), "cliente": {"id": cliente.id, "nome": cliente.nome or '', "cpf": cliente.cpf or '', "cnpj": cliente.cnpj or '', "email": cliente.email or '', "telefone": cliente.telefone or tel}, "enderecos": [endereco_cliente_payload(e) for e in enderecos[:10]]})


@app.post("/painel/contrato-novo")
def contrato_novo_salvar(
        request: Request,
        nome: str = Form(""),
        telefone: str = Form(""),
        whatsapp_brasil: str = Form(""),
        cpf: str = Form(""),
        cnpj: str = Form(""),
        email: str = Form(""),
        endereco: str = Form(""),
        numero: str = Form(""),
        complemento: str = Form(""),
        bairro: str = Form(""),
        cidade: str = Form(""),
        estado: str = Form(""),
        cep: str = Form(""),
        produto_id: str = Form(""),
        contrato_id: str = Form(""),
        data_evento: str = Form(""),
        hora_inicio: str = Form(""),
        retirada_obrigatoria: str = Form(""),
        retirada_data: str = Form(""),
        retirada_hora: str = Form(""),
        valor: str = Form("0"),
        sinal: str = Form("0"),
        local_nome: str = Form(""),
        local: str = Form(""),
        acesso_local: str = Form(""),
        local_responsavel_nome: str = Form(""),
        local_responsavel_telefone: str = Form(""),
        observacoes: str = Form(""),
        modo_criacao: str = Form("whatsapp"),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    telefone_limpo = limpar_identificador(telefone)
    cpf_limpo = limpar_identificador(cpf)
    cnpj_limpo = limpar_identificador(cnpj)
    identificador = cpf_limpo or cnpj_limpo or telefone_limpo or uuid.uuid4().hex[:12]

    produtos = db.query(ProdutoServico).filter_by(empresa_id=empresa.id, ativo=True).order_by(ProdutoServico.nome).all()
    contratos = db.query(Contrato).filter_by(empresa_id=empresa.id, ativo=True).order_by(Contrato.nome).all()
    form = {
        "nome": nome, "telefone": telefone, "whatsapp_brasil": whatsapp_brasil,
        "cpf": cpf, "cnpj": cnpj, "email": email, "endereco": endereco,
        "numero": numero, "complemento": complemento, "bairro": bairro,
        "cidade": cidade, "estado": estado, "cep": cep, "produto_id": produto_id,
        "contrato_id": contrato_id, "data_evento": data_evento, "hora_inicio": hora_inicio,
        "retirada_obrigatoria": retirada_obrigatoria, "retirada_data": retirada_data,
        "retirada_hora": retirada_hora, "valor": valor, "sinal": sinal,
        "local_nome": local_nome, "local": local, "acesso_local": acesso_local,
        "local_responsavel_nome": local_responsavel_nome,
        "local_responsavel_telefone": local_responsavel_telefone,
        "observacoes": observacoes, "modo_criacao": modo_criacao,
    }

    def render_erro(mensagem: str):
        return templates.TemplateResponse("admin/contrato_novo.html", {
            "request": request,
            "empresa": empresa,
            "produtos": produtos,
            "contratos": contratos,
            "erro": mensagem,
            "form": form
        }, status_code=400)

    if not nome.strip():
        return render_erro("Informe o nome do cliente.")
    if not telefone.strip():
        return render_erro("Informe o WhatsApp ou telefone do cliente.")
    if whatsapp_brasil and not celular_brasileiro_valido(telefone):
        return render_erro("Informe um WhatsApp brasileiro válido no formato (DD) 9XXXX-XXXX.")
    if cpf_limpo and not cpf_valido(cpf_limpo):
        return render_erro("CPF inválido.")
    if cnpj_limpo and not cnpj_valido(cnpj_limpo):
        return render_erro("CNPJ inválido.")
    if not endereco.strip() or not numero.strip() or not bairro.strip():
        return render_erro("Informe o endereço, número e bairro.")

    cadastro_cliente = modo_criacao == "cadastro"
    if not cadastro_cliente and not celular_brasileiro_valido(local_responsavel_telefone):
        return render_erro("Informe um WhatsApp brasileiro válido para o responsável no local.")
    if not cadastro_cliente and not local_responsavel_nome.strip():
        return render_erro("Informe o nome do responsável no local.")
    if not cadastro_cliente and (not data_evento or not hora_inicio):
        return render_erro("Informe a data e a hora do evento.")
    if not cadastro_cliente and not hora_meia_em_meia_valida(hora_inicio):
        return render_erro("A hora precisa estar em intervalo de 30 minutos. Exemplo: 18:00 ou 18:30.")

    data_evento_obj = datetime.strptime(data_evento, "%Y-%m-%d").date() if data_evento else None
    duplicado_q = None
    if not cadastro_cliente:
        duplicado_q = db.query(Solicitacao).join(Cliente, Solicitacao.cliente_id == Cliente.id).filter(
            Solicitacao.empresa_id == empresa.id,
            Solicitacao.data_evento == data_evento_obj,
            ~Solicitacao.status.in_(["cancelada", "cancelado_cliente", "rejeitada"])
        )
    condicoes = []
    if telefone_limpo:
        condicoes.append(Cliente.telefone == telefone_limpo)
        condicoes.append(Cliente.identificador == telefone_limpo)
    if cpf_limpo:
        condicoes.append(Cliente.cpf == cpf_limpo)
        condicoes.append(Cliente.identificador == cpf_limpo)
    if cnpj_limpo:
        condicoes.append(Cliente.cnpj == cnpj_limpo)
        condicoes.append(Cliente.identificador == cnpj_limpo)
    from sqlalchemy import or_
    if condicoes and duplicado_q is not None:
        duplicado = duplicado_q.filter(or_(*condicoes)).first()
        if duplicado:
            return render_erro(
                f"Já existe uma reserva/contrato para este telefone/CPF/CNPJ nesta data: #{duplicado.id} - {duplicado.cliente.nome}.")

    cliente = None
    if telefone_limpo:
        cliente = db.query(Cliente).filter(Cliente.empresa_id == empresa.id, or_(Cliente.telefone == telefone_limpo, Cliente.identificador == telefone_limpo)).first()
    if not cliente:
        cliente = db.query(Cliente).filter_by(empresa_id=empresa.id, identificador=identificador).first()
    if not cliente:
        cliente = Cliente(empresa_id=empresa.id, identificador=identificador)
        db.add(cliente)

    cliente.nome = nome.strip()
    cliente.telefone = telefone_limpo or telefone.strip()
    cliente.cpf = cpf_limpo
    cliente.cnpj = cnpj_limpo
    cliente.email = email.strip()
    cliente.endereco = endereco.strip()
    cliente.numero = numero.strip()
    cliente.complemento = complemento.strip()
    cliente.bairro = bairro.strip()
    cliente.cidade = cidade.strip()
    cliente.estado = estado.strip()
    cliente.cep = cep.strip()
    cliente.observacoes = observacoes.strip()
    db.flush()
    salvar_endereco_cliente(
        db, empresa.id, cliente.id, endereco, numero, complemento, bairro, cidade, estado, cep,
        apelido=local_nome,
    )

    if cadastro_cliente:
        db.commit()
        return RedirectResponse(f"/painel/cliente/{cliente.id}?cadastro=salvo", status_code=303)

    produto = db.get(ProdutoServico, int(produto_id)) if produto_id else None
    if produto and produto.empresa_id != empresa.id:
        raise HTTPException(404)
    if modo_criacao == "manual" and not produto:
        return render_erro("No contrato manual, informe pelo menos um item principal.")

    inicio_obj = datetime.strptime(hora_inicio, "%H:%M").time()
    retirada_obrigatoria_bool = bool(retirada_obrigatoria)
    retirada_data_obj = datetime.strptime(retirada_data, "%Y-%m-%d").date() if retirada_data else data_evento_obj
    retirada_hora_obj = datetime.strptime(retirada_hora, "%H:%M").time() if retirada_hora else None
    valor_float = texto_para_float(valor)
    sinal_float = texto_para_float(sinal)
    manual = modo_criacao == "manual"

    item = Solicitacao(
        empresa_id=empresa.id,
        cliente_id=cliente.id,
        produto_id=produto.id if produto else None,
        contrato_id=int(contrato_id) if contrato_id else (produto.contrato_id if produto and produto.contrato_id else None),
        data_evento=data_evento_obj,
        hora_inicio=inicio_obj,
        hora_fim=somar_minutos(inicio_obj, produto.duracao_minutos or 240) if produto else None,
        retirada_obrigatoria=retirada_obrigatoria_bool,
        retirada_data=retirada_data_obj if retirada_obrigatoria_bool else None,
        retirada_hora=retirada_hora_obj,
        bairro=bairro.strip(),
        local=endereco.strip(),
        local_numero=numero.strip(),
        local_complemento=complemento.strip(),
        local_cidade=cidade.strip(),
        local_estado=estado.strip(),
        local_cep=cep.strip(),
        local_nome=local_nome.strip(),
        local_responsavel_nome=local_responsavel_nome.strip(),
        local_responsavel_telefone=limpar_identificador(
            local_responsavel_telefone) or local_responsavel_telefone.strip(),
        acesso_local=acesso_local.strip(),
        valor=valor_float,
        sinal=sinal_float,
        observacoes=observacoes.strip(),
        status="reserva_confirmada" if manual else ("aguardando_aceite" if (contrato_id or (produto and produto.contrato_id)) and produto else "pre_reserva"),
        aprovado_em=agora_utc() if manual else None,
        aceite_em=agora_utc() if manual else None,
        sinal_recebido=True if manual and sinal_float > 0 else False,
        valor_pago=sinal_float if manual and sinal_float > 0 else 0,
        pagamento_confirmado_em=agora_utc() if manual and sinal_float > 0 else None
    )
    if item.retirada_obrigatoria and not item.retirada_hora:
        item.retirada_hora = item.hora_fim or item.hora_inicio

    db.add(item)
    db.flush()

    if produto:
        db.add(ReservaItem(
            empresa_id=empresa.id,
            solicitacao_id=item.id,
            produto_id=produto.id,
            nome=produto.nome,
            descricao=produto.descricao,
            quantidade=1,
            valor_unitario=valor_float,
            valor_total=valor_float
        ))

    if manual:
        _processar_humiat_aceite(db, empresa, item)

    if manual and sinal_float > 0:
        db.add(Pagamento(
            empresa_id=empresa.id,
            solicitacao_id=item.id,
            data_pagamento=date.today(),
            valor=sinal_float,
            forma_pagamento="pix",
            comprovante_no_nome_cliente=True,
            nome_comprovante=cliente.nome,
            observacoes="Sinal informado no contrato manual.",
            usuario_registro=request.session.get("usuario_sistema", "Usuário")
        ))

    db.commit()
    return RedirectResponse(f"/painel/solicitacao/{item.id}", status_code=303)


def awaitable_form_fallback(request: Request) -> dict:
    # Em rotas síncronas o FastAPI já consumiu os campos do Form.
    # Mantemos um dicionário vazio apenas para o template não quebrar em caso de erro.
    return {}


def form_solicitacao_completo(item: Solicitacao) -> dict:
    """Monta o formulário único usando o endereço congelado do próprio contrato."""
    cliente = item.cliente
    endereco_evento = dados_endereco_solicitacao(item)
    return {
        "nome": cliente.nome if cliente else "",
        "telefone": cliente.telefone if cliente else "",
        "cpf": cliente.cpf if cliente else "",
        "cnpj": cliente.cnpj if cliente else "",
        "email": cliente.email if cliente else "",
        "endereco": endereco_evento["endereco"],
        "numero": endereco_evento["numero"],
        "complemento": endereco_evento["complemento"],
        "bairro": endereco_evento["bairro"],
        "cidade": endereco_evento["cidade"],
        "estado": endereco_evento["estado"],
        "cep": endereco_evento["cep"],
        "data_evento": item.data_evento.isoformat() if item.data_evento else "",
        "hora_inicio": item.hora_inicio.strftime("%H:%M") if item.hora_inicio else "",
        "retirada_obrigatoria": "1" if retirada_obrigatoria_ativa(item) else "",
        "retirada_data": item.retirada_data.isoformat() if item.retirada_data else (item.data_evento.isoformat() if item.data_evento else ""),
        "retirada_hora": item.retirada_hora.strftime("%H:%M") if item.retirada_hora else (item.hora_fim.strftime("%H:%M") if item.hora_fim else ""),
        "produto_id": str(item.produto_id or ""),
        "contrato_id": str(item.contrato_id or ""),
        "valor": moeda_br(item.valor or 0),
        "sinal": moeda_br(item.sinal or 0),
        "local_nome": item.local_nome or "",
        "local": item.local or "",
        "acesso_local": item.acesso_local or "",
        "local_responsavel_nome": item.local_responsavel_nome or "",
        "local_responsavel_telefone": item.local_responsavel_telefone or "",
        "observacoes": item.observacoes or "",
        "modo_criacao": "manual",
    }


@app.get("/painel/solicitacao/{solicitacao_id}/editar-completo", response_class=HTMLResponse)
def editar_solicitacao_completa(
        solicitacao_id: int,
        request: Request,
        copiado: str = "",
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)

    produtos = db.query(ProdutoServico).filter_by(empresa_id=empresa.id, ativo=True).order_by(ProdutoServico.nome).all()
    contratos = db.query(Contrato).filter_by(empresa_id=empresa.id, ativo=True).order_by(Contrato.nome).all()
    return templates.TemplateResponse("admin/contrato_novo.html", {
        "request": request,
        "empresa": empresa,
        "produtos": produtos,
        "contratos": contratos,
        "erro": "",
        "form": form_solicitacao_completo(item),
        "modo_edicao": True,
        "item": item,
        "copia_endereco": bool(copiado),
    })


@app.post("/painel/solicitacao/{solicitacao_id}/editar-completo")
def salvar_solicitacao_completa(
        solicitacao_id: int,
        request: Request,
        nome: str = Form(""),
        telefone: str = Form(""),
        cpf: str = Form(""),
        cnpj: str = Form(""),
        email: str = Form(""),
        endereco: str = Form(""),
        numero: str = Form(""),
        complemento: str = Form(""),
        bairro: str = Form(""),
        cidade: str = Form(""),
        estado: str = Form(""),
        cep: str = Form(""),
        produto_id: str = Form(""),
        contrato_id: str = Form(""),
        data_evento: str = Form(""),
        hora_inicio: str = Form(""),
        retirada_obrigatoria: str = Form(""),
        retirada_data: str = Form(""),
        retirada_hora: str = Form(""),
        valor: str = Form("0"),
        sinal: str = Form("0"),
        local_nome: str = Form(""),
        local: str = Form(""),
        acesso_local: str = Form(""),
        local_responsavel_nome: str = Form(""),
        local_responsavel_telefone: str = Form(""),
        observacoes: str = Form(""),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id or not item.cliente:
        raise HTTPException(404)

    produtos = db.query(ProdutoServico).filter_by(empresa_id=empresa.id, ativo=True).order_by(ProdutoServico.nome).all()
    contratos = db.query(Contrato).filter_by(empresa_id=empresa.id, ativo=True).order_by(Contrato.nome).all()
    form = dict(
        nome=nome, telefone=telefone, cpf=cpf, cnpj=cnpj, email=email, endereco=endereco,
        numero=numero, complemento=complemento, bairro=bairro, cidade=cidade, estado=estado,
        cep=cep, produto_id=produto_id, contrato_id=contrato_id, data_evento=data_evento,
        hora_inicio=hora_inicio, retirada_obrigatoria=retirada_obrigatoria,
        retirada_data=retirada_data, retirada_hora=retirada_hora,
        valor=valor, sinal=sinal, local_nome=local_nome, local=local,
        acesso_local=acesso_local, local_responsavel_nome=local_responsavel_nome,
        local_responsavel_telefone=local_responsavel_telefone, observacoes=observacoes,
        modo_criacao="manual"
    )

    def render_erro(mensagem: str):
        return templates.TemplateResponse("admin/contrato_novo.html", {
            "request": request, "empresa": empresa, "produtos": produtos, "contratos": contratos,
            "erro": mensagem, "form": form, "modo_edicao": True, "item": item
        }, status_code=400)

    telefone_limpo = limpar_identificador(telefone)
    cpf_limpo = limpar_identificador(cpf)
    cnpj_limpo = limpar_identificador(cnpj)

    if not nome.strip():
        return render_erro("Informe o nome do cliente.")
    if not telefone_limpo and not cpf_limpo and not cnpj_limpo:
        return render_erro("Informe pelo menos telefone, CPF ou CNPJ.")
    if not hora_meia_em_meia_valida(hora_inicio):
        return render_erro("A hora precisa estar em intervalo de 30 minutos. Exemplo: 18:00 ou 18:30.")
    if cpf_limpo and not cpf_valido(cpf_limpo):
        return render_erro("CPF inválido.")
    if cnpj_limpo and not cnpj_valido(cnpj_limpo):
        return render_erro("CNPJ inválido.")

    produto = db.get(ProdutoServico, int(produto_id)) if produto_id else None
    if produto and produto.empresa_id != empresa.id:
        raise HTTPException(404)

    cliente = item.cliente
    cliente.nome = nome.strip()
    cliente.telefone = telefone_limpo or telefone.strip()
    cliente.cpf = cpf_limpo
    cliente.cnpj = cnpj_limpo
    cliente.email = email.strip()
    cliente.endereco = endereco.strip()
    cliente.numero = numero.strip()
    cliente.complemento = complemento.strip()
    cliente.bairro = bairro.strip()
    cliente.cidade = cidade.strip()
    cliente.estado = estado.strip()
    cliente.cep = cep.strip()

    inicio_obj = datetime.strptime(hora_inicio, "%H:%M").time()
    data_evento_obj = datetime.strptime(data_evento, "%Y-%m-%d").date()
    retirada_obrigatoria_bool = bool(retirada_obrigatoria)
    retirada_data_obj = datetime.strptime(retirada_data, "%Y-%m-%d").date() if retirada_data else data_evento_obj
    retirada_hora_obj = datetime.strptime(retirada_hora, "%H:%M").time() if retirada_hora else None
    valor_float = texto_para_float(valor)
    sinal_float = texto_para_float(sinal)

    item.produto_id = produto.id if produto else None
    item.contrato_id = int(contrato_id) if contrato_id else (produto.contrato_id if produto and produto.contrato_id else None)
    item.data_evento = data_evento_obj
    item.hora_inicio = inicio_obj
    item.hora_fim = somar_minutos(inicio_obj, produto.duracao_minutos or 240) if produto else item.hora_fim
    item.retirada_obrigatoria = retirada_obrigatoria_bool
    item.retirada_data = retirada_data_obj if retirada_obrigatoria_bool else None
    item.retirada_hora = retirada_hora_obj or (item.hora_fim or item.hora_inicio if retirada_obrigatoria_bool else None)
    item.bairro = bairro.strip()
    item.local = endereco.strip() or local.strip()
    item.local_numero = numero.strip()
    item.local_complemento = complemento.strip()
    item.local_cidade = cidade.strip()
    item.local_estado = estado.strip()
    item.local_cep = cep.strip()
    item.local_nome = local_nome.strip()
    item.local_responsavel_nome = local_responsavel_nome.strip()
    item.local_responsavel_telefone = limpar_identificador(
        local_responsavel_telefone) or local_responsavel_telefone.strip()
    item.acesso_local = acesso_local.strip()
    item.valor = valor_float
    item.sinal = sinal_float
    item.observacoes = observacoes.strip()
    salvar_endereco_cliente(
        db, empresa.id, cliente.id, item.local, item.local_numero, item.local_complemento,
        item.bairro, item.local_cidade, item.local_estado, item.local_cep, apelido=item.local_nome,
    )
    _invalidar_geocodificacao(item)

    if produto:
        item_principal = item.itens[0] if item.itens else None
        if not item_principal:
            item_principal = ReservaItem(empresa_id=empresa.id, solicitacao_id=item.id, quantidade=1)
            db.add(item_principal)
        item_principal.produto_id = produto.id
        item_principal.nome = produto.nome
        item_principal.descricao = produto.descricao
        item_principal.valor_unitario = valor_float
        item_principal.valor_total = valor_float

    if item.status == "aguardando_nova_data":
        retirar_solicitacao_da_operacao(db, item)
    else:
        criar_eventos_operacionais(db, item)
    db.commit()
    _tentar_geocodificar_solicitacao(db, item)
    return RedirectResponse(f"/painel/solicitacao/{item.id}", status_code=303)


@app.get("/painel/clientes", response_class=HTMLResponse)
def clientes(request: Request, busca: str = "", db: Session = Depends(get_db),
             empresa: Empresa = Depends(empresa_logada)):
    termo_texto = (busca or "").strip()
    termo_limpo = limpar_identificador(busca)
    itens = []
    if termo_texto:
        condicoes = [
            Cliente.nome.ilike(f"%{termo_texto}%"),
            Cliente.email.ilike(f"%{termo_texto}%"),
        ]
        if termo_limpo:
            condicoes.extend([
                Cliente.cpf.contains(termo_limpo),
                Cliente.cnpj.contains(termo_limpo),
                Cliente.telefone.contains(termo_limpo),
                Cliente.identificador.contains(termo_limpo),
            ])
        itens = (
            db.query(Cliente)
            .filter(Cliente.empresa_id == empresa.id)
            .filter(or_(*condicoes))
            .order_by(Cliente.nome)
            .all()
        )
    return templates.TemplateResponse("admin/clientes.html",
                                      {"request": request, "empresa": empresa, "itens": itens, "busca": busca})


@app.get("/painel/cliente/{cliente_id}", response_class=HTMLResponse)
def cliente_detalhe(cliente_id: int, request: Request, db: Session = Depends(get_db),
                    empresa: Empresa = Depends(empresa_logada)):
    cliente = db.get(Cliente, cliente_id)
    if not cliente or cliente.empresa_id != empresa.id:
        raise HTTPException(404)
    equipamentos = db.query(EquipamentoCliente).filter_by(empresa_id=empresa.id, cliente_id=cliente.id).order_by(
        EquipamentoCliente.nome).all()
    solicitacoes = db.query(Solicitacao).filter_by(empresa_id=empresa.id, cliente_id=cliente.id).order_by(
        Solicitacao.criado_em.desc()).all()
    produtos = db.query(ProdutoServico).filter_by(empresa_id=empresa.id, ativo=True).order_by(ProdutoServico.nome).all()
    contratos = db.query(Contrato).filter_by(empresa_id=empresa.id, ativo=True).order_by(Contrato.nome).all()
    return templates.TemplateResponse("admin/cliente_detalhe.html",
                                      {"request": request, "empresa": empresa, "cliente": cliente,
                                       "equipamentos": equipamentos, "solicitacoes": solicitacoes, "produtos": produtos,
                                       "contratos": contratos})


@app.post("/painel/solicitacao/{solicitacao_id}/usar-como-base")
def usar_solicitacao_como_base(
        solicitacao_id: int,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    origem = db.get(Solicitacao, solicitacao_id)
    if not origem or origem.empresa_id != empresa.id:
        raise HTTPException(404)

    nova = Solicitacao(
        empresa_id=empresa.id,
        cliente_id=origem.cliente_id,
        produto_id=origem.produto_id,
        contrato_id=origem.contrato_id,
        data_evento=origem.data_evento,
        hora_inicio=origem.hora_inicio,
        hora_fim=origem.hora_fim,
        bairro=origem.bairro,
        local=origem.local,
        local_numero=origem.local_numero,
        local_complemento=origem.local_complemento,
        local_cidade=origem.local_cidade,
        local_estado=origem.local_estado,
        local_cep=origem.local_cep,
        local_nome=origem.local_nome,
        local_responsavel_nome=origem.local_responsavel_nome,
        local_responsavel_telefone=origem.local_responsavel_telefone,
        acesso_local=origem.acesso_local,
        valor=origem.valor,
        sinal=origem.sinal,
        valor_pago=0,
        sinal_recebido=False,
        observacoes=origem.observacoes,
        status="pre_reserva",
    )
    db.add(nova)
    db.flush()

    for it in origem.itens:
        db.add(ReservaItem(
            empresa_id=empresa.id,
            solicitacao_id=nova.id,
            produto_id=it.produto_id,
            nome=it.nome,
            descricao=it.descricao,
            quantidade=it.quantidade,
            valor_unitario=it.valor_unitario,
            valor_total=it.valor_total,
        ))

    db.commit()
    return RedirectResponse(f"/painel/solicitacao/{nova.id}/editar-completo?copiado=1", status_code=303)


@app.post("/painel/cliente/{cliente_id}/dados")
def atualizar_cliente_dados(
        cliente_id: int,
        nome: str = Form(""),
        telefone: str = Form(""),
        cpf: str = Form(""),
        cnpj: str = Form(""),
        email: str = Form(""),
        endereco: str = Form(""),
        numero: str = Form(""),
        complemento: str = Form(""),
        bairro: str = Form(""),
        cidade: str = Form(""),
        estado: str = Form(""),
        cep: str = Form(""),
        observacoes: str = Form(""),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    cliente = db.get(Cliente, cliente_id)
    if not cliente or cliente.empresa_id != empresa.id:
        raise HTTPException(404)

    cliente.nome = nome.strip() or cliente.nome
    cliente.telefone = limpar_identificador(telefone)
    cliente.cpf = limpar_identificador(cpf)
    cliente.cnpj = limpar_identificador(cnpj)
    cliente.email = email.strip()
    cliente.endereco = endereco.strip()
    cliente.numero = numero.strip()
    cliente.complemento = complemento.strip()
    cliente.bairro = bairro.strip()
    cliente.cidade = cidade.strip()
    cliente.estado = estado.strip()
    cliente.cep = limpar_identificador(cep)
    cliente.observacoes = observacoes.strip()

    if empresa.identificador_principal == "cpf" and cliente.cpf:
        cliente.identificador = cliente.cpf
    elif empresa.identificador_principal == "cnpj" and cliente.cnpj:
        cliente.identificador = cliente.cnpj
    elif cliente.telefone:
        cliente.identificador = cliente.telefone

    db.commit()
    return RedirectResponse(f"/painel/cliente/{cliente.id}", status_code=303)


@app.post("/painel/cliente/{cliente_id}/pre-reserva-rapida")
def criar_pre_reserva_rapida(
        cliente_id: int,
        produto_id: str = Form(""),
        contrato_id: str = Form(""),
        data_evento: str = Form(""),
        hora_inicio: str = Form(""),
        valor: str = Form("0"),
        sinal: str = Form("0"),
        local_nome: str = Form(""),
        local: str = Form(""),
        local_responsavel_nome: str = Form(""),
        local_responsavel_telefone: str = Form(""),
        observacoes: str = Form(""),
        acao: str = Form("salvar"),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    cliente = db.get(Cliente, cliente_id)
    if not cliente or cliente.empresa_id != empresa.id:
        raise HTTPException(404)
    produto = db.get(ProdutoServico, int(produto_id)) if produto_id else None
    if produto and produto.empresa_id != empresa.id:
        raise HTTPException(404)
    inicio_obj = datetime.strptime(hora_inicio, "%H:%M").time()
    endereco_texto = (local or cliente.endereco or "").strip()
    historico = None
    if endereco_texto:
        historico = (
            db.query(EnderecoCliente)
            .filter(
                EnderecoCliente.empresa_id == empresa.id,
                EnderecoCliente.cliente_id == cliente.id,
                func.lower(EnderecoCliente.endereco) == endereco_texto.lower(),
            )
            .order_by(EnderecoCliente.atualizado_em.desc(), EnderecoCliente.id.desc())
            .first()
        )
    usar_atual = bool(cliente.endereco and cliente.endereco.strip().casefold() == endereco_texto.casefold())
    numero_evento = (historico.numero if historico else cliente.numero if usar_atual else "") or ""
    complemento_evento = (historico.complemento if historico else cliente.complemento if usar_atual else "") or ""
    bairro_evento = (historico.bairro if historico else cliente.bairro if usar_atual else "") or ""
    cidade_evento = (historico.cidade if historico else cliente.cidade if usar_atual else "") or ""
    estado_evento = (historico.estado if historico else cliente.estado if usar_atual else "") or ""
    cep_evento = (historico.cep if historico else cliente.cep if usar_atual else "") or ""
    item = Solicitacao(
        empresa_id=empresa.id,
        cliente_id=cliente.id,
        produto_id=produto.id if produto else None,
        contrato_id=int(contrato_id) if contrato_id else (produto.contrato_id if produto and produto.contrato_id else None),
        data_evento=datetime.strptime(data_evento, "%Y-%m-%d").date(),
        hora_inicio=inicio_obj,
        hora_fim=somar_minutos(inicio_obj, produto.duracao_minutos or 240) if produto else None,
        bairro=bairro_evento.strip(),
        local=endereco_texto,
        local_numero=numero_evento.strip(),
        local_complemento=complemento_evento.strip(),
        local_cidade=cidade_evento.strip(),
        local_estado=estado_evento.strip(),
        local_cep=cep_evento.strip(),
        local_nome=local_nome,
        local_responsavel_nome=local_responsavel_nome,
        local_responsavel_telefone=local_responsavel_telefone,
        valor=texto_para_float(valor),
        sinal=texto_para_float(sinal),
        observacoes=observacoes,
        status="aguardando_aceite"
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    if produto:
        db.add(ReservaItem(
            empresa_id=empresa.id,
            solicitacao_id=item.id,
            produto_id=produto.id,
            nome=produto.nome,
            descricao=produto.descricao,
            quantidade=1,
            valor_unitario=texto_para_float(valor),
            valor_total=texto_para_float(valor)
        ))
        db.commit()
    return RedirectResponse(f"/painel/solicitacao/{item.id}", status_code=303)


@app.post("/painel/cliente/{cliente_id}/equipamentos")
def salvar_equipamento_cliente(
        cliente_id: int,
        nome: str = Form(...), marca: str = Form(""), modelo: str = Form(""), numero_serie: str = Form(""),
        observacoes: str = Form(""),
        acao: str = Form("salvar"),
        db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)
):
    cliente = db.get(Cliente, cliente_id)
    if not cliente or cliente.empresa_id != empresa.id:
        raise HTTPException(404)
    db.add(EquipamentoCliente(
        empresa_id=empresa.id, cliente_id=cliente.id, nome=nome.strip(), marca=marca.strip(),
        modelo=modelo.strip(), numero_serie=numero_serie.strip(), observacoes=observacoes.strip()
    ))
    db.commit()
    return RedirectResponse(f"/painel/cliente/{cliente_id}", status_code=303)


def usuario_pode_financeiro(request: Request, empresa: Empresa, db: Session) -> bool:
    usuario_sistema = request.session.get("usuario_sistema")
    if usuario_sistema and empresa.usuario_admin and usuario_sistema.lower() == empresa.usuario_admin.lower():
        return True
    usuario = db.query(UsuarioEmpresa).filter_by(empresa_id=empresa.id,
                                                 usuario=usuario_sistema).first() if usuario_sistema else None
    return True if not usuario else bool(getattr(usuario, "visualiza_financeiro", True))


def garantir_contas_financeiras(db: Session, empresa_id: int):
    # Caminho normal: uma única consulta. A criação automática só ocorre em base nova.
    contas = (
        db.query(ContaFinanceira)
        .filter_by(empresa_id=empresa_id, ativa=True)
        .order_by(ContaFinanceira.id)
        .all()
    )
    if contas:
        return contas
    existe_alguma = db.query(ContaFinanceira.id).filter_by(empresa_id=empresa_id).first()
    if not existe_alguma:
        for nome, tipo in [("Banco Principal", "banco"), ("Dinheiro", "dinheiro"), ("Cartão", "cartao")]:
            db.add(ContaFinanceira(empresa_id=empresa_id, nome=nome, tipo=tipo, saldo_inicial=0))
        db.commit()
        return (
            db.query(ContaFinanceira)
            .filter_by(empresa_id=empresa_id, ativa=True)
            .order_by(ContaFinanceira.id)
            .all()
        )
    return []


def parse_valor_banco(valor) -> float:
    if valor is None:
        return 0.0
    texto = str(valor).strip().replace("R$", "").replace(" ", "")
    if not texto:
        return 0.0
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except Exception:
        return 0.0


def parse_data_banco(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, fmt).date()
        except Exception:
            pass
    return None


def texto_normalizado_financeiro(valor: str) -> str:
    texto = (valor or "").strip().lower()
    trocas = str.maketrans("áàâãäéèêëíìîïóòôõöúùûüçñ", "aaaaaeeeeiiiiooooouuuucn")
    texto = texto.translate(trocas)
    texto = re.sub(r"\s+", " ", texto)
    return texto


def categoria_sugerida(historico: str, valor: float) -> str:
    h = texto_normalizado_financeiro(historico)
    if any(p in h for p in
           ["uber", "tim", "claro", "vivo", "light", "enel", "internet", "telefone", "google", "meta", "facebook",
            "conta azul", "mei", "simples", "taxa", "tarifa", "maquininha", "stone", "mercado pago", "nic br",
            "hospedagem", "dominio"]):
        return "empresa"
    if any(p in h for p in
           ["mercado", "farmacia", "padaria", "ifood", "restaurante", "posto", "combustivel", "condominio",
            "aluguel casa"]):
        return "casa"
    if any(p in h for p in ["agua", "aguas", "manut", "reparo", "peca", "assistencia"]):
        return "manutencao"
    if valor > 0 or "pix recebido" in h or "pix devolvido" in h:
        return "aluguel"
    return "empresa"


def hash_lancamento_banco(empresa_id: int, conta_id: int, data_lanc, historico: str, documento: str, valor: float,
                          saldo: float) -> str:
    base = "|".join([
        str(empresa_id), str(conta_id), str(data_lanc),
        texto_normalizado_financeiro(historico), texto_normalizado_financeiro(documento),
        f"{float(valor or 0):.2f}", f"{float(saldo or 0):.2f}"
    ])
    return hashlib.sha256(base.encode("utf-8")).hexdigest()


def melhores_vinculos_para_banco(lancamento, pagamentos, limite=5):
    return melhores_vinculos_financeiros(
        data_lanc=lancamento.data,
        texto_lanc=lancamento.historico,
        valor_lanc=lancamento.valor,
        pagamentos=pagamentos,
        limite=limite
    )


def melhores_vinculos_para_manual(lancamento, pagamentos, limite=5):
    return melhores_vinculos_financeiros(
        data_lanc=lancamento.data,
        texto_lanc=lancamento.descricao,
        valor_lanc=lancamento.valor,
        pagamentos=pagamentos,
        limite=limite
    )


def melhores_vinculos_financeiros(data_lanc, texto_lanc, valor_lanc, pagamentos, limite=5):
    if (valor_lanc or 0) <= 0:
        return []
    hist = texto_normalizado_financeiro(texto_lanc)
    candidatos = []
    for p in pagamentos:
        nome = texto_normalizado_financeiro(
            getattr(p.solicitacao.cliente, "nome", "")) if p.solicitacao and p.solicitacao.cliente else ""
        diff_valor = abs(float(valor_lanc or 0) - float(p.valor or 0))
        diff_dias = abs((data_lanc - p.data_pagamento).days) if data_lanc and p.data_pagamento else 99
        nome_score = SequenceMatcher(None, hist, nome).ratio() if nome else 0
        if nome and nome in hist:
            nome_score = max(nome_score, 0.95)
        score = 0
        if diff_valor < 0.01:
            score += 100
        else:
            score += max(0, 45 - min(diff_valor, 45))
        score += max(0, 30 - min(diff_dias, 30))
        score += nome_score * 40
        if diff_valor <= 10 or diff_dias <= 3 or nome_score >= .55:
            candidatos.append({"pagamento": p, "score": score, "diff_valor": diff_valor, "diff_dias": diff_dias})
    return sorted(candidatos, key=lambda x: (-x["score"], x["diff_valor"], x["diff_dias"]))[:limite]


def ler_extrato_upload(upload: UploadFile):
    nome = upload.filename or "extrato"
    conteudo = upload.file.read()
    linhas = []
    if nome.lower().endswith(".csv"):
        texto = conteudo.decode("utf-8-sig", errors="ignore")
        amostra = texto[:2048]
        delimitador = ";" if amostra.count(";") > amostra.count(",") else ","
        leitor = csv.reader(StringIO(texto), delimiter=delimitador)
        linhas = [linha for linha in leitor]
    else:
        try:
            from openpyxl import load_workbook
        except Exception:
            raise HTTPException(400, "Para importar XLSX, instale openpyxl ou envie o extrato em CSV.")
        wb = load_workbook(BytesIO(conteudo), data_only=True)
        ws = wb.active
        linhas = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]

    cabecalho_idx = None
    for idx, linha in enumerate(linhas):
        normal = [str(c or "").strip().lower() for c in linha]
        if "data" in normal and any("hist" in c for c in normal) and any("valor" in c for c in normal):
            cabecalho_idx = idx
            break
    if cabecalho_idx is None:
        raise HTTPException(400, "Não encontrei as colunas Data, Histórico, Valor e Saldo no extrato.")

    cab = [str(c or "").strip().lower() for c in linhas[cabecalho_idx]]

    def achar(nome):
        for i, c in enumerate(cab):
            if nome in c:
                return i
        return -1

    i_data = achar("data")
    i_hist = next((i for i, c in enumerate(cab) if "hist" in c), -1)
    i_doc = achar("documento")
    i_valor = achar("valor")
    i_saldo = achar("saldo")
    registros = []
    for linha in linhas[cabecalho_idx + 1:]:
        if not linha or len(linha) <= max(i_data, i_hist, i_valor):
            continue
        data_lanc = parse_data_banco(linha[i_data])
        historico = str(linha[i_hist] or "").strip()
        valor = parse_valor_banco(linha[i_valor])
        if not data_lanc or not historico:
            continue
        saldo = parse_valor_banco(linha[i_saldo]) if i_saldo >= 0 and len(linha) > i_saldo else 0
        documento = str(linha[i_doc] or "").strip() if i_doc >= 0 and len(linha) > i_doc else ""
        registros.append(
            {"data": data_lanc, "historico": historico, "documento": documento, "valor": valor, "saldo": saldo})
    return registros


def garantir_ordem_financeira(db: Session, empresa_id: int):
    # Preenche a ordem dos registros antigos. A ordem fica editável depois pelos botões ↑/↓.
    alterou = False
    for obj in db.query(LancamentoBanco).filter(LancamentoBanco.empresa_id == empresa_id,
                                                (LancamentoBanco.ordem == None) | (LancamentoBanco.ordem == 0)).all():
        obj.ordem = obj.id or 0
        alterou = True
    for obj in db.query(LancamentoManualFinanceiro).filter(LancamentoManualFinanceiro.empresa_id == empresa_id,
                                                           (LancamentoManualFinanceiro.ordem == None) | (
                                                                   LancamentoManualFinanceiro.ordem == 0)).all():
        obj.ordem = obj.id or 0
        alterou = True
    if alterou:
        db.commit()


def mover_lancamento_na_lista(db: Session, modelo, lanc, direcao: str):
    if direcao not in ["cima", "baixo"]:
        raise HTTPException(400, "Direção inválida.")
    base = db.query(modelo).filter(
        modelo.empresa_id == lanc.empresa_id,
        modelo.conta_id == lanc.conta_id,
        modelo.data == lanc.data,
    )
    if hasattr(modelo, "tipo"):
        base = base.filter(modelo.tipo == getattr(lanc, "tipo", "real"))
    linhas = base.order_by(modelo.ordem.asc(), modelo.id.asc()).all()
    pos = next((i for i, item in enumerate(linhas) if item.id == lanc.id), -1)
    if pos < 0:
        return
    destino = pos - 1 if direcao == "cima" else pos + 1
    if destino < 0 or destino >= len(linhas):
        return
    outro = linhas[destino]
    atual_ordem = lanc.ordem or lanc.id or 0
    outra_ordem = outro.ordem or outro.id or 0
    lanc.ordem, outro.ordem = outra_ordem, atual_ordem
    db.commit()


def melhores_vinculos_organiza(lancamento, registros, tipo: str, limite: int = 12):
    """Sugere lançamentos do Organiza compatíveis com o lançamento bancário."""
    if not lancamento or (lancamento.valor or 0) <= 0:
        return []
    historico_origem = getattr(lancamento, "historico", None) or getattr(lancamento, "descricao", None) or ""
    historico = texto_normalizado_financeiro(historico_origem)
    candidatos = []
    tipo_normalizado = (tipo or "").strip().lower().replace("ç", "c").replace("ã", "a")
    if tipo_normalizado == "manutencao":
        tipo_normalizado = "manutencao"

    for item in registros:
        item_tipo = (item.tipo or "").strip().lower().replace("ç", "c").replace("ã", "a")
        if item_tipo != tipo_normalizado:
            continue
        nome = texto_normalizado_financeiro(item.cliente or "")
        descricao = texto_normalizado_financeiro(item.descricao or "")
        alvo = " ".join(x for x in [nome, descricao] if x)
        diff_valor = abs(float(lancamento.valor or 0) - float(item.valor or 0))
        diff_dias = abs((lancamento.data - item.data_pagamento).days) if lancamento.data and item.data_pagamento else 99
        score_nome = max(
            SequenceMatcher(None, historico, nome).ratio() if nome else 0,
            SequenceMatcher(None, historico, alvo).ratio() if alvo else 0,
        )
        # Valor e proximidade de data têm peso maior; nome ajuda a ordenar.
        score = (1 / (1 + diff_valor)) * 4 + (1 / (1 + diff_dias)) * 2 + score_nome
        candidatos.append({
            "item": item,
            "diff_valor": diff_valor,
            "diff_dias": diff_dias,
            "score": score,
        })
    candidatos.sort(key=lambda c: (-c["score"], c["diff_valor"], c["diff_dias"]))
    return candidatos[:limite]


@app.get("/painel/financeiro", response_class=HTMLResponse)
def financeiro(
        request: Request,
        conta_id: int = 0,
        data_inicial: str = "",
        data_final: str = "",
        categoria: str = "",
        busca: str = "",
        status_sistema: str = "pendente",
        mes_cards: str = "",
        semana_cards: str = "",
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    if not usuario_pode_financeiro(request, empresa, db):
        raise HTTPException(403, "Usuário sem permissão para visualizar o financeiro.")

    # GET financeiro é somente leitura. A correção de ordem antiga não roda em toda abertura.
    contas = garantir_contas_financeiras(db, empresa.id)
    conta = next((c for c in contas if c.id == conta_id), None) if conta_id else (contas[0] if contas else None)

    hoje = date.today()
    data_inicial = data_inicial or hoje.replace(day=1).isoformat()
    data_final = data_final or hoje.isoformat()
    inicio = datetime.strptime(data_inicial, "%Y-%m-%d").date()
    fim = datetime.strptime(data_final, "%Y-%m-%d").date()

    # Períodos independentes dos cards: mês vigente no topo e semana vigente (segunda a domingo) no rodapé.
    def primeiro_dia_mes(valor: date) -> date:
        return valor.replace(day=1)

    def avancar_mes(valor: date, quantidade: int) -> date:
        indice = (valor.year * 12 + valor.month - 1) + quantidade
        return date(indice // 12, indice % 12 + 1, 1)

    mes_vigente = primeiro_dia_mes(hoje)
    try:
        mes_cards_inicio = datetime.strptime(mes_cards, "%Y-%m").date().replace(day=1) if mes_cards else mes_vigente
    except ValueError:
        mes_cards_inicio = mes_vigente
    mes_cards_fim = avancar_mes(mes_cards_inicio, 1) - timedelta(days=1)
    meses_cards = [avancar_mes(mes_vigente, deslocamento) for deslocamento in range(3)]

    # Semanas do mês selecionado. A primeira e a última podem ser parciais,
    # garantindo que todos os contratos do mês apareçam em exatamente uma semana.
    semanas_cards = []
    cursor_semana = mes_cards_inicio
    while cursor_semana <= mes_cards_fim:
        dias_ate_domingo = 6 - cursor_semana.weekday()
        fim_periodo = min(cursor_semana + timedelta(days=dias_ate_domingo), mes_cards_fim)
        semanas_cards.append({"inicio": cursor_semana, "fim": fim_periodo})
        cursor_semana = fim_periodo + timedelta(days=1)

    semana_cards_inicio_solicitada = None
    try:
        if semana_cards:
            semana_cards_inicio_solicitada = datetime.strptime(semana_cards, "%Y-%m-%d").date()
    except ValueError:
        semana_cards_inicio_solicitada = None

    semana_selecionada = next(
        (periodo for periodo in semanas_cards
         if periodo["inicio"] == semana_cards_inicio_solicitada),
        None
    )
    if not semana_selecionada:
        semana_selecionada = next(
            (periodo for periodo in semanas_cards
             if periodo["inicio"] <= hoje <= periodo["fim"]),
            semanas_cards[0]
        )

    semana_cards_inicio = semana_selecionada["inicio"]
    semana_cards_fim = semana_selecionada["fim"]

    q_banco = db.query(LancamentoBanco).options(
        joinedload(LancamentoBanco.conta),
        joinedload(LancamentoBanco.repasse_solicitacao).joinedload(Solicitacao.cliente),
        joinedload(LancamentoBanco.pagamento)
        .joinedload(Pagamento.solicitacao)
        .joinedload(Solicitacao.cliente),
        joinedload(LancamentoBanco.organiza_lancamento),
        joinedload(LancamentoBanco.vinculos_repasse)
        .joinedload(VinculoRepasseBanco.solicitacao)
        .joinedload(Solicitacao.cliente),
        joinedload(LancamentoBanco.vinculos_repasse)
        .joinedload(VinculoRepasseBanco.solicitacao)
        .joinedload(Solicitacao.empresa_transferida),
    ).filter(LancamentoBanco.empresa_id == empresa.id)
    q_manual_real = db.query(LancamentoManualFinanceiro).options(
        joinedload(LancamentoManualFinanceiro.conta),
        joinedload(LancamentoManualFinanceiro.repasse_solicitacao).joinedload(Solicitacao.cliente),
        joinedload(LancamentoManualFinanceiro.pagamento)
        .joinedload(Pagamento.solicitacao)
        .joinedload(Solicitacao.cliente),
        joinedload(LancamentoManualFinanceiro.organiza_lancamento),
    ).filter(
        LancamentoManualFinanceiro.empresa_id == empresa.id,
        LancamentoManualFinanceiro.tipo == "real"
    )
    q_receber = db.query(LancamentoManualFinanceiro).filter(
        LancamentoManualFinanceiro.empresa_id == empresa.id,
        LancamentoManualFinanceiro.tipo == "receber",
        LancamentoManualFinanceiro.recebido == False
    )
    if conta:
        q_banco = q_banco.filter(LancamentoBanco.conta_id == conta.id)
        q_manual_real = q_manual_real.filter(LancamentoManualFinanceiro.conta_id == conta.id)
        q_receber = q_receber.filter(LancamentoManualFinanceiro.conta_id == conta.id)
    if data_inicial:
        q_banco = q_banco.filter(LancamentoBanco.data >= inicio)
        q_manual_real = q_manual_real.filter(LancamentoManualFinanceiro.data >= inicio)
        q_receber = q_receber.filter(LancamentoManualFinanceiro.data >= inicio)
    if data_final:
        q_banco = q_banco.filter(LancamentoBanco.data <= fim)
        q_manual_real = q_manual_real.filter(LancamentoManualFinanceiro.data <= fim)
        q_receber = q_receber.filter(LancamentoManualFinanceiro.data <= fim)
    if categoria == "sem_categoria":
        q_banco = q_banco.filter(or_(LancamentoBanco.categoria == None, LancamentoBanco.categoria == ""))
        q_manual_real = q_manual_real.filter(or_(
            LancamentoManualFinanceiro.categoria == None,
            LancamentoManualFinanceiro.categoria == ""
        ))
        q_receber = q_receber.filter(or_(
            LancamentoManualFinanceiro.categoria == None,
            LancamentoManualFinanceiro.categoria == ""
        ))
    elif categoria:
        q_banco = q_banco.filter(LancamentoBanco.categoria == categoria)
        q_manual_real = q_manual_real.filter(LancamentoManualFinanceiro.categoria == categoria)
        q_receber = q_receber.filter(LancamentoManualFinanceiro.categoria == categoria)
    valor_busca = None
    if busca:
        termo_busca = busca.strip()
        like = f"%{termo_busca}%"
        # O campo de procura aceita texto e também valor em formato brasileiro (ex.: 100,00 ou R$ 100,00).
        if re.fullmatch(r"[Rr$\s0-9.,-]+", termo_busca):
            try:
                valor_busca = texto_para_float(termo_busca.replace("R$", "").replace("r$", "").strip())
            except Exception:
                valor_busca = None

        filtro_banco = [LancamentoBanco.historico.ilike(like)]
        filtro_manual = [LancamentoManualFinanceiro.descricao.ilike(like)]
        if valor_busca is not None:
            filtro_banco.append(func.abs(LancamentoBanco.valor - valor_busca) < 0.01)
            filtro_manual.append(func.abs(LancamentoManualFinanceiro.valor - valor_busca) < 0.01)
        q_banco = q_banco.filter(or_(*filtro_banco))
        q_manual_real = q_manual_real.filter(or_(*filtro_manual))
        q_receber = q_receber.filter(or_(*filtro_manual))

    banco = q_banco.order_by(LancamentoBanco.data.desc(), LancamentoBanco.ordem.asc(), LancamentoBanco.id.asc()).all()
    manuais_reais = q_manual_real.order_by(LancamentoManualFinanceiro.data.desc(),
                                           LancamentoManualFinanceiro.ordem.asc(),
                                           LancamentoManualFinanceiro.id.asc()).all()
    receber = q_receber.order_by(LancamentoManualFinanceiro.data.asc(), LancamentoManualFinanceiro.id.asc()).all()

    q_contratos_receber = db.query(Solicitacao).options(
        joinedload(Solicitacao.cliente)
    ).join(Cliente).filter(
        Solicitacao.empresa_id == empresa.id,
        Solicitacao.cancelado_em == None,
        # O saldo do contrato só vira conta a receber depois do aceite/aprovação.
        # Pagamentos já lançados em rascunhos continuam disponíveis para conciliação,
        # mas o restante não é cobrado enquanto o contrato não estiver aprovado.
        Solicitacao.status.in_(STATUS_CONTRATO_APROVADO),
        (func.coalesce(Solicitacao.valor, 0) - func.coalesce(Solicitacao.valor_pago, 0)) > 0.009
    )
    if data_inicial:
        q_contratos_receber = q_contratos_receber.filter(Solicitacao.data_evento >= inicio)
    if data_final:
        q_contratos_receber = q_contratos_receber.filter(Solicitacao.data_evento <= fim)
    if busca:
        like = f"%{busca.strip()}%"
        filtros_contrato = [Cliente.nome.ilike(like)]
        if valor_busca is not None:
            filtros_contrato.extend([
                func.abs(Solicitacao.valor - valor_busca) < 0.01,
                func.abs((func.coalesce(Solicitacao.valor, 0) - func.coalesce(Solicitacao.valor_pago, 0)) - valor_busca) < 0.01,
            ])
        q_contratos_receber = q_contratos_receber.filter(or_(*filtros_contrato))
    contratos_receber = q_contratos_receber.order_by(Solicitacao.data_evento.asc(), Solicitacao.id.asc()).all()
    total_contratos_receber = sum(max((c.valor or 0) - (c.valor_pago or 0), 0) for c in contratos_receber)

    hoje = date.today()
    contratos_vencidos = [c for c in contratos_receber if c.data_evento and c.data_evento < hoje]
    contratos_em_dia = [c for c in contratos_receber if not c.data_evento or c.data_evento >= hoje]
    total_contratos_vencidos = sum(max((c.valor or 0) - (c.valor_pago or 0), 0) for c in contratos_vencidos)
    total_contratos_em_dia = sum(max((c.valor or 0) - (c.valor_pago or 0), 0) for c in contratos_em_dia)

    q_pagamentos_sistema = db.query(Pagamento).options(
        joinedload(Pagamento.solicitacao).joinedload(Solicitacao.cliente)
    ).join(Solicitacao).join(Cliente).filter(
        Pagamento.empresa_id == empresa.id
    )
    if data_inicial:
        q_pagamentos_sistema = q_pagamentos_sistema.filter(Pagamento.data_pagamento >= inicio)
    if data_final:
        q_pagamentos_sistema = q_pagamentos_sistema.filter(Pagamento.data_pagamento <= fim)
    if busca:
        like = f"%{busca.strip()}%"
        filtros_pagamento = [
            Cliente.nome.ilike(like),
            Pagamento.nome_comprovante.ilike(like),
        ]
        if valor_busca is not None:
            filtros_pagamento.append(func.abs(Pagamento.valor - valor_busca) < 0.01)
        q_pagamentos_sistema = q_pagamentos_sistema.filter(or_(*filtros_pagamento))

    pagamentos_sistema_mes = q_pagamentos_sistema.order_by(Pagamento.data_pagamento.desc(), Pagamento.id.desc()).all()
    total_contratos_pagos_mes = sum(float(p.valor or 0) for p in pagamentos_sistema_mes)

    if status_sistema == "vinculado":
        q_pagamentos_sistema = q_pagamentos_sistema.filter(Pagamento.conciliado_em != None)
    elif status_sistema != "todos":
        status_sistema = "pendente"
        q_pagamentos_sistema = q_pagamentos_sistema.filter(Pagamento.conciliado_em == None)

    pagamentos_sistema = q_pagamentos_sistema.order_by(Pagamento.data_pagamento.desc(), Pagamento.id.desc()).all()

    pagamentos_pendentes_vinculo = db.query(Pagamento).options(
        joinedload(Pagamento.solicitacao).joinedload(Solicitacao.cliente)
    ).join(Solicitacao).join(Cliente).filter(
        Pagamento.empresa_id == empresa.id,
        Pagamento.conciliado_em == None
    ).all()

    # Cards superiores: sempre obedecem somente ao pequeno seletor de mês.
    q_banco_cards = db.query(LancamentoBanco).filter(
        LancamentoBanco.empresa_id == empresa.id,
        LancamentoBanco.data >= mes_cards_inicio,
        LancamentoBanco.data <= mes_cards_fim
    )
    q_manual_cards = db.query(LancamentoManualFinanceiro).filter(
        LancamentoManualFinanceiro.empresa_id == empresa.id,
        LancamentoManualFinanceiro.data >= mes_cards_inicio,
        LancamentoManualFinanceiro.data <= mes_cards_fim
    )
    if conta:
        q_banco_cards = q_banco_cards.filter(LancamentoBanco.conta_id == conta.id)
        q_manual_cards = q_manual_cards.filter(LancamentoManualFinanceiro.conta_id == conta.id)

    banco_cards = q_banco_cards.all()
    manuais_cards = q_manual_cards.all()
    entradas = sum(float(l.valor or 0) for l in banco_cards if (l.valor or 0) > 0) + sum(
        float(l.valor or 0) for l in manuais_cards if l.tipo == "real" and (l.valor or 0) > 0)
    saidas = sum(abs(float(l.valor or 0)) for l in banco_cards if (l.valor or 0) < 0) + sum(
        abs(float(l.valor or 0)) for l in manuais_cards if l.tipo == "real" and (l.valor or 0) < 0)
    saldo_real = entradas - saidas
    total_receber = sum(
        max(float(l.valor or 0), 0) for l in manuais_cards if l.tipo == "receber" and not l.recebido)

    # Uma única consulta mensal alimenta os cards, o relatório semanal e a semana selecionada.
    contratos_mes = db.query(Solicitacao).filter(
        Solicitacao.empresa_id == empresa.id,
        Solicitacao.cancelado_em == None,
        Solicitacao.data_evento >= mes_cards_inicio,
        Solicitacao.data_evento <= mes_cards_fim,
    ).all()
    contratos_cards = [c for c in contratos_mes if c.status in STATUS_CONTRATO_APROVADO]
    contratos_cards_proprios = [c for c in contratos_cards if not c.empresa_transferida_id]
    contratos_cards_transferidos = [c for c in contratos_cards if c.empresa_transferida_id]
    quantidade_contratos_cards = len(contratos_cards)
    quantidade_contratos_cards_proprios = len(contratos_cards_proprios)
    quantidade_contratos_cards_transferidos = len(contratos_cards_transferidos)
    total_contratos_receber_cards = sum(
        max(float(c.valor or 0) - float(c.valor_pago or 0), 0) for c in contratos_cards)
    total_repasse_cards = sum(float(c.valor_repasse or 0) for c in contratos_cards_transferidos)

    # Acumulado do banco: duas consultas agrupadas para todas as contas.
    # Evita executar duas somas separadas para cada conta financeira.
    inicio_ano = hoje.replace(month=1, day=1)
    totais_banco_por_conta = {
        conta_id_resultado: float(total or 0)
        for conta_id_resultado, total in db.query(
            LancamentoBanco.conta_id,
            func.coalesce(func.sum(LancamentoBanco.valor), 0),
        ).filter(
            LancamentoBanco.empresa_id == empresa.id,
            LancamentoBanco.data >= inicio_ano,
            LancamentoBanco.data <= hoje,
        ).group_by(LancamentoBanco.conta_id).all()
    }
    totais_manuais_por_conta = {
        conta_id_resultado: float(total or 0)
        for conta_id_resultado, total in db.query(
            LancamentoManualFinanceiro.conta_id,
            func.coalesce(func.sum(LancamentoManualFinanceiro.valor), 0),
        ).filter(
            LancamentoManualFinanceiro.empresa_id == empresa.id,
            LancamentoManualFinanceiro.tipo == "real",
            LancamentoManualFinanceiro.data >= inicio_ano,
            LancamentoManualFinanceiro.data <= hoje,
        ).group_by(LancamentoManualFinanceiro.conta_id).all()
    }

    def saldo_real_conta(conta_calculo):
        if not conta_calculo:
            return 0.0
        return (
            totais_banco_por_conta.get(conta_calculo.id, 0.0)
            + totais_manuais_por_conta.get(conta_calculo.id, 0.0)
        )

    saldo_banco = saldo_real_conta(conta)
    saldo_todos = sum(saldo_real_conta(c) for c in contas if c.ativa)

    # Relatório mensal por semana calculado em memória sobre a única consulta mensal.
    relatorio_semanal = []
    for indice, periodo in enumerate(semanas_cards, start=1):
        contratos_periodo = [
            c for c in contratos_mes
            if c.data_evento and periodo["inicio"] <= c.data_evento <= periodo["fim"]
        ]
        valor_total_periodo = sum(float(c.valor or 0) for c in contratos_periodo)
        valor_recebido_periodo = sum(min(float(c.valor_pago or 0), float(c.valor or 0)) for c in contratos_periodo)
        valor_receber_periodo = sum(
            max(float(c.valor or 0) - float(c.valor_pago or 0), 0) for c in contratos_periodo
        )
        contratos_proprios_periodo = [c for c in contratos_periodo if not c.empresa_transferida_id]
        contratos_transferidos_periodo = [c for c in contratos_periodo if c.empresa_transferida_id]
        relatorio_semanal.append({
            "numero": indice,
            "inicio": periodo["inicio"],
            "fim": periodo["fim"],
            "quantidade": len(contratos_periodo),
            "quantidade_proprios": len(contratos_proprios_periodo),
            "quantidade_transferidos": len(contratos_transferidos_periodo),
            "valor_total": valor_total_periodo,
            "valor_total_proprios": sum(float(c.valor or 0) for c in contratos_proprios_periodo),
            "valor_total_transferidos": sum(float(c.valor or 0) for c in contratos_transferidos_periodo),
            "valor_repasse": sum(float(c.valor_repasse or 0) for c in contratos_transferidos_periodo),
            "valor_recebido": valor_recebido_periodo,
            "valor_receber": valor_receber_periodo,
        })

    relatorio_total = {
        "quantidade": sum(item["quantidade"] for item in relatorio_semanal),
        "valor_total": sum(item["valor_total"] for item in relatorio_semanal),
        "valor_recebido": sum(item["valor_recebido"] for item in relatorio_semanal),
        "valor_receber": sum(item["valor_receber"] for item in relatorio_semanal),
        "quantidade_proprios": sum(item["quantidade_proprios"] for item in relatorio_semanal),
        "quantidade_transferidos": sum(item["quantidade_transferidos"] for item in relatorio_semanal),
        "valor_total_proprios": sum(item["valor_total_proprios"] for item in relatorio_semanal),
        "valor_total_transferidos": sum(item["valor_total_transferidos"] for item in relatorio_semanal),
        "valor_repasse": sum(item["valor_repasse"] for item in relatorio_semanal),
    }

    saldo_previsto = saldo_real + total_receber + total_contratos_receber_cards

    # Cards inferiores: reutilizam os contratos já carregados para o mês.
    contratos_semana = [
        c for c in contratos_mes
        if c.data_evento and semana_cards_inicio <= c.data_evento <= semana_cards_fim
    ]
    contratos_semana_proprios = [c for c in contratos_semana if not c.empresa_transferida_id]
    contratos_semana_transferidos = [c for c in contratos_semana if c.empresa_transferida_id]
    quantidade_contratos_semana = len(contratos_semana)
    quantidade_contratos_semana_proprios = len(contratos_semana_proprios)
    quantidade_contratos_semana_transferidos = len(contratos_semana_transferidos)
    valor_total_contratos_semana = sum(float(c.valor or 0) for c in contratos_semana)
    valor_total_contratos_semana_proprios = sum(float(c.valor or 0) for c in contratos_semana_proprios)
    valor_total_contratos_semana_transferidos = sum(float(c.valor or 0) for c in contratos_semana_transferidos)
    valor_receber_contratos_semana = sum(
        max(float(c.valor or 0) - float(c.valor_pago or 0), 0) for c in contratos_semana)
    valor_receber_contratos_semana_proprios = sum(
        max(float(c.valor or 0) - float(c.valor_pago or 0), 0) for c in contratos_semana_proprios)
    valor_receber_contratos_semana_transferidos = sum(
        max(float(c.valor or 0) - float(c.valor_pago or 0), 0) for c in contratos_semana_transferidos)
    valor_repasse_semana = sum(float(c.valor_repasse or 0) for c in contratos_semana_transferidos)

    q_repasses = db.query(Solicitacao).options(
        joinedload(Solicitacao.cliente),
        joinedload(Solicitacao.empresa_transferida),
    ).join(Cliente).filter(
        Solicitacao.empresa_id == empresa.id,
        Solicitacao.cancelado_em == None,
        Solicitacao.empresa_transferida_id != None,
        func.coalesce(Solicitacao.valor_repasse, 0) > 0
    )
    if data_inicial:
        q_repasses = q_repasses.filter(Solicitacao.data_evento >= inicio)
    if data_final:
        q_repasses = q_repasses.filter(Solicitacao.data_evento <= fim)
    if busca:
        like_repasse = f"%{busca.strip()}%"
        filtros_repasse = [Cliente.nome.ilike(like_repasse)]
        if valor_busca is not None:
            filtros_repasse.append(func.abs(Solicitacao.valor_repasse - abs(valor_busca)) < 0.01)
        q_repasses = q_repasses.filter(or_(*filtros_repasse))
    repasses_base = q_repasses.order_by(Solicitacao.data_evento.desc(), Solicitacao.id.desc()).all()

    # O status do repasse é calculado pelo total efetivamente vinculado no banco.
    vinculos_repasse_todos = db.query(VinculoRepasseBanco).options(
        joinedload(VinculoRepasseBanco.solicitacao).joinedload(Solicitacao.cliente),
        joinedload(VinculoRepasseBanco.solicitacao).joinedload(Solicitacao.empresa_transferida),
    ).filter(
        VinculoRepasseBanco.empresa_id == empresa.id
    ).all()
    valor_vinculado_por_repasse = {}
    vinculos_por_banco = {}
    for vr in vinculos_repasse_todos:
        valor_vinculado_por_repasse[vr.solicitacao_id] = valor_vinculado_por_repasse.get(vr.solicitacao_id, 0.0) + float(vr.valor or 0)
        vinculos_por_banco.setdefault(vr.lancamento_banco_id, []).append(vr)

    def status_repasse(item):
        pago = valor_vinculado_por_repasse.get(item.id, 0.0)
        total = float(item.valor_repasse or 0)
        if pago >= total - 0.01:
            return "vinculado"
        if pago > 0.01:
            return "parcial"
        return "pendente"

    if status_sistema == "vinculado":
        repasses_sistema = [r for r in repasses_base if status_repasse(r) == "vinculado"]
    elif status_sistema == "pendente":
        repasses_sistema = [r for r in repasses_base if status_repasse(r) != "vinculado"]
    else:
        repasses_sistema = repasses_base

    repasses_pendentes = [
        r for r in repasses_base
        if valor_vinculado_por_repasse.get(r.id, 0.0) < float(r.valor_repasse or 0) - 0.01
    ]

    # Ao marcar uma saída como "Repasse", o vínculo passa a ser feito pelo lado do Banco.
    candidatos_repasse_por_banco = {}
    saldo_repasse_por_banco = {}
    for l in banco:
        if (l.valor or 0) >= 0 or l.categoria != "repasse" or l.pagamento_id:
            continue
        usado = sum(float(v.valor or 0) for v in vinculos_por_banco.get(l.id, []))
        saldo_disponivel = max(abs(float(l.valor or 0)) - usado, 0)
        saldo_repasse_por_banco[l.id] = saldo_disponivel
        if saldo_disponivel > 0.01:
            candidatos_repasse_por_banco[l.id] = sorted(
                repasses_pendentes,
                key=lambda r: (
                    abs((float(r.valor_repasse or 0) - valor_vinculado_por_repasse.get(r.id, 0.0)) - saldo_disponivel),
                    abs((l.data - r.data_evento).days if l.data and r.data_evento else 9999)
                )
            )[:20]

    candidatos_vinculo = {
        l.id: melhores_vinculos_para_banco(l, pagamentos_pendentes_vinculo)
        for l in banco
        if not l.pagamento_id and not getattr(l, "organiza_lancamento_id", None) and l.categoria == "aluguel"
    }

    # Organiza: carrega vínculos e registros uma única vez.
    lancamentos_banco_organiza = db.query(LancamentoBanco).options(
        joinedload(LancamentoBanco.organiza_lancamento)
    ).filter(
        LancamentoBanco.empresa_id == empresa.id,
        LancamentoBanco.organiza_lancamento_id != None,
    ).all()
    lancamentos_manuais_organiza = db.query(LancamentoManualFinanceiro).options(
        joinedload(LancamentoManualFinanceiro.organiza_lancamento)
    ).filter(
        LancamentoManualFinanceiro.empresa_id == empresa.id,
        LancamentoManualFinanceiro.organiza_lancamento_id != None,
    ).all()

    ids_organiza_vinculados = {
        l.organiza_lancamento_id
        for l in (*lancamentos_banco_organiza, *lancamentos_manuais_organiza)
        if l.organiza_lancamento_id
    }
    todos_lancamentos_organiza = (
        db.query(LancamentoOrganiza)
        .filter(LancamentoOrganiza.empresa_id == empresa.id)
        .order_by(LancamentoOrganiza.data_pagamento.desc(), LancamentoOrganiza.id.desc())
        .all()
    )
    registros_organiza_disponiveis = [
        item for item in todos_lancamentos_organiza
        if item.id not in ids_organiza_vinculados
    ]
    candidatos_organiza = {
        l.id: melhores_vinculos_organiza(l, registros_organiza_disponiveis, l.categoria)
        for l in banco
        if not l.pagamento_id
        and not getattr(l, "organiza_lancamento_id", None)
        and l.categoria in ("venda", "manutencao")
        and (l.valor or 0) > 0
    }
    bancos_por_organiza = {
        l.organiza_lancamento_id: l
        for l in (*lancamentos_banco_organiza, *lancamentos_manuais_organiza)
        if l.organiza_lancamento_id
    }

    candidatos_manual = {
        m.id: melhores_vinculos_para_manual(m, pagamentos_pendentes_vinculo)
        for m in manuais_reais
        if not getattr(m, "pagamento_id", None) and m.categoria == "aluguel" and (m.valor or 0) > 0
    }
    candidatos_manual_organiza = {
        m.id: melhores_vinculos_organiza(m, registros_organiza_disponiveis, m.categoria)
        for m in manuais_reais
        if not getattr(m, "pagamento_id", None)
        and not getattr(m, "organiza_lancamento_id", None)
        and m.categoria in ("venda", "manutencao")
        and (m.valor or 0) > 0
    }

    # Organiza fica separado dos lançamentos nativos do Connect.
    # Reutiliza a consulta já realizada e limita somente a exibição.
    lancamentos_organiza_financeiro = todos_lancamentos_organiza[:500]

    # Transferências internas recebidas: na empresa de destino o valor já pago pelo
    # cliente deixa de ser "a receber do cliente" e passa a ser "a receber da empresa de origem".
    copias_transferencia = db.query(Solicitacao).options(
        joinedload(Solicitacao.cliente)
    ).filter(
        Solicitacao.empresa_id == empresa.id,
        Solicitacao.cancelado_em == None,
        Solicitacao.transferencia_origem_id != None,
    ).all()
    origens_ids = [c.transferencia_origem_id for c in copias_transferencia if c.transferencia_origem_id]
    origens_transferencia = {}
    empresas_origem = {}
    pagos_por_origem = {}
    if origens_ids:
        origens = db.query(Solicitacao).filter(Solicitacao.id.in_(origens_ids)).all()
        origens_transferencia = {o.id: o for o in origens}
        empresa_ids_origem = {o.empresa_id for o in origens}
        if empresa_ids_origem:
            empresas_origem = {
                e.id: e for e in db.query(Empresa).filter(Empresa.id.in_(empresa_ids_origem)).all()
            }
        totais_pago = db.query(
            VinculoRepasseBanco.solicitacao_id,
            func.coalesce(func.sum(VinculoRepasseBanco.valor), 0),
        ).filter(
            VinculoRepasseBanco.solicitacao_id.in_(origens_ids)
        ).group_by(VinculoRepasseBanco.solicitacao_id).all()
        pagos_por_origem = {sid: float(total or 0) for sid, total in totais_pago}

    repasses_receber_interempresa = []
    for copia in copias_transferencia:
        origem = origens_transferencia.get(copia.transferencia_origem_id)
        if not origem:
            continue
        total = max(float(origem.valor_repasse or 0), 0)
        pago = min(pagos_por_origem.get(origem.id, 0.0), total)
        repasses_receber_interempresa.append({
            "copia": copia,
            "origem": origem,
            "empresa_origem": empresas_origem.get(origem.empresa_id),
            "total": total,
            "pago": pago,
            "saldo": max(total - pago, 0),
        })

    return templates.TemplateResponse("admin/financeiro.html", {
        "request": request, "empresa": empresa, "contas": contas, "conta": conta,
        "data_inicial": data_inicial, "data_final": data_final, "categoria": categoria, "busca": busca,
        "status_sistema": status_sistema,
        "mes_cards": mes_cards_inicio.strftime("%Y-%m"), "meses_cards": meses_cards,
        "mes_cards_inicio": mes_cards_inicio, "mes_cards_fim": mes_cards_fim,
        "semana_cards": semana_cards_inicio.isoformat(), "semanas_cards": semanas_cards,
        "semana_cards_inicio": semana_cards_inicio, "semana_cards_fim": semana_cards_fim,
        "timedelta": timedelta,
        "banco": banco, "manuais_reais": manuais_reais, "receber": receber, "pagamentos_sistema": pagamentos_sistema,
        "contratos_receber": contratos_receber, "total_contratos_receber": total_contratos_receber,
        "quantidade_contratos_cards": quantidade_contratos_cards,
        "quantidade_contratos_cards_proprios": quantidade_contratos_cards_proprios,
        "quantidade_contratos_cards_transferidos": quantidade_contratos_cards_transferidos,
        "total_contratos_receber_cards": total_contratos_receber_cards,
        "total_repasse_cards": total_repasse_cards,
        "quantidade_contratos_semana": quantidade_contratos_semana,
        "quantidade_contratos_semana_proprios": quantidade_contratos_semana_proprios,
        "quantidade_contratos_semana_transferidos": quantidade_contratos_semana_transferidos,
        "valor_total_contratos_semana": valor_total_contratos_semana,
        "valor_total_contratos_semana_proprios": valor_total_contratos_semana_proprios,
        "valor_total_contratos_semana_transferidos": valor_total_contratos_semana_transferidos,
        "valor_receber_contratos_semana": valor_receber_contratos_semana,
        "valor_receber_contratos_semana_proprios": valor_receber_contratos_semana_proprios,
        "valor_receber_contratos_semana_transferidos": valor_receber_contratos_semana_transferidos,
        "valor_repasse_semana": valor_repasse_semana,
        "contratos_vencidos": contratos_vencidos, "contratos_em_dia": contratos_em_dia,
        "total_contratos_vencidos": total_contratos_vencidos, "total_contratos_em_dia": total_contratos_em_dia,
        "pagamentos_sistema_mes": pagamentos_sistema_mes, "total_contratos_pagos_mes": total_contratos_pagos_mes,
        "entradas": entradas, "saidas": saidas, "saldo_real": saldo_real, "total_receber": total_receber,
        "saldo_previsto": saldo_previsto, "saldo_banco": saldo_banco, "saldo_todos": saldo_todos,
        "relatorio_semanal": relatorio_semanal, "relatorio_total": relatorio_total,
        "candidatos_vinculo": candidatos_vinculo,
        "candidatos_manual": candidatos_manual,
        "candidatos_manual_organiza": candidatos_manual_organiza,
        "candidatos_organiza": candidatos_organiza,
        "bancos_por_organiza": bancos_por_organiza,
        "repasses_sistema": repasses_sistema,
        "repasses_receber_interempresa": repasses_receber_interempresa,
                "valor_vinculado_por_repasse": valor_vinculado_por_repasse,
        "vinculos_por_banco": vinculos_por_banco,
        "candidatos_repasse_por_banco": candidatos_repasse_por_banco,
        "saldo_repasse_por_banco": saldo_repasse_por_banco,
        "lancamentos_organiza_financeiro": lancamentos_organiza_financeiro,
        "categorias": [("casa", "Casa"), ("empresa", "Empresa"), ("aluguel", "Aluguel"), ("venda", "Venda"), ("manutencao", "Manutenção"), ("repasse", "Repasse")]
    })



def _relatorio_financeiro_mensal(db: Session, empresa_id: int, mes_ref: str):
    try:
        inicio_mes = datetime.strptime(mes_ref, "%Y-%m").date().replace(day=1)
    except ValueError:
        raise HTTPException(400, "Mês inválido.")
    indice = inicio_mes.year * 12 + inicio_mes.month
    fim_mes = date(indice // 12, indice % 12 + 1, 1) - timedelta(days=1)

    semanas = []
    cursor = inicio_mes
    numero = 1
    while cursor <= fim_mes:
        fim_semana = min(cursor + timedelta(days=6 - cursor.weekday()), fim_mes)
        contratos = db.query(Solicitacao).filter(
            Solicitacao.empresa_id == empresa_id,
            Solicitacao.cancelado_em == None,
            Solicitacao.data_evento >= cursor,
            Solicitacao.data_evento <= fim_semana
        ).all()
        contratos_proprios = [c for c in contratos if not c.empresa_transferida_id]
        contratos_transferidos = [c for c in contratos if c.empresa_transferida_id]
        valor_total = sum(float(c.valor or 0) for c in contratos)
        recebido = sum(min(float(c.valor_pago or 0), float(c.valor or 0)) for c in contratos)
        receber = sum(max(float(c.valor or 0) - float(c.valor_pago or 0), 0) for c in contratos)
        semanas.append({
            "numero": numero,
            "inicio": cursor,
            "fim": fim_semana,
            "quantidade": len(contratos),
            "quantidade_proprios": len(contratos_proprios),
            "quantidade_transferidos": len(contratos_transferidos),
            "valor_total": valor_total,
            "valor_total_proprios": sum(float(c.valor or 0) for c in contratos_proprios),
            "valor_total_transferidos": sum(float(c.valor or 0) for c in contratos_transferidos),
            "valor_repasse": sum(float(c.valor_repasse or 0) for c in contratos_transferidos),
            "valor_recebido": recebido,
            "valor_receber": receber,
        })
        cursor = fim_semana + timedelta(days=1)
        numero += 1

    total = {
        "quantidade": sum(s["quantidade"] for s in semanas),
        "valor_total": sum(s["valor_total"] for s in semanas),
        "valor_recebido": sum(s["valor_recebido"] for s in semanas),
        "valor_receber": sum(s["valor_receber"] for s in semanas),
        "quantidade_proprios": sum(s["quantidade_proprios"] for s in semanas),
        "quantidade_transferidos": sum(s["quantidade_transferidos"] for s in semanas),
        "valor_total_proprios": sum(s["valor_total_proprios"] for s in semanas),
        "valor_total_transferidos": sum(s["valor_total_transferidos"] for s in semanas),
        "valor_repasse": sum(s["valor_repasse"] for s in semanas),
    }
    return inicio_mes, fim_mes, semanas, total


def _xlsx_relatorio_financeiro(inicio_mes, semanas, total):
    # Gera um XLSX simples e válido sem dependência adicional.
    linhas = [
        ["Relatório financeiro mensal", "", "", "", "", "", "", "", "", "", ""],
        [inicio_mes.strftime("%m/%Y"), "", "", "", "", "", "", "", "", "", ""],
        ["Semana", "Período", "Qtd. total", "Qtd. próprios", "Qtd. transferidos", "Valor total", "Valor próprios", "Valor transferidos", "Repasse a pagar", "Recebido", "A receber"],
    ]
    for item in semanas:
        linhas.append([
            f"Semana {item['numero']}",
            f"{item['inicio'].strftime('%d/%m/%Y')} a {item['fim'].strftime('%d/%m/%Y')}",
            item["quantidade"],
            item["quantidade_proprios"],
            item["quantidade_transferidos"],
            item["valor_total"],
            item["valor_total_proprios"],
            item["valor_total_transferidos"],
            item["valor_repasse"],
            item["valor_recebido"],
            item["valor_receber"],
        ])
    linhas.append([
        "TOTAL DO MÊS", "",
        total["quantidade"], total["quantidade_proprios"], total["quantidade_transferidos"],
        total["valor_total"], total["valor_total_proprios"], total["valor_total_transferidos"],
        total["valor_repasse"], total["valor_recebido"], total["valor_receber"],
    ])

    def coluna_excel(numero):
        resultado = ""
        while numero:
            numero, resto = divmod(numero - 1, 26)
            resultado = chr(65 + resto) + resultado
        return resultado

    cells = []
    for r, linha in enumerate(linhas, start=1):
        for c, valor in enumerate(linha, start=1):
            ref = f"{coluna_excel(c)}{r}"
            if isinstance(valor, (int, float)):
                estilo = ' s="2"' if c >= 6 else ' s="1"'
                cells.append(f'<c r="{ref}"{estilo}><v>{valor}</v></c>')
            else:
                estilo = ' s="3"' if r == 1 else (' s="4"' if r in (3, len(linhas)) else '')
                cells.append(f'<c r="{ref}" t="inlineStr"{estilo}><is><t>{xml_escape(str(valor))}</t></is></c>')

    rows_xml = []
    idx = 0
    for r, linha in enumerate(linhas, start=1):
        quantidade = len(linha)
        rows_xml.append(f'<row r="{r}">' + "".join(cells[idx:idx+quantidade]) + '</row>')
        idx += quantidade

    sheet_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<cols><col min="1" max="1" width="18" customWidth="1"/><col min="2" max="2" width="28" customWidth="1"/>
<col min="3" max="5" width="16" customWidth="1"/><col min="6" max="11" width="18" customWidth="1"/></cols>
<sheetData>{''.join(rows_xml)}</sheetData>
<mergeCells count="2"><mergeCell ref="A1:K1"/><mergeCell ref="A2:K2"/></mergeCells>
</worksheet>'''
    styles_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<numFmts count="1"><numFmt numFmtId="164" formatCode="R$ #,##0.00"/></numFmts>
<fonts count="2"><font><sz val="11"/><name val="Calibri"/></font><font><b/><sz val="12"/><name val="Calibri"/></font></fonts>
<fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>
<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>
<cellXfs count="5">
<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
<xf numFmtId="164" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>
<xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0"/>
<xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0"/>
</cellXfs></styleSheet>'''
    arquivos = {
        "[Content_Types].xml": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>''',
        "_rels/.rels": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>''',
        "xl/workbook.xml": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
<sheets><sheet name="Relatório mensal" sheetId="1" r:id="rId1"/></sheets></workbook>''',
        "xl/_rels/workbook.xml.rels": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>''',
        "xl/worksheets/sheet1.xml": sheet_xml,
        "xl/styles.xml": styles_xml,
    }
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as pacote:
        for nome, conteudo in arquivos.items():
            pacote.writestr(nome, conteudo)
    return buffer.getvalue()


@app.get("/painel/financeiro/relatorio-mensal.xlsx")
def financeiro_relatorio_excel(
        request: Request,
        mes: str,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    if not usuario_pode_financeiro(request, empresa, db):
        raise HTTPException(403, "Usuário sem permissão para visualizar o financeiro.")
    inicio_mes, _, semanas, total = _relatorio_financeiro_mensal(db, empresa.id, mes)
    conteudo = _xlsx_relatorio_financeiro(inicio_mes, semanas, total)
    nome = f"relatorio-financeiro-{inicio_mes.strftime('%Y-%m')}.xlsx"
    return Response(
        conteudo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'}
    )


@app.get("/painel/financeiro/relatorio-mensal.pdf")
def financeiro_relatorio_pdf(
        request: Request,
        mes: str,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    if not usuario_pode_financeiro(request, empresa, db):
        raise HTTPException(403, "Usuário sem permissão para visualizar o financeiro.")
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except Exception:
        raise HTTPException(500, "Para gerar PDF, instale a dependência reportlab.")

    inicio_mes, _, semanas, total = _relatorio_financeiro_mensal(db, empresa.id, mes)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=28, leftMargin=28, topMargin=28, bottomMargin=28)
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(f"Relatório financeiro mensal - {inicio_mes.strftime('%m/%Y')}", estilos["Title"]),
        Spacer(1, 14),
    ]
    dados = [["Semana", "Período", "Qtd.", "Próprios", "Transf.", "Valor total", "Valor próprios", "Valor transf.", "Repasse", "Recebido", "A receber"]]
    moeda = lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    for item in semanas:
        dados.append([
            f"Semana {item['numero']}",
            f"{item['inicio'].strftime('%d/%m/%Y')} a {item['fim'].strftime('%d/%m/%Y')}",
            str(item["quantidade"]),
            str(item["quantidade_proprios"]),
            str(item["quantidade_transferidos"]),
            moeda(item["valor_total"]),
            moeda(item["valor_total_proprios"]),
            moeda(item["valor_total_transferidos"]),
            moeda(item["valor_repasse"]),
            moeda(item["valor_recebido"]),
            moeda(item["valor_receber"]),
        ])
    dados.append([
        "TOTAL DO MÊS", "", str(total["quantidade"]), str(total["quantidade_proprios"]),
        str(total["quantidade_transferidos"]), moeda(total["valor_total"]), moeda(total["valor_total_proprios"]),
        moeda(total["valor_total_transferidos"]), moeda(total["valor_repasse"]), moeda(total["valor_recebido"]),
        moeda(total["valor_receber"])
    ])
    tabela = Table(dados, colWidths=[58, 100, 38, 48, 42, 68, 68, 68, 68, 68, 68], repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAF2FF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F8FAFC")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F8FAFC")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
    ]))
    elementos.append(tabela)
    doc.build(elementos)
    nome = f"relatorio-financeiro-{inicio_mes.strftime('%Y-%m')}.pdf"
    return Response(
        buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'}
    )


@app.post("/painel/financeiro/conta")
def financeiro_salvar_conta(
        request: Request,
        conta_id: int = Form(0),
        saldo_inicial: str = Form("0"),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    conta = db.get(ContaFinanceira, conta_id)
    if not conta or conta.empresa_id != empresa.id:
        raise HTTPException(404)
    conta.saldo_inicial = texto_para_float(saldo_inicial)
    db.commit()
    return redirect_preservando_filtros(request, f"/painel/financeiro?conta_id={conta.id}")


@app.post("/painel/financeiro/importar")
def financeiro_importar_extrato(
        request: Request,
        conta_id: int = Form(...),
        arquivo: UploadFile = File(...),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    conta = db.get(ContaFinanceira, conta_id)
    if not conta or conta.empresa_id != empresa.id:
        raise HTTPException(404)
    registros = ler_extrato_upload(arquivo)
    importados = 0
    duplicados = 0
    conciliados = 0
    hashes_do_arquivo = set()
    proxima_ordem = int(db.query(func.coalesce(func.max(LancamentoBanco.ordem), 0)).filter_by(empresa_id=empresa.id,
                                                                                              conta_id=conta.id).scalar() or 0) + 1
    for idx_registro, r in enumerate(registros):
        h = hash_lancamento_banco(empresa.id, conta.id, r["data"], r["historico"], r["documento"], r["valor"],
                                  r["saldo"])
        if h in hashes_do_arquivo:
            duplicados += 1
            continue
        hashes_do_arquivo.add(h)
        existe = db.query(LancamentoBanco).filter(
            LancamentoBanco.empresa_id == empresa.id,
            LancamentoBanco.conta_id == conta.id,
            (LancamentoBanco.hash_importacao == h) | (
                    (LancamentoBanco.data == r["data"]) &
                    (LancamentoBanco.historico == r["historico"]) &
                    (LancamentoBanco.documento == r["documento"]) &
                    (LancamentoBanco.valor == r["valor"]) &
                    (LancamentoBanco.saldo == r["saldo"])
            )
        ).first()
        if existe:
            if existe.pagamento_id:
                conciliados += 1
            else:
                duplicados += 1
            if not getattr(existe, "hash_importacao", None):
                existe.hash_importacao = h
            continue
        db.add(LancamentoBanco(
            empresa_id=empresa.id,
            conta_id=conta.id,
            data=r["data"],
            historico=r["historico"],
            documento=r["documento"],
            valor=r["valor"],
            saldo=r["saldo"],
            categoria=categoria_sugerida(r["historico"], r["valor"]),
            categoria_confirmada=False,
            hash_importacao=h,
            origem_importacao=arquivo.filename,
            ordem=proxima_ordem + idx_registro
        ))
        importados += 1
    db.commit()
    return redirect_preservando_filtros(request, f"/painel/financeiro?conta_id={conta.id}",
                                        {"importados": importados, "duplicados": duplicados,
                                         "conciliados": conciliados})


@app.post("/painel/financeiro/banco/{lancamento_id}/categoria")
def financeiro_categoria_banco(
        request: Request,
        lancamento_id: int,
        categoria: str = Form(...),
        confirmado: str = Form("0"),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoBanco, lancamento_id)
    if not lanc or lanc.empresa_id != empresa.id:
        raise HTTPException(404)
    if categoria not in ["casa", "empresa", "aluguel", "venda", "manutencao", "repasse"]:
        raise HTTPException(400, "Categoria inválida.")
    if categoria != "repasse":
        possui_rateio = db.query(VinculoRepasseBanco).filter(
            VinculoRepasseBanco.lancamento_banco_id == lanc.id
        ).first()
        if possui_rateio:
            raise HTTPException(400, "Desvincule os repasses deste lançamento antes de alterar a categoria.")
    lanc.categoria = categoria
    lanc.categoria_confirmada = confirmado == "1"
    db.commit()
    return RedirectResponse(request.headers.get("referer") or "/painel/financeiro", status_code=303)


@app.post("/painel/financeiro/banco/{lancamento_id}/vincular-repasse")
def financeiro_vincular_repasse_banco(
        request: Request,
        lancamento_id: int,
        solicitacao_id: int = Form(...),
        modo: str = Form("total"),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoBanco, lancamento_id)
    repasse = db.get(Solicitacao, solicitacao_id)
    if not lanc or lanc.empresa_id != empresa.id or not repasse or repasse.empresa_id != empresa.id or not repasse.empresa_transferida_id:
        raise HTTPException(404)
    if (lanc.valor or 0) >= 0 or lanc.pagamento_id or lanc.categoria != "repasse":
        raise HTTPException(400, "Marque esta saída com a categoria Repasse antes de vincular.")

    usado_banco = db.query(func.coalesce(func.sum(VinculoRepasseBanco.valor), 0)).filter(
        VinculoRepasseBanco.lancamento_banco_id == lanc.id
    ).scalar() or 0
    saldo_banco = max(abs(float(lanc.valor or 0)) - float(usado_banco), 0)

    ja_vinculado_repasse = db.query(func.coalesce(func.sum(VinculoRepasseBanco.valor), 0)).filter(
        VinculoRepasseBanco.solicitacao_id == repasse.id
    ).scalar() or 0
    saldo_repasse = max(float(repasse.valor_repasse or 0) - float(ja_vinculado_repasse), 0)

    if saldo_banco <= 0.01:
        raise HTTPException(400, "Este lançamento bancário não possui saldo disponível.")
    if saldo_repasse <= 0.01:
        raise HTTPException(400, "Este repasse já está totalmente pago.")

    if modo == "total":
        if saldo_banco + 0.01 < saldo_repasse:
            raise HTTPException(400, "O saldo deste lançamento é menor que o total pendente do repasse. Use a opção 'Usar saldo'.")
        valor_vinculo = saldo_repasse
    elif modo == "saldo":
        valor_vinculo = min(saldo_banco, saldo_repasse)
    else:
        raise HTTPException(400, "Modo de vínculo inválido.")

    existente = db.query(VinculoRepasseBanco).filter(
        VinculoRepasseBanco.lancamento_banco_id == lanc.id,
        VinculoRepasseBanco.solicitacao_id == repasse.id
    ).first()
    if existente:
        existente.valor = float(existente.valor or 0) + valor_vinculo
    else:
        db.add(VinculoRepasseBanco(
            empresa_id=empresa.id,
            lancamento_banco_id=lanc.id,
            solicitacao_id=repasse.id,
            valor=valor_vinculo,
            criado_por=request.session.get("usuario_nome") or "Financeiro"
        ))

    total_apos = float(ja_vinculado_repasse) + valor_vinculo
    if total_apos >= float(repasse.valor_repasse or 0) - 0.01:
        repasse.repasse_pago_em = agora_utc()
        repasse.repasse_pago_por = request.session.get("usuario_nome") or "Financeiro"
    else:
        repasse.repasse_pago_em = None
        repasse.repasse_pago_por = None

    # Campo legado fica vazio; a partir daqui o vínculo aceita vários contratos por lançamento.
    lanc.repasse_solicitacao_id = None
    db.commit()
    return RedirectResponse(request.headers.get("referer") or "/painel/financeiro", status_code=303)


@app.post("/painel/financeiro/banco/{lancamento_id}/repasse/{solicitacao_id}/desvincular")
def financeiro_desvincular_repasse_banco(
        request: Request,
        lancamento_id: int,
        solicitacao_id: int,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoBanco, lancamento_id)
    repasse = db.get(Solicitacao, solicitacao_id)
    if not lanc or lanc.empresa_id != empresa.id or not repasse or repasse.empresa_id != empresa.id:
        raise HTTPException(404)

    vinculo = db.query(VinculoRepasseBanco).filter(
        VinculoRepasseBanco.lancamento_banco_id == lanc.id,
        VinculoRepasseBanco.solicitacao_id == repasse.id,
        VinculoRepasseBanco.empresa_id == empresa.id
    ).first()
    if vinculo:
        db.delete(vinculo)
    db.flush()

    total_restante = db.query(func.coalesce(func.sum(VinculoRepasseBanco.valor), 0)).filter(
        VinculoRepasseBanco.solicitacao_id == repasse.id
    ).scalar() or 0
    if float(total_restante) < float(repasse.valor_repasse or 0) - 0.01:
        repasse.repasse_pago_em = None
        repasse.repasse_pago_por = None
    db.commit()
    return RedirectResponse(request.headers.get("referer") or "/painel/financeiro", status_code=303)


@app.post("/painel/financeiro/banco/{lancamento_id}/vincular-organiza")
def financeiro_vincular_organiza(
        request: Request,
        lancamento_id: int,
        organiza_id: int = Form(...),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoBanco, lancamento_id)
    item = db.get(LancamentoOrganiza, organiza_id)
    if not lanc or lanc.empresa_id != empresa.id or not item or item.empresa_id != empresa.id:
        raise HTTPException(404)
    if lanc.pagamento_id or lanc.organiza_lancamento_id:
        raise HTTPException(400, "Este lançamento bancário já está vinculado.")
    if lanc.categoria not in ("venda", "manutencao") or lanc.categoria != (item.tipo or "").lower():
        raise HTTPException(400, "O tipo do banco deve corresponder ao lançamento do Organiza.")
    ja_usado = db.query(LancamentoBanco).filter(
        LancamentoBanco.empresa_id == empresa.id,
        LancamentoBanco.organiza_lancamento_id == item.id
    ).first()
    if ja_usado:
        raise HTTPException(400, "Este lançamento do Organiza já está vinculado a outro movimento bancário.")
    lanc.organiza_lancamento_id = item.id
    lanc.categoria_confirmada = True
    db.commit()
    return RedirectResponse(request.headers.get("referer") or "/painel/financeiro", status_code=303)


@app.post("/painel/financeiro/banco/{lancamento_id}/desvincular-organiza")
def financeiro_desvincular_organiza(
        request: Request,
        lancamento_id: int,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoBanco, lancamento_id)
    if not lanc or lanc.empresa_id != empresa.id:
        raise HTTPException(404)
    lanc.organiza_lancamento_id = None
    db.commit()
    return RedirectResponse(request.headers.get("referer") or "/painel/financeiro", status_code=303)


@app.post("/painel/financeiro/banco/{lancamento_id}/vincular")
def financeiro_vincular_banco(
        request: Request,
        lancamento_id: int,
        pagamento_id: int = Form(...),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoBanco, lancamento_id)
    pagamento = db.get(Pagamento, pagamento_id)
    if not lanc or lanc.empresa_id != empresa.id or not pagamento or pagamento.empresa_id != empresa.id:
        raise HTTPException(404)
    lanc.pagamento_id = pagamento.id
    lanc.categoria = "aluguel"
    pagamento.conciliado_em = agora_utc()
    pagamento.conciliado_por = request.session.get("usuario_nome") or "Financeiro"
    db.commit()
    return RedirectResponse(request.headers.get("referer") or "/painel/financeiro", status_code=303)


@app.post("/painel/financeiro/banco/{lancamento_id}/desvincular")
def financeiro_desvincular_banco(
        request: Request,
        lancamento_id: int,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoBanco, lancamento_id)
    if not lanc or lanc.empresa_id != empresa.id:
        raise HTTPException(404)
    if lanc.pagamento:
        lanc.pagamento.conciliado_em = None
        lanc.pagamento.conciliado_por = None
    lanc.pagamento_id = None
    db.commit()
    return RedirectResponse(request.headers.get("referer") or "/painel/financeiro", status_code=303)


@app.post("/painel/financeiro/banco/{lancamento_id}/excluir")
def financeiro_excluir_banco(
        request: Request,
        lancamento_id: int,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoBanco, lancamento_id)
    if not lanc or lanc.empresa_id != empresa.id:
        raise HTTPException(404)
    if lanc.pagamento:
        lanc.pagamento.conciliado_em = None
        lanc.pagamento.conciliado_por = None
    if getattr(lanc, "repasse_solicitacao", None):
        lanc.repasse_solicitacao.repasse_pago_em = None
        lanc.repasse_solicitacao.repasse_pago_por = None
    db.delete(lanc)
    db.commit()
    return RedirectResponse(request.headers.get("referer") or "/painel/financeiro", status_code=303)


@app.post("/painel/financeiro/banco/{lancamento_id}/mover")
def financeiro_mover_banco(
        request: Request,
        lancamento_id: int,
        direcao: str = Form(...),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoBanco, lancamento_id)
    if not lanc or lanc.empresa_id != empresa.id:
        raise HTTPException(404)
    mover_lancamento_na_lista(db, LancamentoBanco, lanc, direcao)
    return RedirectResponse(request.headers.get("referer") or "/painel/financeiro", status_code=303)


@app.post("/painel/financeiro/sistema/{pagamento_id}/lancar")
def financeiro_lancar_pagamento_sistema(
        request: Request,
        pagamento_id: int,
        conta_id: int = Form(...),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    conta = db.get(ContaFinanceira, conta_id)
    pagamento = db.get(Pagamento, pagamento_id)
    if not conta or conta.empresa_id != empresa.id or not pagamento or pagamento.empresa_id != empresa.id:
        raise HTTPException(404)

    existente_banco = db.query(LancamentoBanco).filter_by(empresa_id=empresa.id, pagamento_id=pagamento.id).first()
    existente_manual = db.query(LancamentoManualFinanceiro).filter_by(empresa_id=empresa.id, pagamento_id=pagamento.id, tipo="real").first()
    if not existente_banco and not existente_manual:
        proxima_ordem = int(db.query(func.coalesce(func.max(LancamentoManualFinanceiro.ordem), 0)).filter_by(
            empresa_id=empresa.id, conta_id=conta.id).scalar() or 0) + 1
        cliente_nome = pagamento.solicitacao.cliente.nome if pagamento.solicitacao and pagamento.solicitacao.cliente else "Cliente"
        forma = (pagamento.forma_pagamento or "pagamento").strip()
        db.add(LancamentoManualFinanceiro(
            empresa_id=empresa.id,
            conta_id=conta.id,
            data=pagamento.data_pagamento,
            descricao=f"{cliente_nome} - {forma}",
            valor=pagamento.valor or 0,
            categoria="aluguel",
            tipo="real",
            recebido=False,
            pagamento_id=pagamento.id,
            ordem=proxima_ordem
        ))
    pagamento.conciliado_em = agora_utc()
    pagamento.conciliado_por = request.session.get("usuario_nome") or "Financeiro"
    db.commit()
    return RedirectResponse(request.headers.get("referer") or f"/painel/financeiro?conta_id={conta.id}", status_code=303)


@app.post("/painel/financeiro/manual")
def financeiro_lancamento_manual(
        request: Request,
        conta_id: int = Form(...),
        data: str = Form(...),
        descricao: str = Form(...),
        valor: str = Form("0"),
        categoria: str = Form("empresa"),
        tipo: str = Form("real"),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    conta = db.get(ContaFinanceira, conta_id)
    if not conta or conta.empresa_id != empresa.id:
        raise HTTPException(404)
    if categoria not in ["casa", "empresa", "aluguel", "venda", "manutencao", "repasse"]:
        raise HTTPException(400, "Categoria inválida.")
    if tipo not in ["real", "receber"]:
        raise HTTPException(400, "Tipo inválido.")
    valor_float = texto_para_float(valor)
    proxima_ordem = int(
        db.query(func.coalesce(func.max(LancamentoManualFinanceiro.ordem), 0)).filter_by(empresa_id=empresa.id,
                                                                                         conta_id=conta.id).scalar() or 0) + 1
    if tipo == "receber" and valor_float < 0:
        valor_float = abs(valor_float)
    db.add(LancamentoManualFinanceiro(
        empresa_id=empresa.id,
        conta_id=conta.id,
        data=datetime.strptime(data, "%Y-%m-%d").date(),
        descricao=descricao.strip(),
        valor=valor_float,
        categoria=categoria,
        tipo=tipo,
        recebido=False,
        ordem=proxima_ordem
    ))
    db.commit()
    return redirect_preservando_filtros(request, f"/painel/financeiro?conta_id={conta.id}")


@app.post("/painel/financeiro/manual/{lancamento_id}/mover")
def financeiro_mover_manual(
        request: Request,
        lancamento_id: int,
        direcao: str = Form(...),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoManualFinanceiro, lancamento_id)
    if not lanc or lanc.empresa_id != empresa.id or lanc.tipo != "real":
        raise HTTPException(404)
    mover_lancamento_na_lista(db, LancamentoManualFinanceiro, lanc, direcao)
    return RedirectResponse(request.headers.get("referer") or f"/painel/financeiro?conta_id={lanc.conta_id}",
                            status_code=303)


@app.post("/painel/financeiro/manual/{lancamento_id}/editar")
def financeiro_editar_manual(
        request: Request,
        lancamento_id: int,
        data: str = Form(...),
        descricao: str = Form(...),
        valor: str = Form("0"),
        categoria: str = Form("empresa"),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoManualFinanceiro, lancamento_id)
    if not lanc or lanc.empresa_id != empresa.id:
        raise HTTPException(404)
    if categoria not in ["casa", "empresa", "aluguel", "venda", "manutencao", "repasse"]:
        raise HTTPException(400, "Categoria inválida.")
    lanc.data = datetime.strptime(data, "%Y-%m-%d").date()
    lanc.descricao = descricao.strip()
    lanc.valor = texto_para_float(valor)
    lanc.categoria = categoria
    if categoria != "aluguel" and getattr(lanc, "pagamento_id", None):
        pagamento = db.get(Pagamento, lanc.pagamento_id)
        if pagamento:
            pagamento.conciliado_em = None
        lanc.pagamento_id = None
    if categoria not in ("venda", "manutencao") and getattr(lanc, "organiza_lancamento_id", None):
        lanc.organiza_lancamento_id = None
    db.commit()
    return redirect_preservando_filtros(request, f"/painel/financeiro?conta_id={lanc.conta_id}")


@app.post("/painel/financeiro/manual/{lancamento_id}/vincular-organiza")
def financeiro_vincular_manual_organiza(
        request: Request,
        lancamento_id: int,
        organiza_id: int = Form(...),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoManualFinanceiro, lancamento_id)
    item = db.get(LancamentoOrganiza, organiza_id)
    if not lanc or lanc.empresa_id != empresa.id or lanc.tipo != "real" or not item or item.empresa_id != empresa.id:
        raise HTTPException(404)
    if lanc.pagamento_id or getattr(lanc, "organiza_lancamento_id", None):
        raise HTTPException(400, "Este lançamento manual já está vinculado.")
    if lanc.categoria not in ("venda", "manutencao") or lanc.categoria != (item.tipo or "").lower():
        raise HTTPException(400, "O tipo deve corresponder ao lançamento do Organiza.")

    usado_banco = db.query(LancamentoBanco).filter(
        LancamentoBanco.empresa_id == empresa.id,
        LancamentoBanco.organiza_lancamento_id == item.id
    ).first()
    usado_manual = db.query(LancamentoManualFinanceiro).filter(
        LancamentoManualFinanceiro.empresa_id == empresa.id,
        LancamentoManualFinanceiro.organiza_lancamento_id == item.id
    ).first()
    if usado_banco or usado_manual:
        raise HTTPException(400, "Este lançamento do Organiza já está vinculado.")

    lanc.organiza_lancamento_id = item.id
    db.commit()
    return RedirectResponse(
        request.headers.get("referer") or f"/painel/financeiro?conta_id={lanc.conta_id}",
        status_code=303
    )


@app.post("/painel/financeiro/manual/{lancamento_id}/desvincular-organiza")
def financeiro_desvincular_manual_organiza(
        request: Request,
        lancamento_id: int,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoManualFinanceiro, lancamento_id)
    if not lanc or lanc.empresa_id != empresa.id:
        raise HTTPException(404)
    lanc.organiza_lancamento_id = None
    db.commit()
    return RedirectResponse(
        request.headers.get("referer") or f"/painel/financeiro?conta_id={lanc.conta_id}",
        status_code=303
    )


@app.post("/painel/financeiro/manual/{lancamento_id}/vincular")
def financeiro_vincular_manual(
        request: Request,
        lancamento_id: int,
        pagamento_id: int = Form(...),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoManualFinanceiro, lancamento_id)
    pagamento = db.get(Pagamento, pagamento_id)
    if not lanc or lanc.empresa_id != empresa.id or lanc.tipo != "real" or not pagamento or pagamento.empresa_id != empresa.id:
        raise HTTPException(404)
    lanc.pagamento_id = pagamento.id
    lanc.categoria = "aluguel"
    pagamento.conciliado_em = agora_utc()
    pagamento.conciliado_por = request.session.get("usuario_nome") or "Financeiro"
    db.commit()
    return RedirectResponse(request.headers.get("referer") or f"/painel/financeiro?conta_id={lanc.conta_id}",
                            status_code=303)


@app.post("/painel/financeiro/manual/{lancamento_id}/desvincular")
def financeiro_desvincular_manual(
        request: Request,
        lancamento_id: int,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoManualFinanceiro, lancamento_id)
    if not lanc or lanc.empresa_id != empresa.id or lanc.tipo != "real":
        raise HTTPException(404)
    if lanc.pagamento:
        lanc.pagamento.conciliado_em = None
        lanc.pagamento.conciliado_por = None
    lanc.pagamento_id = None
    db.commit()
    return RedirectResponse(request.headers.get("referer") or f"/painel/financeiro?conta_id={lanc.conta_id}",
                            status_code=303)


@app.post("/painel/financeiro/manual/{lancamento_id}/excluir")
def financeiro_excluir_manual(
        request: Request,
        lancamento_id: int,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoManualFinanceiro, lancamento_id)
    if not lanc or lanc.empresa_id != empresa.id or lanc.tipo != "real":
        raise HTTPException(404)
    conta_id = lanc.conta_id
    if lanc.pagamento:
        lanc.pagamento.conciliado_em = None
        lanc.pagamento.conciliado_por = None
    db.delete(lanc)
    db.commit()
    return RedirectResponse(request.headers.get("referer") or f"/painel/financeiro?conta_id={conta_id}",
                            status_code=303)


@app.post("/painel/financeiro/receber/{lancamento_id}/receber")
def financeiro_marcar_recebido(
        request: Request,
        lancamento_id: int,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    lanc = db.get(LancamentoManualFinanceiro, lancamento_id)
    if not lanc or lanc.empresa_id != empresa.id or lanc.tipo != "receber":
        raise HTTPException(404)
    lanc.recebido = True
    db.add(LancamentoManualFinanceiro(
        empresa_id=empresa.id,
        conta_id=lanc.conta_id,
        data=date.today(),
        descricao=f"Recebido: {lanc.descricao}",
        valor=abs(lanc.valor or 0),
        categoria=lanc.categoria,
        tipo="real",
        recebido=True
    ))
    db.commit()
    return redirect_preservando_filtros(request, f"/painel/financeiro?conta_id={lanc.conta_id}")


@app.post("/painel/solicitacao/{solicitacao_id}/pagamento")
def confirmar_pagamento(
        request: Request,
        solicitacao_id: int,
        data_pagamento: str = Form(""),
        valor_pago: str = Form("0"),
        forma_pagamento: str = Form("pix"),
        comprovante_no_nome_cliente: str = Form("sim"),
        nome_comprovante: str = Form(""),
        observacoes_pagamento: str = Form(""),
        retorno: str = Form(""),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)
    # Pagamento é opcional e independente do aceite.
    # Pode ser informado antes ou depois do contrato aceito; ele apenas gera o lançamento financeiro.
    valor = texto_para_float(valor_pago)
    if valor <= 0:
        return RedirectResponse(retorno or f"/painel/solicitacao/{solicitacao_id}", status_code=303)
    total_atual = sum((p.valor or 0) for p in getattr(item, "pagamentos", []) or [])
    validar_total_pagamentos(item, total_atual + valor)
    data_ref = datetime.strptime(data_pagamento, "%Y-%m-%d").date() if data_pagamento else date.today()
    no_nome = comprovante_no_nome_cliente == "sim"
    pagamento = Pagamento(
        empresa_id=empresa.id,
        solicitacao_id=item.id,
        data_pagamento=data_ref,
        valor=valor,
        forma_pagamento=forma_pagamento,
        comprovante_no_nome_cliente=no_nome,
        nome_comprovante=item.cliente.nome if no_nome else nome_comprovante.strip(),
        observacoes=observacoes_pagamento.strip(),
        usuario_registro=request.session.get("usuario_sistema", "Usuário")
    )
    db.add(pagamento)
    db.flush()
    recalcular_pagamento_solicitacao(db, item)
    # Não altera aceite do contrato. Pagamento muda apenas o resumo financeiro.
    db.commit()
    return RedirectResponse(retorno or f"/painel/solicitacao/{solicitacao_id}", status_code=303)


@app.post("/painel/solicitacao/{solicitacao_id}/pagamento/{pagamento_id}/excluir")
def excluir_pagamento_solicitacao(
        solicitacao_id: int,
        pagamento_id: int,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    item = db.get(Solicitacao, solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)

    pagamento = db.get(Pagamento, pagamento_id)
    if not pagamento or pagamento.empresa_id != empresa.id or pagamento.solicitacao_id != item.id:
        raise HTTPException(404)

    db.delete(pagamento)
    db.flush()
    recalcular_pagamento_solicitacao(db, item)

    # Excluir pagamento não altera aceite nem status do contrato.
    db.commit()
    return RedirectResponse(f"/painel/solicitacao/{solicitacao_id}", status_code=303)


@app.post("/painel/pagamento/{pagamento_id}/editar")
def editar_pagamento_financeiro(
        request: Request,
        pagamento_id: int,
        data_pagamento: str = Form(""),
        valor_pago: str = Form("0"),
        forma_pagamento: str = Form("pix"),
        nome_comprovante: str = Form(""),
        observacoes_pagamento: str = Form(""),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    pagamento = db.get(Pagamento, pagamento_id)
    if not pagamento or pagamento.empresa_id != empresa.id:
        raise HTTPException(404)
    item = db.get(Solicitacao, pagamento.solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)

    valor = texto_para_float(valor_pago)
    if valor <= 0:
        raise HTTPException(400, "O valor do pagamento precisa ser maior que zero.")

    total_sem_este = sum((p.valor or 0) for p in db.query(Pagamento).filter(
        Pagamento.empresa_id == empresa.id,
        Pagamento.solicitacao_id == item.id,
        Pagamento.id != pagamento.id
    ).all())
    validar_total_pagamentos(item, total_sem_este + valor)

    pagamento.data_pagamento = datetime.strptime(data_pagamento, "%Y-%m-%d").date() if data_pagamento else date.today()
    pagamento.valor = valor
    pagamento.forma_pagamento = forma_pagamento
    pagamento.nome_comprovante = nome_comprovante.strip() or (item.cliente.nome if item.cliente else "")
    pagamento.observacoes = observacoes_pagamento.strip()
    recalcular_pagamento_solicitacao(db, item)
    db.commit()
    voltar = request.headers.get("referer") or "/painel/financeiro"
    return RedirectResponse(voltar, status_code=303)


@app.post("/painel/pagamento/{pagamento_id}/excluir")
def excluir_pagamento_financeiro(
        request: Request,
        pagamento_id: int,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    pagamento = db.get(Pagamento, pagamento_id)
    if not pagamento or pagamento.empresa_id != empresa.id:
        raise HTTPException(404)
    item = db.get(Solicitacao, pagamento.solicitacao_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)
    db.delete(pagamento)
    db.flush()
    recalcular_pagamento_solicitacao(db, item)
    db.commit()
    voltar = request.headers.get("referer") or "/painel/financeiro"
    return RedirectResponse(voltar, status_code=303)


@app.post("/painel/pagamento/{pagamento_id}/conciliar")
def conciliar_pagamento_financeiro(
        request: Request,
        pagamento_id: int,
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    pagamento = db.get(Pagamento, pagamento_id)
    if not pagamento or pagamento.empresa_id != empresa.id:
        raise HTTPException(404)
    if not pagamento.conciliado_em:
        pagamento.conciliado_em = agora_utc()
        pagamento.conciliado_por = request.session.get("usuario_sistema") or request.session.get(
            "usuario_nome") or "Usuário"
        db.commit()
    voltar = request.headers.get("referer") or "/painel/financeiro"
    return RedirectResponse(voltar, status_code=303)


@app.get("/painel/disponibilidade", response_class=HTMLResponse)
def disponibilidade(request: Request, data: str = "", produto_id: int = 0, db: Session = Depends(get_db),
                    empresa: Empresa = Depends(empresa_logada)):
    data_consulta = datetime.strptime(data, "%Y-%m-%d").date() if data else date.today()

    produtos = db.query(ProdutoServico).filter_by(empresa_id=empresa.id, ativo=True).order_by(ProdutoServico.nome).all()

    # Reservas consideradas: locações ativas na data escolhida.
    status_ignorados = ["cancelada", "rejeitada"]
    reservas_do_dia = (
        db.query(Solicitacao)
        .options(
            joinedload(Solicitacao.cliente),
            selectinload(Solicitacao.itens),
        )
        .filter(Solicitacao.empresa_id == empresa.id)
        .filter(Solicitacao.data_evento == data_consulta)
        .filter(~Solicitacao.status.in_(status_ignorados))
        .all()
    )

    alugado_por_produto = {}
    locais_por_produto = {}
    for reserva in reservas_do_dia:
        for item in reserva.itens:
            chave = item.produto_id
            if not chave:
                continue
            alugado_por_produto[chave] = alugado_por_produto.get(chave, 0) + (item.quantidade or 1)
            locais_por_produto.setdefault(chave, []).append({
                "cliente": reserva.cliente.nome if reserva.cliente else "Cliente",
                "hora": reserva.hora_inicio.strftime("%H:%M") if reserva.hora_inicio else "-",
                "hora_ordenacao": reserva.hora_inicio or time.min,
                "bairro": (dados_endereco_solicitacao(reserva).get("bairro") or "-"),
                "quantidade": item.quantidade or 1,
                "reserva_id": reserva.id,
                "observacoes": ((reserva.observacoes or "") or (reserva.cliente.observacoes if reserva.cliente else "") or "").strip(),
                "retirada_obrigatoria": retirada_obrigatoria_ativa(reserva),
                "retirada_data": reserva.retirada_data,
                "retirada_hora": reserva.retirada_hora,
            })

    itens = []
    produto_selecionado = None
    for produto in produtos:
        total = produto.quantidade_disponivel or 0
        alugados = alugado_por_produto.get(produto.id, 0)
        disponiveis = max(total - alugados, 0)
        conflito = alugados > total
        status = "conflito" if conflito else ("disponivel" if disponiveis > 1 else ("atencao" if disponiveis == 1 else "indisponivel"))
        locais_ordenados = sorted(
            locais_por_produto.get(produto.id, []),
            key=lambda loc: (loc.get("hora_ordenacao") or time.min, loc.get("reserva_id") or 0)
        )
        dados = {
            "produto": produto,
            "total": total,
            "alugados": alugados,
            "disponiveis": disponiveis,
            "status": status,
            "conflito": conflito,
            "locais": locais_ordenados,
        }
        itens.append(dados)
        if produto.id == produto_id:
            produto_selecionado = dados

    return templates.TemplateResponse(
        "admin/disponibilidade.html",
        {
            "request": request,
            "empresa": empresa,
            "data_consulta": data_consulta,
            "itens": itens,
            "produto_selecionado": produto_selecionado,
        },
    )


def _responsavel_contrato_exibicao(item: Solicitacao) -> str:
    """Responsável comercial sem criar histórico visual novo.

    Novos contratos usam o primeiro envio para aceite. Para contratos antigos,
    preserva a regra validada: o usuário do primeiro pagamento/sinal registrado.
    """
    if getattr(item, "responsavel_contrato", None):
        return item.responsavel_contrato
    pagamentos = sorted(
        list(getattr(item, "pagamentos", []) or []),
        key=lambda p: (p.criado_em or datetime.min, p.id or 0),
    )
    for pagamento in pagamentos:
        if pagamento.usuario_registro:
            return pagamento.usuario_registro
    return ""


def _anexar_responsaveis_exibicao(itens):
    for item in itens or []:
        item.responsavel_contrato_exibicao = _responsavel_contrato_exibicao(item)
    return itens


@app.get("/painel/agenda", response_class=HTMLResponse)
def agenda(
        request: Request,
        data_inicial: str = "",
        data_final: str = "",
        ativos: str = "1",
        credito: str = "",
        cancelados: str = "",
        equipe_id: int = 0,
        situacao_rota: str = "todos",
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    # Agenda em modo consulta não executa auditoria nem manutenção automática.
    equipes = equipes_visiveis_usuario(request, db, empresa.id)
    ids_equipes = {e.id for e in equipes}
    if equipe_id and equipe_id not in ids_equipes:
        equipe_id = 0
    inicio, fim = periodo_semana_atual()

    # Checkbox desmarcado não é enviado no GET. Quando existe query string,
    # usa exatamente os filtros presentes para permitir visualizar somente crédito.
    if request.query_params:
        ativos = "1" if "ativos" in request.query_params else ""
        credito = "1" if "credito" in request.query_params else ""
        cancelados = "1" if "cancelados" in request.query_params else ""

    # Mantém o último filtro usado na agenda para a equipe não precisar refazer a busca.
    filtro_salvo = request.session.get("agenda_filtro", {}) if not request.query_params else {}
    data_inicial = data_inicial or filtro_salvo.get("data_inicial") or inicio.isoformat()
    data_final = data_final or filtro_salvo.get("data_final") or fim.isoformat()
    if not request.query_params and filtro_salvo:
        ativos = filtro_salvo.get("ativos", "1")
        credito = filtro_salvo.get("credito", "")
        cancelados = filtro_salvo.get("cancelados", "")

    request.session["agenda_filtro"] = {
        "data_inicial": data_inicial,
        "data_final": data_final,
        "ativos": "1" if ativos else "",
        "credito": "1" if credito else "",
        "cancelados": "1" if cancelados else "",
    }

    status_credito = {"aguardando_nova_data"}
    status_cancelados = {"cancelada", "cancelado_cliente", "rejeitada"}
    status_inativos = status_credito | status_cancelados

    filtros_status = []
    if ativos:
        filtros_status.append("ativos")
    if credito:
        filtros_status.append("credito")
    if cancelados:
        filtros_status.append("cancelados")

    q = db.query(Solicitacao).filter_by(empresa_id=empresa.id)
    if data_inicial:
        q = q.filter(Solicitacao.data_evento >= datetime.strptime(data_inicial, "%Y-%m-%d").date())
    if data_final:
        q = q.filter(Solicitacao.data_evento <= datetime.strptime(data_final, "%Y-%m-%d").date())

    # Agenda deve mostrar todas as locações do período.
    # O filtro de rascunho/contrato sem aceite fica somente na tela inicial (/painel).
    solicitacoes = (
        q.options(
            joinedload(Solicitacao.cliente),
            joinedload(Solicitacao.produto),
            selectinload(Solicitacao.itens),
            selectinload(Solicitacao.pagamentos),
        )
        .order_by(
            Solicitacao.data_evento.asc(),
            Solicitacao.hora_inicio.asc(),
            Solicitacao.id.asc(),
        )
        .all()
    )

    agenda_operacional = (
        db.query(Agenda)
        .options(joinedload(Agenda.equipe))
        .filter(Agenda.empresa_id == empresa.id, Agenda.solicitacao_id.in_([x.id for x in solicitacoes] or [-1]))
        .all()
    )
    rotas_por_solicitacao = {}
    for rota in agenda_operacional:
        atual = rotas_por_solicitacao.get(rota.solicitacao_id)
        if not atual or (rota.roteirizado and not atual.roteirizado):
            rotas_por_solicitacao[rota.solicitacao_id] = rota

    itens = []
    for s in solicitacoes:
        rota = rotas_por_solicitacao.get(s.id)
        if equipe_id and (not rota or rota.equipe_id != equipe_id):
            continue
        if situacao_rota == "roteirizado" and (not rota or not rota.roteirizado):
            continue
        if situacao_rota == "nao_roteirizado" and rota and rota.roteirizado:
            continue
        status_atual = s.status or ""
        eh_credito = status_atual in status_credito
        eh_cancelado = status_atual in status_cancelados
        eh_ativo = status_atual not in status_inativos

        if (eh_ativo and "ativos" in filtros_status) or (eh_credito and "credito" in filtros_status) or (
                eh_cancelado and "cancelados" in filtros_status):
            itens.append(s)

    _anexar_responsaveis_exibicao(itens)
    mensagens = mensagens_empresa(empresa)
    return templates.TemplateResponse("admin/agenda.html", {
        "request": request,
        "itens": itens,
        "total_itens": len(itens),
        "empresa": empresa,
        "data_inicial": data_inicial,
        "data_final": data_final,
        "filtro_ativos": bool(ativos),
        "filtro_credito": bool(credito),
        "filtro_cancelados": bool(cancelados),
        "equipes": equipes, "equipe_id": equipe_id, "situacao_rota": situacao_rota,
        "rotas_por_solicitacao": rotas_por_solicitacao,
        "mensagens": mensagens,
    })


@app.post("/painel/solicitacao/{solicitacao_id}/responsavel-retirada")
def salvar_responsavel_retirada(solicitacao_id: int, request: Request, retirada_responsavel_nome: str = Form(""), retirada_responsavel_telefone: str = Form(""), db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    item = db.query(Solicitacao).filter_by(id=solicitacao_id, empresa_id=empresa.id).first()
    if not item:
        raise HTTPException(404)
    if retirada_responsavel_telefone and not celular_brasileiro_valido(retirada_responsavel_telefone):
        raise HTTPException(400, "Informe um WhatsApp brasileiro válido para o responsável pela retirada.")
    item.retirada_responsavel_nome = retirada_responsavel_nome.strip()
    item.retirada_responsavel_telefone = limpar_identificador(retirada_responsavel_telefone) or retirada_responsavel_telefone.strip()
    db.commit()
    return RedirectResponse(request.headers.get("referer") or "/painel/preparar", status_code=303)

@app.post("/painel/agenda/{agenda_id}/roteiro")
def atualizar_roteiro(
        request: Request,
        agenda_id: int,
        direcao: str = Form(""),
        previsao_entrega: str = Form(""),
        data_evento: str = Form(""),
        data_operacao: str = Form(""),
        status_operacional: str = Form("pendente"),
        equipe_id: int = Form(...),
        link_localizacao: str = Form(""),
        db: Session = Depends(get_db),
        empresa: Empresa = Depends(empresa_logada)
):
    item = db.get(Agenda, agenda_id)
    if not item or item.empresa_id != empresa.id:
        raise HTTPException(404)

    equipe = db.query(Equipe).filter_by(id=equipe_id, empresa_id=empresa.id, ativa=True).first()
    permitidas = {e.id for e in equipes_visiveis_usuario(request, db, empresa.id)}
    if not equipe or equipe.id not in permitidas:
        raise HTTPException(403, "Equipe não permitida para este usuário.")

    status_anterior = item.status_operacional
    data_anterior = item.data
    hora_anterior = item.hora_inicio

    novo_status = status_operacional if status_operacional in {"pendente", "concluido"} else "pendente"
    falta_pagamento = 0
    if item.solicitacao:
        falta_pagamento = max((item.solicitacao.valor or 0) - (item.solicitacao.valor_pago or 0), 0)
    if item.tipo_evento == "retirada" and novo_status == "concluido" and falta_pagamento > 0.009:
        destino = request.headers.get("referer") or "/painel/reservas"
        partes = urlparse(destino)
        qs = dict(parse_qsl(partes.query, keep_blank_values=True))
        qs["op_erro"] = f"Não é possível encerrar a busca: falta receber R$ {falta_pagamento:,.2f}.".replace(",", "X").replace(".", ",").replace("X", ".")
        destino = urlunparse((partes.scheme, partes.netloc, partes.path, partes.params, urlencode(qs), partes.fragment))
        return RedirectResponse(destino, status_code=303)

    item.previsao_entrega = previsao_entrega
    item.equipe_id = equipe.id
    item.roteirizado = True
    item.link_localizacao = link_localizacao
    item.status_operacional = novo_status

    retirada_bloqueada = bool(item.tipo_evento == "retirada" and item.solicitacao and retirada_obrigatoria_ativa(item.solicitacao))

    nova_data = None
    data_informada = (data_operacao or data_evento or "").strip()
    if not retirada_bloqueada and data_informada:
        try:
            nova_data = datetime.strptime(data_informada, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(400, "Data da operação inválida.")

        data_limite_evento = item.solicitacao.data_evento if item.solicitacao else None
        if item.tipo_evento == "entrega" and data_limite_evento and nova_data > data_limite_evento:
            raise HTTPException(400, "A data da entrega não pode ser posterior à data do contrato.")
        if item.tipo_evento == "retirada" and data_limite_evento and nova_data < data_limite_evento:
            raise HTTPException(400, "A data da busca não pode ser anterior à data do contrato.")
        item.data = nova_data

    previsao_entrega = (previsao_entrega or "").strip()
    if not retirada_bloqueada and previsao_entrega:
        try:
            nova_hora = datetime.strptime(previsao_entrega, "%H:%M").time()
        except ValueError:
            raise HTTPException(400, "Hora da operação inválida.")
        item.previsao_entrega = previsao_entrega
        item.hora_inicio = nova_hora
    elif retirada_bloqueada:
        item.data = item.solicitacao.retirada_data or item.solicitacao.data_evento
        item.hora_inicio = item.solicitacao.retirada_hora or item.solicitacao.hora_fim or item.solicitacao.hora_inicio
        item.previsao_entrega = item.hora_inicio.strftime("%H:%M") if item.hora_inicio else ""

    # Marca visualmente que este card já foi roteirizado.
    # A cor cinza clara da tela usa este marcador dentro do histórico.
    usuario = request.session.get("usuario_nome") or request.session.get("usuario") or "Usuário"
    registro = (
        f"[{datetime.now().strftime('%d/%m/%Y %H:%M')}] "
        f"Roteirização salva por {usuario}. "
        f"Entrega: {data_anterior.strftime('%d/%m/%Y') if data_anterior else '-'} "
        f"{hora_anterior.strftime('%H:%M') if hora_anterior else '-'} → "
        f"{item.data.strftime('%d/%m/%Y') if item.data else '-'} "
        f"{item.hora_inicio.strftime('%H:%M') if item.hora_inicio else '-'}."
    )
    item.observacoes_operacionais = ((item.observacoes_operacionais or "") + "\n" + registro).strip()

    if item.tipo_evento == "entrega" and status_anterior != "concluido" and item.status_operacional == "concluido":
        criar_retirada_apos_entrega(db, item)

    # A Operação precisa permanecer rápida e independente da Inteligência.
    # Rotas inteligentes abertas não são recalculadas durante o atendimento
    # operacional; elas serão atualizadas somente dentro do próprio módulo de
    # Inteligência quando o usuário solicitar gerar/recalcular a rota.
    db.commit()
    destino = request.headers.get("referer") or "/painel/reservas"
    return RedirectResponse(destino, status_code=303)


@app.post("/painel/agenda/{agenda_id}/assumir-comunicacao-operacao")
def assumir_comunicacao_operacao(
    agenda_id: int,
    request: Request,
    db: Session = Depends(get_db),
    empresa: Empresa = Depends(empresa_logada),
):
    agenda_item = db.query(Agenda).filter_by(id=agenda_id, empresa_id=empresa.id).first()
    if not agenda_item or not agenda_item.solicitacao:
        raise HTTPException(404)
    solicitacao = agenda_item.solicitacao
    if not solicitacao.responsavel_operacao:
        solicitacao.responsavel_operacao = (
            request.session.get("usuario_nome")
            or request.session.get("usuario_sistema")
            or request.session.get("usuario")
            or "Usuário"
        )
        db.commit()
    return {"ok": True, "responsavel": solicitacao.responsavel_operacao}


@app.get("/e/{slug}", response_class=HTMLResponse)
def portal_empresa(slug: str, request: Request, db: Session = Depends(get_db)):
    empresa = db.query(Empresa).filter_by(slug=slug, ativa=True).first()
    if not empresa:
        raise HTTPException(404, "Empresa não encontrada")
    return templates.TemplateResponse("publico/identificar.html", {"request": request, "empresa": empresa})


@app.post("/e/{slug}/buscar")
def buscar_cliente(slug: str, identificador: str = Form(...), db: Session = Depends(get_db)):
    empresa = db.query(Empresa).filter_by(slug=slug, ativa=True).first()
    if not empresa:
        raise HTTPException(404)
    ident = limpar_identificador(identificador)
    return RedirectResponse(f"/e/{slug}/cadastro?identificador={ident}", status_code=303)


def _url_confirmacao_whatsapp(empresa: Empresa, item: Solicitacao, tipo: str) -> str | None:
    telefone = _limpar_tel_whatsapp(getattr(empresa, "whatsapp_retorno", "") or "")
    if not telefone:
        return None
    cliente = item.cliente.nome if item.cliente and item.cliente.nome else "Cliente"
    data = item.data_evento.strftime("%d/%m/%Y") if item.data_evento else ""
    if tipo == "pre_contrato":
        texto = (f"Olá, sou {cliente}. Acabei de preencher meu pré-contrato "
                 f"para o evento do dia {data}. Solicitação #{item.id}.")
    else:
        texto = (f"Olá, sou {cliente}. Confirmo o aceite do contrato #{item.id} "
                 f"referente ao evento do dia {data}.")
    return f"https://wa.me/{telefone}?text={quote(texto)}"


@app.get("/e/{slug}/pre-contrato", response_class=HTMLResponse)
def pre_contrato_cliente(slug: str, request: Request, erro: str = "", db: Session = Depends(get_db)):
    empresa = db.query(Empresa).filter_by(slug=slug, ativa=True).first()
    if not empresa:
        raise HTTPException(404)
    return templates.TemplateResponse("publico/cadastro.html", {
        "request": request, "empresa": empresa, "cliente": None, "identificador": "",
        "cliente_encontrado": False, "cpf_confirmacao": "", "erro": erro,
        "campos_cfg": {ce.campo.chave: ce for ce in
                       db.query(CampoEmpresa).join(CampoGlobal).filter(CampoEmpresa.empresa_id == empresa.id).all()}
    })


@app.get("/e/{slug}/cadastro", response_class=HTMLResponse)
def cadastro_cliente(slug: str, request: Request, identificador: str = "", cpf_confirmacao: str = "", erro: str = "",
                     db: Session = Depends(get_db)):
    empresa = db.query(Empresa).filter_by(slug=slug, ativa=True).first()
    ident = limpar_identificador(identificador)
    cliente_encontrado = db.query(Cliente).filter_by(empresa_id=empresa.id, identificador=ident).first()
    cliente = None
    cpf_limpo = limpar_identificador(cpf_confirmacao)
    if cliente_encontrado and cpf_limpo and limpar_identificador(cliente_encontrado.cpf) == cpf_limpo:
        cliente = cliente_encontrado
    return templates.TemplateResponse("publico/cadastro.html", {
        "request": request, "empresa": empresa, "cliente": cliente, "identificador": ident,
        "cliente_encontrado": bool(cliente_encontrado), "cpf_confirmacao": cpf_confirmacao, "erro": erro,
        "campos_cfg": {ce.campo.chave: ce for ce in
                       db.query(CampoEmpresa).join(CampoGlobal).filter(CampoEmpresa.empresa_id == empresa.id).all()}
    })


@app.post("/e/{slug}/reserva")
@app.post("/e/{slug}/pre-cadastro")
def salvar_pre_cadastro(
        request: Request, slug: str, identificador: str = Form(...), tipo_pessoa: str = Form("fisica"),
        nome: str = Form(""), data_nascimento: str = Form(""), telefone: str = Form(""), cpf: str = Form(""),
        cnpj: str = Form(""), email: str = Form(""), endereco: str = Form(""), numero: str = Form(""),
        complemento: str = Form(""),
        bairro: str = Form(""), cidade: str = Form(""), estado: str = Form(""), cep: str = Form(""),
        local: str = Form(""),
        local_nome: str = Form(""), acesso_local: str = Form(""), local_responsavel_nome: str = Form(""),
        local_responsavel_telefone: str = Form(""),
        data_evento: str = Form(...), hora_inicio: str = Form(...), observacoes: str = Form(""),
        acao: str = Form("salvar"),
        db: Session = Depends(get_db)
):
    empresa = db.query(Empresa).filter_by(slug=slug, ativa=True).first()
    cpf_limpo = limpar_identificador(cpf)
    cnpj_limpo = limpar_identificador(cnpj)
    telefone_limpo = limpar_identificador(telefone)
    ident = limpar_identificador(identificador)
    if not ident or ident == "novo":
        # Pré-contrato em branco: o identificador nasce dos dados reais enviados.
        # Prioridade: CPF/CNPJ quando existir; senão celular; senão código temporário.
        ident = cpf_limpo or cnpj_limpo or telefone_limpo or uuid.uuid4().hex[:12]
    campos_empresa = {
        ce.campo.chave: ce for ce in
        db.query(CampoEmpresa).join(CampoGlobal).filter(CampoEmpresa.empresa_id == empresa.id).all()
    }

    def campo_obrigatorio(chave: str) -> bool:
        ce = campos_empresa.get(chave)
        return bool(ce and ce.visivel and ce.obrigatorio)

    form_data = {
        "tipo_pessoa": tipo_pessoa, "nome": nome, "data_nascimento": data_nascimento, "telefone": telefone,
        "cpf": cpf, "cnpj": cnpj, "email": email, "endereco": endereco, "numero": numero, "complemento": complemento,
        "bairro": bairro, "cidade": cidade, "estado": estado, "cep": cep, "local": local, "local_nome": local_nome,
        "acesso_local": acesso_local, "local_responsavel_nome": local_responsavel_nome,
        "local_responsavel_telefone": local_responsavel_telefone, "data_evento": data_evento,
        "hora_inicio": hora_inicio, "observacoes": observacoes
    }

    def render_erro(codigo: str):
        cliente_encontrado = db.query(Cliente).filter_by(empresa_id=empresa.id, identificador=ident).first()
        return templates.TemplateResponse("publico/cadastro.html", {
            "request": request, "empresa": empresa, "cliente": None, "identificador": ident,
            "cliente_encontrado": bool(cliente_encontrado), "cpf_confirmacao": "", "erro": codigo,
            "campos_cfg": campos_empresa, "form": form_data
        }, status_code=400)

    if not celular_brasileiro_valido(telefone):
        return render_erro("whatsapp_invalido")
    if not celular_brasileiro_valido(local_responsavel_telefone):
        return render_erro("responsavel_whatsapp_invalido")
    if not local_responsavel_nome.strip():
        return render_erro("responsavel_whatsapp_invalido")
    if cpf_limpo and cnpj_limpo:
        return render_erro("cpf_cnpj")
    if tipo_pessoa == "fisica" and cpf_limpo and not cpf_valido(cpf_limpo):
        return render_erro("cpf_invalido")
    if tipo_pessoa == "fisica" and campo_obrigatorio("cpf") and not cpf_limpo:
        return render_erro("cpf_invalido")
    if tipo_pessoa == "juridica" and cnpj_limpo and not cnpj_valido(cnpj_limpo):
        return render_erro("cnpj_invalido")
    if tipo_pessoa == "juridica" and campo_obrigatorio("cnpj") and not cnpj_limpo:
        return render_erro("cnpj_invalido")
    cliente = db.query(Cliente).filter_by(empresa_id=empresa.id, identificador=ident).first()
    if not cliente:
        cliente = Cliente(empresa_id=empresa.id, identificador=ident)
        db.add(cliente)
    cliente.nome = nome
    cliente.data_nascimento = datetime.strptime(data_nascimento, "%Y-%m-%d").date() if data_nascimento else None
    cliente.telefone = telefone_limpo or telefone
    cliente.cpf = cpf_limpo
    cliente.cnpj = cnpj_limpo
    cliente.email = email
    cliente.endereco = endereco
    cliente.numero = numero
    cliente.complemento = complemento
    cliente.bairro = bairro
    cliente.cidade = cidade
    cliente.estado = estado
    cliente.cep = cep
    cliente.observacoes = observacoes
    db.flush()
    salvar_endereco_cliente(
        db, empresa.id, cliente.id, endereco, numero, complemento, bairro, cidade, estado, cep,
        apelido=local_nome,
    )
    db.commit()
    db.refresh(cliente)

    if not hora_meia_em_meia_valida(hora_inicio):
        return render_erro("hora_invalida")
    data_obj = datetime.strptime(data_evento, "%Y-%m-%d").date()
    rascunho_existente = (
        db.query(Solicitacao)
        .join(Cliente, Solicitacao.cliente_id == Cliente.id)
        .filter(
            Solicitacao.empresa_id == empresa.id,
            Solicitacao.data_evento == data_obj,
            Solicitacao.status.in_(["reserva", "pre_reserva", "contrato_enviado", "aguardando_aceite"]),
            Cliente.telefone == (telefone_limpo or telefone)
        )
        .first()
    )
    if rascunho_existente:
        return render_erro("rascunho_duplicado")
    inicio_obj = datetime.strptime(hora_inicio, "%H:%M").time()
    # O pré-contrato público usa a duração padrão de 4 horas até que os itens sejam definidos.
    fim_obj = somar_minutos(inicio_obj, 240)
    solicitacao = Solicitacao(
        empresa_id=empresa.id, cliente_id=cliente.id, data_evento=data_obj, hora_inicio=inicio_obj,
        hora_fim=fim_obj, bairro=bairro.strip(), local=endereco.strip(), local_numero=numero.strip(),
        local_complemento=complemento.strip(), local_cidade=cidade.strip(), local_estado=estado.strip(),
        local_cep=cep.strip(), local_nome=local_nome.strip(),
        local_responsavel_nome=local_responsavel_nome, local_responsavel_telefone=local_responsavel_telefone,
        acesso_local=acesso_local, observacoes=observacoes, status="pre_reserva"
    )
    db.add(solicitacao)
    db.commit()
    db.refresh(solicitacao)
    url_whatsapp = _url_confirmacao_whatsapp(empresa, solicitacao, "pre_contrato")
    if url_whatsapp:
        return RedirectResponse(url_whatsapp, status_code=303)
    return RedirectResponse(f"/e/{slug}/obrigado/{solicitacao.id}", status_code=303)


def _wrap_pdf_text(c, texto, x, y, largura, leading=14, fonte="Helvetica", tamanho=10):
    """Quebra texto respeitando margem inferior para não invadir o rodapé."""
    margem_inferior = 110
    margem_superior = c._pagesize[1] - 70

    def nova_pagina_se_precisar(y_atual):
        if y_atual < margem_inferior:
            c.showPage()
            c.setFont(fonte, tamanho)
            return margem_superior
        return y_atual

    c.setFont(fonte, tamanho)
    for paragrafo in (texto or "").splitlines():
        palavras = paragrafo.split()
        if not palavras:
            y -= leading
            y = nova_pagina_se_precisar(y)
            continue
        linha = ""
        for palavra in palavras:
            teste = (linha + " " + palavra).strip()
            if c.stringWidth(teste, fonte, tamanho) <= largura:
                linha = teste
            else:
                y = nova_pagina_se_precisar(y)
                c.drawString(x, y, linha)
                y -= leading
                linha = palavra
        if linha:
            y = nova_pagina_se_precisar(y)
            c.drawString(x, y, linha)
            y -= leading
    return y


@app.get("/e/{slug}/contrato/{solicitacao_id}.pdf")
def contrato_cliente_pdf(slug: str, solicitacao_id: int, request: Request, db: Session = Depends(get_db)):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader
    except Exception:
        raise HTTPException(500, "Para gerar PDF, instale a dependência: reportlab")

    empresa = db.query(Empresa).filter_by(slug=slug, ativa=True).first()
    item = db.get(Solicitacao, solicitacao_id)
    if not empresa or not item or item.empresa_id != empresa.id:
        raise HTTPException(404)
    contrato = db.get(Contrato, item.contrato_id) if item.contrato_id else None
    produto = item.produto
    itens_reserva = db.query(ReservaItem).filter_by(empresa_id=empresa.id, solicitacao_id=item.id).all()

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    w, h = A4
    y = h - 70

    logo = empresa.logo_url or empresa.logo_idb_url
    if logo and logo.startswith("/static/"):
        logo_path = Path(logo.lstrip("/"))
        if logo_path.exists():
            try:
                c.drawImage(ImageReader(str(logo_path)), 40, y - 42, width=70, height=42, preserveAspectRatio=True,
                            mask='auto')
            except Exception:
                pass
    c.setFont("Helvetica-Bold", 16)
    c.drawString(120, y, empresa.nome or "Contrato")
    c.setFont("Helvetica-Bold", 13)
    c.drawString(40, y - 62, f"Contrato / Reserva #{item.id}")
    y -= 88

    c.setFont("Helvetica-Bold", 11);
    c.drawString(40, y, "Dados preenchidos");
    y -= 16
    for linha in linhas_informacoes_preenchidas_contrato(item, formato="texto"):
        if y < 110:
            c.showPage();
            y = h - 70
        y = _wrap_pdf_text(c, linha, 40, y, w - 80, leading=13, tamanho=9)
    y -= 8

    c.setFont("Helvetica-Bold", 11);
    c.drawString(40, y, "Itens");
    y -= 16
    if itens_reserva:
        for ri in itens_reserva:
            c.drawString(50, y, f"{ri.quantidade or 1}x {ri.nome} - R$ {moeda_br(ri.valor_total or 0)}")
            y -= 14
    elif produto:
        c.drawString(50, y, produto.nome);
        y -= 14
    y -= 10

    c.setFont("Helvetica-Bold", 11);
    c.drawString(40, y, contrato.nome if contrato else "Contrato");
    y -= 16
    y = _wrap_pdf_text(c, contrato.clausulas if contrato else (item.observacoes or ""), 40, y, w - 80)
    y -= 24
    if y < 120:
        c.showPage();
        y = h - 70
    c.setFont("Helvetica", 10)
    c.drawString(40, y, "Declaro estar ciente e de acordo com as condições desta locação.")
    y -= 42
    c.line(40, y, 330, y)
    y -= 14
    c.drawString(40, y, "Assinatura do cliente")
    y -= 28
    c.drawString(40, y, "Data: ____/____/________")
    y -= 20
    c.setFont("Helvetica", 9)
    if y < 90:
        c.showPage()
        y = h - 70
    c.drawString(40, y, f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} - {empresa.nome}")
    c.save()
    buffer.seek(0)
    nome_pdf = f"contrato_{empresa.slug}_{item.id}.pdf"
    return Response(buffer.read(), media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="{nome_pdf}"'})


@app.get("/e/{slug}/contrato/{solicitacao_id}/clausulas", response_class=HTMLResponse)
def contrato_cliente_clausulas(slug: str, solicitacao_id: int, request: Request, db: Session = Depends(get_db)):
    empresa = db.query(Empresa).filter_by(slug=slug, ativa=True).first()
    item = db.get(Solicitacao, solicitacao_id)
    if not empresa or not item or item.empresa_id != empresa.id:
        raise HTTPException(404)
    contrato_ids = set()
    if item.contrato_id:
        contrato_ids.add(item.contrato_id)
    itens_reserva = db.query(ReservaItem).filter_by(empresa_id=empresa.id, solicitacao_id=item.id).all()
    for ri in itens_reserva:
        if ri.produto_id:
            produto = db.get(ProdutoServico, ri.produto_id)
            if produto and produto.contrato_id:
                contrato_ids.add(produto.contrato_id)
    contratos_clausulas = []
    for cid in contrato_ids:
        c = db.get(Contrato, cid)
        if c and c.empresa_id == empresa.id and c.ativo:
            contratos_clausulas.append(c)
    if not contratos_clausulas and item.observacoes:
        contratos_clausulas = []
    return templates.TemplateResponse("publico/clausulas.html", {
        "request": request,
        "empresa": empresa,
        "item": item,
        "contratos_clausulas": contratos_clausulas,
    })


@app.get("/e/{slug}/contrato/{solicitacao_id}", response_class=HTMLResponse)
def contrato_cliente(slug: str, solicitacao_id: int, request: Request, db: Session = Depends(get_db)):
    empresa = db.query(Empresa).filter_by(slug=slug, ativa=True).first()
    item = db.get(Solicitacao, solicitacao_id)
    if not empresa or not item or item.empresa_id != empresa.id:
        raise HTTPException(404)

    # Contrato em crédito continua existindo, porém o aceite fica indisponível
    # até o responsável reativá-lo e revisar o novo rascunho.
    if item.status == "aguardando_nova_data":
        return templates.TemplateResponse("publico/contrato_indisponivel.html", {
            "request": request,
            "empresa": empresa,
            "item": item,
        })

    if item.status not in ["pre_reserva", "reserva", "aguardando_aceite",
                           "contrato_enviado", "aceito", "aguardando_pagamento",
                           "reserva_confirmada", "cancelado_cliente"]:
        raise HTTPException(404)
    contrato = db.get(Contrato, item.contrato_id) if item.contrato_id else None
    produto = db.get(ProdutoServico, item.produto_id) if item.produto_id else None
    itens_reserva = db.query(ReservaItem).filter_by(empresa_id=empresa.id, solicitacao_id=item.id).all()
    return templates.TemplateResponse("publico/contrato.html",
                                      {"request": request, "empresa": empresa, "item": item, "contrato": contrato,
                                       "produto": produto, "itens_reserva": itens_reserva})


@app.get("/e/{slug}/contrato/{solicitacao_id}/editar", response_class=HTMLResponse)
def editar_dados_contrato_cliente(slug: str, solicitacao_id: int, request: Request, db: Session = Depends(get_db)):
    empresa = db.query(Empresa).filter_by(slug=slug, ativa=True).first()
    item = db.get(Solicitacao, solicitacao_id)
    if not empresa or not item or item.empresa_id != empresa.id or not item.cliente:
        raise HTTPException(404)
    if status_reserva_confirmada(item.status) or item.status == "cancelado_cliente":
        return RedirectResponse(f"/e/{slug}/contrato/{item.id}", status_code=303)

    endereco_evento = dados_endereco_solicitacao(item)
    return templates.TemplateResponse("publico/cadastro.html", {
        "request": request,
        "empresa": empresa,
        "cliente": item.cliente,
        "identificador": item.cliente.identificador or item.cliente.telefone or item.cliente.cpf or "",
        "cliente_encontrado": True,
        "cpf_confirmacao": "",
        "erro": "",
        "modo_edicao_contrato": True,
        "item": item,
        "form": {
            "tipo_pessoa": "juridica" if item.cliente.cnpj else "fisica",
            "nome": item.cliente.nome or "",
            "telefone": item.cliente.telefone or "",
            "cpf": item.cliente.cpf or "",
            "cnpj": item.cliente.cnpj or "",
            "email": item.cliente.email or "",
            "endereco": endereco_evento["endereco"],
            "numero": endereco_evento["numero"],
            "complemento": endereco_evento["complemento"],
            "bairro": endereco_evento["bairro"],
            "cidade": endereco_evento["cidade"],
            "estado": endereco_evento["estado"],
            "cep": endereco_evento["cep"],
            "local": item.local or "",
            "local_nome": item.local_nome or "",
            "acesso_local": item.acesso_local or "",
            "local_responsavel_nome": item.local_responsavel_nome or "",
            "local_responsavel_telefone": item.local_responsavel_telefone or "",
            "data_evento": item.data_evento.isoformat() if item.data_evento else "",
            "hora_inicio": item.hora_inicio.strftime("%H:%M") if item.hora_inicio else "",
            "observacoes": item.observacoes or item.cliente.observacoes or "",
        },
        "campos_cfg": {ce.campo.chave: ce for ce in
                       db.query(CampoEmpresa).join(CampoGlobal).filter(CampoEmpresa.empresa_id == empresa.id).all()}
    })


@app.post("/e/{slug}/contrato/{solicitacao_id}/editar")
def salvar_dados_contrato_cliente(
        slug: str, solicitacao_id: int,
        identificador: str = Form(""), tipo_pessoa: str = Form("fisica"),
        nome: str = Form(""), data_nascimento: str = Form(""), telefone: str = Form(""), cpf: str = Form(""),
        cnpj: str = Form(""), email: str = Form(""), endereco: str = Form(""), numero: str = Form(""),
        complemento: str = Form(""), bairro: str = Form(""), cidade: str = Form(""), estado: str = Form(""),
        cep: str = Form(""), local: str = Form(""), local_nome: str = Form(""), acesso_local: str = Form(""),
        local_responsavel_nome: str = Form(""), local_responsavel_telefone: str = Form(""),
        data_evento: str = Form(...), hora_inicio: str = Form(...), observacoes: str = Form(""),
        db: Session = Depends(get_db)
):
    empresa = db.query(Empresa).filter_by(slug=slug, ativa=True).first()
    item = db.get(Solicitacao, solicitacao_id)
    if not empresa or not item or item.empresa_id != empresa.id or not item.cliente:
        raise HTTPException(404)
    if status_reserva_confirmada(item.status) or item.status == "cancelado_cliente":
        return RedirectResponse(f"/e/{slug}/contrato/{item.id}", status_code=303)

    cliente = item.cliente
    cliente.nome = nome.strip()
    cliente.data_nascimento = datetime.strptime(data_nascimento, "%Y-%m-%d").date() if data_nascimento else None
    cliente.telefone = limpar_identificador(telefone) or telefone.strip()
    cliente.cpf = limpar_identificador(cpf)
    cliente.cnpj = limpar_identificador(cnpj)
    cliente.email = email.strip()
    cliente.endereco = endereco.strip()
    cliente.numero = numero.strip()
    cliente.complemento = complemento.strip()
    cliente.bairro = bairro.strip()
    cliente.cidade = cidade.strip()
    cliente.estado = estado.strip()
    cliente.cep = limpar_identificador(cep) or cep.strip()
    cliente.observacoes = observacoes.strip()
    cliente.identificador = cliente.cpf or cliente.cnpj or cliente.telefone or limpar_identificador(identificador) or cliente.identificador

    item.data_evento = datetime.strptime(data_evento, "%Y-%m-%d").date()
    item.hora_inicio = datetime.strptime(hora_inicio, "%H:%M").time()
    item.bairro = bairro.strip()
    item.local = endereco.strip() or local.strip()
    item.local_numero = numero.strip()
    item.local_complemento = complemento.strip()
    item.local_cidade = cidade.strip()
    item.local_estado = estado.strip()
    item.local_cep = (limpar_identificador(cep) or cep.strip())
    item.local_nome = local_nome.strip()
    item.acesso_local = acesso_local.strip()
    item.local_responsavel_nome = local_responsavel_nome.strip()
    item.local_responsavel_telefone = limpar_identificador(local_responsavel_telefone) or local_responsavel_telefone.strip()
    item.observacoes = observacoes.strip()
    salvar_endereco_cliente(
        db, empresa.id, cliente.id, item.local, item.local_numero, item.local_complemento,
        item.bairro, item.local_cidade, item.local_estado, item.local_cep, apelido=item.local_nome,
    )
    _invalidar_geocodificacao(item)

    db.commit()
    _tentar_geocodificar_solicitacao(db, item)
    return RedirectResponse(f"/e/{slug}/contrato/{item.id}", status_code=303)


@app.post("/e/{slug}/cancelar/{solicitacao_id}")
def cancelar_contrato(slug: str, solicitacao_id: int, db: Session = Depends(get_db)):
    empresa = db.query(Empresa).filter_by(slug=slug, ativa=True).first()
    item = db.get(Solicitacao, solicitacao_id)
    if not empresa or not item or item.empresa_id != empresa.id:
        raise HTTPException(404)
    if item.status != "cancelado_cliente":
        item.status = "cancelado_cliente"
        item.cancelado_em = agora_utc()
        db.commit()
    return RedirectResponse(f"/e/{slug}/obrigado/{solicitacao_id}", status_code=303)


@app.post("/e/{slug}/aceitar/{solicitacao_id}")
def aceitar_contrato(slug: str, solicitacao_id: int, aceite: Optional[str] = Form(None), db: Session = Depends(get_db)):
    empresa = db.query(Empresa).filter_by(slug=slug, ativa=True).first()
    item = db.get(Solicitacao, solicitacao_id)
    if not empresa or not item or item.empresa_id != empresa.id:
        raise HTTPException(404)
    itens_reserva = db.query(ReservaItem).filter_by(empresa_id=empresa.id, solicitacao_id=item.id).count()
    aceite_registrado_agora = False
    if item.status in ["aguardando_aceite", "contrato_enviado"] and item.contrato_id and itens_reserva > 0:
        aceite_registrado_agora = True
        item.status = "aguardando_pagamento" if (item.sinal or 0) > 0 else "reserva_confirmada"
        item.aceite_em = agora_utc()
        item.aprovado_em = item.aceite_em
        fim_obj = item.hora_fim or (somar_minutos(item.hora_inicio,
                                                  item.produto.duracao_minutos) if item.produto and item.produto.duracao_minutos else None)
        item.hora_fim = fim_obj
        criar_eventos_operacionais(db, item)
        _processar_humiat_aceite(db, empresa, item)
        db.commit()
    # Todo aceite gravado deve abrir o WhatsApp, independentemente de sinal/pagamento pendente.
    url_whatsapp = _url_confirmacao_whatsapp(empresa, item, "aceite") if aceite_registrado_agora else None
    if url_whatsapp:
        return RedirectResponse(url_whatsapp, status_code=303)
    return RedirectResponse(f"/e/{slug}/obrigado/{solicitacao_id}", status_code=303)


@app.get("/e/{slug}/confirmar-whatsapp/{solicitacao_id}", response_class=HTMLResponse)
def confirmar_whatsapp(slug: str, solicitacao_id: int, request: Request, tipo: str = "pre_contrato", db: Session = Depends(get_db)):
    empresa = db.query(Empresa).filter_by(slug=slug, ativa=True).first()
    solicitacao = db.get(Solicitacao, solicitacao_id)
    if not empresa or not solicitacao or solicitacao.empresa_id != empresa.id:
        raise HTTPException(404)
    tipo_confirmacao = "aceite" if tipo == "aceite" else "pre_contrato"
    url_whatsapp = _url_confirmacao_whatsapp(empresa, solicitacao, tipo_confirmacao)
    return templates.TemplateResponse("publico/confirmar_whatsapp.html", {
        "request": request,
        "empresa": empresa,
        "solicitacao": solicitacao,
        "url_whatsapp": url_whatsapp,
        "tipo": tipo_confirmacao,
    })


@app.get("/e/{slug}/obrigado/{solicitacao_id}", response_class=HTMLResponse)
def obrigado(slug: str, solicitacao_id: int, request: Request, db: Session = Depends(get_db)):
    empresa = db.query(Empresa).filter_by(slug=slug).first()
    solicitacao = db.get(Solicitacao, solicitacao_id)
    return templates.TemplateResponse("publico/obrigado.html",
                                      {"request": request, "empresa": empresa, "solicitacao": solicitacao})

# ============================================================
# CENTRAL DE INTELIGÊNCIA LOGÍSTICA (módulo premium independente)
# ============================================================

def _coord_de_link(link: str | None):
    if not link:
        return (None, None)
    texto = str(link)
    padroes = [r'@(-?\d+\.\d+),(-?\d+\.\d+)', r'query=(-?\d+\.\d+),(-?\d+\.\d+)', r'q=(-?\d+\.\d+),(-?\d+\.\d+)']
    for padrao in padroes:
        achou = re.search(padrao, texto)
        if achou:
            try:
                return float(achou.group(1)), float(achou.group(2))
            except Exception:
                pass
    return (None, None)


def _distancia_km(lat1, lon1, lat2, lon2):
    if None in (lat1, lon1, lat2, lon2):
        return None
    raio = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return raio * (2 * math.atan2(math.sqrt(a), math.sqrt(1 - a)))


_CACHE_TRECHOS_RODOVIARIOS: dict[tuple, tuple[float, int, str]] = {}


def _trecho_rodoviario(lat1, lon1, lat2, lon2, cfg=None):
    """Retorna distância por ruas e duração operacional do trecho.

    Usa OSRM para calcular a rota viária real. Como o serviço público não fornece
    trânsito em tempo real, aplica uma margem urbana configurável. Se o provedor
    estiver indisponível, usa uma estimativa conservadora, nunca linha reta pura.
    """
    if None in (lat1, lon1, lat2, lon2):
        return None, None, "sem_coordenadas"

    try:
        chave = tuple(round(float(v), 5) for v in (lat1, lon1, lat2, lon2))
    except Exception:
        return None, None, "coordenadas_invalidas"
    if chave in _CACHE_TRECHOS_RODOVIARIOS:
        return _CACHE_TRECHOS_RODOVIARIOS[chave]

    fator_trafego = max(1.0, min(2.5, float(os.getenv("ROTA_FATOR_TRAFICO", "1.35") or 1.35)))
    margem_urbana = max(0, min(20, int(os.getenv("ROTA_MARGEM_URBANA_MIN", "3") or 3)))
    url = (
        "https://router.project-osrm.org/route/v1/driving/"
        f"{float(lon1):.6f},{float(lat1):.6f};{float(lon2):.6f},{float(lat2):.6f}"
        "?overview=false&steps=false&alternatives=false"
    )
    try:
        req = UrlRequest(url, headers={
            "User-Agent": "Conect-Inteligencia-Logistica/1.0",
            "Accept": "application/json",
        })
        with urlopen(req, timeout=10) as resposta:
            dados = json.loads(resposta.read().decode("utf-8"))
        rotas = dados.get("routes") or []
        if dados.get("code") == "Ok" and rotas:
            distancia_km = max(0.0, float(rotas[0].get("distance") or 0) / 1000.0)
            minutos_base = max(1.0, float(rotas[0].get("duration") or 0) / 60.0)
            minutos = max(1, int(math.ceil(minutos_base * fator_trafego + margem_urbana)))
            resultado = (round(distancia_km, 1), minutos, "osrm")
            _CACHE_TRECHOS_RODOVIARIOS[chave] = resultado
            logger.info("[ROTA] osrm distancia=%.1fkm base=%.1fmin operacional=%smin", distancia_km, minutos_base, minutos)
            return resultado
    except Exception as exc:
        logger.warning("[ROTA] OSRM indisponível; usando estimativa conservadora: %s", exc)

    linha_reta = _distancia_km(lat1, lon1, lat2, lon2)
    if linha_reta is None:
        return None, None, "falha"
    # Ruas raramente seguem linha reta. O fator 1,30 aproxima a malha viária e a
    # velocidade urbana é limitada a 25 km/h para não produzir previsões irreais.
    distancia_km = max(0.1, linha_reta * 1.30)
    velocidade_cfg = float(getattr(cfg, "velocidade_media_kmh", 25) or 25) if cfg else 25.0
    velocidade_urbana = max(12.0, min(25.0, velocidade_cfg))
    minutos = max(5, int(math.ceil((distancia_km / velocidade_urbana) * 60 * fator_trafego + margem_urbana)))
    resultado = (round(distancia_km, 1), minutos, "estimativa")
    _CACHE_TRECHOS_RODOVIARIOS[chave] = resultado
    return resultado


def _config_rota(db: Session, empresa_id: int):
    cfg = db.query(ConfiguracaoRotaInteligente).filter_by(empresa_id=empresa_id).first()
    if not cfg:
        cfg = ConfiguracaoRotaInteligente(empresa_id=empresa_id)
        db.add(cfg)
        db.flush()
    return cfg


def _endereco_operacao(agenda: Agenda):
    """Usa exatamente a mesma origem de endereço do botão Iniciar rota."""
    sol = agenda.solicitacao
    if not sol:
        return agenda.bairro or "Endereço não informado"
    return endereco_rota_solicitacao(sol) or agenda.bairro or "Endereço não informado"


def _hora_somar(hora_base: time | None, minutos: int, padrao: time = time(8, 0)) -> time:
    """Soma minutos a um horário sem depender da data real do evento."""
    base = hora_base or padrao
    return (datetime.combine(date.today(), base) + timedelta(minutes=max(0, int(minutos or 0)))).time()


def _respeitar_horario_minimo_cliente(instante: datetime, cfg: ConfiguracaoRotaInteligente, tipo: str | None = None) -> datetime:
    """Impede chegada ao cliente antes do horário mínimo contratual configurado.

    A equipe pode sair da loja antes desse horário, mas aguarda para chegar ao primeiro
    cliente somente a partir do limite definido nas configurações da Inteligência.
    Paradas técnicas de loja não recebem essa restrição.
    """
    if tipo == "loja":
        return instante
    minimo = getattr(cfg, "horario_minimo_cliente", None) or time(8, 0)
    limite = datetime.combine(instante.date(), minimo)
    return max(instante, limite)


def _produtos_quantidades(sol: Solicitacao | None) -> dict[int, int]:
    produtos: dict[int, int] = {}
    if not sol:
        return produtos
    for item in sol.itens or []:
        produto_id = getattr(item, "produto_id", None) or (item.produto.id if item.produto else None)
        if produto_id:
            produtos[int(produto_id)] = produtos.get(int(produto_id), 0) + max(1, int(item.quantidade or 1))
    if not produtos and sol.produto_id:
        produtos[int(sol.produto_id)] = 1
    return produtos


def _limite_operacao(agenda: Agenda, cfg: ConfiguracaoRotaInteligente, tipo: str | None = None,
                     data_operacao: date | None = None, hora_saida: time | None = None):
    """Retorna o compromisso usado pela Inteligência.

    Uma retirada comum e ainda não roteirizada é flexível: os horários existentes no
    contrato/operação não funcionam como trava. O horário só é protegido quando o
    operador salvou a roteirização ou quando a retirada é obrigatória.
    """
    sol = agenda.solicitacao
    tipo = tipo or agenda.tipo_evento
    if tipo == "retirada":
        horario_protegido = bool(agenda.roteirizado or (sol and sol.retirada_obrigatoria))
        if horario_protegido:
            # Roteirização manual prevalece sobre qualquer sugestão da Inteligência.
            if agenda.roteirizado and agenda.hora_inicio:
                return agenda.hora_inicio, False
            if sol and sol.retirada_hora:
                return sol.retirada_hora, False
            return agenda.hora_inicio or hora_saida or time(8, 0), False

        # Retirada não roteirizada e não obrigatória não possui compromisso de hora.
        # 23:59 é apenas um limite técnico neutro; a posição será escolhida por
        # proximidade, reaproveitamento e impacto nas entregas.
        return time(23, 59), False

    inicio = sol.hora_inicio if sol and sol.hora_inicio else agenda.hora_inicio
    base = datetime.combine(date.today(), inicio)
    return (base - timedelta(minutes=max(0, int(cfg.antecedencia_entrega or 60)))).time(), False


def _pontos_carga_solicitacao(sol: Solicitacao | None) -> int:
    if not sol:
        return 1
    total = 0
    for item in sol.itens or []:
        pontos = int(getattr(item.produto, "carga_pontos", 1) or 1) if item.produto else 1
        total += max(1, int(item.quantidade or 1)) * max(1, pontos)
    if total == 0 and sol.produto:
        total = max(1, int(getattr(sol.produto, "carga_pontos", 1) or 1))
    return max(1, total)


def _unidades_carga_por_veiculo(produtos_quantidades: dict[int, int], veiculo: VeiculoLogistico | None):
    """Cria unidades usando exclusivamente o perfil do veículo, nunca o cadastro do item."""
    unidades = []
    perfis = {p.produto_id: p for p in (veiculo.perfis_carga if veiculo else []) if p.ativo}
    for produto_id, quantidade in (produtos_quantidades or {}).items():
        perfil = perfis.get(int(produto_id))
        if not perfil:
            return None, f"Produto {produto_id} sem perfil de carga para o veículo selecionado"
        locais = tuple(local for local, permitido in (
            ("interno", perfil.permite_interno), ("mala", perfil.permite_mala), ("teto", perfil.permite_teto)
        ) if permitido)
        if not locais:
            return None, f"Produto {produto_id} não possui compartimento permitido neste veículo"
        for _ in range(max(1, int(quantidade or 1)) * max(1, int(perfil.volumes or 1))):
            unidades.append({"produto_id": int(produto_id), "locais": locais})
    return unidades, None


def _acomodar_unidades(unidades, capacidades):
    ocupacao = {"interno": 0, "mala": 0, "teto": 0}
    ordenadas = sorted(unidades, key=lambda u: len(u.get("locais") or ("interno",)))
    def tentar(i):
        if i >= len(ordenadas):
            return True
        unidade = ordenadas[i]
        for local in sorted(unidade.get("locais") or ("interno",), key=lambda l: capacidades.get(l, 0) - ocupacao.get(l, 0)):
            if ocupacao.get(local, 0) < capacidades.get(local, 0):
                ocupacao[local] += 1
                if tentar(i + 1):
                    return True
                ocupacao[local] -= 1
        return False
    return (True, ocupacao.copy()) if tentar(0) else (False, ocupacao)


def _inserir_retornos_por_compartimento(ordenados, veiculo, cfg):
    if not veiculo:
        raise ValueError("Selecione um veículo para a Inteligência calcular a carga")
    capacidades = {"interno": int(veiculo.capacidade_interno or 0), "mala": int(veiculo.capacidade_mala or 0), "teto": int(veiculo.capacidade_teto or 0)}
    for c in ordenados:
        unidades, erro = _unidades_carga_por_veiculo(c.get("produtos") or {}, veiculo)
        if erro:
            raise ValueError(erro)
        c["unidades_carga"] = unidades or []
    resultado, carga = [], []
    entregas_pendentes = [c for c in ordenados if c["tipo"] == "entrega"]
    carregadas = set()
    def carregar_na_loja(inicio=0):
        nonlocal carga
        carga = []  # retiradas são descarregadas e entregas futuras são carregadas
        # Créditos representam equipamentos que serão recolhidos antes da entrega.
        # Assim a loja não carrega uma segunda unidade desnecessariamente.
        creditos_retirada: dict[int, int] = {}
        for futura in ordenados[inicio:]:
            if futura["tipo"] == "retirada":
                for pid, qtd in (futura.get("produtos") or {}).items():
                    creditos_retirada[int(pid)] = creditos_retirada.get(int(pid), 0) + int(qtd or 0)
                continue
            if futura["tipo"] != "entrega":
                continue
            chave = futura["agenda"].id if futura.get("agenda") else id(futura)
            if chave in carregadas:
                continue
            # Se uma retirada anterior na sequência abastece esta entrega, carrega
            # na loja apenas o saldo não coberto por esse reaproveitamento.
            faltantes_por_produto = dict(futura.get("produtos") or {})
            for pid, qtd in list(faltantes_por_produto.items()):
                credito = min(int(qtd or 0), creditos_retirada.get(int(pid), 0))
                if credito:
                    faltantes_por_produto[int(pid)] = int(qtd or 0) - credito
                    creditos_retirada[int(pid)] -= credito
            faltantes_por_produto = {pid: qtd for pid, qtd in faltantes_por_produto.items() if qtd > 0}
            unidades_loja, erro = _unidades_carga_por_veiculo(faltantes_por_produto, veiculo)
            if erro:
                raise ValueError(erro)
            teste = carga + (unidades_loja or [])
            cabe, _ = _acomodar_unidades(teste, capacidades)
            if cabe:
                carga = teste
                carregadas.add(chave)
                futura["unidades_carregadas_loja"] = unidades_loja or []
                futura["abastecida_por_retirada"] = len(unidades_loja or []) < len(futura.get("unidades_carga", []))
    carregar_na_loja(0)
    for idx, c in enumerate(ordenados):
        chave = c["agenda"].id if c.get("agenda") else id(c)
        unidades = c.get("unidades_carga", [])
        precisa_loja = False
        if c["tipo"] == "entrega" and chave not in carregadas:
            precisa_loja = True
        elif c["tipo"] == "retirada":
            precisa_loja = not _acomodar_unidades(carga + unidades, capacidades)[0]
        if precisa_loja:
            resultado.append({"agenda": None, "tipo": "loja", "titulo": "Retorno automático à loja",
                "endereco": cfg.endereco_loja or "Loja", "lat": cfg.latitude_loja, "lon": cfg.longitude_loja,
                "limite": None, "servico": max(0, int(cfg.minutos_parada_loja or 20)), "pontos": 0,
                "carga_movimento": -len(carga), "carga_apos": 0, "risco": "normal",
                "motivo": "Retorno automático: redistribuição da carga por compartimentos"})
            carregar_na_loja(idx)
        if c["tipo"] == "entrega":
            for unidade in unidades:
                for pos, atual in enumerate(carga):
                    if atual.get("produto_id") == unidade.get("produto_id"):
                        carga.pop(pos); break
            movimento = -len(unidades)
        else:
            carga.extend(unidades); movimento = len(unidades)
        cabe, ocupacao = _acomodar_unidades(carga, capacidades)
        c["carga_movimento"] = movimento
        c["carga_apos"] = len(carga)
        c["ocupacao_compartimentos"] = ocupacao
        resultado.append(c)
    return resultado



def _ajustar_ordem_inicial_por_capacidade(ordenados, veiculo, cfg):
    """Evita começar a missão com uma retirada que não cabe no veículo.

    Se a simulação exigir um retorno à loja antes da primeira operação, a ordem
    escolhida é fisicamente impossível: a missão já começa na loja. Nesse caso,
    antecipa a primeira entrega disponível para liberar espaço e testa novamente.
    """
    if not veiculo or not ordenados:
        return ordenados
    base = list(ordenados)
    for _ in range(len(base)):
        simulada = _inserir_retornos_por_compartimento([_copiar_candidato_rota(c) for c in base], veiculo, cfg)
        if not simulada or simulada[0].get("tipo") != "loja":
            return base
        if base[0].get("tipo") != "retirada":
            return base
        idx_entrega = next((i for i, c in enumerate(base[1:], 1) if c.get("tipo") == "entrega"), None)
        if idx_entrega is None:
            return base
        entrega = base.pop(idx_entrega)
        base.insert(0, entrega)
        entrega["motivo"] = (entrega.get("motivo") or "Entrega priorizada") + " • Antecipada para liberar volume no veículo antes da retirada"
    return base


def _texto_ocupacao_compartimentos(ocupacao, capacidades):
    ocupacao = ocupacao or {"interno": 0, "mala": 0, "teto": 0}
    capacidades = capacidades or {"interno": 0, "mala": 0, "teto": 0}
    return (
        f"Interno {ocupacao.get('interno', 0)}/{capacidades.get('interno', 0)} · "
        f"Mala {ocupacao.get('mala', 0)}/{capacidades.get('mala', 0)} · "
        f"Teto/outros {ocupacao.get('teto', 0)}/{capacidades.get('teto', 0)}"
    )


def _anotar_ocupacao_rota_salva(paradas, veiculo):
    """Reconstrói o volume antes/depois de cada card para auditoria visual."""
    if not veiculo:
        return
    capacidades = {
        "interno": int(veiculo.capacidade_interno or 0),
        "mala": int(veiculo.capacidade_mala or 0),
        "teto": int(veiculo.capacidade_teto or 0),
    }
    grupos, atual = [], []
    for p in paradas:
        atual.append(p)
        if p.tipo == "loja":
            grupos.append(atual); atual = []
    if atual:
        grupos.append(atual)

    for grupo in grupos:
        carga = []
        creditos = {}
        # Monta exatamente a carga prevista na saída desta missão.
        for p in grupo:
            if p.tipo == "loja":
                continue
            produtos = _produtos_quantidades(p.solicitacao)
            if p.tipo == "retirada":
                for pid, qtd in produtos.items():
                    creditos[int(pid)] = creditos.get(int(pid), 0) + int(qtd or 0)
                continue
            faltantes = dict(produtos)
            for pid, qtd in list(faltantes.items()):
                usar = min(int(qtd or 0), creditos.get(int(pid), 0))
                faltantes[int(pid)] = int(qtd or 0) - usar
                creditos[int(pid)] = creditos.get(int(pid), 0) - usar
            faltantes = {pid: qtd for pid, qtd in faltantes.items() if qtd > 0}
            unidades, _ = _unidades_carga_por_veiculo(faltantes, veiculo)
            carga.extend(unidades or [])

        cabe_saida, ocup_saida = _acomodar_unidades(carga, capacidades)
        primeira = next((p for p in grupo if p.tipo != "loja"), None)
        if primeira:
            primeira.ocupacao_saida_view = _texto_ocupacao_compartimentos(ocup_saida, capacidades)
            primeira.carga_saida_excede_view = not cabe_saida

        for p in grupo:
            cabe_antes, ocup_antes = _acomodar_unidades(carga, capacidades)
            p.ocupacao_antes_view = _texto_ocupacao_compartimentos(ocup_antes, capacidades)
            p.carga_antes_excede_view = not cabe_antes
            if p.tipo == "loja":
                carga = []
            else:
                unidades, _ = _unidades_carga_por_veiculo(_produtos_quantidades(p.solicitacao), veiculo)
                unidades = unidades or []
                if p.tipo == "entrega":
                    for unidade in unidades:
                        for pos, atual_u in enumerate(carga):
                            if atual_u.get("produto_id") == unidade.get("produto_id"):
                                carga.pop(pos); break
                elif p.tipo == "retirada":
                    carga.extend(unidades)
            cabe_depois, ocup_depois = _acomodar_unidades(carga, capacidades)
            p.ocupacao_depois_view = _texto_ocupacao_compartimentos(ocup_depois, capacidades)
            p.carga_depois_excede_view = not cabe_depois
            p.capacidade_compartimentos_view = _texto_ocupacao_compartimentos({}, capacidades)

def _retiradas_planejadas_inteligencia(db: Session, empresa_id: int, ignorar_rota_id: int | None = None):
    """Retorna a primeira data planejada de cada retirada ainda ativa.

    Uma previsão aceita pela Inteligência fica reservada para aquele dia e não
    reaparece automaticamente nos dias seguintes. Se a parada não for concluída,
    adiada ou removida, a sincronização do próprio dia poderá recolocá-la depois.
    """
    q = (
        db.query(RotaInteligenteParada, RotaInteligente)
        .join(RotaInteligente, RotaInteligenteParada.rota_id == RotaInteligente.id)
        .filter(
            RotaInteligente.empresa_id == empresa_id,
            RotaInteligente.status != "concluida",
            RotaInteligenteParada.tipo == "retirada",
            RotaInteligenteParada.status != "concluido",
            RotaInteligenteParada.solicitacao_id.isnot(None),
        )
    )
    if ignorar_rota_id:
        q = q.filter(RotaInteligente.id != ignorar_rota_id)
    planejadas = {}
    for parada, rota in q.order_by(RotaInteligente.data_operacao.asc()).all():
        sid = int(parada.solicitacao_id)
        planejadas.setdefault(sid, rota.data_operacao)
    return planejadas


def _estoque_fisico_projetado(db: Session, empresa: Empresa, data_operacao: date,
                               retiradas_planejadas: dict[int, date]):
    """Calcula o que estará fisicamente na loja no início do dia.

    quantidade_disponivel representa o total operacional cadastrado. Descontamos
    contratos entregues e ainda na rua. Uma retirada planejada para dia anterior
    é considerada realizada; retirada do próprio dia continua na rua até a parada.
    """
    totais = {
        p.id: max(0, int(p.quantidade_disponivel or 0))
        for p in db.query(ProdutoServico).filter(ProdutoServico.empresa_id == empresa.id).all()
    }
    na_rua = {pid: 0 for pid in totais}
    entregas = (
        db.query(Agenda)
        .join(Solicitacao, Agenda.solicitacao_id == Solicitacao.id)
        .filter(
            Agenda.empresa_id == empresa.id,
            Agenda.tipo_evento == "entrega",
            Agenda.data < data_operacao,
            Agenda.status_operacional == "concluido",
            Solicitacao.status.in_(STATUS_CONTRATO_APROVADO),
        ).all()
    )
    for entrega in entregas:
        sid = int(entrega.solicitacao_id)
        retirada_concluida = db.query(Agenda.id).filter(
            Agenda.empresa_id == empresa.id,
            Agenda.solicitacao_id == sid,
            Agenda.tipo_evento == "retirada",
            Agenda.status_operacional == "concluido",
        ).first()
        if retirada_concluida:
            continue
        data_planejada = retiradas_planejadas.get(sid)
        if data_planejada and data_planejada < data_operacao:
            continue
        for pid, qtd in _produtos_quantidades(entrega.solicitacao).items():
            na_rua[int(pid)] = na_rua.get(int(pid), 0) + int(qtd or 0)
    na_loja = {pid: max(0, total - na_rua.get(pid, 0)) for pid, total in totais.items()}
    return totais, na_rua, na_loja


def _montar_candidatos_rota(db: Session, empresa: Empresa, data_operacao: date, equipe_id: int | None, cfg,
                             hora_saida: time | None = None, ignorar_rota_id: int | None = None):
    """Monta operações independentes de entrega e retirada.

    Inclui entregas do dia, retiradas do dia, retiradas vencidas de dias anteriores e
    cria uma retirada sintética quando o contrato tem retirada prevista mas não existe
    um card específico de retirada na agenda.
    """
    retiradas_planejadas = _retiradas_planejadas_inteligencia(db, empresa.id, ignorar_rota_id)

    q = db.query(Agenda).join(Solicitacao, Agenda.solicitacao_id == Solicitacao.id).filter(
        Agenda.empresa_id == empresa.id,
        Solicitacao.status.in_(STATUS_CONTRATO_APROVADO),
    )
    # A equipe restringe operações atribuídas a outra equipe, mas mantém as ainda não atribuídas.
    if equipe_id:
        q = q.filter(or_(Agenda.equipe_id == equipe_id, Agenda.equipe_id.is_(None)))
    agendas = q.order_by(Agenda.data.asc(), Agenda.hora_inicio.asc(), Agenda.id.asc()).all()

    candidatos = []
    retiradas_existentes: set[int] = set()
    entregas_do_dia = []

    def adicionar(ag: Agenda, tipo: str, *, sintetica: bool = False, data_retirada: date | None = None):
        sol = ag.solicitacao
        lat_link, lon_link = _coord_de_link(ag.link_localizacao)
        lat = sol.latitude if sol and sol.latitude is not None else lat_link
        lon = sol.longitude if sol and sol.longitude is not None else lon_link
        limite, horario_calculado = _limite_operacao(ag, cfg, tipo, data_operacao, hora_saida)
        servico = int(cfg.minutos_desmontagem if tipo == "retirada" else cfg.minutos_montagem)
        pontos = _pontos_carga_solicitacao(sol)
        produtos = _produtos_quantidades(sol)
        vencida = bool(tipo == "retirada" and (data_retirada or (sol.retirada_data if sol else None) or ag.data) < data_operacao)
        obrigatoria = bool(tipo == "retirada" and (vencida or (sol and sol.retirada_obrigatoria)))
        titulo_base = ag.titulo
        if tipo == "retirada" and not str(titulo_base).lower().startswith("retirada"):
            titulo_base = f"Retirada - {titulo_base}"
        candidatos.append({
            "agenda": ag, "agenda_id": None if sintetica else ag.id,
            "solicitacao_id": ag.solicitacao_id,
            "tipo": tipo, "titulo": titulo_base,
            "endereco": _endereco_operacao(ag), "lat": lat, "lon": lon,
            "limite": limite, "servico": max(0, servico), "pontos": pontos,
            "sem_coordenadas": lat is None or lon is None,
            "horario_calculado": horario_calculado,
            "retirada_vencida": vencida,
            "retirada_obrigatoria": obrigatoria,
            "retirada_obrigatoria_estoque": False,
            "retirada_roteirizada_manual": bool(tipo == "retirada" and ag.roteirizado),
            "retirada_horario_protegido": bool(tipo == "retirada" and (ag.roteirizado or obrigatoria)),
            "sintetica": sintetica,
            "produtos": produtos,
            "bairro": ((sol.bairro if sol else None) or ag.bairro or "").strip(),
            "hora_evento_inicio": sol.hora_inicio if sol else ag.hora_inicio,
            "hora_evento_fim": sol.hora_fim if sol else ag.hora_fim,
            "prioridade_grupo": 1 if tipo == "entrega" else 2,
            "reaproveitamento": False,
            "reaproveita_para": None,
        })

    # Cards reais da agenda.
    for ag in agendas:
        sol = ag.solicitacao
        tipo_agenda = ag.tipo_evento if ag.tipo_evento in ("entrega", "retirada") else "entrega"
        if tipo_agenda == "entrega" and ag.data == data_operacao and ag.status_operacional != "concluido":
            adicionar(ag, "entrega")
            entregas_do_dia.append(candidatos[-1])
        elif tipo_agenda == "retirada" and ag.data <= data_operacao and ag.status_operacional != "concluido":
            data_reservada = retiradas_planejadas.get(int(ag.solicitacao_id))
            if data_reservada and data_reservada < data_operacao:
                continue
            adicionar(ag, "retirada", data_retirada=ag.data)
            retiradas_existentes.add(ag.solicitacao_id)

    # Retiradas previstas, inclusive antes de a entrega ser concluída.
    # A Agenda tradicional só cria o card de busca após encerrar a entrega, mas a
    # Inteligência precisa enxergar essa operação futura para planejar a semana.
    entregas_por_solicitacao = {
        ag.solicitacao_id: ag for ag in agendas if ag.tipo_evento == "entrega"
    }
    for solicitacao_id, entrega in entregas_por_solicitacao.items():
        sol = entrega.solicitacao
        if not sol or solicitacao_id in retiradas_existentes:
            continue

        retirada_concluida = any(
            outro.solicitacao_id == solicitacao_id
            and outro.tipo_evento == "retirada"
            and outro.status_operacional == "concluido"
            for outro in agendas
        )
        if retirada_concluida:
            continue

        data_prevista, _ = _previsao_retirada_operacional(sol, entrega)
        data_reservada = retiradas_planejadas.get(int(solicitacao_id))
        if data_reservada and data_reservada < data_operacao:
            continue
        if data_prevista <= data_operacao:
            adicionar(entrega, "retirada", sintetica=True, data_retirada=data_prevista)
            retiradas_existentes.add(solicitacao_id)

    # Cruza a disponibilidade cadastrada com as entregas. Quando faltar unidade,
    # promove apenas as retiradas necessárias. As demais continuam opcionais, mas
    # são marcadas como oportunidade de reaproveitamento quando compartilham produto.
    retiradas = [c for c in candidatos if c["tipo"] == "retirada"]
    demanda: dict[int, int] = {}
    for entrega in entregas_do_dia:
        for produto_id, qtd in entrega["produtos"].items():
            demanda[produto_id] = demanda.get(produto_id, 0) + qtd
    estoque_total, estoque_na_rua, disponibilidade = _estoque_fisico_projetado(
        db, empresa, data_operacao, retiradas_planejadas
    )
    deficit = {pid: max(0, qtd - disponibilidade.get(pid, 0)) for pid, qtd in demanda.items()}
    for candidato in candidatos:
        candidato["estoque_total"] = dict(estoque_total)
        candidato["estoque_na_rua"] = dict(estoque_na_rua)
        candidato["estoque_loja"] = dict(disponibilidade)
        candidato["demanda_dia"] = dict(demanda)

    for entrega in entregas_do_dia:
        for retirada in retiradas:
            comuns = set(entrega["produtos"]) & set(retirada["produtos"])
            if not comuns:
                continue
            retirada["reaproveitamento"] = True
            retirada["reaproveita_para"] = entrega["titulo"]
            for produto_id in comuns:
                if deficit.get(produto_id, 0) > 0:
                    retirada["retirada_obrigatoria"] = True
                    retirada["retirada_obrigatoria_estoque"] = True
                    # A retirada é necessária para liberar equipamento, mas não ganha
                    # prioridade fixa. O otimizador decide sua posição sem atrasar entregas.
                    deficit[produto_id] = max(0, deficit[produto_id] - retirada["produtos"].get(produto_id, 1))
                    break
            if retirada["retirada_obrigatoria"]:
                break

    return candidatos


def _normalizar_localidade(valor: str | None) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or "")).encode("ascii", "ignore").decode("ascii").lower().strip()
    return re.sub(r"[^a-z0-9]+", " ", texto).strip()


def _deslocamento_estimado_sem_coordenadas(origem_bairro: str | None, destino_bairro: str | None) -> int:
    origem = _normalizar_localidade(origem_bairro)
    destino = _normalizar_localidade(destino_bairro)
    if origem and destino and origem == destino:
        return 10
    if origem and destino and (origem in destino or destino in origem):
        return 15
    return 30



def _copiar_candidato_rota(c):
    """Cópia rasa segura para simulações; preserva referências ORM sem alterar o plano real."""
    novo = dict(c)
    novo["produtos"] = dict(c.get("produtos") or {})
    novo["historico_calculo"] = list(c.get("historico_calculo") or [])
    novo.pop("unidades_carga", None)
    novo.pop("ocupacao_compartimentos", None)
    return novo


def _ordem_cauda_aproximada(restantes, lat, lon, cfg):
    """Monta uma cauda rápida para comparar missões completas, sempre terminando na loja."""
    pendentes = list(restantes)
    ordem = []
    lat_atual, lon_atual = lat, lon
    while pendentes:
        melhor = None
        melhor_chave = None
        for c in pendentes:
            _, minutos, _ = _trecho_rodoviario(lat_atual, lon_atual, c.get("lat"), c.get("lon"), cfg)
            if minutos is None:
                minutos = _deslocamento_estimado_sem_coordenadas("", c.get("bairro"))
            # Entregas com menor limite continuam protegidas, mas a distância participa.
            limite_min = c["limite"].hour * 60 + c["limite"].minute if c.get("limite") else 24 * 60
            protegida = 0 if c.get("retirada_horario_protegido") else 1
            chave = (0 if c["tipo"] == "entrega" else protegida, limite_min, int(minutos or 0))
            if melhor_chave is None or chave < melhor_chave:
                melhor_chave, melhor = chave, c
        ordem.append(melhor)
        if melhor.get("lat") is not None:
            lat_atual, lon_atual = melhor.get("lat"), melhor.get("lon")
        pendentes.remove(melhor)
    return ordem


def _avaliar_missao_completa(ordem, data_operacao, hora_saida, cfg, veiculo, origem_lat, origem_lon):
    """Avalia local atual -> todas as paradas -> loja, incluindo carga, retornos e horários."""
    simulados = [_copiar_candidato_rota(c) for c in ordem]
    try:
        if veiculo:
            simulados = _ajustar_ordem_inicial_por_capacidade(simulados, veiculo, cfg)
            simulados = _inserir_retornos_por_compartimento(simulados, veiculo, cfg)
    except ValueError:
        return {"inviavel": True, "atraso": 9999, "minutos": 99999, "km": 99999.0, "retornos": 99}

    atual = datetime.combine(data_operacao, hora_saida)
    lat_atual, lon_atual = origem_lat, origem_lon
    km_total = 0.0
    atraso_total = 0
    maior_atraso = 0
    retornos = 0
    for p in simulados:
        dist, mins, _ = _trecho_rodoviario(lat_atual, lon_atual, p.get("lat"), p.get("lon"), cfg)
        if mins is None:
            mins = _deslocamento_estimado_sem_coordenadas("", p.get("bairro"))
        atual += timedelta(minutes=int(mins or 0))
        atual = _respeitar_horario_minimo_cliente(atual, cfg, p.get("tipo"))
        km_total += float(dist or 0)
        if p.get("tipo") == "retirada" and p.get("retirada_horario_protegido") and p.get("limite"):
            limite = datetime.combine(data_operacao, p["limite"])
            if atual < limite:
                atual = limite
        if p.get("tipo") == "entrega" and p.get("limite"):
            limite = datetime.combine(data_operacao, p["limite"])
            atraso = max(0, int((atual - limite).total_seconds() / 60))
            atraso_total += atraso
            maior_atraso = max(maior_atraso, atraso)
        atual += timedelta(minutes=max(0, int(p.get("servico") or 0)))
        if p.get("tipo") == "loja":
            retornos += 1
        if p.get("lat") is not None:
            lat_atual, lon_atual = p.get("lat"), p.get("lon")

    # O fim da missão é sempre a loja, nunca o último cliente.
    dist_volta, min_volta, _ = _trecho_rodoviario(
        lat_atual, lon_atual, cfg.latitude_loja, cfg.longitude_loja, cfg
    )
    if min_volta is None:
        min_volta = 30
    atual += timedelta(minutes=int(min_volta or 0))
    km_total += float(dist_volta or 0)
    duracao = max(0, int((atual - datetime.combine(data_operacao, hora_saida)).total_seconds() / 60))
    return {
        "inviavel": atraso_total > 0,
        "atraso": atraso_total,
        "maior_atraso": maior_atraso,
        "minutos": duracao,
        "km": round(km_total, 1),
        "retornos": retornos,
        "volta_loja_min": int(min_volta or 0),
    }


def _ordenar_inteligente(candidatos, data_operacao, hora_saida, cfg, origem_lat=None, origem_lon=None, veiculo=None):
    """Ordena por compromisso de entrega, deslocamento e reaproveitamento real.

    Retirada vencida não é colocada automaticamente em primeiro lugar. Ela sobe
    somente quando libera produto para uma entrega e pode ser encaixada sem criar
    atraso relevante. Após uma retirada, entregas do mesmo produto recebem forte
    vantagem para permitir retirada -> entrega direta, sem retorno à loja.
    """
    restantes = list(candidatos)
    # Histórico técnico usado pela tela "Entender o cálculo". Ele registra por que
    # uma parada perdeu prioridade em cada etapa antes de finalmente ser escolhida.
    for candidato in restantes:
        candidato["historico_calculo"] = []
    resultado = []
    atual = datetime.combine(data_operacao, hora_saida)
    lat_atual, lon_atual = origem_lat, origem_lon
    bairro_atual = ""
    velocidade = max(5.0, float(cfg.velocidade_media_kmh or 30))
    produtos_recolhidos: dict[int, int] = {}

    while restantes:
        melhor = None
        melhor_chave = None
        for c in restantes:
            distancia, desloc, fonte_trecho = _trecho_rodoviario(lat_atual, lon_atual, c["lat"], c["lon"], cfg)
            if desloc is None:
                desloc = _deslocamento_estimado_sem_coordenadas(bairro_atual, c.get("bairro"))
            chegada = atual + timedelta(minutes=desloc)
            limite_dt = datetime.combine(data_operacao, c["limite"])

            # Retirada roteirizada pelo operador ou obrigatória tem horário protegido.
            # Se a equipe chegar antes, considera a espera até o horário definido.
            espera_protegida = 0
            if c["tipo"] == "retirada" and c.get("retirada_horario_protegido") and chegada < limite_dt:
                espera_protegida = int((limite_dt - chegada).total_seconds() / 60)
                chegada = limite_dt

            atraso = (chegada - limite_dt).total_seconds() / 60
            folga = (limite_dt - chegada).total_seconds() / 60

            comuns_recolhidos = sum(
                min(qtd, produtos_recolhidos.get(pid, 0))
                for pid, qtd in c.get("produtos", {}).items()
            ) if c["tipo"] == "entrega" else 0
            entregas_dependentes = [
                e for e in restantes
                if e is not c and e["tipo"] == "entrega"
                and set(e.get("produtos", {})) & set(c.get("produtos", {}))
            ] if c["tipo"] == "retirada" else []
            libera_entrega = bool(entregas_dependentes and c.get("retirada_obrigatoria"))

            # Oportunidade geográfica: mesmo sem déficit de estoque, uma retirada pode
            # abastecer diretamente uma entrega posterior quando o desvio é pequeno.
            # Ex.: loja -> retirada próxima -> entrega, evitando levar outra unidade da loja.
            desvio_reaproveitamento = None
            if c["tipo"] == "retirada" and entregas_dependentes:
                for entrega_dep in entregas_dependentes:
                    _, direto_min, _ = _trecho_rodoviario(
                        lat_atual, lon_atual, entrega_dep.get("lat"), entrega_dep.get("lon"), cfg
                    )
                    _, retirada_entrega_min, _ = _trecho_rodoviario(
                        c.get("lat"), c.get("lon"), entrega_dep.get("lat"), entrega_dep.get("lon"), cfg
                    )
                    if direto_min is None:
                        direto_min = _deslocamento_estimado_sem_coordenadas(bairro_atual, entrega_dep.get("bairro"))
                    if retirada_entrega_min is None:
                        retirada_entrega_min = _deslocamento_estimado_sem_coordenadas(c.get("bairro"), entrega_dep.get("bairro"))
                    desvio = max(0, int(desloc) + int(c["servico"]) + int(retirada_entrega_min) - int(direto_min))
                    if desvio_reaproveitamento is None or desvio < desvio_reaproveitamento:
                        desvio_reaproveitamento = desvio
            oportunidade_geografica = bool(
                c["tipo"] == "retirada"
                and c.get("reaproveitamento")
                and desvio_reaproveitamento is not None
            )

            # Retirada estratégica: na primeira decisão, avalia a proximidade real
            # da loja, o reaproveitamento e o desvio máximo configurado. A decisão
            # só será priorizada depois que a missão completa confirmar que nenhuma
            # entrega ficará atrasada e que a capacidade do veículo é válida.
            distancia_loja_retirada = None
            if c["tipo"] == "retirada" and c.get("lat") is not None and cfg.latitude_loja is not None:
                distancia_loja_retirada, _, _ = _trecho_rodoviario(
                    cfg.latitude_loja, cfg.longitude_loja, c.get("lat"), c.get("lon"), cfg
                )
            raio_estrategico = max(0.0, float(getattr(cfg, "raio_retirada_estrategica_km", 10) or 10))
            desvio_max_estrategico = max(0, int(getattr(cfg, "desvio_max_retirada_estrategica_min", 60) or 60))

            # Compara a missão inteira a partir do local atual. A análise não termina
            # no próximo cliente: inclui todas as paradas, retornos exigidos pela carga
            # e a volta final obrigatória à loja.
            cauda = _ordem_cauda_aproximada(
                [x for x in restantes if x is not c], c.get("lat"), c.get("lon"), cfg
            )
            avaliacao_missao = _avaliar_missao_completa(
                [c] + cauda, data_operacao, atual.time(), cfg, veiculo,
                lat_atual, lon_atual
            )
            inviabiliza_entrega = bool(avaliacao_missao.get("inviavel"))
            maior_atraso_futuro = int(avaliacao_missao.get("maior_atraso") or 0)
            retirada_estrategica = bool(
                len(resultado) == 0
                and c["tipo"] == "retirada"
                and c.get("reaproveitamento")
                and distancia_loja_retirada is not None
                and float(distancia_loja_retirada) <= raio_estrategico
                and desvio_reaproveitamento is not None
                and int(desvio_reaproveitamento) <= desvio_max_estrategico
                and not inviabiliza_entrega
            )

            # Menor chave vence. Atraso de entrega domina qualquer economia.
            if c["tipo"] == "entrega":
                score_tipo = 0
                score_prazo = max(0, atraso) * 10000 + max(0, 90 - folga) * 40
                bonus_reuso = -2500 * comuns_recolhidos
            else:
                # Retirada flexível é estratégica, não prioridade absoluta. Quando foi
                # roteirizada manualmente ou é obrigatória, o horário passa a ser trava.
                score_tipo = 3500
                if c.get("retirada_horario_protegido"):
                    score_prazo = max(0, atraso) * 10000 + espera_protegida * 20
                else:
                    score_prazo = 0
                bonus_reuso = -3200 if libera_entrega else (-1800 if c.get("reaproveitamento") else 0)
                if oportunidade_geografica:
                    # Supera a preferência padrão por entrega quando a retirada está perto
                    # da origem/rota e pode ser usada diretamente na entrega seguinte.
                    bonus_reuso -= max(3200, 5200 - int(desvio_reaproveitamento or 0) * 50)
                # Se há entrega que já ficaria apertada, não desvia para retirada agora.
                entregas_criticas = 0
                for e in restantes:
                    if e["tipo"] != "entrega":
                        continue
                    lim_e = datetime.combine(data_operacao, e["limite"])
                    if (lim_e - atual).total_seconds() / 60 <= desloc + c["servico"] + 45:
                        entregas_criticas += 1
                score_tipo += entregas_criticas * 5000

            score_dist = desloc * 10
            penalidade_inviavel = (1_000_000 + int(avaliacao_missao.get("atraso") or 0) * 100_000) if inviabiliza_entrega else 0
            score_missao = (
                int(avaliacao_missao.get("minutos") or 0) * 20
                + int(float(avaliacao_missao.get("km") or 0) * 10)
                + int(avaliacao_missao.get("retornos") or 0) * 600
            )
            # Reaproveitamento válido ganha valor por reduzir carga inicial/retorno,
            # mas nunca supera atraso real de entrega.
            if c["tipo"] == "retirada" and c.get("reaproveitamento"):
                score_missao -= 1800
            pontuacao_total = penalidade_inviavel + score_prazo + score_tipo + score_dist + score_missao + bonus_reuso
            # A prioridade operacional vem antes da pontuação matemática.
            # 0: retirada obrigatória viável; 1: retirada estratégica viável;
            # 2: entrega; 3: demais retiradas. Atrasos continuam eliminatórios.
            if inviabiliza_entrega:
                prioridade_operacional = 9
            elif c["tipo"] == "retirada" and c.get("retirada_obrigatoria"):
                prioridade_operacional = 0
            elif retirada_estrategica:
                prioridade_operacional = 1
            elif c["tipo"] == "entrega":
                prioridade_operacional = 2
            else:
                prioridade_operacional = 3
            chave = (
                prioridade_operacional,
                pontuacao_total,
                limite_dt,
                c["agenda"].id if c.get("agenda") else 0,
            )
            diagnostico_etapa = {
                "etapa": len(resultado) + 1,
                "hora_base": atual.strftime("%H:%M"),
                "deslocamento": int(desloc or 0),
                "chegada": chegada.strftime("%H:%M"),
                "pontuacao": int(pontuacao_total),
                "score_tipo": int(score_tipo),
                "score_prazo": int(score_prazo),
                "score_distancia": int(score_dist),
                "bonus_reuso": int(bonus_reuso),
                "penalidade_inviavel": int(penalidade_inviavel),
                "score_missao": int(score_missao),
                "missao_minutos": int(avaliacao_missao.get("minutos") or 0),
                "missao_km": float(avaliacao_missao.get("km") or 0),
                "missao_retornos": int(avaliacao_missao.get("retornos") or 0),
                "volta_loja_min": int(avaliacao_missao.get("volta_loja_min") or 0),
                "inviabiliza_entrega": bool(inviabiliza_entrega),
                "maior_atraso_futuro": int(maior_atraso_futuro),
                "desvio_reaproveitamento": None if desvio_reaproveitamento is None else int(desvio_reaproveitamento),
                "oportunidade_geografica": bool(oportunidade_geografica),
                "retirada_estrategica": bool(retirada_estrategica),
                "prioridade_operacional": int(prioridade_operacional),
                "distancia_loja_retirada": None if distancia_loja_retirada is None else float(distancia_loja_retirada),
                "raio_estrategico": float(raio_estrategico),
                "desvio_max_estrategico": int(desvio_max_estrategico),
                "retirada_protegida": bool(c.get("retirada_horario_protegido")),
                "fonte_trecho": fonte_trecho or "estimativa",
            }
            c.setdefault("historico_calculo", []).append(diagnostico_etapa)
            if melhor_chave is None or chave < melhor_chave:
                melhor_chave = chave
                melhor = (c, distancia, desloc, chegada, limite_dt, libera_entrega, comuns_recolhidos,
                          oportunidade_geografica, desvio_reaproveitamento, retirada_estrategica, diagnostico_etapa)

        c, distancia, desloc, chegada, limite_dt, libera_entrega, comuns_recolhidos, oportunidade_geografica, desvio_reaproveitamento, retirada_estrategica, diagnostico_escolhido = melhor
        atraso_min = int((chegada - limite_dt).total_seconds() / 60)
        folga_min = int((limite_dt - chegada).total_seconds() / 60)
        risco = "atrasado" if atraso_min > 0 else ("atencao" if folga_min <= 20 else "normal")
        motivo = []
        if c["tipo"] == "entrega":
            motivo.append("Entrega priorizada pelo horário de início e pela menor folga")
            if comuns_recolhidos:
                motivo.append("Usa equipamento recolhido anteriormente, sem retorno à loja")
        elif retirada_estrategica:
            motivo.append("Retirada estratégica priorizada antes da primeira entrega")
            motivo.append(
                f"Próxima da loja: {diagnostico_escolhido.get('distancia_loja_retirada', 0):.1f} km "
                f"(raio configurado: {diagnostico_escolhido.get('raio_estrategico', 0):.1f} km)"
            )
            motivo.append("Missão completa validada sem atraso nas entregas")
            motivo.append("Aumenta o estoque disponível para uma entrega posterior")
        elif libera_entrega:
            motivo.append("Retirada obrigatória: libera equipamento para entrega seguinte")
        elif oportunidade_geografica:
            motivo.append("Retirada antecipada por proximidade da rota e reaproveitamento direto")
            motivo.append(f"Desvio estimado de apenas {int(desvio_reaproveitamento or 0)} min")
        elif c.get("reaproveitamento"):
            motivo.append("Retirada encaixada por oportunidade de reaproveitamento")
        else:
            motivo.append("Retirada encaixada sem comprometer as entregas")
        if c.get("retirada_vencida"):
            motivo.append("Retirada pendente de dia anterior")
        if c.get("retirada_roteirizada_manual"):
            motivo.append(f"Horário definido manualmente e protegido: {c['limite'].strftime('%H:%M')}")
        elif c.get("retirada_obrigatoria_estoque"):
            motivo.append("Retirada obrigatória por falta de estoque físico na loja; horário escolhido pela Inteligência")
            for pid, qtd in (c.get("produtos") or {}).items():
                falta = max(0, int((c.get("demanda_dia") or {}).get(pid, 0)) - int((c.get("estoque_loja") or {}).get(pid, 0)))
                if falta:
                    motivo.append(
                        f"Estoque operacional: total {(c.get('estoque_total') or {}).get(pid, 0)}, "
                        f"na loja {(c.get('estoque_loja') or {}).get(pid, 0)}, "
                        f"na rua {(c.get('estoque_na_rua') or {}).get(pid, 0)}, "
                        f"demanda do dia {(c.get('demanda_dia') or {}).get(pid, 0)}; falta {falta}"
                    )
                    break
        elif c.get("retirada_obrigatoria"):
            motivo.append(f"Horário obrigatório protegido: {c['limite'].strftime('%H:%M')}")
        elif c["tipo"] == "retirada":
            motivo.append("Retirada flexível: horário escolhido pela Inteligência")
        if c.get("reaproveitamento"):
            motivo.append(f"Pode atender {c.get('reaproveita_para')}")
        motivo.append(f"Atraso previsto de {atraso_min} min" if risco == "atrasado" else
                      f"Folga de {max(0, folga_min)} min")
        motivo.append(f"Trecho de {distancia:.1f} km" if distancia is not None else
                      f"Sem coordenadas: deslocamento estimado em {desloc} min")
        motivo.append(f"Carga: {c['pontos']} ponto(s)")

        # Explicação auditável: mostra o que efetivamente pesou na decisão e, nas
        # retiradas, registra também por que ela não foi escolhida nas etapas anteriores.
        detalhe = diagnostico_escolhido
        motivo.append(
            "Cálculo: total {total} = prazo {prazo} + tipo {tipo} + distância {dist} "
            "+ missão completa {missao} + penalidade {pen} + bônus de reaproveitamento {bonus}".format(
                total=detalhe["pontuacao"], prazo=detalhe["score_prazo"],
                tipo=detalhe["score_tipo"], dist=detalhe["score_distancia"],
                missao=detalhe.get("score_missao", 0),
                pen=detalhe["penalidade_inviavel"], bonus=detalhe["bonus_reuso"],
            )
        )
        motivo.append(
            f"Missão simulada: {detalhe.get('missao_km', 0):.1f} km, "
            f"{detalhe.get('missao_minutos', 0)} min, "
            f"{detalhe.get('missao_retornos', 0)} retorno(s) intermediário(s) e "
            f"{detalhe.get('volta_loja_min', 0)} min do último ponto até a loja"
        )
        if c["tipo"] == "retirada":
            if detalhe.get("retirada_estrategica"):
                motivo.append(
                    "Classificação operacional: estratégica; entrou antes da pontuação por estar próxima da loja, "
                    "reaproveitar equipamento e não atrasar nenhuma entrega"
                )
            # Na etapa final a entrega vinculada pode já ter sido removida da lista,
            # portanto o desvio atual pode ficar vazio mesmo havendo coordenadas.
            # Para a auditoria, usa o melhor desvio calculado nas etapas anteriores.
            desvios_validos = [
                a.get("desvio_reaproveitamento")
                for a in c.get("historico_calculo", [])
                if a.get("desvio_reaproveitamento") is not None
            ]
            desvio_auditoria = detalhe["desvio_reaproveitamento"]
            if desvio_auditoria is None and desvios_validos:
                desvio_auditoria = min(desvios_validos)
            if desvio_auditoria is None:
                motivo.append("Reaproveitamento geográfico não aplicável nesta etapa: a entrega vinculada já foi escolhida")
            else:
                motivo.append(
                    f"Acréscimo local estimado: {int(desvio_auditoria)} min; "
                    "a decisão final usa a missão completa com volta à loja"
                )
            if detalhe["inviabiliza_entrega"]:
                motivo.append(
                    f"Antecipar neste ponto atrasaria uma entrega em até "
                    f"{detalhe['maior_atraso_futuro']} min"
                )
            avaliacoes_anteriores = c.get("historico_calculo", [])[:-1]
            for avaliacao in avaliacoes_anteriores[-4:]:
                causa = []
                if avaliacao.get("inviabiliza_entrega"):
                    causa.append(f"risco de atraso de {avaliacao.get('maior_atraso_futuro', 0)} min")
                desvio_ant = avaliacao.get("desvio_reaproveitamento")
                if desvio_ant is not None:
                    causa.append(f"acréscimo local de {desvio_ant} min; missão completa ficou pior")
                if not causa:
                    causa.append(f"pontuação {avaliacao.get('pontuacao', 0)} maior que a parada escolhida")
                motivo.append(
                    f"Na etapa {avaliacao.get('etapa')}, não foi escolhida: " + ", ".join(causa)
                )
        saida = chegada + timedelta(minutes=c["servico"])
        resultado.append({**c, "distancia": float(distancia or 0), "desloc": desloc, "chegada": chegada,
                          "saida": saida, "risco": risco, "motivo": " • ".join(motivo),
                          "folga_min": folga_min})

        if c["tipo"] == "retirada":
            for pid, qtd in c.get("produtos", {}).items():
                produtos_recolhidos[pid] = produtos_recolhidos.get(pid, 0) + qtd
        else:
            for pid, qtd in c.get("produtos", {}).items():
                produtos_recolhidos[pid] = max(0, produtos_recolhidos.get(pid, 0) - qtd)

        atual = saida
        bairro_atual = c.get("bairro") or bairro_atual
        if c["lat"] is not None:
            lat_atual, lon_atual = c["lat"], c["lon"]
        restantes.remove(c)
    return resultado


def _inserir_retornos_capacidade(ordenados, capacidade: int | None, cfg):
    if not capacidade or capacidade <= 0:
        for c in ordenados:
            c["carga_movimento"] = -c["pontos"] if c["tipo"] == "entrega" else c["pontos"]
            c["carga_apos"] = 0
        return ordenados
    resultado, carga = [], 0
    pendentes_entrega = [c for c in ordenados if c["tipo"] == "entrega"]
    carregados = set()
    for c in pendentes_entrega:
        chave_agenda = c["agenda"].id if c.get("agenda") else id(c)
        if carga + c["pontos"] <= capacidade:
            carga += c["pontos"]
            carregados.add(chave_agenda)
    for idx, c in enumerate(ordenados):
        chave_agenda = c["agenda"].id if c.get("agenda") else id(c)
        movimento = -c["pontos"] if c["tipo"] == "entrega" else c["pontos"]
        # Retirada reaproveitada pode alimentar a próxima entrega do mesmo produto sem loja.
        reaproveita = bool(c.get("reaproveitamento"))
        precisa_loja = (
            (c["tipo"] == "entrega" and chave_agenda not in carregados and carga < c["pontos"])
            or (c["tipo"] == "retirada" and not reaproveita and carga + c["pontos"] > capacidade)
        )
        if precisa_loja:
            resultado.append({"agenda": None, "tipo": "loja", "titulo": "Retorno automático à loja",
                "endereco": cfg.endereco_loja or "Loja", "lat": cfg.latitude_loja, "lon": cfg.longitude_loja,
                "limite": None, "servico": max(0, int(cfg.minutos_parada_loja or 20)), "pontos": 0,
                "carga_movimento": -carga, "carga_apos": 0, "risco": "normal",
                "motivo": "Retorno inserido automaticamente pela capacidade do veículo"})
            carga = 0
            if c["tipo"] == "entrega":
                for futura in ordenados[idx:]:
                    if futura["tipo"] != "entrega":
                        continue
                    futura_chave = futura["agenda"].id if futura.get("agenda") else id(futura)
                    if futura_chave not in carregados and carga + futura["pontos"] <= capacidade:
                        carga += futura["pontos"]
                        carregados.add(futura_chave)
        c["carga_movimento"] = movimento
        carga = max(0, carga + movimento)
        c["carga_apos"] = carga
        resultado.append(c)
    return resultado

def _garantir_retorno_final_loja(ordenados, cfg):
    """Fecha cada plano na loja e diferencia retorno final de retorno entre missões.

    Como toda missão começa fisicamente na loja, um retorno inserido na primeira
    posição é redundante e criava uma "1ª rota" vazia. Também compactamos retornos
    consecutivos para não exibir missões sem nenhuma operação entre eles.
    """
    resultado = list(ordenados)

    while resultado and resultado[0].get("tipo") == "loja":
        resultado.pop(0)
    compactado = []
    for parada in resultado:
        if parada.get("tipo") == "loja" and compactado and compactado[-1].get("tipo") == "loja":
            compactado[-1] = parada
        else:
            compactado.append(parada)
    resultado = compactado

    # Retornos já existentes no meio do plano encerram uma missão e iniciam outra.
    for idx, parada in enumerate(resultado):
        if parada.get("tipo") != "loja":
            continue
        tem_operacao_depois = any(p.get("tipo") != "loja" for p in resultado[idx + 1:])
        if tem_operacao_depois:
            parada["titulo"] = "Retornar à loja — preparar próxima missão"
            parada["motivo"] = (
                parada.get("motivo")
                or "Retorno intermediário para descarregar, reorganizar a carga e iniciar a próxima missão"
            )
            parada["retorno_final"] = False
        else:
            parada["titulo"] = "Retornar à loja — finalizar operação"
            parada["motivo"] = "Última etapa obrigatória: a rota termina na loja, não no cliente"
            parada["retorno_final"] = True

    # O último card deve ser sempre a volta à loja.
    if not resultado or resultado[-1].get("tipo") != "loja":
        resultado.append({
            "tipo": "loja",
            "titulo": "Retornar à loja — finalizar operação",
            "endereco": cfg.endereco_loja or "Loja",
            "bairro": "",
            "lat": cfg.latitude_loja,
            "lon": cfg.longitude_loja,
            "limite": None,
            "servico": 0,
            "pontos": 0,
            "produtos": {},
            "carga_movimento": 0,
            "carga_apos": 0,
            "risco": "normal",
            "retorno_final": True,
            "motivo": "Última etapa obrigatória: retornar à loja para finalizar a operação",
            "historico_calculo": [],
        })
    return resultado


def _simular_sequencia_rota(ordenados, data_operacao, hora_saida, cfg, origem_lat=None, origem_lon=None):
    """Recalcula a linha do tempo e informa se alguma entrega chega após o limite."""
    atual = datetime.combine(data_operacao, hora_saida)
    lat_atual, lon_atual = origem_lat, origem_lon
    bairro_atual = ""
    velocidade = max(5.0, float(cfg.velocidade_media_kmh or 30))
    atrasos = []
    for c in ordenados:
        distancia, desloc, _ = _trecho_rodoviario(lat_atual, lon_atual, c.get("lat"), c.get("lon"), cfg)
        if desloc is None:
            desloc = _deslocamento_estimado_sem_coordenadas(bairro_atual, c.get("bairro"))
        chegada = atual + timedelta(minutes=desloc)
        chegada = _respeitar_horario_minimo_cliente(chegada, cfg, c.get("tipo"))
        saida = chegada + timedelta(minutes=max(0, int(c.get("servico") or 0)))
        c["distancia"] = float(distancia or 0)
        c["desloc"] = desloc
        c["chegada"] = chegada
        c["saida"] = saida
        if c.get("tipo") == "entrega" and c.get("limite"):
            limite_dt = datetime.combine(data_operacao, c["limite"])
            folga = int((limite_dt - chegada).total_seconds() / 60)
            c["folga_min"] = folga
            c["risco"] = "atrasado" if folga < 0 else ("atencao" if folga <= 20 else "normal")
            if folga < 0:
                atrasos.append({"titulo": c.get("titulo") or "Entrega", "minutos": -folga})
        atual = saida
        bairro_atual = c.get("bairro") or bairro_atual
        if c.get("lat") is not None and c.get("lon") is not None:
            lat_atual, lon_atual = c["lat"], c["lon"]
    return atrasos


def _recalcular_rota_salva(db: Session, rota: RotaInteligente):
    cfg = _config_rota(db, rota.empresa_id)
    paradas = db.query(RotaInteligenteParada).filter_by(rota_id=rota.id).order_by(RotaInteligenteParada.ordem).all()
    if not paradas:
        return
    atual = datetime.combine(rota.data_operacao, rota.horario_saida)
    lat_atual, lon_atual = cfg.latitude_loja, cfg.longitude_loja
    carga_inicial = 0
    for inicial in paradas:
        if inicial.tipo == "loja": break
        if int(inicial.carga_movimento or 0) < 0: carga_inicial += abs(int(inicial.carga_movimento or 0))
    distancia_total = 0.0; retornos = 0; carga = carga_inicial; carga_max = carga_inicial
    velocidade = max(5.0, float(cfg.velocidade_media_kmh or 30))
    for idx, p in enumerate(paradas, 1):
        p.ordem = idx
        if p.status == "concluido" and p.chegada_real:
            # chegada_real registra o instante em que o operador encerrou a parada.
            # A partir daqui, toda a previsão usa o relógio real, absorvendo atrasos
            # ou adiantamentos ocorridos durante a execução.
            atual = p.chegada_real
        else:
            distancia, desloc, _ = _trecho_rodoviario(lat_atual, lon_atual, p.latitude, p.longitude, cfg)
            if desloc is None:
                desloc = 30
            p.distancia_anterior_km = float(distancia or 0); p.deslocamento_anterior_min = desloc
            p.chegada_prevista = atual + timedelta(minutes=desloc)
            p.chegada_prevista = _respeitar_horario_minimo_cliente(p.chegada_prevista, cfg, p.tipo)
            p.saida_prevista = p.chegada_prevista + timedelta(minutes=p.servico_min or 0)
            if p.horario_limite:
                limite = datetime.combine(rota.data_operacao, p.horario_limite)
                folga = int((limite - p.chegada_prevista).total_seconds() / 60)
                p.risco = "atrasado" if folga < 0 else ("atencao" if folga <= 20 else "normal")
            else: p.risco = "normal"
            atual = p.saida_prevista
            distancia_total += float(distancia or 0)
        if p.tipo == "loja":
            retornos += 1; carga = 0
        else:
            carga = max(0, carga + int(p.carga_movimento or 0))
        p.carga_apos_parada = carga; carga_max = max(carga_max, carga)
        if p.latitude is not None: lat_atual, lon_atual = p.latitude, p.longitude
    duracao = max(0, int((atual - datetime.combine(rota.data_operacao, rota.horario_saida)).total_seconds() / 60))
    rota.distancia_total_km = round(distancia_total, 1); rota.duracao_total_min = duracao
    rota.retornos_loja = retornos; rota.carga_maxima_pontos = carga_max
    rota.custo_estimado = round(distancia_total * float(cfg.custo_km or 0) + (duracao / 60) * float(cfg.custo_hora_equipe or 0), 2)
    entregas_atrasadas = [p for p in paradas if p.tipo == "entrega" and p.risco == "atrasado"]
    if paradas and all(p.status == "concluido" for p in paradas):
        rota.status = "concluida"
    elif any(p.status in ("em_andamento", "concluido") for p in paradas):
        rota.status = "em_execucao"
    else:
        rota.status = "invalida_atraso" if entregas_atrasadas else "planejada"
    rota.versao_calculo = int(rota.versao_calculo or 0) + 1



def _reconstruir_rota_salva(db: Session, rota: RotaInteligente):
    """Reconstrói as paradas com os dados atuais da Operação, sem novo consumo.

    Preserva paradas concluídas e recria todas as demais a partir das agendas e
    retiradas vigentes. Isso permite incluir contratos adicionados depois da
    geração inicial e remover operações canceladas ou já concluídas.
    """
    empresa = db.get(Empresa, rota.empresa_id)
    if not empresa:
        return 0
    cfg = _config_rota(db, rota.empresa_id)
    candidatos = _montar_candidatos_rota(
        db, empresa, rota.data_operacao, rota.equipe_id, cfg, rota.horario_saida, ignorar_rota_id=rota.id
    )
    # A ordem deve ser escolhida sem ficar presa ao horário de saída salvo.
    # Primeiro otimiza a missão com uma base neutra; depois recalcula apenas os
    # horários. Isso evita rejeitar uma retirada estratégica por um "atraso"
    # artificial causado por uma saída calculada para a ordem antiga.
    hora_base = time(8, 0)
    candidatos = _montar_candidatos_rota(
        db, empresa, rota.data_operacao, rota.equipe_id, cfg, hora_base, ignorar_rota_id=rota.id
    )
    ordenados = _ordenar_inteligente(
        candidatos, rota.data_operacao, hora_base, cfg,
        cfg.latitude_loja, cfg.longitude_loja,
        db.query(VeiculoLogistico).filter_by(
            id=rota.veiculo_id, empresa_id=rota.empresa_id, ativo=True
        ).first() if rota.veiculo_id else None
    )
    veiculo = db.query(VeiculoLogistico).filter_by(
        id=rota.veiculo_id, empresa_id=rota.empresa_id, ativo=True
    ).first() if rota.veiculo_id else None
    ordenados = _ajustar_ordem_inicial_por_capacidade(ordenados, veiculo, cfg)
    ordenados = _inserir_retornos_por_compartimento(ordenados, veiculo, cfg)
    ordenados = _garantir_retorno_final_loja(ordenados, cfg)
    rota.horario_saida = _calcular_horario_saida_ideal(
        ordenados, rota.data_operacao, cfg, cfg.latitude_loja, cfg.longitude_loja
    )
    _simular_sequencia_rota(
        ordenados, rota.data_operacao, rota.horario_saida, cfg,
        cfg.latitude_loja, cfg.longitude_loja
    )

    existentes = db.query(RotaInteligenteParada).filter_by(rota_id=rota.id).order_by(
        RotaInteligenteParada.ordem
    ).all()
    concluidas = [p for p in existentes if p.status == "concluido"]
    chaves_concluidas = {
        (p.agenda_id, p.solicitacao_id, p.tipo)
        for p in concluidas if p.tipo != "loja"
    }

    for p in existentes:
        if p.status != "concluido":
            db.delete(p)
    db.flush()

    ordem = 1
    for p in concluidas:
        p.ordem = ordem
        ordem += 1

    adicionadas = 0
    for c in ordenados:
        ag = c.get("agenda")
        agenda_id = c.get("agenda_id")
        solicitacao_id = c.get("solicitacao_id") or (ag.solicitacao_id if ag else None)
        chave = (agenda_id, solicitacao_id, c["tipo"])
        if c["tipo"] != "loja" and chave in chaves_concluidas:
            continue
        db.add(RotaInteligenteParada(
            rota_id=rota.id, agenda_id=agenda_id, solicitacao_id=solicitacao_id,
            ordem=ordem, tipo=c["tipo"], titulo=c["titulo"], endereco=c["endereco"],
            latitude=c.get("lat"), longitude=c.get("lon"), horario_limite=c.get("limite"),
            chegada_prevista=c.get("chegada"), saida_prevista=c.get("saida"),
            distancia_anterior_km=c.get("distancia", 0),
            deslocamento_anterior_min=c.get("desloc", 0), servico_min=c["servico"],
            risco=c.get("risco", "normal"), motivo_prioridade=c.get("motivo"),
            retorno_loja=c["tipo"] == "loja",
            carga_movimento=int(c.get("carga_movimento", 0)),
            carga_apos_parada=int(c.get("carga_apos", 0)),
        ))
        ordem += 1
        adicionadas += 1

    db.flush()
    _recalcular_rota_salva(db, rota)
    return adicionadas


def _sincronizar_rotas_inteligentes_ativas(db: Session, empresa_id: int, ignorar_rota_id: int | None = None) -> int:
    """Relê a Operação e reconstrói rotas ainda abertas sem consumir Humiat."""
    q = db.query(RotaInteligente).filter(
        RotaInteligente.empresa_id == empresa_id,
        RotaInteligente.status != "concluida",
    )
    if ignorar_rota_id:
        q = q.filter(RotaInteligente.id != ignorar_rota_id)

    atualizadas = 0
    for rota_ativa in q.order_by(RotaInteligente.data_operacao.asc()).all():
        _reconstruir_rota_salva(db, rota_ativa)
        atualizadas += 1
    return atualizadas


def _normalizar_endereco_geocodificacao(valor: str) -> str:
    texto = unicodedata.normalize("NFKC", str(valor or ""))
    texto = re.sub(r"\bCEP\s*[:\-]?\s*", "", texto, flags=re.IGNORECASE)
    texto = re.sub(r"\s+", " ", texto).strip(" ,.-")
    return texto


def _partes_unicas_endereco(*partes: str) -> str:
    """Monta uma consulta sem repetir bairro, cidade ou estado já presentes."""
    resultado = []
    vistos = set()
    for parte in partes:
        valor = _normalizar_endereco_geocodificacao(parte)
        if not valor:
            continue
        chave = unicodedata.normalize("NFKD", valor).encode("ascii", "ignore").decode().casefold()
        if chave in vistos:
            continue
        # Evita acrescentar uma parte simples que já aparece inteira em outra parte.
        if any(chave == existente or chave in existente for existente in vistos):
            continue
        vistos.add(chave)
        resultado.append(valor)
    return ", ".join(resultado)


def _consultar_nominatim(consulta: str):
    params = urlencode({
        "q": consulta,
        "format": "jsonv2",
        "limit": 1,
        "countrycodes": "br",
        "addressdetails": 1,
    })
    req = UrlRequest(
        "https://nominatim.openstreetmap.org/search?" + params,
        headers={
            "User-Agent": "Conect-Humiat/1.0 (geocodificacao operacional; contato: suporte@humiat.com.br)",
            "Accept-Language": "pt-BR,pt;q=0.9",
            "Referer": "https://humiat.com.br/",
        },
    )
    with urlopen(req, timeout=15) as resp:
        status = getattr(resp, "status", 200)
        corpo = resp.read().decode("utf-8", errors="replace")
    dados = json.loads(corpo)
    if not dados:
        return None, None, "ZERO_RESULTS"
    lat = float(dados[0]["lat"])
    lon = float(dados[0]["lon"])
    return lat, lon, dados[0].get("display_name") or consulta


def _consultar_photon(consulta: str):
    """Segundo provedor sem chave, usado somente quando o Nominatim não encontra."""
    params = urlencode({"q": consulta, "limit": 1, "lang": "pt"})
    req = UrlRequest(
        "https://photon.komoot.io/api/?" + params,
        headers={"User-Agent": "Conect-Humiat/1.0", "Accept-Language": "pt-BR,pt;q=0.9"},
    )
    with urlopen(req, timeout=15) as resp:
        corpo = resp.read().decode("utf-8", errors="replace")
    dados = json.loads(corpo)
    features = dados.get("features") or []
    if not features:
        return None, None, "ZERO_RESULTS"
    coords = (features[0].get("geometry") or {}).get("coordinates") or []
    if len(coords) < 2:
        return None, None, "RESPOSTA_SEM_COORDENADAS"
    lon, lat = float(coords[0]), float(coords[1])
    props = features[0].get("properties") or {}
    descricao = ", ".join(str(props.get(k)) for k in ("name", "street", "district", "city", "state", "country") if props.get(k))
    return lat, lon, descricao or consulta


def _coordenadas_validas_brasil(lat, lon) -> bool:
    return lat is not None and lon is not None and -34.0 <= float(lat) <= 6.0 and -74.0 <= float(lon) <= -32.0


def _geocodificar_consultas(consultas, identificador="endereco"):
    """Tenta variações e dois provedores, registrando no Render o motivo real."""
    unicas = []
    vistos = set()
    for consulta in consultas:
        consulta = _normalizar_endereco_geocodificacao(consulta)
        if len(consulta) < 8:
            continue
        chave = consulta.casefold()
        if chave not in vistos:
            vistos.add(chave)
            unicas.append(consulta)

    if not unicas:
        geo_logger.warning("[GEO] id=%s endereco insuficiente", identificador)
        return None, None, "endereço insuficiente"

    ultimo_motivo = "não encontrado"
    for indice, consulta in enumerate(unicas, 1):
        for provedor, funcao in (("nominatim", _consultar_nominatim), ("photon", _consultar_photon)):
            geo_logger.info("[GEO] id=%s tentativa=%s provedor=%s consulta=%r", identificador, indice, provedor, consulta)
            try:
                lat, lon, detalhe = funcao(consulta)
                if _coordenadas_validas_brasil(lat, lon):
                    geo_logger.info("[GEO] id=%s localizado provedor=%s lat=%.7f lon=%.7f resultado=%r", identificador, provedor, lat, lon, detalhe)
                    return lat, lon, detalhe
                ultimo_motivo = detalhe or "resultado inválido"
                geo_logger.warning("[GEO] id=%s sem resultado provedor=%s motivo=%s", identificador, provedor, ultimo_motivo)
            except HTTPError as exc:
                corpo = ""
                try:
                    corpo = exc.read().decode("utf-8", errors="replace")[:500]
                except Exception:
                    pass
                ultimo_motivo = f"HTTP {exc.code}"
                geo_logger.warning("[GEO] id=%s provedor=%s erro_http=%s resposta=%r", identificador, provedor, exc.code, corpo)
            except (URLError, TimeoutError) as exc:
                ultimo_motivo = type(exc).__name__
                geo_logger.warning("[GEO] id=%s provedor=%s falha_rede=%r", identificador, provedor, exc)
            except Exception as exc:
                ultimo_motivo = type(exc).__name__
                geo_logger.exception("[GEO] id=%s provedor=%s erro inesperado", identificador, provedor)
            time_module.sleep(1.05 if provedor == "nominatim" else 0.25)
    return None, None, ultimo_motivo


def _variacoes_endereco_solicitacao(sol: Solicitacao):
    dados = dados_endereco_solicitacao(sol)
    rua = dados["endereco"]
    numero = dados["numero"]
    bairro = dados["bairro"]
    cidade = dados["cidade"] or "Rio de Janeiro"
    estado = dados["estado"] or "RJ"
    cep = re.sub(r"\D", "", dados["cep"] or "")

    consultas = [
        _partes_unicas_endereco(rua, numero, bairro, cidade, estado, cep, "Brasil"),
        _partes_unicas_endereco(rua, numero, cidade, estado, cep, "Brasil"),
        _partes_unicas_endereco(cep, numero, cidade, estado, "Brasil") if cep else "",
        _partes_unicas_endereco(rua, numero, bairro, cidade, "Brasil"),
    ]
    return consultas


def _geocodificar_endereco_nominatim(endereco: str, bairro: str = ""):
    """Compatibilidade com chamadas antigas; agora possui variações e provedor alternativo."""
    consulta = _partes_unicas_endereco(endereco, bairro, "Brasil")
    return _geocodificar_consultas([consulta], identificador="endereco-avulso")


def _geocodificar_solicitacao(sol: Solicitacao):
    return _geocodificar_consultas(_variacoes_endereco_solicitacao(sol), identificador=f"solicitacao-{sol.id}")


def _geocodificar_solicitacoes_automaticamente(db: Session, empresa_id: int, solicitacoes):
    """Geocodifica em lote, reutilizando coordenadas para endereços iguais."""
    resultado = {"localizados": 0, "ja_prontos": 0, "pendentes": 0, "erros": []}
    cache = {}
    for sol, endereco, bairro in solicitacoes:
        if sol.latitude is not None and sol.longitude is not None and sol.status_geocodificacao == "localizado":
            resultado["ja_prontos"] += 1
            continue
        consultas = _variacoes_endereco_solicitacao(sol)
        chave = tuple(_normalizar_endereco_geocodificacao(q).casefold() for q in consultas if q)
        if chave in cache:
            lat, lon, detalhe = cache[chave]
        else:
            lat, lon, detalhe = _geocodificar_solicitacao(sol)
            cache[chave] = (lat, lon, detalhe)
        sol.data_geocodificacao = agora_utc()
        if lat is None or lon is None:
            sol.status_geocodificacao = "revisar"
            resultado["pendentes"] += 1
            nome = getattr(sol.cliente, "nome", None) or f"Contrato {sol.id}"
            resultado["erros"].append(f"{nome}: {detalhe}")
            geo_logger.error("[GEO] solicitacao=%s cliente=%r falhou motivo=%s", sol.id, nome, detalhe)
            continue
        sol.latitude, sol.longitude = lat, lon
        sol.status_geocodificacao = "localizado"
        resultado["localizados"] += 1
    db.commit()
    return resultado


def _invalidar_geocodificacao(item: Solicitacao):
    item.latitude = None
    item.longitude = None
    item.status_geocodificacao = "pendente"
    item.data_geocodificacao = None


def _tentar_geocodificar_solicitacao(db: Session, item: Solicitacao, *, commit: bool = True) -> bool:
    """Tenta localizar sem nunca impedir o salvamento do contrato."""
    try:
        lat, lon, detalhe = _geocodificar_solicitacao(item)
        item.data_geocodificacao = agora_utc()
        if lat is None or lon is None:
            item.status_geocodificacao = "revisar"
            geo_logger.warning("[GEO] solicitacao=%s pendente motivo=%s", item.id, detalhe)
            if commit:
                db.commit()
            return False
        item.latitude, item.longitude = lat, lon
        item.status_geocodificacao = "localizado"
        if commit:
            db.commit()
        return True
    except Exception:
        geo_logger.exception("[GEO] solicitacao=%s falha inesperada", getattr(item, "id", None))
        item.status_geocodificacao = "revisar"
        if commit:
            try:
                db.commit()
            except Exception:
                db.rollback()
        return False


def _garantir_localizacao_loja(db: Session, cfg) -> bool:
    if cfg.latitude_loja is not None and cfg.longitude_loja is not None:
        return True
    endereco = (cfg.endereco_loja or "").strip()
    if not endereco:
        geo_logger.warning("[GEO] loja sem endereço configurado")
        return False
    consultas = [
        _partes_unicas_endereco(endereco, "Brasil"),
        _partes_unicas_endereco(endereco, "Rio de Janeiro", "RJ", "Brasil"),
    ]
    lat, lon, detalhe = _geocodificar_consultas(consultas, identificador="loja")
    if lat is None or lon is None:
        geo_logger.error("[GEO] loja não localizada motivo=%s endereco=%r", detalhe, endereco)
        return False
    cfg.latitude_loja, cfg.longitude_loja = lat, lon
    db.commit()
    return True


def _solicitacoes_sem_coordenadas(candidatos):
    """Deduplica contratos sem coordenadas que participam da operação consultada."""
    pendentes = []
    vistos = set()
    for c in candidatos:
        ag = c.get("agenda")
        sol = ag.solicitacao if ag else None
        if not sol or sol.id in vistos:
            continue
        if sol.latitude is None or sol.longitude is None:
            vistos.add(sol.id)
            pendentes.append({
                "id": sol.id,
                "cliente": sol.cliente.nome if sol.cliente else c.get("titulo") or f"Contrato {sol.id}",
                "endereco": c.get("endereco") or sol.local or "Endereço não informado",
                "bairro": sol.bairro or "",
                "latitude": sol.latitude,
                "longitude": sol.longitude,
            })
    return pendentes


@app.get("/painel/solicitacao/{solicitacao_id}/iniciar-rota")
def iniciar_rota_solicitacao(
    solicitacao_id: int, provedor: str = "maps",
    db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)
):
    item = db.query(Solicitacao).filter_by(id=solicitacao_id, empresa_id=empresa.id).first()
    if not item:
        raise HTTPException(404)
    # A navegação operacional usa exatamente o mesmo texto exibido em ``Destino``
    # no card da Operação. Nesta versão diagnóstica não usamos latitude/longitude:
    # isso permite comparar o endereço enviado pelo Conect com o resultado escolhido
    # pelo Waze/Maps sem uma segunda fonte de destino escondida.
    destino = endereco_rota_solicitacao(item)
    if not destino:
        raise HTTPException(400, "Endereço não informado para iniciar a rota.")
    if provedor == "waze":
        url = "https://waze.com/ul?" + urlencode({"q": destino, "navigate": "yes"})
    else:
        url = "https://www.google.com/maps/search/?api=1&" + urlencode({"query": destino})
    return RedirectResponse(url, status_code=303)


@app.post("/painel/inteligencia-logistica/localizacoes/automaticas")
def calcular_localizacoes_automaticas(
    data_operacao: str = Form(""), equipe_id: int = Form(0),
    db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)
):
    try:
        data_filtro = datetime.strptime(data_operacao, "%Y-%m-%d").date() if data_operacao else date.today()
    except Exception:
        data_filtro = date.today()
    cfg = _config_rota(db, empresa.id)
    candidatos = _montar_candidatos_rota(db, empresa, data_filtro, equipe_id or None, cfg, time(7, 0))
    itens = []
    vistos = set()
    for c in candidatos:
        ag = c.get("agenda")
        sol = ag.solicitacao if ag else None
        if not sol or sol.id in vistos:
            continue
        vistos.add(sol.id)
        endereco = c.get("endereco") or sol.local or ""
        itens.append((sol, endereco, sol.bairro or c.get("bairro") or ""))
    if not itens:
        return RedirectResponse(
            f"/painel/inteligencia-logistica/nova?data={data_filtro.isoformat()}&equipe_id={equipe_id}&erro=" + quote("Nenhum endereço disponível para calcular"),
            status_code=303,
        )
    resultado = _geocodificar_solicitacoes_automaticamente(db, empresa.id, itens)
    msg = f"Localizações: {resultado['localizados']} calculadas, {resultado['ja_prontos']} já prontas"
    if resultado["pendentes"]:
        msg += f", {resultado['pendentes']} precisam de revisão"
    return RedirectResponse(
        f"/painel/inteligencia-logistica/nova?data={data_filtro.isoformat()}&equipe_id={equipe_id}&sucesso=" + quote(msg) + ("&abrir_geo=1" if resultado["pendentes"] else ""),
        status_code=303,
    )


@app.post("/painel/inteligencia-logistica/localizacao/{solicitacao_id}")
def salvar_localizacao_inteligencia(
    solicitacao_id: int, request: Request, latitude: str = Form(""), longitude: str = Form(""),
    retorno: str = Form("/painel/inteligencia-logistica/nova"),
    db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)
):
    sol = db.query(Solicitacao).filter_by(id=solicitacao_id, empresa_id=empresa.id).first()
    if not sol:
        raise HTTPException(status_code=404)
    try:
        lat = float(str(latitude).replace(",", ".").strip())
        lon = float(str(longitude).replace(",", ".").strip())
        if not (-90 <= lat <= 90 and -180 <= lon <= 180):
            raise ValueError
    except Exception:
        sep = "&" if "?" in retorno else "?"
        return RedirectResponse(f"{retorno}{sep}erro=Latitude ou longitude inválida", status_code=303)
    sol.latitude = lat
    sol.longitude = lon
    sol.status_geocodificacao = "localizado"
    sol.data_geocodificacao = agora_utc()
    db.commit()
    sep = "&" if "?" in retorno else "?"
    return RedirectResponse(f"{retorno}{sep}sucesso=Localização atualizada", status_code=303)


@app.post("/painel/inteligencia-logistica/localizacao/{solicitacao_id}/limpar")
def limpar_localizacao_inteligencia(
    solicitacao_id: int, retorno: str = Form("/painel/inteligencia-logistica/nova"),
    db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)
):
    sol = db.query(Solicitacao).filter_by(id=solicitacao_id, empresa_id=empresa.id).first()
    if not sol:
        raise HTTPException(status_code=404)
    sol.latitude = None
    sol.longitude = None
    sol.status_geocodificacao = "pendente"
    sol.data_geocodificacao = None
    db.commit()
    sep = "&" if "?" in retorno else "?"
    return RedirectResponse(f"{retorno}{sep}sucesso=Localização marcada como pendente", status_code=303)


@app.get("/painel/inteligencia-logistica", response_class=HTMLResponse)
def inteligencia_logistica(request: Request, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    cfg = _config_rota(db, empresa.id)
    rotas = db.query(RotaInteligente).filter_by(empresa_id=empresa.id).order_by(
        RotaInteligente.data_operacao.desc(), RotaInteligente.id.desc()
    ).limit(8).all()
    total_rotas = db.query(func.count(RotaInteligente.id)).filter_by(empresa_id=empresa.id).scalar() or 0
    totais = db.query(
        func.coalesce(func.sum(RotaInteligente.distancia_total_km), 0),
        func.coalesce(func.sum(RotaInteligente.duracao_total_min), 0),
        func.coalesce(func.sum(RotaInteligente.retornos_loja), 0),
    ).filter_by(empresa_id=empresa.id).first()
    riscos = db.query(func.count(RotaInteligenteParada.id)).join(RotaInteligente).filter(
        RotaInteligente.empresa_id == empresa.id,
        RotaInteligenteParada.risco.in_(["atencao", "atrasado"]),
        RotaInteligenteParada.status != "concluido",
    ).scalar() or 0
    configurada = bool(cfg.endereco_loja and cfg.latitude_loja is not None and cfg.longitude_loja is not None)
    return templates.TemplateResponse("admin/inteligencia_dashboard.html", {
        "request": request, "empresa": empresa, "rotas": rotas, "config": cfg,
        "total_rotas": total_rotas, "distancia_total": float(totais[0] or 0),
        "duracao_total": int(totais[1] or 0), "retornos_total": int(totais[2] or 0),
        "riscos_abertos": int(riscos), "configurada": configurada,
        "erro": request.query_params.get("erro"), "sucesso": request.query_params.get("sucesso")
    })


@app.get("/painel/inteligencia-logistica/nova", response_class=HTMLResponse)
def nova_inteligencia(request: Request, data: str = "", equipe_id: int = 0, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    garantir_agenda_reservas(db, empresa.id)
    try:
        data_filtro = datetime.strptime(data, "%Y-%m-%d").date() if data else date.today()
    except Exception:
        data_filtro = date.today()
    cfg = _config_rota(db, empresa.id)
    equipes = equipes_visiveis_usuario(request, db, empresa.id)
    veiculos = db.query(VeiculoLogistico).filter_by(empresa_id=empresa.id, ativo=True).order_by(VeiculoLogistico.nome).all()
    candidatos = _montar_candidatos_rota(db, empresa, data_filtro, equipe_id or None, cfg, time(7, 0))
    existente = db.query(RotaInteligente).filter(
        RotaInteligente.empresa_id == empresa.id,
        RotaInteligente.data_operacao == data_filtro,
        RotaInteligente.equipe_id == (equipe_id or None),
    ).order_by(RotaInteligente.id.desc()).first()
    localizacoes_pendentes = _solicitacoes_sem_coordenadas(candidatos)
    return templates.TemplateResponse("admin/inteligencia_nova.html", {
        "request": request, "empresa": empresa, "config": cfg, "equipes": equipes,
        "veiculos": veiculos, "data_filtro": data_filtro, "equipe_id": equipe_id,
        "candidatos": candidatos, "existente": existente,
        "localizacoes_pendentes": localizacoes_pendentes,
        "erro": request.query_params.get("erro"), "sucesso": request.query_params.get("sucesso")
    })


@app.get("/painel/inteligencia-logistica/configuracoes", response_class=HTMLResponse)
def configuracoes_inteligencia(request: Request, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    cfg = _config_rota(db, empresa.id)
    veiculos = db.query(VeiculoLogistico).filter_by(empresa_id=empresa.id, ativo=True).order_by(VeiculoLogistico.nome).all()
    produtos = db.query(ProdutoServico).filter_by(empresa_id=empresa.id, ativo=True).order_by(ProdutoServico.nome).all()
    return templates.TemplateResponse("admin/inteligencia_configuracoes.html", {
        "request": request, "empresa": empresa, "config": cfg, "veiculos": veiculos, "produtos": produtos,
        "erro": request.query_params.get("erro"), "sucesso": request.query_params.get("sucesso")
    })


@app.get("/painel/inteligencia-logistica/historico", response_class=HTMLResponse)
def historico_inteligencia(request: Request, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    rotas = db.query(RotaInteligente).filter_by(empresa_id=empresa.id).order_by(
        RotaInteligente.data_operacao.desc(), RotaInteligente.id.desc()
    ).all()
    return templates.TemplateResponse("admin/inteligencia_historico.html", {
        "request": request, "empresa": empresa, "rotas": rotas
    })


@app.post("/painel/inteligencia-logistica/configuracao")
def salvar_config_inteligencia(request: Request, endereco_loja: str = Form(""), latitude_loja: str = Form(""), longitude_loja: str = Form(""), minutos_montagem: int = Form(30), minutos_desmontagem: int = Form(20), antecedencia_entrega: int = Form(60), horario_minimo_cliente: str = Form("08:00"), raio_retirada_estrategica_km: str = Form("10"), desvio_max_retirada_estrategica_min: int = Form(60), minutos_parada_loja: int = Form(20), velocidade_media_kmh: str = Form("30"), custo_km: str = Form("0"), custo_hora_equipe: str = Form("0"), db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    cfg = _config_rota(db, empresa.id)
    endereco_anterior = (cfg.endereco_loja or "").strip()
    novo_endereco = endereco_loja.strip()
    cfg.endereco_loja = novo_endereco or None
    # Coordenadas da loja são técnicas e calculadas automaticamente pelo endereço.
    if endereco_anterior != novo_endereco:
        cfg.latitude_loja = None
        cfg.longitude_loja = None
    cfg.minutos_montagem = max(0, minutos_montagem); cfg.minutos_desmontagem = max(0, minutos_desmontagem)
    cfg.antecedencia_entrega = max(0, antecedencia_entrega); cfg.minutos_parada_loja = max(0, minutos_parada_loja)
    try:
        cfg.horario_minimo_cliente = datetime.strptime((horario_minimo_cliente or "08:00").strip(), "%H:%M").time()
    except Exception:
        cfg.horario_minimo_cliente = time(8, 0)
    try:
        cfg.raio_retirada_estrategica_km = max(0, float((raio_retirada_estrategica_km or "10").replace(",", ".")))
    except Exception:
        cfg.raio_retirada_estrategica_km = 10
    cfg.desvio_max_retirada_estrategica_min = max(0, int(desvio_max_retirada_estrategica_min or 0))
    try: cfg.velocidade_media_kmh = max(5, float(velocidade_media_kmh.replace(",", ".")))
    except Exception: cfg.velocidade_media_kmh = 30
    try: cfg.custo_km = max(0, float(custo_km.replace(",", ".")))
    except Exception: cfg.custo_km = 0
    try: cfg.custo_hora_equipe = max(0, float(custo_hora_equipe.replace(",", ".")))
    except Exception: cfg.custo_hora_equipe = 0
    db.commit()
    _garantir_localizacao_loja(db, cfg)
    return RedirectResponse("/painel/inteligencia-logistica/configuracoes?sucesso=Configuração salva", status_code=303)


@app.post("/painel/inteligencia-logistica/veiculos")
def criar_veiculo_logistico(nome: str = Form(...), capacidade_interno: int = Form(4), capacidade_mala: int = Form(1),
                            capacidade_teto: int = Form(3), db: Session = Depends(get_db),
                            empresa: Empresa = Depends(empresa_logada)):
    nome = nome.strip()
    if nome and not db.query(VeiculoLogistico).filter_by(empresa_id=empresa.id, nome=nome).first():
        db.add(VeiculoLogistico(empresa_id=empresa.id, nome=nome,
            capacidade_interno=max(0, capacidade_interno), capacidade_mala=max(0, capacidade_mala),
            capacidade_teto=max(0, capacidade_teto)))
        db.commit()
    return RedirectResponse("/painel/inteligencia-logistica/configuracoes?sucesso=Veículo adicionado", status_code=303)


def _calcular_horario_saida_ideal(ordenados, data_operacao, cfg, origem_lat=None, origem_lon=None):
    """Calcula o horário mais tarde possível para sair da loja sem violar prazos."""
    acumulado = 0
    lat_atual, lon_atual = origem_lat, origem_lon
    bairro_atual = ""
    velocidade = max(5.0, float(cfg.velocidade_media_kmh or 30))
    limites = []
    for c in ordenados:
        distancia, desloc, _ = _trecho_rodoviario(lat_atual, lon_atual, c.get("lat"), c.get("lon"), cfg)
        if desloc is None:
            desloc = _deslocamento_estimado_sem_coordenadas(bairro_atual, c.get("bairro"))
        acumulado += desloc
        if c.get("limite") and (c.get("tipo") == "entrega" or c.get("retirada_obrigatoria")):
            limite_dt = datetime.combine(data_operacao, c["limite"])
            limites.append(limite_dt - timedelta(minutes=acumulado))
        acumulado += max(0, int(c.get("servico") or 0))
        if c.get("lat") is not None and c.get("lon") is not None:
            lat_atual, lon_atual = c["lat"], c["lon"]
        bairro_atual = c.get("bairro") or bairro_atual
    if not limites:
        return time(8, 0)
    saida = min(limites)
    # margem operacional de 10 minutos e arredondamento para 5 minutos
    saida -= timedelta(minutes=10)
    minutos = max(0, saida.hour * 60 + saida.minute)
    minutos = (minutos // 5) * 5
    return time(minutos // 60, minutos % 60)



@app.post("/painel/inteligencia-logistica/veiculos/{veiculo_id}/editar")
def editar_veiculo_logistico(veiculo_id: int, nome: str = Form(...), capacidade_interno: int = Form(0),
                              capacidade_mala: int = Form(0), capacidade_teto: int = Form(0),
                              db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    veiculo = db.query(VeiculoLogistico).filter_by(id=veiculo_id, empresa_id=empresa.id, ativo=True).first()
    if not veiculo:
        raise HTTPException(status_code=404)
    nome = (nome or "").strip()
    if not nome:
        return RedirectResponse("/painel/inteligencia-logistica/configuracoes?erro=Informe o nome do veículo", status_code=303)
    duplicado = db.query(VeiculoLogistico).filter(
        VeiculoLogistico.empresa_id == empresa.id,
        VeiculoLogistico.nome == nome,
        VeiculoLogistico.id != veiculo.id,
    ).first()
    if duplicado:
        return RedirectResponse("/painel/inteligencia-logistica/configuracoes?erro=Já existe outro veículo com esse nome", status_code=303)
    veiculo.nome = nome
    veiculo.capacidade_interno = max(0, int(capacidade_interno or 0))
    veiculo.capacidade_mala = max(0, int(capacidade_mala or 0))
    veiculo.capacidade_teto = max(0, int(capacidade_teto or 0))
    db.commit()
    return RedirectResponse("/painel/inteligencia-logistica/configuracoes?sucesso=Veículo atualizado", status_code=303)


@app.post("/painel/inteligencia-logistica/veiculos/{veiculo_id}/perfil/{perfil_id}/editar")
def editar_perfil_carga_veiculo(veiculo_id: int, perfil_id: int, volumes: int = Form(1),
                                 permite_interno: bool = Form(False), permite_mala: bool = Form(False),
                                 permite_teto: bool = Form(False), db: Session = Depends(get_db),
                                 empresa: Empresa = Depends(empresa_logada)):
    veiculo = db.query(VeiculoLogistico).filter_by(id=veiculo_id, empresa_id=empresa.id, ativo=True).first()
    perfil = db.query(VeiculoPerfilCarga).filter_by(id=perfil_id, veiculo_id=veiculo_id).first()
    if not veiculo or not perfil:
        raise HTTPException(status_code=404)
    if not (permite_interno or permite_mala or permite_teto):
        return RedirectResponse("/painel/inteligencia-logistica/configuracoes?erro=Selecione ao menos um compartimento", status_code=303)
    perfil.volumes = max(1, int(volumes or 1))
    perfil.permite_interno = bool(permite_interno)
    perfil.permite_mala = bool(permite_mala)
    perfil.permite_teto = bool(permite_teto)
    perfil.ativo = True
    db.commit()
    return RedirectResponse("/painel/inteligencia-logistica/configuracoes?sucesso=Equipamento atualizado", status_code=303)


@app.post("/painel/inteligencia-logistica/veiculos/{veiculo_id}/perfil/{perfil_id}/remover")
def remover_perfil_carga_veiculo(veiculo_id: int, perfil_id: int, db: Session = Depends(get_db),
                                  empresa: Empresa = Depends(empresa_logada)):
    veiculo = db.query(VeiculoLogistico).filter_by(id=veiculo_id, empresa_id=empresa.id, ativo=True).first()
    perfil = db.query(VeiculoPerfilCarga).filter_by(id=perfil_id, veiculo_id=veiculo_id).first()
    if not veiculo or not perfil:
        raise HTTPException(status_code=404)
    perfil.ativo = False
    db.commit()
    return RedirectResponse("/painel/inteligencia-logistica/configuracoes?sucesso=Equipamento removido do veículo", status_code=303)


@app.post("/painel/inteligencia-logistica/veiculos/{veiculo_id}/perfil")
def salvar_perfil_carga_veiculo(veiculo_id: int, produto_id: int = Form(...), volumes: int = Form(1),
                                permite_interno: bool = Form(False), permite_mala: bool = Form(False),
                                permite_teto: bool = Form(False), db: Session = Depends(get_db),
                                empresa: Empresa = Depends(empresa_logada)):
    veiculo = db.query(VeiculoLogistico).filter_by(id=veiculo_id, empresa_id=empresa.id).first()
    produto = db.query(ProdutoServico).filter_by(id=produto_id, empresa_id=empresa.id).first()
    if not veiculo or not produto:
        raise HTTPException(status_code=404)
    if not (permite_interno or permite_mala or permite_teto):
        return RedirectResponse("/painel/inteligencia-logistica/configuracoes?erro=Selecione ao menos um compartimento", status_code=303)
    perfil = db.query(VeiculoPerfilCarga).filter_by(veiculo_id=veiculo.id, produto_id=produto.id).first()
    if not perfil:
        perfil = VeiculoPerfilCarga(veiculo_id=veiculo.id, produto_id=produto.id)
        db.add(perfil)
    perfil.volumes = max(1, int(volumes or 1)); perfil.permite_interno = bool(permite_interno)
    perfil.permite_mala = bool(permite_mala); perfil.permite_teto = bool(permite_teto); perfil.ativo = True
    db.commit()
    return RedirectResponse("/painel/inteligencia-logistica/configuracoes?sucesso=Perfil de carga salvo", status_code=303)

@app.post("/painel/inteligencia-logistica/gerar")
def gerar_rota_inteligente(request: Request, data_operacao: str = Form(...), equipe_id: int = Form(0), veiculo_id: int = Form(0), db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    garantir_agenda_reservas(db, empresa.id)
    try:
        data_op = datetime.strptime(data_operacao, "%Y-%m-%d").date()
    except Exception:
        return RedirectResponse("/painel/inteligencia-logistica/nova?erro=Data inválida", status_code=303)
    veiculo = db.query(VeiculoLogistico).filter_by(id=veiculo_id, empresa_id=empresa.id, ativo=True).first() if veiculo_id else None
    if not veiculo:
        return RedirectResponse("/painel/inteligencia-logistica/nova?erro=Selecione um veículo configurado", status_code=303)
    hora_provisoria = time(8, 0)
    chave = f"{empresa.id}:{data_op.isoformat()}:{equipe_id or 0}:{veiculo_id or 0}"
    existente = db.query(RotaInteligente).filter_by(empresa_id=empresa.id, chave_consumo=chave).first()
    if existente:
        _reconstruir_rota_salva(db, existente)
        db.commit()
        return RedirectResponse(
            f"/painel/inteligencia-logistica/rota/{existente.id}?sucesso=Rota atualizada com os dados atuais sem novo consumo",
            status_code=303
        )
    if int(empresa.humiat_saldo or 0) < 1:
        return RedirectResponse("/painel/inteligencia-logistica/nova?erro=Saldo Humiat insuficiente para gerar a rota", status_code=303)
    cfg = _config_rota(db, empresa.id)
    candidatos = _montar_candidatos_rota(db, empresa, data_op, equipe_id or None, cfg, hora_provisoria)
    if not candidatos:
        return RedirectResponse("/painel/inteligencia-logistica/nova?erro=Nenhuma entrega ou retirada aprovada disponível para essa data/equipe", status_code=303)

    # Primeiro clique em Calcular: localiza silenciosamente e salva para todo o sistema.
    itens_geo, vistos_geo = [], set()
    for candidato in candidatos:
        ag = candidato.get("agenda")
        sol = ag.solicitacao if ag else None
        if sol and sol.id not in vistos_geo and (sol.latitude is None or sol.longitude is None):
            vistos_geo.add(sol.id)
            itens_geo.append((sol, endereco_rota_solicitacao(sol), sol.bairro or ""))
    if itens_geo:
        _geocodificar_solicitacoes_automaticamente(db, empresa.id, itens_geo)
    if not _garantir_localizacao_loja(db, cfg):
        return RedirectResponse("/painel/inteligencia-logistica/nova?erro=" + quote("Não foi possível localizar o endereço da loja. Confira o endereço nas configurações da Inteligência."), status_code=303)

    candidatos = _montar_candidatos_rota(db, empresa, data_op, equipe_id or None, cfg, hora_provisoria)
    pendentes_geo = _solicitacoes_sem_coordenadas(candidatos)
    if pendentes_geo:
        nomes = ", ".join(p["cliente"] for p in pendentes_geo[:3])
        complemento = f" e mais {len(pendentes_geo)-3}" if len(pendentes_geo) > 3 else ""
        return RedirectResponse("/painel/inteligencia-logistica/nova?erro=" + quote(f"Não foi possível localizar automaticamente: {nomes}{complemento}. Confira rua, número, bairro e cidade no contrato. O contrato permanece salvo."), status_code=303)

    ordenados = _ordenar_inteligente(candidatos, data_op, hora_provisoria, cfg, cfg.latitude_loja, cfg.longitude_loja, veiculo)
    try:
        ordenados = _ajustar_ordem_inicial_por_capacidade(ordenados, veiculo, cfg)
        ordenados = _inserir_retornos_por_compartimento(ordenados, veiculo, cfg)
        ordenados = _garantir_retorno_final_loja(ordenados, cfg)
    except ValueError as exc:
        return RedirectResponse("/painel/inteligencia-logistica/nova?erro=" + quote(str(exc)), status_code=303)
    hora = _calcular_horario_saida_ideal(ordenados, data_op, cfg, cfg.latitude_loja, cfg.longitude_loja)
    # IMPORTANTE: não reordena depois de calcular a saída. O horário foi derivado
    # desta sequência; reotimizar usando esse mesmo horário criava um ciclo em que
    # a retirada anterior parecia atrasar a entrega, embora bastasse sair antes.
    # Aqui recalculamos somente a linha do tempo da ordem já escolhida.
    atrasos = _simular_sequencia_rota(ordenados, data_op, hora, cfg, cfg.latitude_loja, cfg.longitude_loja)
    if atrasos:
        resumo = ", ".join(f"{a['titulo']} ({a['minutos']} min)" for a in atrasos[:3])
        return RedirectResponse(
            "/painel/inteligencia-logistica/nova?erro=" + quote(
                f"Rota inválida: há entrega(s) após o limite: {resumo}. Antecipe a saída, redistribua a equipe/veículo ou faça entrega no dia anterior. Nenhum Humiat foi consumido."
            ),
            status_code=303
        )
    codigo = f"IL-{data_op.strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
    rota = RotaInteligente(empresa_id=empresa.id, equipe_id=equipe_id or None, veiculo_id=veiculo_id or None,
        codigo=codigo, chave_consumo=chave, data_operacao=data_op, horario_saida=hora, humiat_consumido=True,
        custo_humiat=1, criado_por=request.session.get("usuario_nome") or "Usuário")
    db.add(rota); db.flush()
    for ordem, c in enumerate(ordenados, 1):
        ag = c.get("agenda")
        db.add(RotaInteligenteParada(rota_id=rota.id, agenda_id=c.get("agenda_id"), solicitacao_id=c.get("solicitacao_id") or (ag.solicitacao_id if ag else None),
            ordem=ordem, tipo=c["tipo"], titulo=c["titulo"], endereco=c["endereco"], latitude=c.get("lat"), longitude=c.get("lon"),
            horario_limite=c.get("limite"), chegada_prevista=c.get("chegada"), saida_prevista=c.get("saida"),
            distancia_anterior_km=c.get("distancia", 0), deslocamento_anterior_min=c.get("desloc", 0), servico_min=c["servico"],
            risco=c.get("risco", "normal"), motivo_prioridade=c.get("motivo"), retorno_loja=c["tipo"] == "loja",
            carga_movimento=int(c.get("carga_movimento", 0)), carga_apos_parada=int(c.get("carga_apos", 0))))
    _registrar_movimento_humiat(db, empresa, -1, "consumo_rota_inteligente", f"Rota Inteligente {codigo}",
        observacao=f"Data {data_op.strftime('%d/%m/%Y')} | Equipe {equipe_id or 'não definida'}", usuario=rota.criado_por)
    db.flush(); _recalcular_rota_salva(db, rota); db.commit()
    return RedirectResponse(f"/painel/inteligencia-logistica/rota/{rota.id}?sucesso=Rota gerada com 1 Humiat", status_code=303)



def _nomes_produtos_solicitacao(sol: Solicitacao | None) -> list[str]:
    if not sol:
        return []
    nomes = []
    for item in sol.itens or []:
        nome = item.produto.nome if item.produto else "Equipamento"
        qtd = max(1, int(item.quantidade or 1))
        nomes.append(f"{qtd}x {nome}" if qtd > 1 else nome)
    if not nomes and sol.produto:
        nomes.append(sol.produto.nome)
    return nomes


def _resumo_rotas_do_dia(paradas: list[RotaInteligenteParada]):
    """Divide o plano em saídas da loja e resume somente o que a operação precisa ver.

    Uma rota do dia é o trecho entre a saída da loja e o próximo retorno à loja.
    O resumo identifica o que sai da loja, o que é retirado e reaproveitado em uma
    entrega posterior e o que deve ser deixado ao retornar.
    """
    grupos, atual = [], []
    for parada in paradas:
        atual.append(parada)
        if parada.tipo == "loja":
            grupos.append(atual)
            atual = []
    if atual:
        grupos.append(atual)

    resumos = []
    for numero, grupo in enumerate(grupos, 1):
        disponiveis_retirados: dict[int, int] = {}
        sair_com, reaproveitar, deixar = [], [], []
        retiradas_no_trecho = []

        for parada in grupo:
            if parada.tipo == "loja":
                continue
            sol = parada.solicitacao
            produtos = _produtos_quantidades(sol)
            nomes = _nomes_produtos_solicitacao(sol)
            cliente = sol.cliente.nome if sol and sol.cliente else parada.titulo
            if parada.tipo == "retirada":
                retiradas_no_trecho.append((parada, produtos, nomes, cliente))
                for pid, qtd in produtos.items():
                    disponiveis_retirados[pid] = disponiveis_retirados.get(pid, 0) + qtd
                continue

            usados_retirada = False
            for pid, qtd in produtos.items():
                usar = min(qtd, disponiveis_retirados.get(pid, 0))
                if usar:
                    usados_retirada = True
                    disponiveis_retirados[pid] -= usar
            texto = f"{', '.join(nomes) or 'Equipamento'} → {cliente}"
            if usados_retirada:
                reaproveitar.append(texto)
            else:
                sair_com.append(texto)

        # Tudo que foi retirado e não reaproveitado antes do retorno fica para deixar na loja.
        saldo = dict(disponiveis_retirados)
        for parada, produtos, nomes, cliente in retiradas_no_trecho:
            restante = sum(min(qtd, saldo.get(pid, 0)) for pid, qtd in produtos.items())
            if restante > 0:
                deixar.append(f"{', '.join(nomes) or 'Equipamento'} ← {cliente}")
                for pid, qtd in produtos.items():
                    saldo[pid] = max(0, saldo.get(pid, 0) - qtd)

        operacoes = [p for p in grupo if p.tipo != "loja"]
        retorno = next((p for p in reversed(grupo) if p.tipo == "loja"), None)
        resumos.append({
            "numero": numero,
            "paradas": grupo,
            "operacoes": operacoes,
            "sair_com": sair_com,
            "reaproveitar": reaproveitar,
            "deixar": deixar,
            "retorno": retorno,
            "inicio": operacoes[0].chegada_prevista if operacoes else None,
            "fim": (retorno.chegada_prevista if retorno else (operacoes[-1].saida_prevista if operacoes else None)),
            "concluida": bool(grupo and all(p.status == "concluido" for p in grupo)),
            "em_execucao": any(p.status == "em_andamento" for p in grupo),
        })
    return resumos


@app.get("/painel/inteligencia-logistica/rota/{rota_id}", response_class=HTMLResponse)
def visualizar_rota_inteligente(rota_id: int, request: Request, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    rota = db.query(RotaInteligente).filter_by(id=rota_id, empresa_id=empresa.id).first()
    if not rota: raise HTTPException(status_code=404)
    # Paradas concluídas permanecem no histórico do banco, mas somem da operação ativa.
    paradas = (
        db.query(RotaInteligenteParada)
        .filter(
            RotaInteligenteParada.rota_id == rota.id,
            RotaInteligenteParada.status != "concluido",
        )
        .order_by(RotaInteligenteParada.ordem)
        .all()
    )

    em_execucao = rota.status == "em_execucao" or any(p.status == "em_andamento" for p in paradas)
    nomes_dias = ("Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Sábado", "Domingo")

    for p in paradas:
        sol = p.solicitacao
        p.hora_inicio_evento_view = sol.hora_inicio if sol else None
        p.hora_fim_evento_view = sol.hora_fim if sol else None
        p.retirada_hora_view = sol.retirada_hora if sol else None
        p.data_operacao_view = p.agenda.data if p.agenda and p.agenda.data else rota.data_operacao
        p.dia_semana_view = nomes_dias[p.data_operacao_view.weekday()]
        p.contrato_url_view = f"/painel/solicitacao/{p.solicitacao_id}" if p.solicitacao_id else None

        # Uma retirada sintética pode participar do planejamento semanal antes da
        # entrega, mas não pode ser iniciada nem concluída.
        p.aguardando_entrega_view = False
        if p.tipo == "retirada" and p.solicitacao_id:
            entrega = (
                db.query(Agenda)
                .filter_by(
                    empresa_id=empresa.id,
                    solicitacao_id=p.solicitacao_id,
                    tipo_evento="entrega",
                )
                .first()
            )
            p.aguardando_entrega_view = bool(
                entrega and entrega.status_operacional != "concluido"
            )

        p.calculo_view = [trecho.strip() for trecho in str(p.motivo_prioridade or "").split(" • ") if trecho.strip()]
        p.folga_min_view = None
        # Atraso só existe durante a execução. No planejamento mostramos apenas
        # o horário recomendado.
        if em_execucao and p.chegada_prevista and p.horario_limite:
            limite_dt = datetime.combine(rota.data_operacao, p.horario_limite)
            p.folga_min_view = int((limite_dt - p.chegada_prevista).total_seconds() / 60)

    inicio_semana = rota.data_operacao - timedelta(days=rota.data_operacao.weekday())
    fim_semana = inicio_semana + timedelta(days=6)
    rotas_semana = (
        db.query(RotaInteligente)
        .filter(
            RotaInteligente.empresa_id == empresa.id,
            RotaInteligente.data_operacao >= inicio_semana,
            RotaInteligente.data_operacao <= fim_semana,
            RotaInteligente.status != "concluida",
        )
        .order_by(RotaInteligente.data_operacao, RotaInteligente.horario_saida)
        .all()
    )

    _anotar_ocupacao_rota_salva(paradas, rota.veiculo)
    resumos_rotas = _resumo_rotas_do_dia(paradas)
    cfg_inteligencia = _config_rota(db, empresa.id)
    return templates.TemplateResponse("admin/inteligencia_rota.html", {
        "request": request, "empresa": empresa, "rota": rota, "paradas": paradas,
        "resumos_rotas": resumos_rotas, "em_execucao": em_execucao,
        "inicio_semana": inicio_semana, "fim_semana": fim_semana,
        "rotas_semana": rotas_semana, "cfg_inteligencia": cfg_inteligencia,
        "sucesso": request.query_params.get("sucesso"), "erro": request.query_params.get("erro")
    })



@app.post("/painel/inteligencia-logistica/rota/{rota_id}/aplicar-operacao")
def aplicar_rota_inteligente_na_operacao(
    rota_id: int,
    request: Request,
    db: Session = Depends(get_db),
    empresa: Empresa = Depends(empresa_logada),
):
    """Copia somente data e hora planejadas pela Inteligência para a Operação.

    Não cria cards, não altera equipe, status operacional ou qualquer outra
    regra da Operação. O operador continua livre para editar a roteirização.
    """
    rota = db.query(RotaInteligente).filter_by(id=rota_id, empresa_id=empresa.id).first()
    if not rota:
        raise HTTPException(status_code=404)

    paradas = (
        db.query(RotaInteligenteParada)
        .filter(
            RotaInteligenteParada.rota_id == rota.id,
            RotaInteligenteParada.tipo.in_(("entrega", "retirada")),
            RotaInteligenteParada.status != "concluido",
        )
        .order_by(RotaInteligenteParada.ordem)
        .all()
    )

    atualizadas = 0
    ignoradas = 0
    usuario = request.session.get("usuario_nome") or request.session.get("usuario") or "Usuário"
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")

    for parada in paradas:
        if not parada.chegada_prevista:
            ignoradas += 1
            continue

        agenda = None
        if parada.agenda_id:
            agenda = db.query(Agenda).filter_by(
                id=parada.agenda_id,
                empresa_id=empresa.id,
            ).first()

        # Retiradas previstas pela Inteligência podem ter sido criadas de forma
        # sintética. Nesse caso, usa o card operacional já existente do contrato.
        if not agenda and parada.solicitacao_id:
            agenda = db.query(Agenda).filter_by(
                empresa_id=empresa.id,
                solicitacao_id=parada.solicitacao_id,
                tipo_evento=parada.tipo,
            ).first()

        if not agenda or agenda.status_operacional == "concluido":
            ignoradas += 1
            continue

        data_anterior = agenda.data
        hora_anterior = agenda.hora_inicio
        nova_data = parada.chegada_prevista.date()
        nova_hora = parada.chegada_prevista.time().replace(second=0, microsecond=0)

        agenda.data = nova_data
        agenda.hora_inicio = nova_hora
        agenda.previsao_entrega = nova_hora.strftime("%H:%M")

        # Aplicar a Inteligência deve copiar somente data e hora. Não altera
        # roteirização, equipe ou status do card. A versão anterior marcava
        # roteirizado=True sem definir equipe, fazendo entregas desaparecerem
        # quando a Operação estava filtrada por equipe.
        observacoes_anteriores = agenda.observacoes_operacionais or ""
        aplicado_sem_equipe = (
            agenda.roteirizado
            and not agenda.equipe_id
            and "Data e hora aplicadas pela Inteligência" in observacoes_anteriores
        )
        if aplicado_sem_equipe:
            # Repara automaticamente os cards afetados pela versão anterior.
            agenda.roteirizado = False

        registro = (
            f"[{agora}] Data e hora aplicadas pela Inteligência por {usuario}. "
            f"{data_anterior.strftime('%d/%m/%Y') if data_anterior else '-'} "
            f"{hora_anterior.strftime('%H:%M') if hora_anterior else '-'} → "
            f"{nova_data.strftime('%d/%m/%Y')} {nova_hora.strftime('%H:%M')}."
        )
        agenda.observacoes_operacionais = (
            (agenda.observacoes_operacionais or "") + "\n" + registro
        ).strip()
        atualizadas += 1

    db.commit()

    if atualizadas == 0:
        mensagem = "Nenhuma entrega ou retirada disponível para aplicar na Operação"
        return RedirectResponse(
            f"/painel/inteligencia-logistica/rota/{rota.id}?erro={quote(mensagem)}",
            status_code=303,
        )

    mensagem = f"{atualizadas} operação(ões) atualizada(s) somente com data e hora"
    if ignoradas:
        mensagem += f"; {ignoradas} item(ns) sem vínculo ou já concluído(s) foram ignorados"
    return RedirectResponse(
        f"/painel/inteligencia-logistica/rota/{rota.id}?sucesso={quote(mensagem)}",
        status_code=303,
    )


@app.post("/painel/inteligencia-logistica/rota/{rota_id}/horario-saida")
def alterar_horario_saida_rota(
    rota_id: int, horario_saida: str = Form(...), db: Session = Depends(get_db),
    empresa: Empresa = Depends(empresa_logada)
):
    rota = db.query(RotaInteligente).filter_by(id=rota_id, empresa_id=empresa.id).first()
    if not rota:
        raise HTTPException(status_code=404)
    try:
        nova_hora = datetime.strptime(horario_saida, "%H:%M").time()
    except Exception:
        return RedirectResponse(f"/painel/inteligencia-logistica/rota/{rota_id}?erro=Horário inválido", status_code=303)
    rota.horario_saida = nova_hora
    _recalcular_rota_salva(db, rota)
    db.commit()
    return RedirectResponse(
        f"/painel/inteligencia-logistica/rota/{rota.id}?sucesso=Horário alterado e previsões recalculadas",
        status_code=303,
    )


@app.post("/painel/inteligencia-logistica/rota/{rota_id}/iniciar/{parada_id}")
def iniciar_parada_inteligente(
    rota_id: int, parada_id: int, db: Session = Depends(get_db),
    empresa: Empresa = Depends(empresa_logada)
):
    rota = db.query(RotaInteligente).filter_by(id=rota_id, empresa_id=empresa.id).first()
    parada = db.query(RotaInteligenteParada).filter_by(id=parada_id, rota_id=rota_id).first()
    if not rota or not parada:
        raise HTTPException(status_code=404)
    if parada.status == "concluido":
        return RedirectResponse(f"/painel/inteligencia-logistica/rota/{rota.id}?erro=Esta parada já foi concluída", status_code=303)

    if parada.tipo == "retirada" and parada.solicitacao_id:
        entrega = db.query(Agenda).filter_by(
            empresa_id=empresa.id,
            solicitacao_id=parada.solicitacao_id,
            tipo_evento="entrega",
        ).first()
        if entrega and entrega.status_operacional != "concluido":
            return RedirectResponse(
                f"/painel/inteligencia-logistica/rota/{rota.id}?erro={quote('Esta retirada ainda aguarda a conclusão da entrega.')}",
                status_code=303,
            )

    # O operador pode iniciar qualquer pedido. Ele vira a próxima parada, sem apagar
    # o plano anterior nem as conclusões já registradas.
    pendentes = db.query(RotaInteligenteParada).filter(
        RotaInteligenteParada.rota_id == rota.id,
        RotaInteligenteParada.status != "concluido",
    ).order_by(RotaInteligenteParada.ordem).all()
    concluidas = db.query(RotaInteligenteParada).filter(
        RotaInteligenteParada.rota_id == rota.id,
        RotaInteligenteParada.status == "concluido",
    ).order_by(RotaInteligenteParada.ordem).all()
    for p in pendentes:
        if p.id != parada.id and p.status == "em_andamento":
            p.status = "pendente"
    parada.status = "em_andamento"
    nova_ordem = concluidas + [parada] + [p for p in pendentes if p.id != parada.id]
    for ordem, p in enumerate(nova_ordem, 1):
        p.ordem = ordem

    # Ao começar o trabalho, o relógio real passa a comandar as próximas previsões.
    agora = agora_utc()
    if rota.data_operacao == agora.date() and not concluidas:
        rota.horario_saida = agora.time().replace(second=0, microsecond=0)
    rota.status = "em_execucao"
    _recalcular_rota_salva(db, rota)
    db.commit()
    return RedirectResponse(
        f"/painel/inteligencia-logistica/rota/{rota.id}?sucesso={quote('Parada iniciada. As próximas previsões foram recalculadas.')}",
        status_code=303,
    )


@app.post("/painel/inteligencia-logistica/rota/{rota_id}/mover/{parada_id}")
def mover_parada_inteligente(
    rota_id: int,
    parada_id: int,
    direcao: str = Form(...),
    db: Session = Depends(get_db),
    empresa: Empresa = Depends(empresa_logada),
):
    """Move um card manualmente e recalcula somente distâncias e horários.

    A sequência escolhida pelo operador é preservada; não executa novamente o
    otimizador e não consome Humiat.
    """
    rota = db.query(RotaInteligente).filter_by(id=rota_id, empresa_id=empresa.id).first()
    parada = db.query(RotaInteligenteParada).filter_by(id=parada_id, rota_id=rota_id).first()
    if not rota or not parada:
        raise HTTPException(status_code=404)
    if direcao not in ("subir", "descer"):
        raise HTTPException(status_code=400, detail="Direção inválida")
    if parada.status in ("concluido", "em_andamento"):
        return RedirectResponse(
            f"/painel/inteligencia-logistica/rota/{rota.id}?erro={quote('Uma parada iniciada ou concluída não pode ser movida.')}",
            status_code=303,
        )
    if parada.fixada:
        return RedirectResponse(
            f"/painel/inteligencia-logistica/rota/{rota.id}?erro={quote('Destrave este card antes de alterar sua posição.')}",
            status_code=303,
        )

    paradas = db.query(RotaInteligenteParada).filter_by(rota_id=rota.id).order_by(
        RotaInteligenteParada.ordem.asc(), RotaInteligenteParada.id.asc()
    ).all()
    indice = next((i for i, item in enumerate(paradas) if item.id == parada.id), None)
    destino = (indice - 1) if direcao == "subir" else (indice + 1)
    if indice is None or destino < 0 or destino >= len(paradas):
        return RedirectResponse(f"/painel/inteligencia-logistica/rota/{rota.id}", status_code=303)

    vizinha = paradas[destino]
    if vizinha.status in ("concluido", "em_andamento") or vizinha.fixada or vizinha.tipo == "loja":
        return RedirectResponse(
            f"/painel/inteligencia-logistica/rota/{rota.id}?erro={quote('Não é possível atravessar uma parada iniciada, concluída, travada ou um retorno à loja.')}",
            status_code=303,
        )

    parada.ordem, vizinha.ordem = vizinha.ordem, parada.ordem
    db.flush()
    _recalcular_rota_salva(db, rota)
    db.commit()
    return RedirectResponse(
        f"/painel/inteligencia-logistica/rota/{rota.id}?sucesso={quote('Ordem alterada. Somente os horários e deslocamentos foram recalculados.')}",
        status_code=303,
    )


@app.post("/painel/inteligencia-logistica/rota/{rota_id}/fixar/{parada_id}")
def fixar_parada_inteligente(
    rota_id: int,
    parada_id: int,
    db: Session = Depends(get_db),
    empresa: Empresa = Depends(empresa_logada),
):
    """Trava/destrava a posição escolhida pelo operador sem reotimizar a rota."""
    rota = db.query(RotaInteligente).filter_by(id=rota_id, empresa_id=empresa.id).first()
    parada = db.query(RotaInteligenteParada).filter_by(id=parada_id, rota_id=rota_id).first()
    if not rota or not parada:
        raise HTTPException(status_code=404)
    if parada.tipo == "loja" or parada.status in ("concluido", "em_andamento"):
        return RedirectResponse(
            f"/painel/inteligencia-logistica/rota/{rota.id}?erro={quote('Este card não pode ser travado ou destravado agora.')}",
            status_code=303,
        )
    parada.fixada = not bool(parada.fixada)
    _recalcular_rota_salva(db, rota)
    db.commit()
    mensagem = "Card travado. A Inteligência não poderá mudar esta posição." if parada.fixada else "Card destravado. Ele pode ser movimentado novamente."
    return RedirectResponse(
        f"/painel/inteligencia-logistica/rota/{rota.id}?sucesso={quote(mensagem)}",
        status_code=303,
    )


@app.post("/painel/inteligencia-logistica/rota/{rota_id}/recalcular")
def recalcular_rota_inteligente(rota_id: int, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    rota = db.query(RotaInteligente).filter_by(id=rota_id, empresa_id=empresa.id).first()
    if not rota: raise HTTPException(status_code=404)
    quantidade = _reconstruir_rota_salva(db, rota)
    db.commit()
    return RedirectResponse(
        f"/painel/inteligencia-logistica/rota/{rota.id}?sucesso=Rota reconstruída com {quantidade} operação(ões) atual(is), sem novo consumo",
        status_code=303
    )


@app.post("/painel/inteligencia-logistica/rota/{rota_id}/carro-cheio/{parada_id}")
def carro_cheio_rota(rota_id: int, parada_id: int, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    rota = db.query(RotaInteligente).filter_by(id=rota_id, empresa_id=empresa.id).first()
    parada = db.query(RotaInteligenteParada).filter_by(id=parada_id, rota_id=rota_id).first()
    if not rota or not parada: raise HTTPException(status_code=404)
    cfg = _config_rota(db, empresa.id)
    posteriores = db.query(RotaInteligenteParada).filter(RotaInteligenteParada.rota_id == rota.id, RotaInteligenteParada.ordem > parada.ordem).all()
    for p in posteriores: p.ordem += 1
    db.add(RotaInteligenteParada(rota_id=rota.id, ordem=parada.ordem + 1, tipo="loja", titulo="Retorno à loja — carro cheio",
        endereco=cfg.endereco_loja or "Loja", latitude=cfg.latitude_loja, longitude=cfg.longitude_loja,
        servico_min=max(0, int(cfg.minutos_parada_loja or 20)), retorno_loja=True, motivo_prioridade="Retorno inserido manualmente porque o veículo atingiu a capacidade"))
    db.flush(); _recalcular_rota_salva(db, rota); db.commit()
    return RedirectResponse(f"/painel/inteligencia-logistica/rota/{rota.id}?sucesso=Retorno à loja inserido e rota recalculada", status_code=303)


@app.post("/painel/inteligencia-logistica/rota/{rota_id}/concluir/{parada_id}")
def concluir_parada_inteligente(rota_id: int, parada_id: int, db: Session = Depends(get_db), empresa: Empresa = Depends(empresa_logada)):
    rota = db.query(RotaInteligente).filter_by(id=rota_id, empresa_id=empresa.id).first()
    parada = db.query(RotaInteligenteParada).filter_by(id=parada_id, rota_id=rota_id).first()
    if not rota or not parada:
        raise HTTPException(status_code=404)

    if parada.tipo == "retirada" and parada.solicitacao_id:
        entrega = db.query(Agenda).filter_by(
            empresa_id=empresa.id,
            solicitacao_id=parada.solicitacao_id,
            tipo_evento="entrega",
        ).first()
        if entrega and entrega.status_operacional != "concluido":
            return RedirectResponse(
                f"/painel/inteligencia-logistica/rota/{rota.id}?erro={quote('Não é possível concluir a retirada antes da entrega.')}",
                status_code=303,
            )

    parada.status = "concluido"
    parada.chegada_real = agora_utc()

    # Sincroniza o encerramento com a Operação. Retirada prevista não usa o card
    # da entrega; ela é materializada como BUSCAR e então marcada como concluída.
    if parada.tipo == "entrega" and parada.agenda:
        parada.agenda.status_operacional = "concluido"
        criar_retirada_apos_entrega(db, parada.agenda)
    elif parada.tipo == "retirada" and parada.solicitacao:
        entrega = (
            db.query(Agenda)
            .filter_by(
                empresa_id=empresa.id,
                solicitacao_id=parada.solicitacao_id,
                tipo_evento="entrega",
            )
            .first()
        )
        retirada = _criar_ou_obter_retirada_prevista(db, parada.solicitacao, entrega)
        retirada.status_operacional = "concluido"
        parada.agenda_id = retirada.id
    elif parada.agenda:
        parada.agenda.status_operacional = "concluido"

    db.flush()
    _recalcular_rota_salva(db, rota)
    _sincronizar_rotas_inteligentes_ativas(db, empresa.id, ignorar_rota_id=rota.id)
    if all(p.status == "concluido" for p in rota.paradas):
        rota.status = "concluida"
    else:
        rota.status = "em_execucao"
    db.commit()
    return RedirectResponse(f"/painel/inteligencia-logistica/rota/{rota.id}?sucesso=Parada concluída, Operação atualizada e próximas recalculadas", status_code=303)
